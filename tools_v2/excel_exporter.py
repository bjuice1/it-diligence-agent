"""
Excel Exporter - Multi-Sheet Export for Granular Facts

This module provides Excel and CSV export functionality for all
data extracted during the multi-pass extraction process.

Exports:
- System Registry (Pass 1)
- Granular Facts by Domain (Pass 2)
- Validation Results (Pass 3)
- Summary Dashboard
"""

import csv
from pathlib import Path
from typing import Dict, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel export will be limited")


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Colors for status cells
STATUS_COLORS = {
    "pass": "C6EFCE",   # Light green
    "warn": "FFEB9C",   # Light yellow
    "fail": "FFC7CE",   # Light red
    "active": "C6EFCE",
    "deprecated": "FFC7CE",
    "planned": "DDEBF7",  # Light blue
    "unknown": "F2F2F2"   # Light gray
}

# Header style
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if OPENPYXL_AVAILABLE else None
HEADER_FONT = Font(color="FFFFFF", bold=True) if OPENPYXL_AVAILABLE else None

# Border style
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
) if OPENPYXL_AVAILABLE else None


# =============================================================================
# CSV EXPORTER
# =============================================================================

class CSVExporter:
    """
    Export data to CSV format.

    Works without any additional dependencies.
    """

    @staticmethod
    def export_granular_facts(
        granular_facts_store,
        output_path: Path,
        domain_filter: Optional[str] = None
    ) -> Path:
        """
        Export granular facts to CSV.

        Args:
            granular_facts_store: GranularFactsStore instance
            output_path: Output file path
            domain_filter: Optional domain to filter by

        Returns:
            Path to created file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        rows = granular_facts_store.to_rows(domain=domain_filter)

        if not rows:
            logger.warning("No facts to export")
            return output_path

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"Exported {len(rows)} facts to {output_path}")
        return output_path

    @staticmethod
    def export_systems(
        system_registry,
        output_path: Path
    ) -> Path:
        """
        Export system registry to CSV.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        rows = system_registry.to_rows()

        if not rows:
            logger.warning("No systems to export")
            return output_path

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"Exported {len(rows)} systems to {output_path}")
        return output_path

    @staticmethod
    def export_validation(
        validation_report,
        output_path: Path
    ) -> Path:
        """
        Export validation results to CSV.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        rows = validation_report.to_rows()

        if not rows:
            logger.warning("No validation results to export")
            return output_path

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"Exported {len(rows)} validation results to {output_path}")
        return output_path


# =============================================================================
# EXCEL EXPORTER
# =============================================================================

class ExcelExporter:
    """
    Export data to multi-sheet Excel workbook.

    Requires openpyxl to be installed.
    """

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

    def create_full_workbook(
        self,
        system_registry,
        granular_facts_store,
        validation_report,
        output_path: Path,
        include_charts: bool = True,
        work_items: list = None
    ) -> Path:
        """
        Create a complete Excel workbook with all data.

        Sheets:
        - Summary: Overview metrics and charts
        - Systems: System registry from Pass 1
        - Infrastructure: Granular facts for infrastructure domain
        - Applications: Granular facts for applications domain
        - Security: Granular facts for cybersecurity domain
        - Network: Granular facts for network domain
        - Identity: Granular facts for identity_access domain
        - Organization: Granular facts for organization domain
        - Validation: Validation results from Pass 3
        - Gaps: Items needing attention
        - Cost Buildup: Cost buildup transparency (if work items provided)

        Args:
            system_registry: SystemRegistry instance
            granular_facts_store: GranularFactsStore instance
            validation_report: ValidationReport instance
            output_path: Output file path
            include_charts: Whether to include charts
            work_items: Optional list of WorkItem objects for cost buildup sheet

        Returns:
            Path to created file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, system_registry, granular_facts_store, validation_report)
        self._create_systems_sheet(wb, system_registry)
        self._create_domain_sheets(wb, granular_facts_store)
        self._create_validation_sheet(wb, validation_report)
        self._create_gaps_sheet(wb, system_registry, validation_report)

        # Add cost buildup transparency sheet if work items provided
        if work_items:
            self._write_cost_buildup_sheet(wb, work_items)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Exported Excel workbook to {output_path}")

        return output_path

    def _create_summary_sheet(self, wb, system_registry, granular_facts_store, validation_report):
        """Create summary overview sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "IT Due Diligence - Extraction Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Generation timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color="666666")

        # Metrics section
        row = 4
        ws[f'A{row}'] = "EXTRACTION METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metrics = [
            ("Total Systems Discovered", system_registry.total_systems),
            ("Total Granular Facts", granular_facts_store.total_facts),
            ("Validation Checks Run", validation_report.total_checks),
            ("Validation Pass Rate", f"{validation_report.pass_rate:.1%}"),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Domain breakdown
        row += 1
        ws[f'A{row}'] = "FACTS BY DOMAIN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        domain_counts = granular_facts_store.count_by_domain()
        for domain, count in sorted(domain_counts.items()):
            ws[f'A{row}'] = domain.replace("_", " ").title()
            ws[f'B{row}'] = count
            row += 1

        # Fact type breakdown
        row += 1
        ws[f'A{row}'] = "FACTS BY TYPE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        type_counts = granular_facts_store.count_by_type()
        for fact_type, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            ws[f'A{row}'] = fact_type.title()
            ws[f'B{row}'] = count
            row += 1

        # Validation summary
        row += 1
        ws[f'A{row}'] = "VALIDATION SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        val_metrics = [
            ("Passed", validation_report.pass_count, "C6EFCE"),
            ("Warnings", validation_report.warn_count, "FFEB9C"),
            ("Failures", validation_report.fail_count, "FFC7CE"),
        ]

        for label, value, color in val_metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_systems_sheet(self, wb, system_registry):
        """Create systems registry sheet."""
        ws = wb.create_sheet("Systems")

        rows = system_registry.to_rows()
        if not rows:
            ws['A1'] = "No systems found"
            return

        # Header
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_num, row_data in enumerate(rows, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=row_data.get(header, ""))

                # Color status cells
                if header == "Status":
                    color = STATUS_COLORS.get(row_data.get(header, "").lower(), "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Auto-fit columns
        self._autofit_columns(ws)

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_domain_sheets(self, wb, granular_facts_store):
        """Create sheets for each domain's granular facts."""
        domains = [
            ("Infrastructure", "infrastructure"),
            ("Applications", "applications"),
            ("Security", "cybersecurity"),
            ("Network", "network"),
            ("Identity", "identity_access"),
            ("Organization", "organization"),
        ]

        for sheet_name, domain in domains:
            ws = wb.create_sheet(sheet_name)

            rows = granular_facts_store.to_rows(domain=domain)

            if not rows:
                ws['A1'] = f"No {domain} facts found"
                continue

            # Header
            headers = list(rows[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')

            # Data rows
            for row_num, row_data in enumerate(rows, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col, value=row_data.get(header, ""))

            # Auto-fit and freeze
            self._autofit_columns(ws)
            ws.freeze_panes = 'A2'

    def _create_validation_sheet(self, wb, validation_report):
        """Create validation results sheet."""
        ws = wb.create_sheet("Validation")

        rows = validation_report.to_rows()
        if not rows:
            ws['A1'] = "No validation results"
            return

        # Header
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_num, row_data in enumerate(rows, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=row_data.get(header, ""))

                # Color status cells
                if header == "Status":
                    status = row_data.get(header, "").lower()
                    color = STATUS_COLORS.get(status, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Auto-fit and freeze
        self._autofit_columns(ws)
        ws.freeze_panes = 'A2'

    def _create_gaps_sheet(self, wb, system_registry, validation_report):
        """Create gaps and issues sheet."""
        ws = wb.create_sheet("Gaps & Issues")

        row = 1

        # Systems without details
        ws[f'A{row}'] = "SYSTEMS WITHOUT GRANULAR FACTS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        systems_without_details = system_registry.get_systems_needing_details()
        if systems_without_details:
            headers = ["System ID", "Name", "Category", "Domain"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            row += 1

            for system in systems_without_details:
                ws.cell(row=row, column=1, value=system.system_id)
                ws.cell(row=row, column=2, value=system.name)
                ws.cell(row=row, column=3, value=system.category)
                ws.cell(row=row, column=4, value=system.domain)
                row += 1
        else:
            ws[f'A{row}'] = "All systems have granular facts"
            ws[f'A{row}'].font = Font(italic=True, color="008000")
            row += 1

        row += 2

        # Validation failures
        ws[f'A{row}'] = "VALIDATION FAILURES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        failures = validation_report.get_failures()
        if failures:
            headers = ["Check ID", "Type", "Severity", "Message", "Action"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True)
            row += 1

            for failure in failures:
                ws.cell(row=row, column=1, value=failure.check_id)
                ws.cell(row=row, column=2, value=failure.check_type)
                ws.cell(row=row, column=3, value=failure.severity)
                ws.cell(row=row, column=4, value=failure.message)
                ws.cell(row=row, column=5, value=failure.suggested_action)
                row += 1
        else:
            ws[f'A{row}'] = "No validation failures"
            ws[f'A{row}'].font = Font(italic=True, color="008000")
            row += 1

        row += 2

        # Validation warnings
        ws[f'A{row}'] = "VALIDATION WARNINGS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        warnings = validation_report.get_warnings()
        if warnings:
            headers = ["Check ID", "Type", "Severity", "Message"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(bold=True)
            row += 1

            for warning in warnings[:20]:  # Limit to 20
                ws.cell(row=row, column=1, value=warning.check_id)
                ws.cell(row=row, column=2, value=warning.check_type)
                ws.cell(row=row, column=3, value=warning.severity)
                ws.cell(row=row, column=4, value=warning.message)
                row += 1

            if len(warnings) > 20:
                ws[f'A{row}'] = f"... and {len(warnings) - 20} more warnings"
                ws[f'A{row}'].font = Font(italic=True)
        else:
            ws[f'A{row}'] = "No validation warnings"
            ws[f'A{row}'].font = Font(italic=True, color="008000")

        # Auto-fit columns
        self._autofit_columns(ws)

    def _write_cost_buildup_sheet(self, wb, work_items):
        """Write cost buildup transparency sheet."""
        ws = wb.create_sheet("Cost Buildup")

        headers = [
            "Work Item ID", "Title", "Domain", "Phase",
            "Cost Estimate (Range)", "Anchor Key", "Anchor Name",
            "Method", "Quantity", "Unit", "Unit Cost Low", "Unit Cost High",
            "Total Low", "Total High", "Confidence", "Source",
            "Assumptions", "Source Facts",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if HEADER_FILL:
                cell.fill = HEADER_FILL
            if HEADER_FONT:
                cell.font = HEADER_FONT
            if OPENPYXL_AVAILABLE:
                cell.alignment = Alignment(horizontal='center')

        row = 2
        for wi in work_items:
            ws.cell(row=row, column=1, value=getattr(wi, 'finding_id', ''))
            ws.cell(row=row, column=2, value=getattr(wi, 'title', ''))
            ws.cell(row=row, column=3, value=getattr(wi, 'domain', ''))
            ws.cell(row=row, column=4, value=getattr(wi, 'phase', ''))
            ws.cell(row=row, column=5, value=getattr(wi, 'cost_estimate', ''))

            cb = getattr(wi, 'cost_buildup', None)
            if cb is not None:
                ws.cell(row=row, column=6, value=cb.anchor_key)
                ws.cell(row=row, column=7, value=cb.anchor_name)
                ws.cell(row=row, column=8, value=cb.estimation_method)
                ws.cell(row=row, column=9, value=cb.quantity)
                ws.cell(row=row, column=10, value=cb.unit_label)
                ws.cell(row=row, column=11, value=cb.unit_cost_low)
                ws.cell(row=row, column=12, value=cb.unit_cost_high)
                ws.cell(row=row, column=13, value=cb.total_low)
                ws.cell(row=row, column=14, value=cb.total_high)
                ws.cell(row=row, column=15, value=cb.confidence)
                ws.cell(row=row, column=16, value=cb.estimation_source)
                ws.cell(row=row, column=17, value="; ".join(cb.assumptions) if cb.assumptions else "")
                ws.cell(row=row, column=18, value=", ".join(cb.source_facts) if cb.source_facts else "")
            else:
                ws.cell(row=row, column=6, value="N/A (range only)")

            row += 1

        # Auto-fit and freeze
        self._autofit_columns(ws)
        ws.freeze_panes = 'A2'

    def _autofit_columns(self, ws, max_width: int = 50):
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def export_to_excel(
    system_registry,
    granular_facts_store,
    validation_report,
    output_path: Path
) -> Path:
    """
    Export all data to Excel workbook.

    Falls back to CSV if openpyxl not available.
    """
    if OPENPYXL_AVAILABLE:
        exporter = ExcelExporter()
        return exporter.create_full_workbook(
            system_registry,
            granular_facts_store,
            validation_report,
            output_path
        )
    else:
        # Fall back to CSV
        logger.warning("openpyxl not available - exporting to CSV instead")
        output_dir = Path(output_path).parent
        CSVExporter.export_granular_facts(
            granular_facts_store,
            output_dir / "granular_facts.csv"
        )
        CSVExporter.export_systems(
            system_registry,
            output_dir / "systems.csv"
        )
        CSVExporter.export_validation(
            validation_report,
            output_dir / "validation.csv"
        )
        return output_dir


def export_to_csv(
    system_registry,
    granular_facts_store,
    validation_report,
    output_dir: Path
) -> Dict[str, Path]:
    """
    Export all data to separate CSV files.

    Returns:
        Dictionary of {name: path} for created files
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    files = {}

    files['systems'] = CSVExporter.export_systems(
        system_registry,
        output_dir / "systems.csv"
    )

    files['granular_facts'] = CSVExporter.export_granular_facts(
        granular_facts_store,
        output_dir / "granular_facts.csv"
    )

    files['validation'] = CSVExporter.export_validation(
        validation_report,
        output_dir / "validation.csv"
    )

    # Also export by domain
    for domain in granular_facts_store.domains:
        files[f'facts_{domain}'] = CSVExporter.export_granular_facts(
            granular_facts_store,
            output_dir / f"facts_{domain}.csv",
            domain_filter=domain
        )

    return files
