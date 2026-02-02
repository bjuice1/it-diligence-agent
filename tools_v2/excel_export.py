"""
Excel Export Module for V2 Architecture

Exports IT Due Diligence findings to client-ready Excel workbooks.
Works with FactStore and ReasoningStore from V2 pipeline.

Features:
- Multiple worksheets (Summary, Facts, Risks, Work Items, VDR Requests)
- Professional formatting with severity/priority coloring
- Evidence citations and fact references
- Sortable/filterable columns
"""

import logging
from typing import Dict, Any, Optional
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Import openpyxl components
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Install with: pip install openpyxl")


def export_to_excel(
    fact_store: Any,
    reasoning_store: Any,
    output_path: Path,
    target_name: str = "Target Company",
    include_vdr: bool = True,
    vdr_pack: Optional[Any] = None
) -> Dict[str, Any]:
    """
    Export V2 analysis results to formatted Excel workbook.

    Args:
        fact_store: FactStore with extracted facts
        reasoning_store: ReasoningStore with findings
        output_path: Path to save Excel file
        target_name: Target company name for headers
        include_vdr: Whether to include VDR requests sheet
        vdr_pack: Optional VDRPack for VDR requests

    Returns:
        Dict with output_path and counts
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    styles = _get_styles()

    # Create worksheets
    _create_summary_sheet(wb, fact_store, reasoning_store, target_name, styles)
    _create_facts_sheet(wb, fact_store, styles)
    _create_risks_sheet(wb, reasoning_store, styles)
    _create_work_items_sheet(wb, reasoning_store, styles)
    _create_cost_buildup_sheet(wb, reasoning_store, styles)  # NEW: Cost transparency
    _create_recommendations_sheet(wb, reasoning_store, styles)
    _create_strategic_sheet(wb, reasoning_store, styles)

    if include_vdr and vdr_pack:
        _create_vdr_sheet(wb, vdr_pack, styles)

    # Remove default empty sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    wb.save(str(output_path))
    logger.info(f"Exported to {output_path}")

    return {
        'output_path': str(output_path),
        'counts': {
            'facts': len(fact_store.facts),
            'gaps': len(fact_store.gaps),
            'risks': len(reasoning_store.risks),
            'work_items': len(reasoning_store.work_items),
            'recommendations': len(reasoning_store.recommendations)
        }
    }


def _get_styles():
    """Define Excel styles."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'header_fill': PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),

        'title_font': Font(bold=True, size=14),
        'section_font': Font(bold=True, size=12),

        'data_alignment': Alignment(vertical="top", wrap_text=True),
        'center_alignment': Alignment(horizontal="center", vertical="top"),
        'border': thin_border,

        # Severity colors
        'critical_fill': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        'critical_font': Font(color="FFFFFF", bold=True),
        'high_fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'medium_fill': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'low_fill': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),

        # Phase colors
        'day1_fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'day100_fill': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'post100_fill': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),

        # Entity colors
        'target_fill': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'buyer_fill': PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),
    }


def _apply_header_style(ws, row, styles, num_cols):
    """Apply header styling."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']


def _apply_severity_formatting(cell, value, styles):
    """Apply severity-based formatting."""
    if not value:
        return
    val = str(value).lower()
    if val == 'critical':
        cell.fill = styles['critical_fill']
        cell.font = styles['critical_font']
    elif val == 'high':
        cell.fill = styles['high_fill']
    elif val == 'medium':
        cell.fill = styles['medium_fill']
    elif val == 'low':
        cell.fill = styles['low_fill']


def _apply_phase_formatting(cell, value, styles):
    """Apply phase-based formatting."""
    if not value:
        return
    val = str(value).lower().replace('_', '')
    if 'day1' in val:
        cell.fill = styles['day1_fill']
    elif 'day100' in val:
        cell.fill = styles['day100_fill']
    elif 'post' in val:
        cell.fill = styles['post100_fill']


def _auto_size_columns(ws, min_width=10, max_width=50):
    """Auto-size columns."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def _create_summary_sheet(wb, fact_store, reasoning_store, target_name, styles):
    """Create Summary worksheet."""
    ws = wb.active
    ws.title = "Summary"

    ws['A1'] = f"IT Due Diligence: {target_name}"
    ws['A1'].font = styles['title_font']
    ws.merge_cells('A1:D1')

    ws['A3'] = "Generated:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Counts
    ws['A5'] = "Findings Summary"
    ws['A5'].font = styles['section_font']

    counts = [
        ("Facts Extracted", len(fact_store.facts)),
        ("Gaps Identified", len(fact_store.gaps)),
        ("Risks", len(reasoning_store.risks)),
        ("Strategic Considerations", len(reasoning_store.strategic_considerations)),
        ("Work Items", len(reasoning_store.work_items)),
        ("Recommendations", len(reasoning_store.recommendations)),
    ]

    for i, (label, count) in enumerate(counts, 6):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=count)

    # Risk severity breakdown
    ws['A14'] = "Risk Severity"
    ws['A14'].font = styles['section_font']

    severity_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    for risk in reasoning_store.risks:
        sev = str(risk.severity).lower()
        if sev in severity_counts:
            severity_counts[sev] += 1

    for i, (sev, count) in enumerate(severity_counts.items(), 15):
        cell = ws.cell(row=i, column=1, value=sev.title())
        ws.cell(row=i, column=2, value=count)
        _apply_severity_formatting(cell, sev, styles)

    # Work items by phase
    ws['A21'] = "Work Items by Phase"
    ws['A21'].font = styles['section_font']

    phase_counts = {'Day_1': 0, 'Day_100': 0, 'Post_100': 0}
    for wi in reasoning_store.work_items:
        if wi.phase in phase_counts:
            phase_counts[wi.phase] += 1

    for i, (phase, count) in enumerate(phase_counts.items(), 22):
        cell = ws.cell(row=i, column=1, value=phase.replace('_', ' '))
        ws.cell(row=i, column=2, value=count)
        _apply_phase_formatting(cell, phase, styles)

    _auto_size_columns(ws)


def _create_facts_sheet(wb, fact_store, styles):
    """Create Facts worksheet."""
    ws = wb.create_sheet("Facts")

    headers = ["Fact ID", "Domain", "Category", "Item", "Entity", "Status", "Details", "Evidence Quote", "Source"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, fact in enumerate(fact_store.facts, 2):
        ws.cell(row=row_num, column=1, value=fact.fact_id)
        ws.cell(row=row_num, column=2, value=fact.domain)
        ws.cell(row=row_num, column=3, value=fact.category)
        ws.cell(row=row_num, column=4, value=fact.item)

        entity_cell = ws.cell(row=row_num, column=5, value=fact.entity)
        if fact.entity == 'target':
            entity_cell.fill = styles['target_fill']
        elif fact.entity == 'buyer':
            entity_cell.fill = styles['buyer_fill']

        ws.cell(row=row_num, column=6, value=fact.status)

        # Format details as string
        if isinstance(fact.details, dict):
            details_str = "; ".join(f"{k}: {v}" for k, v in fact.details.items())
        else:
            details_str = str(fact.details)
        ws.cell(row=row_num, column=7, value=details_str)

        ws.cell(row=row_num, column=8, value=fact.evidence.get('exact_quote', ''))
        ws.cell(row=row_num, column=9, value=fact.source_document)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_risks_sheet(wb, reasoning_store, styles):
    """Create Risks worksheet."""
    ws = wb.create_sheet("Risks")

    headers = ["ID", "Domain", "Title", "Description", "Category", "Severity",
               "Integration Dependent", "Mitigation", "Based On Facts", "Confidence"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, risk in enumerate(reasoning_store.risks, 2):
        ws.cell(row=row_num, column=1, value=risk.finding_id)
        ws.cell(row=row_num, column=2, value=risk.domain)
        ws.cell(row=row_num, column=3, value=risk.title)
        ws.cell(row=row_num, column=4, value=risk.description)
        ws.cell(row=row_num, column=5, value=risk.category)

        sev_cell = ws.cell(row=row_num, column=6, value=risk.severity)
        _apply_severity_formatting(sev_cell, risk.severity, styles)

        ws.cell(row=row_num, column=7, value='Yes' if risk.integration_dependent else 'No')
        ws.cell(row=row_num, column=8, value=risk.mitigation)
        ws.cell(row=row_num, column=9, value=", ".join(risk.based_on_facts))
        ws.cell(row=row_num, column=10, value=risk.confidence)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_work_items_sheet(wb, reasoning_store, styles):
    """Create Work Items worksheet."""
    ws = wb.create_sheet("Work Items")

    headers = ["ID", "Domain", "Title", "Description", "Phase", "Priority",
               "Owner", "Cost Estimate", "Triggered By", "Based On Facts"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, wi in enumerate(reasoning_store.work_items, 2):
        ws.cell(row=row_num, column=1, value=wi.finding_id)
        ws.cell(row=row_num, column=2, value=wi.domain)
        ws.cell(row=row_num, column=3, value=wi.title)
        ws.cell(row=row_num, column=4, value=wi.description)

        phase_cell = ws.cell(row=row_num, column=5, value=wi.phase)
        _apply_phase_formatting(phase_cell, wi.phase, styles)

        pri_cell = ws.cell(row=row_num, column=6, value=wi.priority)
        _apply_severity_formatting(pri_cell, wi.priority, styles)

        ws.cell(row=row_num, column=7, value=wi.owner_type)
        ws.cell(row=row_num, column=8, value=wi.cost_estimate)
        ws.cell(row=row_num, column=9, value=", ".join(wi.triggered_by))
        ws.cell(row=row_num, column=10, value=", ".join(wi.based_on_facts))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_recommendations_sheet(wb, reasoning_store, styles):
    """Create Recommendations worksheet."""
    ws = wb.create_sheet("Recommendations")

    headers = ["ID", "Domain", "Title", "Description", "Action Type", "Urgency",
               "Rationale", "Based On Facts"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, rec in enumerate(reasoning_store.recommendations, 2):
        ws.cell(row=row_num, column=1, value=rec.finding_id)
        ws.cell(row=row_num, column=2, value=rec.domain)
        ws.cell(row=row_num, column=3, value=rec.title)
        ws.cell(row=row_num, column=4, value=rec.description)
        ws.cell(row=row_num, column=5, value=rec.action_type)

        urg_cell = ws.cell(row=row_num, column=6, value=rec.urgency)
        _apply_severity_formatting(urg_cell, rec.urgency, styles)

        ws.cell(row=row_num, column=7, value=rec.rationale)
        ws.cell(row=row_num, column=8, value=", ".join(rec.based_on_facts))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_strategic_sheet(wb, reasoning_store, styles):
    """Create Strategic Considerations worksheet."""
    ws = wb.create_sheet("Strategic")

    headers = ["ID", "Domain", "Title", "Description", "Lens", "Implication",
               "Based On Facts", "Confidence"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, sc in enumerate(reasoning_store.strategic_considerations, 2):
        ws.cell(row=row_num, column=1, value=sc.finding_id)
        ws.cell(row=row_num, column=2, value=sc.domain)
        ws.cell(row=row_num, column=3, value=sc.title)
        ws.cell(row=row_num, column=4, value=sc.description)
        ws.cell(row=row_num, column=5, value=sc.lens)
        ws.cell(row=row_num, column=6, value=sc.implication)
        ws.cell(row=row_num, column=7, value=", ".join(sc.based_on_facts))
        ws.cell(row=row_num, column=8, value=sc.confidence)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_vdr_sheet(wb, vdr_pack, styles):
    """Create VDR Requests worksheet."""
    ws = wb.create_sheet("VDR Requests")

    headers = ["ID", "Domain", "Priority", "Title", "Description", "Category",
               "Source", "Suggested Documents"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, styles, len(headers))

    row_num = 2
    for req in vdr_pack.requests:
        ws.cell(row=row_num, column=1, value=req.request_id)
        ws.cell(row=row_num, column=2, value=req.domain)

        pri_cell = ws.cell(row=row_num, column=3, value=req.priority)
        _apply_severity_formatting(pri_cell, req.priority, styles)

        ws.cell(row=row_num, column=4, value=req.title)
        ws.cell(row=row_num, column=5, value=req.description)
        ws.cell(row=row_num, column=6, value=req.category)
        ws.cell(row=row_num, column=7, value=f"{req.source}: {req.source_id}" if req.source_id else req.source)
        ws.cell(row=row_num, column=8, value=", ".join(req.suggested_documents))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

        row_num += 1

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_cost_buildup_sheet(wb, reasoning_store, styles):
    """
    Create Cost Build-Up worksheet showing detailed cost estimation logic.

    This worksheet provides full transparency into how costs were estimated,
    showing the anchor used, estimation method, quantities, and calculations.
    """
    ws = wb.create_sheet("Cost Build-Up")

    # Get work items with cost build-ups
    work_items_with_buildup = [
        wi for wi in reasoning_store.work_items
        if hasattr(wi, 'cost_buildup') and wi.cost_buildup
    ]

    # Also include work items without build-up (show their simple estimates)
    work_items_without_buildup = [
        wi for wi in reasoning_store.work_items
        if not (hasattr(wi, 'cost_buildup') and wi.cost_buildup)
    ]

    # Title row
    ws.cell(row=1, column=1, value="COST BUILD-UP WORKSHEET")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells('A1:O1')

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d')}")
    ws.cell(row=2, column=1).font = Font(italic=True)

    # Headers - includes Source and Confidence for transparency
    headers = [
        "Work Item ID", "Title", "Phase", "Domain",
        "Cost Anchor", "Source", "Confidence", "Method", "Qty", "Unit",
        "Unit Cost (Low)", "Unit Cost (High)",
        "Total (Low)", "Total (High)",
        "Assumptions"
    ]

    header_row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=header)
    _apply_header_style(ws, header_row, styles, len(headers))

    current_row = header_row + 1

    # Import cost range values
    from tools_v2.reasoning_tools import COST_RANGE_VALUES

    # Group by phase
    phases = ["Day_1", "Day_100", "Post_100"]
    phase_totals = {phase: {"low": 0, "high": 0, "count": 0} for phase in phases}

    for phase in phases:
        phase_items = [wi for wi in work_items_with_buildup if wi.phase == phase]
        phase_items_no_buildup = [wi for wi in work_items_without_buildup if wi.phase == phase]

        if not phase_items and not phase_items_no_buildup:
            continue

        # Phase header
        phase_display = phase.replace("_", " ").upper()
        ws.cell(row=current_row, column=1, value=f"--- {phase_display} ---")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(f'A{current_row}:O{current_row}')
        _apply_phase_formatting(ws.cell(row=current_row, column=1), phase, styles)
        current_row += 1

        # Items with detailed build-up
        for wi in phase_items:
            bu = wi.cost_buildup

            ws.cell(row=current_row, column=1, value=wi.finding_id)
            ws.cell(row=current_row, column=2, value=wi.title[:50] + "..." if len(wi.title) > 50 else wi.title)
            ws.cell(row=current_row, column=3, value=wi.phase)
            ws.cell(row=current_row, column=4, value=wi.domain)
            ws.cell(row=current_row, column=5, value=bu.anchor_name)
            # NEW: Source and Confidence columns
            source = getattr(bu, 'estimation_source', 'benchmark')
            confidence = getattr(bu, 'confidence', 'medium')
            ws.cell(row=current_row, column=6, value=source.replace('_', ' ').title())
            conf_cell = ws.cell(row=current_row, column=7, value=confidence.title())
            # Color-code confidence
            if confidence == 'high':
                conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confidence == 'low':
                conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            ws.cell(row=current_row, column=8, value=bu.estimation_method)
            ws.cell(row=current_row, column=9, value=bu.quantity)
            ws.cell(row=current_row, column=10, value=bu.unit_label)
            ws.cell(row=current_row, column=11, value=bu.unit_cost_low)
            ws.cell(row=current_row, column=12, value=bu.unit_cost_high)
            ws.cell(row=current_row, column=13, value=bu.total_low)
            ws.cell(row=current_row, column=14, value=bu.total_high)
            ws.cell(row=current_row, column=15, value="; ".join(bu.assumptions) if bu.assumptions else "")

            # Format currency columns
            for col in [11, 12, 13, 14]:
                ws.cell(row=current_row, column=col).number_format = '"$"#,##0'

            # Apply borders
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).border = styles['border']
                ws.cell(row=current_row, column=col).alignment = styles['data_alignment']

            # Track totals
            phase_totals[phase]["low"] += bu.total_low
            phase_totals[phase]["high"] += bu.total_high
            phase_totals[phase]["count"] += 1

            current_row += 1

        # Items without detailed build-up (use simple cost_estimate)
        for wi in phase_items_no_buildup:
            cost_range = COST_RANGE_VALUES.get(wi.cost_estimate, {"low": 0, "high": 0})

            ws.cell(row=current_row, column=1, value=wi.finding_id)
            ws.cell(row=current_row, column=2, value=wi.title[:50] + "..." if len(wi.title) > 50 else wi.title)
            ws.cell(row=current_row, column=3, value=wi.phase)
            ws.cell(row=current_row, column=4, value=wi.domain)
            ws.cell(row=current_row, column=5, value="(estimate)")
            ws.cell(row=current_row, column=6, value="Benchmark")  # Source
            ws.cell(row=current_row, column=7, value="Low")  # Confidence - low because no detailed build-up
            ws.cell(row=current_row, column=8, value="range")
            ws.cell(row=current_row, column=9, value=1)
            ws.cell(row=current_row, column=10, value="item")
            ws.cell(row=current_row, column=11, value=cost_range["low"])
            ws.cell(row=current_row, column=12, value=cost_range["high"])
            ws.cell(row=current_row, column=13, value=cost_range["low"])
            ws.cell(row=current_row, column=14, value=cost_range["high"])
            ws.cell(row=current_row, column=15, value=f"Based on {wi.cost_estimate} range")

            # Format currency columns
            for col in [11, 12, 13, 14]:
                ws.cell(row=current_row, column=col).number_format = '"$"#,##0'

            # Light gray background to indicate no detailed build-up
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).border = styles['border']
                ws.cell(row=current_row, column=col).alignment = styles['data_alignment']
                ws.cell(row=current_row, column=col).fill = PatternFill(
                    start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                )

            # Track totals
            phase_totals[phase]["low"] += cost_range["low"]
            phase_totals[phase]["high"] += cost_range["high"]
            phase_totals[phase]["count"] += 1

            current_row += 1

        # Phase subtotal row
        ws.cell(row=current_row, column=1, value=f"SUBTOTAL - {phase_display}")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=9, value=phase_totals[phase]["count"])
        ws.cell(row=current_row, column=10, value="items")
        ws.cell(row=current_row, column=13, value=phase_totals[phase]["low"])
        ws.cell(row=current_row, column=14, value=phase_totals[phase]["high"])

        for col in [13, 14]:
            ws.cell(row=current_row, column=col).number_format = '"$"#,##0'
            ws.cell(row=current_row, column=col).font = Font(bold=True)

        current_row += 2  # Extra space before next phase

    # Grand total section
    current_row += 1
    ws.cell(row=current_row, column=1, value="SUMMARY")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:O{current_row}')
    current_row += 1

    # Summary headers
    summary_headers = ["Phase", "Items", "Total (Low)", "Total (High)", "Midpoint"]
    for col, header in enumerate(summary_headers, 1):
        ws.cell(row=current_row, column=col, value=header)
        ws.cell(row=current_row, column=col).font = Font(bold=True)
        ws.cell(row=current_row, column=col).fill = styles['header_fill']
        ws.cell(row=current_row, column=col).font = styles['header_font']
    current_row += 1

    grand_low = 0
    grand_high = 0

    for phase in phases:
        if phase_totals[phase]["count"] > 0:
            phase_display = phase.replace("_", " ")
            ws.cell(row=current_row, column=1, value=phase_display)
            ws.cell(row=current_row, column=2, value=phase_totals[phase]["count"])
            ws.cell(row=current_row, column=3, value=phase_totals[phase]["low"])
            ws.cell(row=current_row, column=4, value=phase_totals[phase]["high"])
            midpoint = (phase_totals[phase]["low"] + phase_totals[phase]["high"]) / 2
            ws.cell(row=current_row, column=5, value=midpoint)

            for col in [3, 4, 5]:
                ws.cell(row=current_row, column=col).number_format = '"$"#,##0'
                ws.cell(row=current_row, column=col).border = styles['border']

            grand_low += phase_totals[phase]["low"]
            grand_high += phase_totals[phase]["high"]
            current_row += 1

    # Grand total row
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    total_count = sum(pt["count"] for pt in phase_totals.values())
    ws.cell(row=current_row, column=2, value=total_count)
    ws.cell(row=current_row, column=3, value=grand_low)
    ws.cell(row=current_row, column=4, value=grand_high)
    ws.cell(row=current_row, column=5, value=(grand_low + grand_high) / 2)

    for col in [3, 4, 5]:
        ws.cell(row=current_row, column=col).number_format = '"$"#,##0'
        ws.cell(row=current_row, column=col).font = Font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )

    # Add legend
    current_row += 3
    ws.cell(row=current_row, column=1, value="ESTIMATION METHOD LEGEND")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=1, value="• per_user: Cost scales with number of users (licenses, training, etc.)")
    current_row += 1
    ws.cell(row=current_row, column=1, value="• per_app: Cost per application (migration, integration)")
    current_row += 1
    ws.cell(row=current_row, column=1, value="• per_site: Cost per physical location (network, WAN)")
    current_row += 1
    ws.cell(row=current_row, column=1, value="• fixed_by_size: Fixed cost based on organization size tier")
    current_row += 1
    ws.cell(row=current_row, column=1, value="• range: Simple range estimate without detailed calculation")
    current_row += 2
    ws.cell(row=current_row, column=1, value="Gray rows indicate items without detailed cost build-up")

    # Freeze header
    ws.freeze_panes = f'A{header_row + 1}'
    _auto_size_columns(ws, max_width=40)
