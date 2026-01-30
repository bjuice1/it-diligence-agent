"""
Excel Parser

Parses Excel workbooks (.xlsx, .xls) into structured tables.
Handles multi-sheet workbooks, auto-detects header rows,
and trims empty rows/columns.
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass

from tools_v2.parsers.models import ParsedTable, ParseResult
from tools_v2.parsers.type_detector import detect_inventory_type

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel parsing
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel parsing disabled. Install with: pip install openpyxl")

# Try pandas as fallback
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def parse_excel_workbook(file_path: Path) -> ParseResult:
    """
    Parse an Excel workbook into structured tables.

    Extracts all sheets, auto-detects headers, and returns ParsedTable
    objects for each sheet with data.

    Args:
        file_path: Path to .xlsx or .xls file

    Returns:
        ParseResult with tables and metadata
    """
    file_path = Path(file_path)
    result = ParseResult(
        source_file=file_path.name,
        file_type="excel",
    )

    if not file_path.exists():
        result.errors.append(f"File not found: {file_path}")
        return result

    # Try openpyxl first (preferred)
    if OPENPYXL_AVAILABLE:
        return _parse_with_openpyxl(file_path, result)
    elif PANDAS_AVAILABLE:
        return _parse_with_pandas(file_path, result)
    else:
        result.errors.append(
            "No Excel parser available. Install openpyxl: pip install openpyxl"
        )
        return result


def _parse_with_openpyxl(file_path: Path, result: ParseResult) -> ParseResult:
    """Parse using openpyxl library."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Store metadata
        result.metadata = {
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
        }

        table_index = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract data from sheet
            table = _extract_sheet_data(sheet, sheet_name, file_path.name, table_index)

            if table and table.row_count > 0:
                # Detect inventory type
                inv_type, confidence = detect_inventory_type(table.headers)
                table.detected_type = inv_type
                table.detection_confidence = confidence

                result.tables.append(table)
                table_index += 1
            else:
                result.warnings.append(f"Sheet '{sheet_name}' is empty or has no data rows")

        workbook.close()

    except Exception as e:
        result.errors.append(f"Failed to parse Excel file: {str(e)}")
        logger.exception(f"Excel parse error: {file_path}")

    return result


def _extract_sheet_data(
    sheet,
    sheet_name: str,
    source_file: str,
    table_index: int
) -> Optional[ParsedTable]:
    """Extract data from an openpyxl worksheet."""
    # Find the data bounds (skip empty rows/columns)
    min_row, max_row, min_col, max_col = _find_data_bounds(sheet)

    if min_row is None:
        return None

    # First row with data is assumed to be headers
    headers = []
    for col in range(min_col, max_col + 1):
        cell_value = sheet.cell(row=min_row, column=col).value
        header = _normalize_header(cell_value, col)
        headers.append(header)

    # Check if this looks like headers (not data)
    if not _looks_like_headers(headers):
        # Try second row
        if max_row > min_row:
            alt_headers = []
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=min_row + 1, column=col).value
                alt_headers.append(_normalize_header(cell_value, col))

            if _looks_like_headers(alt_headers):
                # Use second row as headers, first row might be title
                headers = alt_headers
                min_row += 1

    # Extract data rows
    rows = []
    for row_num in range(min_row + 1, max_row + 1):
        row_data = {}
        has_data = False

        for col_idx, col_num in enumerate(range(min_col, max_col + 1)):
            cell = sheet.cell(row=row_num, column=col_num)
            value = _extract_cell_value(cell)

            header = headers[col_idx]
            row_data[header] = value

            if value is not None and str(value).strip():
                has_data = True

        # Skip completely empty rows
        if has_data:
            rows.append(row_data)

    if not rows:
        return None

    return ParsedTable(
        headers=headers,
        rows=rows,
        source_file=source_file,
        sheet_name=sheet_name,
        table_index=table_index,
    )


def _find_data_bounds(sheet) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """Find the actual data bounds in a worksheet, skipping empty rows/columns."""
    min_row = None
    max_row = None
    min_col = None
    max_col = None

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                if min_row is None or cell.row < min_row:
                    min_row = cell.row
                if max_row is None or cell.row > max_row:
                    max_row = cell.row
                if min_col is None or cell.column < min_col:
                    min_col = cell.column
                if max_col is None or cell.column > max_col:
                    max_col = cell.column

    return min_row, max_row, min_col, max_col


def _normalize_header(value: Any, col_num: int) -> str:
    """Normalize a header value to lowercase string."""
    if value is None:
        return f"column_{col_num}"

    header = str(value).strip().lower()

    # Remove special characters but keep spaces
    header = ''.join(c if c.isalnum() or c.isspace() or c == '_' else ' ' for c in header)

    # Collapse multiple spaces
    header = ' '.join(header.split())

    if not header:
        return f"column_{col_num}"

    return header


def _looks_like_headers(values: List[str]) -> bool:
    """Check if a row looks like headers vs data."""
    if not values:
        return False

    # Headers are usually text, not numbers
    text_count = 0
    for v in values:
        if v and not v.startswith("column_"):
            # Check if it's not just a number
            try:
                float(v.replace(",", "").replace("$", "").replace("%", ""))
            except (ValueError, AttributeError):
                text_count += 1

    # If most values are text, likely headers
    return text_count >= len(values) * 0.5


def _extract_cell_value(cell) -> Any:
    """Extract and clean a cell value."""
    value = cell.value

    if value is None:
        return None

    # Handle different types
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        return stripped

    # Numbers, dates, etc - return as-is
    return value


def _parse_with_pandas(file_path: Path, result: ParseResult) -> ParseResult:
    """Fallback parser using pandas."""
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)

        result.metadata = {
            "sheet_count": len(excel_file.sheet_names),
            "sheet_names": excel_file.sheet_names,
        }

        table_index = 0
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # Drop completely empty rows and columns
                df = df.dropna(how='all').dropna(axis=1, how='all')

                if df.empty:
                    result.warnings.append(f"Sheet '{sheet_name}' is empty")
                    continue

                # Convert to ParsedTable
                headers = [str(h).lower().strip() for h in df.columns.tolist()]
                rows = df.to_dict('records')

                # Normalize row keys to lowercase
                normalized_rows = []
                for row in rows:
                    normalized_rows.append({
                        str(k).lower().strip(): v for k, v in row.items()
                    })

                table = ParsedTable(
                    headers=headers,
                    rows=normalized_rows,
                    source_file=file_path.name,
                    sheet_name=sheet_name,
                    table_index=table_index,
                )

                # Detect type
                inv_type, confidence = detect_inventory_type(headers)
                table.detected_type = inv_type
                table.detection_confidence = confidence

                result.tables.append(table)
                table_index += 1

            except Exception as e:
                result.warnings.append(f"Error reading sheet '{sheet_name}': {str(e)}")

    except Exception as e:
        result.errors.append(f"Failed to parse Excel file: {str(e)}")
        logger.exception(f"Pandas Excel parse error: {file_path}")

    return result


def parse_csv(file_path: Path) -> ParseResult:
    """
    Parse a CSV file into a single table.

    Args:
        file_path: Path to .csv file

    Returns:
        ParseResult with single table
    """
    file_path = Path(file_path)
    result = ParseResult(
        source_file=file_path.name,
        file_type="csv",
    )

    if not file_path.exists():
        result.errors.append(f"File not found: {file_path}")
        return result

    try:
        import csv

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel  # Default to comma-separated

            reader = csv.DictReader(f, dialect=dialect)

            # Get headers
            headers = [h.lower().strip() for h in reader.fieldnames] if reader.fieldnames else []

            if not headers:
                result.errors.append("CSV file has no headers")
                return result

            # Read rows
            rows = []
            for row in reader:
                # Normalize keys to lowercase
                normalized = {k.lower().strip(): v for k, v in row.items()}
                rows.append(normalized)

            if not rows:
                result.warnings.append("CSV file has no data rows")
                return result

            table = ParsedTable(
                headers=headers,
                rows=rows,
                source_file=file_path.name,
                table_index=0,
            )

            # Detect type
            inv_type, confidence = detect_inventory_type(headers)
            table.detected_type = inv_type
            table.detection_confidence = confidence

            result.tables.append(table)

    except Exception as e:
        result.errors.append(f"Failed to parse CSV file: {str(e)}")
        logger.exception(f"CSV parse error: {file_path}")

    return result
