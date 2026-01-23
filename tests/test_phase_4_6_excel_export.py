"""
Tests for Phase 4-6: Excel Export Tool

Run with: pytest tests/test_phase_4_6_excel_export.py -v
"""
import pytest
import os
import tempfile
from pathlib import Path

from tools.analysis_tools import AnalysisStore

# Check if openpyxl is available
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@pytest.fixture
def populated_store():
    """Create an AnalysisStore with sample data."""
    store = AnalysisStore()

    # Add some applications
    store.execute_tool("record_application", {
        "application_name": "SAP ECC",
        "application_category": "ERP",
        "vendor": "SAP",
        "hosting_model": "On_Premise",
        "business_criticality": "Critical",
        "version": "ECC 6.0",
        "user_count": "500",
        "discovery_source": "App_Inventory_Document",
        "capability_areas_covered": ["finance_accounting"],
        "source_evidence": {
            "exact_quote": "SAP ECC 6.0 deployed on-premise with 500 users",
            "evidence_type": "direct_statement",
            "confidence_level": "high"
        }
    })

    store.execute_tool("record_application", {
        "application_name": "Salesforce",
        "application_category": "CRM",
        "vendor": "Salesforce",
        "hosting_model": "SaaS",
        "business_criticality": "High",
        "discovery_source": "App_Inventory_Document",
        "capability_areas_covered": ["sales_crm"],
        "source_evidence": {
            "exact_quote": "Salesforce CRM for sales team",
            "evidence_type": "direct_statement",
            "confidence_level": "high"
        },
        "fields_not_documented": ["version", "user_count"]
    })

    # Add capability coverage
    store.execute_tool("record_capability_coverage", {
        "capability_area": "finance_accounting",
        "coverage_status": "Fully_Documented",
        "business_relevance": "Critical",
        "relevance_reasoning": "Core financial operations",
        "applications_found": [{"app_name": "SAP ECC", "coverage_quality": "high"}],
        "confidence_level": "high"
    })

    store.execute_tool("record_capability_coverage", {
        "capability_area": "sales_crm",
        "coverage_status": "Partially_Documented",
        "business_relevance": "High",
        "relevance_reasoning": "Sales operations",
        "applications_found": [{"app_name": "Salesforce", "coverage_quality": "medium"}],
        "expected_but_missing": ["CPQ tool"],
        "gap_severity": "medium",
        "follow_up_questions": [
            {"question": "Is there a CPQ solution?", "priority": "medium", "target": "Target_IT_Leadership"}
        ],
        "confidence_level": "medium"
    })

    # Add buyer applications
    store.execute_tool("record_buyer_application", {
        "application_name": "Oracle ERP Cloud",
        "vendor": "Oracle",
        "application_category": "ERP",
        "capability_areas_covered": ["finance_accounting"],
        "information_source": "Buyer_Questionnaire"
    })

    store.execute_tool("record_buyer_application", {
        "application_name": "Salesforce",
        "vendor": "Salesforce",
        "application_category": "CRM",
        "capability_areas_covered": ["sales_crm"],
        "information_source": "Known_Environment"
    })

    # Add overlaps
    store.execute_tool("record_application_overlap", {
        "target_app_name": "SAP ECC",
        "target_app_category": "ERP",
        "buyer_app_name": "Oracle ERP Cloud",
        "overlap_type": "Same_Category_Different_Vendor",
        "considerations": "Both have ERP but different vendors - need to decide consolidation strategy",
        "follow_up_questions": [
            {"question": "What is the buyer's ERP consolidation preference?", "target": "Buyer_IT"}
        ]
    })

    store.execute_tool("record_application_overlap", {
        "target_app_name": "Salesforce",
        "target_app_category": "CRM",
        "buyer_app_name": "Salesforce",
        "overlap_type": "Same_Product",
        "considerations": "Both use Salesforce - consolidation opportunity",
        "follow_up_questions": [
            {"question": "Can instances be merged?", "target": "Buyer_Integration_Team"}
        ]
    })

    return store


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestExportApplicationsToExcel:
    """Test the export_applications_to_excel function."""

    def test_export_creates_file(self, populated_store):
        """Test that export creates an Excel file."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            result = export_applications_to_excel(populated_store, output_path)

            assert os.path.exists(output_path)
            assert result['output_path'] == output_path
            assert result['total_applications'] == 2

    def test_export_has_correct_sheets(self, populated_store):
        """Test that export has all expected worksheets."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            export_applications_to_excel(populated_store, output_path)

            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames

            assert "Applications" in sheet_names
            assert "Capability Coverage" in sheet_names
            assert "Buyer Applications" in sheet_names
            assert "Overlaps" in sheet_names
            assert "Completeness" in sheet_names

    def test_export_applications_sheet_content(self, populated_store):
        """Test that Applications sheet has correct content."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            export_applications_to_excel(populated_store, output_path)

            wb = openpyxl.load_workbook(output_path)
            ws = wb["Applications"]

            # Check header row
            assert ws['A1'].value == "ID"
            assert ws['B1'].value == "Application Name"

            # Check data rows (row 2 should be first app)
            assert ws['B2'].value in ["SAP ECC", "Salesforce"]

    def test_export_without_buyer_comparison(self, populated_store):
        """Test export without buyer comparison tabs."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            result = export_applications_to_excel(
                populated_store, output_path, include_buyer_comparison=False
            )

            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames

            assert "Applications" in sheet_names
            assert "Capability Coverage" in sheet_names
            assert "Buyer Applications" not in sheet_names
            assert "Overlaps" not in sheet_names

    def test_export_returns_counts(self, populated_store):
        """Test that export returns correct counts."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            result = export_applications_to_excel(populated_store, output_path)

            assert result['counts']['Applications'] == 2
            assert result['counts']['Capability Areas'] == 2
            assert result['counts']['Buyer Applications'] == 2
            assert result['counts']['Overlaps'] == 2

    def test_export_empty_store(self):
        """Test export with empty AnalysisStore."""
        from tools.excel_export import export_applications_to_excel

        empty_store = AnalysisStore()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_empty.xlsx")

            result = export_applications_to_excel(empty_store, output_path)

            assert os.path.exists(output_path)
            assert result['total_applications'] == 0


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestCapabilityCoverageSheet:
    """Test the Capability Coverage worksheet."""

    def test_capability_coverage_has_gap_info(self, populated_store):
        """Test that capability coverage sheet shows gaps."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            export_applications_to_excel(populated_store, output_path)

            wb = openpyxl.load_workbook(output_path)
            ws = wb["Capability Coverage"]

            # Find the sales_crm row (should have expected_but_missing)
            found_gap = False
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=2).value == "sales_crm":
                    expected_missing = ws.cell(row=row, column=7).value
                    if expected_missing and "CPQ" in expected_missing:
                        found_gap = True
                    break

            assert found_gap, "Expected to find CPQ in expected_but_missing for sales_crm"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestCompletenessSheet:
    """Test the Completeness worksheet."""

    def test_completeness_shows_metrics(self, populated_store):
        """Test that completeness sheet shows metrics."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            export_applications_to_excel(populated_store, output_path)

            wb = openpyxl.load_workbook(output_path)
            ws = wb["Completeness"]

            # Check title
            assert "Completeness" in ws['A1'].value

            # Check that total applications is shown
            found_total = False
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=1).value
                if cell and "Total Applications" in str(cell):
                    found_total = True
                    break

            assert found_total, "Expected to find Total Applications metric"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestOverlapsSheet:
    """Test the Overlaps worksheet."""

    def test_overlaps_shows_types(self, populated_store):
        """Test that overlaps sheet shows overlap types."""
        from tools.excel_export import export_applications_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_export.xlsx")

            export_applications_to_excel(populated_store, output_path)

            wb = openpyxl.load_workbook(output_path)
            ws = wb["Overlaps"]

            # Check for overlap types
            overlap_types_found = set()
            for row in range(2, ws.max_row + 1):
                ovl_type = ws.cell(row=row, column=5).value
                if ovl_type:
                    overlap_types_found.add(ovl_type)

            assert "Same_Product" in overlap_types_found
            assert "Same_Category_Different_Vendor" in overlap_types_found
