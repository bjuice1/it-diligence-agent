"""
Export Service for IT Due Diligence Agent

Generates human-readable exports (Excel, Markdown, CSV) from analysis data.
Includes AI-enhanced descriptions and assessments for each inventory item.

Phase C: Points 41-70 of cleanup plan.
"""

import os
import csv
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field, asdict

logger = logging.getLogger(__name__)


# =============================================================================
# ENHANCED INVENTORY MODELS
# =============================================================================

@dataclass
class EnhancedApplication:
    """Application with AI-enhanced fields."""
    # Core fields
    name: str
    vendor: str = ""
    category: str = ""
    version: str = ""
    users: str = ""
    criticality: str = ""
    entity: str = "target"

    # AI-enhanced fields
    description: str = ""           # Plain English description of what it does
    age_assessment: str = ""        # "Modern", "Aging", "Legacy - 15+ years"
    tech_debt_notes: str = ""       # Technical debt commentary
    migration_complexity: str = ""  # "Low", "Medium", "High", "Very High"
    integration_risk: str = ""      # Integration/migration risk assessment
    recommendation: str = ""        # "Keep", "Replace", "Consolidate", "Modernize"

    # Source tracking
    fact_ids: List[str] = field(default_factory=list)
    finding_ids: List[str] = field(default_factory=list)


@dataclass
class EnhancedInfrastructure:
    """Infrastructure asset with AI-enhanced fields."""
    # Core fields
    name: str
    asset_type: str = ""            # Server, Storage, Network, Cloud
    environment: str = ""           # Production, DR, Dev, Test
    location: str = ""              # On-prem DC, AWS, Azure, etc.
    specs: str = ""                 # CPU, RAM, storage
    os_version: str = ""
    entity: str = "target"

    # AI-enhanced fields
    description: str = ""           # What this infrastructure does
    age_assessment: str = ""        # Age/EOL status
    health_status: str = ""         # "Healthy", "At Risk", "Critical"
    cloud_readiness: str = ""       # "Ready", "Needs Work", "Not Suitable"
    tech_debt_notes: str = ""       # Technical debt commentary
    recommendation: str = ""        # "Migrate", "Modernize", "Retain", "Decommission"
    estimated_migration_effort: str = ""  # Effort to migrate

    # Source tracking
    fact_ids: List[str] = field(default_factory=list)
    finding_ids: List[str] = field(default_factory=list)


@dataclass
class EnhancedStaffMember:
    """Staff member with AI-enhanced fields."""
    # Core fields
    name: str
    role_title: str = ""
    department: str = ""
    employment_type: str = ""       # FTE, Contractor
    compensation: float = 0
    tenure_years: float = 0
    location: str = ""
    entity: str = "target"

    # AI-enhanced fields
    role_description: str = ""      # Plain English description of role
    market_assessment: str = ""     # Salary vs market rate
    key_person_risk: str = ""       # "High", "Medium", "Low", "None"
    key_person_reason: str = ""     # Why they're key
    retention_risk: str = ""        # Flight risk assessment
    skills_assessment: str = ""     # Skills gap or strength notes
    recommendation: str = ""        # "Retain", "Backfill", "Train", "Transition"

    # Source tracking
    fact_ids: List[str] = field(default_factory=list)


@dataclass
class EnhancedSecurityControl:
    """Security control with AI-enhanced fields."""
    # Core fields
    control_name: str
    domain: str = ""                # Endpoint, Network, Identity, etc.
    current_status: str = ""        # "Implemented", "Partial", "Gap"
    tool_vendor: str = ""
    entity: str = "target"

    # AI-enhanced fields
    description: str = ""           # What this control does
    gap_description: str = ""       # What's missing (if gap)
    risk_impact: str = ""           # Impact if exploited
    risk_level: str = ""            # "Critical", "High", "Medium", "Low"
    remediation_effort: str = ""    # Effort to fix
    remediation_priority: str = ""  # "Immediate", "Short-term", "Medium-term"
    recommendation: str = ""        # Specific remediation steps

    # Source tracking
    fact_ids: List[str] = field(default_factory=list)
    finding_ids: List[str] = field(default_factory=list)


@dataclass
class EnhancedNetworkComponent:
    """Network component with AI-enhanced fields."""
    # Core fields
    name: str
    component_type: str = ""        # WAN, LAN, Firewall, Switch, etc.
    vendor: str = ""
    model: str = ""
    location: str = ""
    entity: str = "target"

    # AI-enhanced fields
    description: str = ""           # What this component does
    age_assessment: str = ""        # Age/EOL status
    modernization_notes: str = ""   # Upgrade recommendations
    integration_complexity: str = "" # Complexity to integrate/migrate
    recommendation: str = ""        # "Keep", "Upgrade", "Replace"

    # Source tracking
    fact_ids: List[str] = field(default_factory=list)
    finding_ids: List[str] = field(default_factory=list)


# =============================================================================
# INVENTORY BUILDER - Converts facts/findings to enhanced inventories
# =============================================================================

class InventoryBuilder:
    """
    Builds enhanced inventories from facts and findings.

    Extracts relevant data and adds AI-generated insights based on
    the analysis findings.
    """

    def __init__(self, facts: List[Dict], findings: List[Dict]):
        """
        Initialize with facts and findings from an analysis run.

        Args:
            facts: List of fact dictionaries
            findings: List of finding dictionaries (risks, work items)
        """
        self.facts = facts
        self.findings = findings

        # Index facts by domain for quick lookup
        self.facts_by_domain = {}
        for fact in facts:
            domain = fact.get('domain', 'general')
            if domain not in self.facts_by_domain:
                self.facts_by_domain[domain] = []
            self.facts_by_domain[domain].append(fact)

        # Index findings by domain
        self.findings_by_domain = {}
        for finding in findings:
            domain = finding.get('domain', 'general')
            if domain not in self.findings_by_domain:
                self.findings_by_domain[domain] = []
            self.findings_by_domain[domain].append(finding)

    def build_applications_inventory(self) -> List[EnhancedApplication]:
        """Build enhanced applications inventory."""
        apps = []
        app_facts = self.facts_by_domain.get('applications', [])
        app_findings = self.findings_by_domain.get('applications', [])

        # Group facts by application name
        apps_map = {}
        for fact in app_facts:
            item = fact.get('item', '')
            if not item:
                continue

            details = fact.get('details', {})

            if item not in apps_map:
                apps_map[item] = EnhancedApplication(name=item)

            app = apps_map[item]
            app.fact_ids.append(fact.get('fact_id', ''))

            # Extract core fields from details
            if details.get('vendor'):
                app.vendor = details['vendor']
            if details.get('category'):
                app.category = details['category']
            if details.get('version'):
                app.version = str(details['version'])
            if details.get('users') or details.get('user_count'):
                app.users = str(details.get('users') or details.get('user_count', ''))
            if details.get('criticality') or details.get('business_criticality'):
                app.criticality = details.get('criticality') or details.get('business_criticality', '')
            if fact.get('entity'):
                app.entity = fact['entity']

            # Build description from evidence
            evidence = fact.get('evidence', {})
            if isinstance(evidence, dict) and evidence.get('exact_quote'):
                if not app.description:
                    app.description = evidence['exact_quote'][:200]

        # Enhance with findings
        for finding in app_findings:
            title = finding.get('title', '').lower()
            description = finding.get('description', '')
            severity = finding.get('severity', '')

            # Match findings to apps
            for app_name, app in apps_map.items():
                if app_name.lower() in title or app_name.lower() in description.lower():
                    app.finding_ids.append(finding.get('finding_id', ''))

                    # Extract insights
                    if 'legacy' in title or 'legacy' in description.lower():
                        app.age_assessment = "Legacy system - consider modernization"
                        app.tech_debt_notes = description[:150]
                    if 'end of life' in description.lower() or 'eol' in description.lower():
                        app.age_assessment = "End of Life - urgent replacement needed"
                    if 'integration' in title:
                        app.integration_risk = f"{severity} risk - {description[:100]}"

                    # Set recommendation based on severity
                    if severity == 'critical':
                        app.recommendation = "Replace - Critical issues identified"
                    elif severity == 'high':
                        app.recommendation = "Modernize - Significant concerns"
                    elif 'technical debt' in description.lower():
                        app.recommendation = "Modernize - Address technical debt"

        # Set defaults for apps without findings
        for app in apps_map.values():
            if not app.age_assessment:
                app.age_assessment = "Current - No significant age concerns"
            if not app.recommendation:
                app.recommendation = "Retain - No issues identified"
            if not app.migration_complexity:
                # Estimate based on criticality and integrations
                if app.criticality and 'critical' in app.criticality.lower():
                    app.migration_complexity = "High - Business critical"
                else:
                    app.migration_complexity = "Medium - Standard migration"

        return list(apps_map.values())

    def build_infrastructure_inventory(self) -> List[EnhancedInfrastructure]:
        """Build enhanced infrastructure inventory."""
        infra = []
        infra_facts = self.facts_by_domain.get('infrastructure', [])
        infra_findings = self.findings_by_domain.get('infrastructure', [])

        # Group facts by item/asset name
        infra_map = {}
        for fact in infra_facts:
            item = fact.get('item', '')
            if not item:
                continue

            details = fact.get('details', {})
            category = fact.get('category', '')

            if item not in infra_map:
                infra_map[item] = EnhancedInfrastructure(name=item)

            asset = infra_map[item]
            asset.fact_ids.append(fact.get('fact_id', ''))

            # Extract fields
            if category:
                asset.asset_type = category.replace('_', ' ').title()
            if details.get('environment'):
                asset.environment = details['environment']
            if details.get('location') or details.get('hosting'):
                asset.location = details.get('location') or details.get('hosting', '')
            if details.get('os') or details.get('os_version'):
                asset.os_version = details.get('os') or details.get('os_version', '')
            if fact.get('entity'):
                asset.entity = fact['entity']

            # Build specs string
            specs_parts = []
            if details.get('cpu') or details.get('vcpu'):
                specs_parts.append(f"CPU: {details.get('cpu') or details.get('vcpu')}")
            if details.get('ram') or details.get('memory'):
                specs_parts.append(f"RAM: {details.get('ram') or details.get('memory')}")
            if details.get('storage'):
                specs_parts.append(f"Storage: {details['storage']}")
            asset.specs = ", ".join(specs_parts) if specs_parts else ""

            # Build description
            evidence = fact.get('evidence', {})
            if isinstance(evidence, dict) and evidence.get('exact_quote'):
                if not asset.description:
                    asset.description = evidence['exact_quote'][:200]

        # Enhance with findings
        for finding in infra_findings:
            title = finding.get('title', '').lower()
            description = finding.get('description', '')
            severity = finding.get('severity', '')

            for asset_name, asset in infra_map.items():
                if asset_name.lower() in title or asset_name.lower() in description.lower():
                    asset.finding_ids.append(finding.get('finding_id', ''))

                    # Extract insights
                    if 'end of life' in description.lower() or 'eol' in description.lower():
                        asset.age_assessment = "End of Life - Requires immediate attention"
                        asset.health_status = "Critical"
                    if 'legacy' in description.lower() or 'aging' in description.lower():
                        asset.age_assessment = "Aging infrastructure"
                        asset.tech_debt_notes = description[:150]
                    if 'cloud' in description.lower():
                        asset.cloud_readiness = "Candidate for cloud migration"

                    if severity == 'critical':
                        asset.recommendation = "Decommission/Replace - Critical issues"
                        asset.health_status = "Critical"
                    elif severity == 'high':
                        asset.recommendation = "Modernize - Significant concerns"
                        asset.health_status = "At Risk"

        # Set defaults
        for asset in infra_map.values():
            if not asset.health_status:
                asset.health_status = "Healthy"
            if not asset.age_assessment:
                asset.age_assessment = "Current"
            if not asset.cloud_readiness:
                asset.cloud_readiness = "Assessment needed"
            if not asset.recommendation:
                asset.recommendation = "Retain - No issues identified"

        return list(infra_map.values())

    def build_organization_inventory(self) -> List[EnhancedStaffMember]:
        """Build enhanced organization/staffing inventory."""
        staff = []
        org_facts = self.facts_by_domain.get('organization', [])

        # Role description templates
        role_descriptions = {
            'director': 'Senior IT leadership responsible for strategic direction and team management',
            'manager': 'Manages team operations, projects, and direct reports',
            'engineer': 'Technical specialist responsible for design, implementation, and maintenance',
            'administrator': 'Manages and maintains systems, handles day-to-day operations',
            'analyst': 'Analyzes requirements, data, or security threats to provide insights',
            'developer': 'Designs and builds software applications and integrations',
            'architect': 'Designs technical solutions and establishes standards',
            'specialist': 'Subject matter expert in a specific technical domain',
            'support': 'Provides technical assistance and troubleshooting',
            'help desk': 'First-line support for end-user technical issues',
        }

        for fact in org_facts:
            if fact.get('category') not in ['leadership', 'central_it', 'app_teams', 'roles', 'key_individuals']:
                continue

            details = fact.get('details', {})
            item = fact.get('item', '')

            member = EnhancedStaffMember(
                name=details.get('name', item),
                role_title=details.get('role', item),
                department=details.get('department', details.get('team', '')),
                employment_type=details.get('employment_type', 'FTE'),
                entity=fact.get('entity', 'target')
            )
            member.fact_ids.append(fact.get('fact_id', ''))

            # Extract compensation
            if details.get('compensation') or details.get('salary'):
                try:
                    comp = details.get('compensation') or details.get('salary')
                    if isinstance(comp, str):
                        comp = float(comp.replace('$', '').replace(',', ''))
                    member.compensation = comp
                except:
                    pass

            # Extract tenure
            if details.get('tenure') or details.get('tenure_years'):
                try:
                    tenure = details.get('tenure') or details.get('tenure_years')
                    if isinstance(tenure, str):
                        tenure = float(tenure.split()[0])
                    member.tenure_years = tenure
                except:
                    pass

            # Generate role description
            role_lower = member.role_title.lower()
            for key, desc in role_descriptions.items():
                if key in role_lower:
                    member.role_description = desc
                    break
            if not member.role_description:
                member.role_description = f"IT professional responsible for {member.role_title.lower()} functions"

            # Key person assessment
            if details.get('is_key_person') or details.get('key_person'):
                member.key_person_risk = "High"
                member.key_person_reason = details.get('key_person_reason', 'Critical knowledge holder')
            elif 'director' in role_lower or 'architect' in role_lower:
                member.key_person_risk = "Medium"
                member.key_person_reason = "Senior technical leadership"
            else:
                member.key_person_risk = "Low"

            # Market assessment (simplified)
            if member.compensation > 0:
                if member.compensation > 180000:
                    member.market_assessment = "Above market - Senior/Leadership level"
                elif member.compensation > 120000:
                    member.market_assessment = "At market - Mid-senior level"
                elif member.compensation > 80000:
                    member.market_assessment = "At market - Mid level"
                else:
                    member.market_assessment = "Below market - Entry/Junior level"

            # Retention risk based on tenure
            if member.tenure_years > 0:
                if member.tenure_years < 1:
                    member.retention_risk = "Medium - New hire, still settling"
                elif member.tenure_years > 10:
                    member.retention_risk = "Low - Long tenure, institutional knowledge"
                else:
                    member.retention_risk = "Standard"

            # Default recommendation
            if member.key_person_risk == "High":
                member.recommendation = "Retain - Document knowledge, identify backup"
            else:
                member.recommendation = "Retain"

            staff.append(member)

        return staff

    def build_security_inventory(self) -> List[EnhancedSecurityControl]:
        """Build enhanced cybersecurity controls inventory."""
        controls = []
        security_facts = self.facts_by_domain.get('cybersecurity', [])
        security_findings = self.findings_by_domain.get('cybersecurity', [])

        controls_map = {}
        for fact in security_facts:
            item = fact.get('item', '')
            if not item:
                continue

            details = fact.get('details', {})
            status = fact.get('status', '')

            if item not in controls_map:
                controls_map[item] = EnhancedSecurityControl(control_name=item)

            control = controls_map[item]
            control.fact_ids.append(fact.get('fact_id', ''))

            control.domain = fact.get('category', '').replace('_', ' ').title()
            control.current_status = status.title() if status else "Unknown"

            if details.get('vendor') or details.get('tool'):
                control.tool_vendor = details.get('vendor') or details.get('tool', '')
            if fact.get('entity'):
                control.entity = fact['entity']

            # Build description from evidence
            evidence = fact.get('evidence', {})
            if isinstance(evidence, dict) and evidence.get('exact_quote'):
                control.description = evidence['exact_quote'][:200]

            # Set gap info if status indicates gap
            if status == 'gap':
                control.gap_description = f"No {item} implementation identified"
                control.risk_level = "High"
                control.remediation_priority = "Short-term"

        # Enhance with findings
        for finding in security_findings:
            title = finding.get('title', '')
            description = finding.get('description', '')
            severity = finding.get('severity', '')

            # Try to match to controls
            for control_name, control in controls_map.items():
                if control_name.lower() in title.lower() or control_name.lower() in description.lower():
                    control.finding_ids.append(finding.get('finding_id', ''))

                    if not control.gap_description and control.current_status != 'Documented':
                        control.gap_description = description[:150]

                    control.risk_impact = description[:100]

                    if severity == 'critical':
                        control.risk_level = "Critical"
                        control.remediation_priority = "Immediate"
                        control.recommendation = "Urgent remediation required"
                    elif severity == 'high':
                        control.risk_level = "High"
                        control.remediation_priority = "Short-term"
                        control.recommendation = "Address within 30 days"
                    elif severity == 'medium':
                        control.risk_level = "Medium"
                        control.remediation_priority = "Medium-term"
                        control.recommendation = "Plan remediation"

        # Set defaults
        for control in controls_map.values():
            if not control.risk_level:
                control.risk_level = "Low" if control.current_status == "Documented" else "Medium"
            if not control.recommendation:
                control.recommendation = "Maintain current controls"

        return list(controls_map.values())

    def build_network_inventory(self) -> List[EnhancedNetworkComponent]:
        """Build enhanced network inventory."""
        components = []
        network_facts = self.facts_by_domain.get('network', [])
        network_findings = self.findings_by_domain.get('network', [])

        network_map = {}
        for fact in network_facts:
            item = fact.get('item', '')
            if not item:
                continue

            details = fact.get('details', {})

            if item not in network_map:
                network_map[item] = EnhancedNetworkComponent(name=item)

            component = network_map[item]
            component.fact_ids.append(fact.get('fact_id', ''))

            component.component_type = fact.get('category', '').replace('_', ' ').upper()
            if details.get('vendor'):
                component.vendor = details['vendor']
            if details.get('model'):
                component.model = details['model']
            if details.get('location'):
                component.location = details['location']
            if fact.get('entity'):
                component.entity = fact['entity']

            # Build description
            evidence = fact.get('evidence', {})
            if isinstance(evidence, dict) and evidence.get('exact_quote'):
                component.description = evidence['exact_quote'][:200]

        # Enhance with findings
        for finding in network_findings:
            title = finding.get('title', '').lower()
            description = finding.get('description', '')
            severity = finding.get('severity', '')

            for comp_name, component in network_map.items():
                if comp_name.lower() in title or comp_name.lower() in description.lower():
                    component.finding_ids.append(finding.get('finding_id', ''))

                    if 'legacy' in description.lower() or 'aging' in description.lower():
                        component.age_assessment = "Aging - Consider upgrade"
                        component.modernization_notes = description[:150]

                    if severity in ['critical', 'high']:
                        component.recommendation = "Upgrade - Issues identified"

                    component.integration_complexity = f"{severity.title()} complexity" if severity else "Medium"

        # Set defaults
        for component in network_map.values():
            if not component.age_assessment:
                component.age_assessment = "Current"
            if not component.recommendation:
                component.recommendation = "Retain"
            if not component.integration_complexity:
                component.integration_complexity = "Standard"

        return list(network_map.values())


# =============================================================================
# EXPORTERS - Generate files in various formats
# =============================================================================

class MarkdownExporter:
    """Exports inventories to Markdown format."""

    @staticmethod
    def export_applications(apps: List[EnhancedApplication], output_path: Path) -> Path:
        """Export applications inventory to Markdown."""
        lines = [
            "# Applications Inventory",
            "",
            f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
            "",
            f"**Total Applications: {len(apps)}**",
            "",
            "---",
            "",
        ]

        # Summary table
        lines.extend([
            "## Summary",
            "",
            "| Application | Vendor | Category | Criticality | Recommendation |",
            "|------------|--------|----------|-------------|----------------|",
        ])

        for app in apps:
            lines.append(
                f"| {app.name} | {app.vendor} | {app.category} | {app.criticality} | {app.recommendation} |"
            )

        lines.extend(["", "---", "", "## Detailed Assessment", ""])

        # Detailed entries
        for app in apps:
            lines.extend([
                f"### {app.name}",
                "",
                f"**Vendor:** {app.vendor or 'Not specified'}",
                f"**Category:** {app.category or 'Not specified'}",
                f"**Version:** {app.version or 'Not specified'}",
                f"**Users:** {app.users or 'Not specified'}",
                f"**Business Criticality:** {app.criticality or 'Not assessed'}",
                "",
                "#### AI Assessment",
                "",
                f"- **Description:** {app.description or 'No description available'}",
                f"- **Age Assessment:** {app.age_assessment}",
                f"- **Migration Complexity:** {app.migration_complexity}",
                f"- **Integration Risk:** {app.integration_risk or 'Not assessed'}",
                f"- **Technical Debt:** {app.tech_debt_notes or 'None identified'}",
                "",
                f"**Recommendation:** {app.recommendation}",
                "",
                "---",
                "",
            ])

        output_path.write_text('\n'.join(lines))
        return output_path

    @staticmethod
    def export_infrastructure(infra: List[EnhancedInfrastructure], output_path: Path) -> Path:
        """Export infrastructure inventory to Markdown."""
        lines = [
            "# Infrastructure Inventory",
            "",
            f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
            "",
            f"**Total Assets: {len(infra)}**",
            "",
            "---",
            "",
            "## Summary",
            "",
            "| Asset | Type | Environment | Health | Recommendation |",
            "|-------|------|-------------|--------|----------------|",
        ]

        for asset in infra:
            lines.append(
                f"| {asset.name} | {asset.asset_type} | {asset.environment} | {asset.health_status} | {asset.recommendation} |"
            )

        lines.extend(["", "---", "", "## Detailed Assessment", ""])

        for asset in infra:
            lines.extend([
                f"### {asset.name}",
                "",
                f"**Type:** {asset.asset_type or 'Not specified'}",
                f"**Environment:** {asset.environment or 'Not specified'}",
                f"**Location:** {asset.location or 'Not specified'}",
                f"**OS/Version:** {asset.os_version or 'Not specified'}",
                f"**Specs:** {asset.specs or 'Not specified'}",
                "",
                "#### AI Assessment",
                "",
                f"- **Description:** {asset.description or 'No description available'}",
                f"- **Health Status:** {asset.health_status}",
                f"- **Age Assessment:** {asset.age_assessment}",
                f"- **Cloud Readiness:** {asset.cloud_readiness}",
                f"- **Technical Debt:** {asset.tech_debt_notes or 'None identified'}",
                f"- **Migration Effort:** {asset.estimated_migration_effort or 'Not estimated'}",
                "",
                f"**Recommendation:** {asset.recommendation}",
                "",
                "---",
                "",
            ])

        output_path.write_text('\n'.join(lines))
        return output_path

    @staticmethod
    def export_organization(staff: List[EnhancedStaffMember], output_path: Path) -> Path:
        """Export organization inventory to Markdown."""
        lines = [
            "# IT Organization & Staffing",
            "",
            f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
            "",
            f"**Total Staff: {len(staff)}**",
            "",
        ]

        # Calculate totals
        total_comp = sum(s.compensation for s in staff if s.compensation)
        fte_count = len([s for s in staff if s.employment_type.upper() == 'FTE'])
        contractor_count = len([s for s in staff if s.employment_type.upper() == 'CONTRACTOR'])
        key_persons = [s for s in staff if s.key_person_risk == 'High']

        lines.extend([
            "## Summary Statistics",
            "",
            f"- **Total Headcount:** {len(staff)}",
            f"- **FTEs:** {fte_count}",
            f"- **Contractors:** {contractor_count}",
            f"- **Total Compensation:** ${total_comp:,.0f}",
            f"- **Key Person Risks:** {len(key_persons)}",
            "",
            "---",
            "",
            "## Staff Roster",
            "",
            "| Name | Role | Department | Type | Compensation | Key Person Risk |",
            "|------|------|------------|------|--------------|-----------------|",
        ])

        for member in staff:
            comp_str = f"${member.compensation:,.0f}" if member.compensation else "N/A"
            lines.append(
                f"| {member.name} | {member.role_title} | {member.department} | {member.employment_type} | {comp_str} | {member.key_person_risk} |"
            )

        # Key person section
        if key_persons:
            lines.extend([
                "",
                "---",
                "",
                "## Key Person Risk Analysis",
                "",
            ])
            for member in key_persons:
                lines.extend([
                    f"### {member.name} - {member.role_title}",
                    "",
                    f"**Risk Level:** {member.key_person_risk}",
                    f"**Reason:** {member.key_person_reason}",
                    f"**Recommendation:** {member.recommendation}",
                    "",
                ])

        lines.extend(["", "---", "", "## Detailed Profiles", ""])

        for member in staff:
            comp_str = f"${member.compensation:,.0f}" if member.compensation else "Not disclosed"
            lines.extend([
                f"### {member.name}",
                "",
                f"**Role:** {member.role_title}",
                f"**Department:** {member.department or 'IT'}",
                f"**Employment Type:** {member.employment_type}",
                f"**Compensation:** {comp_str}",
                f"**Tenure:** {member.tenure_years} years" if member.tenure_years else "",
                "",
                "#### Assessment",
                "",
                f"- **Role Description:** {member.role_description}",
                f"- **Market Assessment:** {member.market_assessment or 'Not assessed'}",
                f"- **Key Person Risk:** {member.key_person_risk}",
                f"- **Retention Risk:** {member.retention_risk or 'Not assessed'}",
                "",
                f"**Recommendation:** {member.recommendation}",
                "",
                "---",
                "",
            ])

        output_path.write_text('\n'.join(lines))
        return output_path

    @staticmethod
    def export_security(controls: List[EnhancedSecurityControl], output_path: Path) -> Path:
        """Export cybersecurity controls to Markdown."""
        lines = [
            "# Cybersecurity Controls Assessment",
            "",
            f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
            "",
            f"**Total Controls Assessed: {len(controls)}**",
            "",
        ]

        # Count by status
        implemented = len([c for c in controls if c.current_status.lower() == 'documented'])
        gaps = len([c for c in controls if c.current_status.lower() == 'gap'])
        partial = len([c for c in controls if c.current_status.lower() == 'partial'])

        lines.extend([
            "## Summary",
            "",
            f"- **Implemented:** {implemented}",
            f"- **Partial:** {partial}",
            f"- **Gaps:** {gaps}",
            "",
            "---",
            "",
            "## Controls Matrix",
            "",
            "| Control | Domain | Status | Risk Level | Priority |",
            "|---------|--------|--------|------------|----------|",
        ])

        for control in controls:
            lines.append(
                f"| {control.control_name} | {control.domain} | {control.current_status} | {control.risk_level} | {control.remediation_priority or 'N/A'} |"
            )

        # Gaps section
        gap_controls = [c for c in controls if c.current_status.lower() in ['gap', 'partial']]
        if gap_controls:
            lines.extend([
                "",
                "---",
                "",
                "## Security Gaps & Remediation",
                "",
            ])
            for control in gap_controls:
                lines.extend([
                    f"### {control.control_name}",
                    "",
                    f"**Status:** {control.current_status}",
                    f"**Risk Level:** {control.risk_level}",
                    f"**Gap Description:** {control.gap_description or 'Not documented'}",
                    f"**Risk Impact:** {control.risk_impact or 'Not assessed'}",
                    f"**Remediation Priority:** {control.remediation_priority or 'Not set'}",
                    "",
                    f"**Recommendation:** {control.recommendation}",
                    "",
                    "---",
                    "",
                ])

        output_path.write_text('\n'.join(lines))
        return output_path

    @staticmethod
    def export_network(components: List[EnhancedNetworkComponent], output_path: Path) -> Path:
        """Export network inventory to Markdown."""
        lines = [
            "# Network Infrastructure Inventory",
            "",
            f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
            "",
            f"**Total Components: {len(components)}**",
            "",
            "---",
            "",
            "## Summary",
            "",
            "| Component | Type | Vendor | Age Status | Recommendation |",
            "|-----------|------|--------|------------|----------------|",
        ]

        for comp in components:
            lines.append(
                f"| {comp.name} | {comp.component_type} | {comp.vendor} | {comp.age_assessment} | {comp.recommendation} |"
            )

        lines.extend(["", "---", "", "## Detailed Assessment", ""])

        for comp in components:
            lines.extend([
                f"### {comp.name}",
                "",
                f"**Type:** {comp.component_type or 'Not specified'}",
                f"**Vendor:** {comp.vendor or 'Not specified'}",
                f"**Model:** {comp.model or 'Not specified'}",
                f"**Location:** {comp.location or 'Not specified'}",
                "",
                "#### Assessment",
                "",
                f"- **Description:** {comp.description or 'No description available'}",
                f"- **Age Status:** {comp.age_assessment}",
                f"- **Modernization Notes:** {comp.modernization_notes or 'None'}",
                f"- **Integration Complexity:** {comp.integration_complexity}",
                "",
                f"**Recommendation:** {comp.recommendation}",
                "",
                "---",
                "",
            ])

        output_path.write_text('\n'.join(lines))
        return output_path


class ExcelExporter:
    """Exports inventories to Excel format with formatting."""

    @staticmethod
    def export_applications(apps: List[EnhancedApplication], output_path: Path) -> Path:
        """Export applications inventory to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        # Headers
        headers = [
            "Application", "Vendor", "Category", "Version", "Users",
            "Criticality", "Description", "Age Assessment", "Migration Complexity",
            "Integration Risk", "Tech Debt Notes", "Recommendation", "Entity"
        ]

        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Write data
        for row, app in enumerate(apps, 2):
            data = [
                app.name, app.vendor, app.category, app.version, app.users,
                app.criticality, app.description, app.age_assessment, app.migration_complexity,
                app.integration_risk, app.tech_debt_notes, app.recommendation, app.entity
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        return output_path

    @staticmethod
    def export_infrastructure(infra: List[EnhancedInfrastructure], output_path: Path) -> Path:
        """Export infrastructure inventory to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Infrastructure"

        headers = [
            "Asset Name", "Type", "Environment", "Location", "OS/Version", "Specs",
            "Description", "Health Status", "Age Assessment", "Cloud Readiness",
            "Tech Debt", "Migration Effort", "Recommendation", "Entity"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        for row, asset in enumerate(infra, 2):
            data = [
                asset.name, asset.asset_type, asset.environment, asset.location,
                asset.os_version, asset.specs, asset.description, asset.health_status,
                asset.age_assessment, asset.cloud_readiness, asset.tech_debt_notes,
                asset.estimated_migration_effort, asset.recommendation, asset.entity
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        return output_path

    @staticmethod
    def export_organization(staff: List[EnhancedStaffMember], output_path: Path) -> Path:
        """Export organization inventory to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Census"

        headers = [
            "Name", "Role Title", "Department", "Employment Type", "Compensation",
            "Tenure (Years)", "Location", "Role Description", "Market Assessment",
            "Key Person Risk", "Key Person Reason", "Retention Risk", "Recommendation", "Entity"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        for row, member in enumerate(staff, 2):
            data = [
                member.name, member.role_title, member.department, member.employment_type,
                member.compensation if member.compensation else None,
                member.tenure_years if member.tenure_years else None,
                member.location, member.role_description, member.market_assessment,
                member.key_person_risk, member.key_person_reason, member.retention_risk,
                member.recommendation, member.entity
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # Format compensation as currency
                if col == 5 and value:
                    cell.number_format = '$#,##0'

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        return output_path

    @staticmethod
    def export_security(controls: List[EnhancedSecurityControl], output_path: Path) -> Path:
        """Export security controls to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Security Controls"

        headers = [
            "Control", "Domain", "Status", "Tool/Vendor", "Description",
            "Gap Description", "Risk Level", "Risk Impact", "Remediation Priority",
            "Recommendation", "Entity"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        for row, control in enumerate(controls, 2):
            data = [
                control.control_name, control.domain, control.current_status,
                control.tool_vendor, control.description, control.gap_description,
                control.risk_level, control.risk_impact, control.remediation_priority,
                control.recommendation, control.entity
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        return output_path

    @staticmethod
    def export_network(components: List[EnhancedNetworkComponent], output_path: Path) -> Path:
        """Export network inventory to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Network"

        headers = [
            "Component", "Type", "Vendor", "Model", "Location",
            "Description", "Age Assessment", "Modernization Notes",
            "Integration Complexity", "Recommendation", "Entity"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        for row, comp in enumerate(components, 2):
            data = [
                comp.name, comp.component_type, comp.vendor, comp.model,
                comp.location, comp.description, comp.age_assessment,
                comp.modernization_notes, comp.integration_complexity,
                comp.recommendation, comp.entity
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        return output_path


class CSVExporter:
    """Exports inventories to CSV format."""

    @staticmethod
    def export_applications(apps: List[EnhancedApplication], output_path: Path) -> Path:
        """Export applications to CSV."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Application", "Vendor", "Category", "Version", "Users",
                "Criticality", "Description", "Age Assessment", "Migration Complexity",
                "Integration Risk", "Tech Debt Notes", "Recommendation", "Entity"
            ])
            for app in apps:
                writer.writerow([
                    app.name, app.vendor, app.category, app.version, app.users,
                    app.criticality, app.description, app.age_assessment, app.migration_complexity,
                    app.integration_risk, app.tech_debt_notes, app.recommendation, app.entity
                ])
        return output_path

    @staticmethod
    def export_infrastructure(infra: List[EnhancedInfrastructure], output_path: Path) -> Path:
        """Export infrastructure to CSV."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Asset Name", "Type", "Environment", "Location", "OS/Version", "Specs",
                "Description", "Health Status", "Age Assessment", "Cloud Readiness",
                "Tech Debt", "Migration Effort", "Recommendation", "Entity"
            ])
            for asset in infra:
                writer.writerow([
                    asset.name, asset.asset_type, asset.environment, asset.location,
                    asset.os_version, asset.specs, asset.description, asset.health_status,
                    asset.age_assessment, asset.cloud_readiness, asset.tech_debt_notes,
                    asset.estimated_migration_effort, asset.recommendation, asset.entity
                ])
        return output_path

    @staticmethod
    def export_organization(staff: List[EnhancedStaffMember], output_path: Path) -> Path:
        """Export organization to CSV."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Name", "Role Title", "Department", "Employment Type", "Compensation",
                "Tenure (Years)", "Location", "Role Description", "Market Assessment",
                "Key Person Risk", "Key Person Reason", "Retention Risk", "Recommendation", "Entity"
            ])
            for member in staff:
                writer.writerow([
                    member.name, member.role_title, member.department, member.employment_type,
                    member.compensation, member.tenure_years, member.location,
                    member.role_description, member.market_assessment, member.key_person_risk,
                    member.key_person_reason, member.retention_risk, member.recommendation, member.entity
                ])
        return output_path

    @staticmethod
    def export_security(controls: List[EnhancedSecurityControl], output_path: Path) -> Path:
        """Export security controls to CSV."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Control", "Domain", "Status", "Tool/Vendor", "Description",
                "Gap Description", "Risk Level", "Risk Impact", "Remediation Priority",
                "Recommendation", "Entity"
            ])
            for control in controls:
                writer.writerow([
                    control.control_name, control.domain, control.current_status,
                    control.tool_vendor, control.description, control.gap_description,
                    control.risk_level, control.risk_impact, control.remediation_priority,
                    control.recommendation, control.entity
                ])
        return output_path

    @staticmethod
    def export_network(components: List[EnhancedNetworkComponent], output_path: Path) -> Path:
        """Export network to CSV."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Component", "Type", "Vendor", "Model", "Location",
                "Description", "Age Assessment", "Modernization Notes",
                "Integration Complexity", "Recommendation", "Entity"
            ])
            for comp in components:
                writer.writerow([
                    comp.name, comp.component_type, comp.vendor, comp.model,
                    comp.location, comp.description, comp.age_assessment,
                    comp.modernization_notes, comp.integration_complexity,
                    comp.recommendation, comp.entity
                ])
        return output_path


# =============================================================================
# MAIN EXPORT SERVICE
# =============================================================================

class ExportService:
    """
    Main export service that orchestrates all exports.

    Usage:
        service = ExportService(facts, findings)
        service.export_all(output_dir)
    """

    def __init__(self, facts: List[Dict], findings: List[Dict]):
        """Initialize with facts and findings."""
        self.facts = facts
        self.findings = findings
        self.builder = InventoryBuilder(facts, findings)

    def export_all(self, output_dir: Path) -> Dict[str, List[Path]]:
        """
        Export all inventories in all formats.

        Args:
            output_dir: Directory to write exports

        Returns:
            Dict mapping domain to list of exported file paths
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        results = {}

        # Applications
        apps = self.builder.build_applications_inventory()
        if apps:
            results['applications'] = [
                MarkdownExporter.export_applications(apps, output_dir / "applications.md"),
                ExcelExporter.export_applications(apps, output_dir / "applications.xlsx"),
                CSVExporter.export_applications(apps, output_dir / "applications.csv"),
            ]
            logger.info(f"Exported {len(apps)} applications")

        # Infrastructure
        infra = self.builder.build_infrastructure_inventory()
        if infra:
            results['infrastructure'] = [
                MarkdownExporter.export_infrastructure(infra, output_dir / "infrastructure.md"),
                ExcelExporter.export_infrastructure(infra, output_dir / "infrastructure.xlsx"),
                CSVExporter.export_infrastructure(infra, output_dir / "infrastructure.csv"),
            ]
            logger.info(f"Exported {len(infra)} infrastructure assets")

        # Organization
        staff = self.builder.build_organization_inventory()
        if staff:
            results['organization'] = [
                MarkdownExporter.export_organization(staff, output_dir / "organization.md"),
                ExcelExporter.export_organization(staff, output_dir / "organization.xlsx"),
                CSVExporter.export_organization(staff, output_dir / "organization.csv"),
            ]
            logger.info(f"Exported {len(staff)} staff members")

        # Cybersecurity
        controls = self.builder.build_security_inventory()
        if controls:
            results['cybersecurity'] = [
                MarkdownExporter.export_security(controls, output_dir / "cybersecurity.md"),
                ExcelExporter.export_security(controls, output_dir / "cybersecurity.xlsx"),
                CSVExporter.export_security(controls, output_dir / "cybersecurity.csv"),
            ]
            logger.info(f"Exported {len(controls)} security controls")

        # Network
        network = self.builder.build_network_inventory()
        if network:
            results['network'] = [
                MarkdownExporter.export_network(network, output_dir / "network.md"),
                ExcelExporter.export_network(network, output_dir / "network.xlsx"),
                CSVExporter.export_network(network, output_dir / "network.csv"),
            ]
            logger.info(f"Exported {len(network)} network components")

        return results

    def export_domain(self, domain: str, output_dir: Path, formats: List[str] = None) -> List[Path]:
        """
        Export a single domain in specified formats.

        Args:
            domain: Domain to export (applications, infrastructure, etc.)
            output_dir: Directory to write exports
            formats: List of formats (md, xlsx, csv). Default: all

        Returns:
            List of exported file paths
        """
        formats = formats or ['md', 'xlsx', 'csv']
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        results = []

        if domain == 'applications':
            apps = self.builder.build_applications_inventory()
            if 'md' in formats:
                results.append(MarkdownExporter.export_applications(apps, output_dir / "applications.md"))
            if 'xlsx' in formats:
                results.append(ExcelExporter.export_applications(apps, output_dir / "applications.xlsx"))
            if 'csv' in formats:
                results.append(CSVExporter.export_applications(apps, output_dir / "applications.csv"))

        elif domain == 'infrastructure':
            infra = self.builder.build_infrastructure_inventory()
            if 'md' in formats:
                results.append(MarkdownExporter.export_infrastructure(infra, output_dir / "infrastructure.md"))
            if 'xlsx' in formats:
                results.append(ExcelExporter.export_infrastructure(infra, output_dir / "infrastructure.xlsx"))
            if 'csv' in formats:
                results.append(CSVExporter.export_infrastructure(infra, output_dir / "infrastructure.csv"))

        elif domain == 'organization':
            staff = self.builder.build_organization_inventory()
            if 'md' in formats:
                results.append(MarkdownExporter.export_organization(staff, output_dir / "organization.md"))
            if 'xlsx' in formats:
                results.append(ExcelExporter.export_organization(staff, output_dir / "organization.xlsx"))
            if 'csv' in formats:
                results.append(CSVExporter.export_organization(staff, output_dir / "organization.csv"))

        elif domain == 'cybersecurity':
            controls = self.builder.build_security_inventory()
            if 'md' in formats:
                results.append(MarkdownExporter.export_security(controls, output_dir / "cybersecurity.md"))
            if 'xlsx' in formats:
                results.append(ExcelExporter.export_security(controls, output_dir / "cybersecurity.xlsx"))
            if 'csv' in formats:
                results.append(CSVExporter.export_security(controls, output_dir / "cybersecurity.csv"))

        elif domain == 'network':
            network = self.builder.build_network_inventory()
            if 'md' in formats:
                results.append(MarkdownExporter.export_network(network, output_dir / "network.md"))
            if 'xlsx' in formats:
                results.append(ExcelExporter.export_network(network, output_dir / "network.xlsx"))
            if 'csv' in formats:
                results.append(CSVExporter.export_network(network, output_dir / "network.csv"))

        return results


# =============================================================================
# CONVENIENCE FUNCTION
# =============================================================================

def export_analysis_run(facts_path: Path, findings_path: Path, output_dir: Path) -> Dict[str, List[Path]]:
    """
    Export all inventories from an analysis run.

    Args:
        facts_path: Path to facts JSON file
        findings_path: Path to findings JSON file
        output_dir: Directory to write exports

    Returns:
        Dict mapping domain to list of exported files
    """
    # Load facts
    with open(facts_path, 'r') as f:
        facts_data = json.load(f)

    # Handle both flat list and nested structure
    if isinstance(facts_data, list):
        facts = facts_data
    elif isinstance(facts_data, dict) and 'facts' in facts_data:
        facts = facts_data['facts']
    else:
        facts = []

    # Load findings
    with open(findings_path, 'r') as f:
        findings_data = json.load(f)

    # Handle various findings structures
    findings = []
    if isinstance(findings_data, list):
        findings = findings_data
    elif isinstance(findings_data, dict):
        # Check for direct lists at root level
        if 'risks' in findings_data and isinstance(findings_data['risks'], list):
            findings.extend(findings_data['risks'])
        if 'work_items' in findings_data and isinstance(findings_data['work_items'], list):
            findings.extend(findings_data['work_items'])
        if 'recommendations' in findings_data and isinstance(findings_data['recommendations'], list):
            findings.extend(findings_data['recommendations'])

        # Also check for nested by domain structure
        if not findings:
            for domain_data in findings_data.values():
                if isinstance(domain_data, dict):
                    if 'risks' in domain_data and isinstance(domain_data['risks'], list):
                        findings.extend(domain_data['risks'])
                    if 'work_items' in domain_data and isinstance(domain_data['work_items'], list):
                        findings.extend(domain_data['work_items'])
                    if 'recommendations' in domain_data and isinstance(domain_data['recommendations'], list):
                        findings.extend(domain_data['recommendations'])

    # Export
    service = ExportService(facts, findings)
    return service.export_all(output_dir)
