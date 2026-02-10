"""
Excel Export for IT DD Synthesis.

Generates formatted Excel workbook with multiple sheets.
"""
from typing import Optional
from .models import SynthesisResult
from .cost_engine import format_currency

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_excel(result: SynthesisResult, output_path: str) -> dict:
    """
    Export synthesis result to Excel workbook.

    Args:
        result: SynthesisResult from synthesizer
        output_path: Path for output .xlsx file

    Returns:
        Dict with status and counts
    """
    if not OPENPYXL_AVAILABLE:
        return {"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    critical_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
    high_fill = PatternFill(start_color="FEEBC8", end_color="FEEBC8", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Remove default sheet
    wb.remove(wb.active)

    # =========================================================================
    # Sheet 1: Executive Summary
    # =========================================================================
    ws = wb.create_sheet("Executive Summary")
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    rows = [
        ("IT Due Diligence Summary", ""),
        ("Target Company", result.target_company),
        ("Generated", result.generated_at[:10]),
        ("", ""),
        ("Overall Assessment", result.overall_assessment),
        ("", ""),
        ("Key Metrics", ""),
        ("Applications Identified", len(result.applications)),
        ("Total Risks", len(result.risks)),
        ("Critical Risks", len([r for r in result.risks if r.severity == "Critical"])),
        ("High Risks", len([r for r in result.risks if r.severity == "High"])),
        ("", ""),
        ("Cost Estimate (Low)", format_currency(result.total_one_time_low)),
        ("Cost Estimate (High)", format_currency(result.total_one_time_high)),
    ]

    for i, (label, value) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if i == 1:
            ws.cell(row=i, column=1).font = Font(bold=True, size=14)

    # =========================================================================
    # Sheet 2: Application Inventory
    # =========================================================================
    ws = wb.create_sheet("Applications")

    headers = ["ID", "Application", "Vendor", "Category", "Hosting", "Criticality", "Version", "EOL Status", "EOL Date", "Users"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row, app in enumerate(result.applications, 2):
        data = [app.id, app.name, app.vendor, app.category, app.hosting, app.criticality, app.version, app.eol_status, app.eol_date, app.user_count]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            # Highlight EOL risks
            if col == 8:  # EOL Status column
                if value == "Past_EOL":
                    cell.fill = critical_fill
                elif value == "Approaching_EOL":
                    cell.fill = high_fill

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # =========================================================================
    # Sheet 3: Risk Register
    # =========================================================================
    ws = wb.create_sheet("Risk Register")

    headers = ["ID", "Severity", "Title", "Domain", "Description", "Mitigation", "Evidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row, risk in enumerate(result.risks, 2):
        data = [risk.id, risk.severity, risk.title, risk.domain, risk.description, risk.mitigation, risk.evidence]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            # Highlight by severity
            if col == 2:
                if value == "Critical":
                    cell.fill = critical_fill
                elif value == "High":
                    cell.fill = high_fill

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40

    # =========================================================================
    # Sheet 4: Cost Estimates
    # =========================================================================
    ws = wb.create_sheet("Cost Estimates")

    headers = ["ID", "Category", "Description", "Timing", "Low Estimate", "High Estimate", "Driver"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row, cost in enumerate(result.cost_items, 2):
        data = [cost.id, cost.category, cost.description, cost.timing, cost.low_estimate, cost.high_estimate, cost.driver]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            # Format currency columns
            if col in [5, 6]:
                cell.number_format = '$#,##0'

    # Total row
    total_row = len(result.cost_items) + 3
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=result.total_one_time_low).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = '$#,##0'
    ws.cell(row=total_row, column=6, value=result.total_one_time_high).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '$#,##0'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # =========================================================================
    # Sheet 5: Follow-Up Questions
    # =========================================================================
    ws = wb.create_sheet("Questions")

    headers = ["ID", "Priority", "Question", "Why It Matters", "Target", "Domain"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row, q in enumerate(result.questions, 2):
        data = [q.id, q.priority, q.question, q.why_it_matters, q.target, q.domain]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

    # Save
    wb.save(output_path)

    return {
        "status": "success",
        "output_path": output_path,
        "sheets": ["Executive Summary", "Applications", "Risk Register", "Cost Estimates", "Questions"],
        "counts": {
            "applications": len(result.applications),
            "risks": len(result.risks),
            "cost_items": len(result.cost_items),
            "questions": len(result.questions),
        }
    }
