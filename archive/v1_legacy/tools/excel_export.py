"""
Excel Export Module

Exports IT Due Diligence findings to client-ready Excel workbooks.

Features:
- Multiple worksheets (Summary, Risks, Gaps, Work Items, Recommendations, etc.)
- Professional formatting with conditional formatting for severity/priority
- Evidence citations and source references
- Sortable/filterable columns
"""

import logging
from typing import Dict, List, Any
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Import openpyxl components at module level
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = ColorScaleRule = FormulaRule = None


def export_findings_to_excel(
    run_id: str,
    repository: Any,
    output_path: str,
    include_evidence: bool = True
) -> Dict[str, Any]:
    """
    Export all findings for a run to a formatted Excel workbook.

    Args:
        run_id: Analysis run ID
        repository: Repository instance for data access
        output_path: Path to save Excel file
        include_evidence: Whether to include source evidence columns

    Returns:
        Dict with output_path, counts per sheet, and total
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ImportError("openpyxl required for Excel export. Run: pip install openpyxl")

    # Fetch all data
    data = _fetch_all_findings(run_id, repository)

    wb = openpyxl.Workbook()

    # Define styles
    styles = _get_styles()

    # Create worksheets
    _create_summary_sheet(wb, data, styles)
    _create_risks_sheet(wb, data['risks'], styles, include_evidence)
    _create_gaps_sheet(wb, data['gaps'], styles, include_evidence)
    _create_work_items_sheet(wb, data['work_items'], styles)
    _create_recommendations_sheet(wb, data['recommendations'], styles)
    _create_current_state_sheet(wb, data['current_state'], styles)
    _create_assumptions_sheet(wb, data['assumptions'], styles)

    # Remove default empty sheet if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    wb.save(output_path)
    logger.info(f"Exported findings to {output_path}")

    # Return summary info
    counts = {
        'Risks': len(data['risks']),
        'Gaps': len(data['gaps']),
        'Work Items': len(data['work_items']),
        'Recommendations': len(data['recommendations']),
        'Current State': len(data['current_state']),
        'Assumptions': len(data['assumptions'])
    }
    total = sum(counts.values())

    return {
        'output_path': output_path,
        'counts': counts,
        'total': total,
        'run_id': run_id
    }


def _fetch_all_findings(run_id: str, repository: Any) -> Dict[str, List]:
    """Fetch all findings for a run from the repository."""
    data = {
        'run_id': run_id,
        'risks': [],
        'gaps': [],
        'work_items': [],
        'recommendations': [],
        'current_state': [],
        'assumptions': [],
        'questions': [],
        'strategic_considerations': []
    }

    try:
        # Get run info
        run = repository.get_run(run_id)
        if run:
            data['run_info'] = run.to_dict() if hasattr(run, 'to_dict') else run

        # Get all findings
        conn = repository.db.connect()

        # Risks
        rows = conn.execute(
            "SELECT * FROM risks WHERE run_id = ? ORDER BY severity DESC, likelihood DESC",
            (run_id,)
        ).fetchall()
        data['risks'] = [dict(row) for row in rows]

        # Gaps
        rows = conn.execute(
            "SELECT * FROM gaps WHERE run_id = ? ORDER BY priority DESC",
            (run_id,)
        ).fetchall()
        data['gaps'] = [dict(row) for row in rows]

        # Work Items
        rows = conn.execute(
            "SELECT * FROM work_items WHERE run_id = ? ORDER BY phase, priority_score DESC",
            (run_id,)
        ).fetchall()
        data['work_items'] = [dict(row) for row in rows]

        # Recommendations
        rows = conn.execute(
            "SELECT * FROM recommendations WHERE run_id = ? ORDER BY priority DESC",
            (run_id,)
        ).fetchall()
        data['recommendations'] = [dict(row) for row in rows]

        # Current State / Inventory
        rows = conn.execute(
            "SELECT * FROM inventory_items WHERE run_id = ? ORDER BY domain, category",
            (run_id,)
        ).fetchall()
        data['current_state'] = [dict(row) for row in rows]

        # Assumptions
        rows = conn.execute(
            "SELECT * FROM assumptions WHERE run_id = ? ORDER BY confidence DESC",
            (run_id,)
        ).fetchall()
        data['assumptions'] = [dict(row) for row in rows]

        # Questions
        rows = conn.execute(
            "SELECT * FROM questions WHERE run_id = ? ORDER BY priority DESC, status",
            (run_id,)
        ).fetchall()
        data['questions'] = [dict(row) for row in rows]

        # Strategic Considerations
        rows = conn.execute(
            "SELECT * FROM strategic_considerations WHERE run_id = ? ORDER BY domain",
            (run_id,)
        ).fetchall()
        data['strategic_considerations'] = [dict(row) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching findings: {e}")
        raise

    return data


def _get_styles():
    """Define Excel styles for consistent formatting."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'header_fill': PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),

        'subheader_font': Font(bold=True, size=10),
        'subheader_fill': PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid"),

        'data_alignment': Alignment(vertical="top", wrap_text=True),
        'center_alignment': Alignment(horizontal="center", vertical="top"),

        'border': thin_border,

        # Severity/Priority colors
        'critical_fill': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        'critical_font': Font(color="FFFFFF", bold=True),
        'high_fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'medium_fill': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'low_fill': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),

        # Phase colors
        'day1_fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'day100_fill': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'post100_fill': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
        'optional_fill': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid"),
    }


def _apply_header_style(ws, row, styles, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']


def _apply_severity_formatting(cell, value, styles):
    """Apply conditional formatting based on severity/priority value."""
    if not value:
        return

    value_lower = str(value).lower()
    if value_lower == 'critical':
        cell.fill = styles['critical_fill']
        cell.font = styles['critical_font']
    elif value_lower == 'high':
        cell.fill = styles['high_fill']
    elif value_lower == 'medium':
        cell.fill = styles['medium_fill']
    elif value_lower == 'low':
        cell.fill = styles['low_fill']


def _apply_phase_formatting(cell, value, styles):
    """Apply conditional formatting based on phase value."""
    if not value:
        return

    value_lower = str(value).lower().replace('_', '')
    if 'day1' in value_lower or value_lower == 'day_1':
        cell.fill = styles['day1_fill']
    elif 'day100' in value_lower or value_lower == 'day_100':
        cell.fill = styles['day100_fill']
    elif 'post' in value_lower:
        cell.fill = styles['post100_fill']
    elif 'optional' in value_lower:
        cell.fill = styles['optional_fill']


def _auto_size_columns(ws, min_width=10, max_width=60):
    """Auto-size columns based on content."""
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_summary_sheet(wb, data: Dict, styles: Dict):
    """Create the Summary worksheet."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws['A1'] = "IT Due Diligence Analysis Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    # Run info
    ws['A3'] = "Run ID:"
    ws['B3'] = data.get('run_id', 'N/A')
    ws['A4'] = "Generated:"
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    run_info = data.get('run_info', {})
    ws['A5'] = "Run Name:"
    ws['B5'] = run_info.get('run_name', 'N/A')
    ws['A6'] = "Status:"
    ws['B6'] = run_info.get('status', 'N/A')

    # Counts summary
    ws['A8'] = "Findings Summary"
    ws['A8'].font = Font(bold=True, size=12)

    counts = [
        ("Risks", len(data['risks'])),
        ("Gaps", len(data['gaps'])),
        ("Work Items", len(data['work_items'])),
        ("Recommendations", len(data['recommendations'])),
        ("Current State Items", len(data['current_state'])),
        ("Assumptions", len(data['assumptions'])),
        ("Questions", len(data['questions'])),
    ]

    row = 9
    for label, count in counts:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Risk severity breakdown
    row += 1
    ws.cell(row=row, column=1, value="Risk Severity Breakdown")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    severity_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    for risk in data['risks']:
        sev = str(risk.get('severity', 'medium')).lower()
        if sev in severity_counts:
            severity_counts[sev] += 1

    for sev, count in severity_counts.items():
        ws.cell(row=row, column=1, value=sev.title())
        ws.cell(row=row, column=2, value=count)
        _apply_severity_formatting(ws.cell(row=row, column=1), sev, styles)
        row += 1

    # Work item phase breakdown
    row += 1
    ws.cell(row=row, column=1, value="Work Items by Phase")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    phase_counts = {'Day_1': 0, 'Day_100': 0, 'Post_100': 0, 'Optional': 0}
    for wi in data['work_items']:
        phase = wi.get('phase', 'Day_100')
        if phase in phase_counts:
            phase_counts[phase] += 1

    for phase, count in phase_counts.items():
        ws.cell(row=row, column=1, value=phase.replace('_', ' '))
        ws.cell(row=row, column=2, value=count)
        _apply_phase_formatting(ws.cell(row=row, column=1), phase, styles)
        row += 1

    _auto_size_columns(ws)


def _create_risks_sheet(wb, risks: List[Dict], styles: Dict, include_evidence: bool):
    """Create the Risks worksheet."""
    ws = wb.create_sheet("Risks")

    # Headers
    headers = ["#", "Domain", "Risk Description", "Severity", "Likelihood",
               "Risk Type", "Integration Dependent", "Deal Impact",
               "Mitigation", "Cost Impact"]

    if include_evidence:
        headers.extend(["Evidence Quote", "Evidence Type", "Confidence"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    # Data rows
    for row_num, risk in enumerate(risks, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=risk.get('domain', ''))
        ws.cell(row=row_num, column=3, value=risk.get('risk_description', ''))

        # Severity with formatting
        sev_cell = ws.cell(row=row_num, column=4, value=risk.get('severity', ''))
        _apply_severity_formatting(sev_cell, risk.get('severity'), styles)

        ws.cell(row=row_num, column=5, value=risk.get('likelihood', ''))
        ws.cell(row=row_num, column=6, value=risk.get('risk_type', ''))
        ws.cell(row=row_num, column=7, value='Yes' if risk.get('integration_dependent') else 'No')

        # Deal impact (may be JSON)
        deal_impact = risk.get('deal_impact', [])
        if isinstance(deal_impact, str):
            try:
                import json
                deal_impact = json.loads(deal_impact)
            except (json.JSONDecodeError, TypeError):
                pass
        if isinstance(deal_impact, list):
            deal_impact = ', '.join(deal_impact)
        ws.cell(row=row_num, column=8, value=deal_impact)

        ws.cell(row=row_num, column=9, value=risk.get('mitigation', ''))
        ws.cell(row=row_num, column=10, value=risk.get('cost_impact_estimate', ''))

        if include_evidence:
            # Parse source_evidence if it's JSON
            evidence = risk.get('source_evidence', {})
            if isinstance(evidence, str):
                try:
                    import json
                    evidence = json.loads(evidence)
                except (json.JSONDecodeError, TypeError):
                    evidence = {}

            ws.cell(row=row_num, column=11, value=evidence.get('exact_quote', '') if evidence else '')
            ws.cell(row=row_num, column=12, value=evidence.get('evidence_type', '') if evidence else '')
            ws.cell(row=row_num, column=13, value=evidence.get('confidence_level', '') if evidence else '')

        # Apply alignment
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    _auto_size_columns(ws)


def _create_gaps_sheet(wb, gaps: List[Dict], styles: Dict, include_evidence: bool):
    """Create the Gaps worksheet."""
    ws = wb.create_sheet("Gaps")

    headers = ["#", "Domain", "Gap Description", "Why Needed", "Priority",
               "Gap Type", "Suggested Source", "Cost Impact", "Question Status"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, gap in enumerate(gaps, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=gap.get('domain', ''))
        ws.cell(row=row_num, column=3, value=gap.get('gap_description', ''))
        ws.cell(row=row_num, column=4, value=gap.get('why_needed', ''))

        # Priority with formatting
        pri_cell = ws.cell(row=row_num, column=5, value=gap.get('priority', ''))
        _apply_severity_formatting(pri_cell, gap.get('priority'), styles)

        ws.cell(row=row_num, column=6, value=gap.get('gap_type', ''))
        ws.cell(row=row_num, column=7, value=gap.get('suggested_source', ''))
        ws.cell(row=row_num, column=8, value=gap.get('cost_impact', ''))
        ws.cell(row=row_num, column=9, value=gap.get('question_status', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_work_items_sheet(wb, work_items: List[Dict], styles: Dict):
    """Create the Work Items worksheet."""
    ws = wb.create_sheet("Work Items")

    headers = ["#", "Domain", "Title", "Description", "Category", "Phase",
               "Phase Rationale", "Effort Estimate", "Cost Estimate",
               "Skills Required", "Dependencies", "Triggered By"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, wi in enumerate(work_items, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=wi.get('domain', ''))
        ws.cell(row=row_num, column=3, value=wi.get('title', ''))
        ws.cell(row=row_num, column=4, value=wi.get('description', ''))
        ws.cell(row=row_num, column=5, value=wi.get('category', ''))

        # Phase with formatting
        phase_cell = ws.cell(row=row_num, column=6, value=wi.get('phase', ''))
        _apply_phase_formatting(phase_cell, wi.get('phase'), styles)

        ws.cell(row=row_num, column=7, value=wi.get('phase_rationale', ''))
        ws.cell(row=row_num, column=8, value=wi.get('effort_estimate', ''))
        ws.cell(row=row_num, column=9, value=wi.get('cost_estimate_range', ''))

        # Parse JSON arrays
        def parse_list(val):
            if isinstance(val, str):
                try:
                    import json
                    val = json.loads(val)
                except (json.JSONDecodeError, TypeError):
                    pass
            if isinstance(val, list):
                return ', '.join(str(v) for v in val)
            return val or ''

        ws.cell(row=row_num, column=10, value=parse_list(wi.get('skills_required')))
        ws.cell(row=row_num, column=11, value=parse_list(wi.get('depends_on')))
        ws.cell(row=row_num, column=12, value=parse_list(wi.get('triggered_by')))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_recommendations_sheet(wb, recommendations: List[Dict], styles: Dict):
    """Create the Recommendations worksheet."""
    ws = wb.create_sheet("Recommendations")

    headers = ["#", "Domain", "Recommendation", "Rationale", "Priority",
               "Timing", "Investment Required", "Driven By"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, rec in enumerate(recommendations, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=rec.get('domain', ''))
        ws.cell(row=row_num, column=3, value=rec.get('recommendation_text', ''))
        ws.cell(row=row_num, column=4, value=rec.get('rationale', ''))

        pri_cell = ws.cell(row=row_num, column=5, value=rec.get('priority', ''))
        _apply_severity_formatting(pri_cell, rec.get('priority'), styles)

        ws.cell(row=row_num, column=6, value=rec.get('timing', ''))
        ws.cell(row=row_num, column=7, value=rec.get('investment_required', ''))

        driven_by = rec.get('driven_by', [])
        if isinstance(driven_by, str):
            try:
                import json
                driven_by = json.loads(driven_by)
            except (json.JSONDecodeError, TypeError):
                pass
        if isinstance(driven_by, list):
            driven_by = ', '.join(driven_by)
        ws.cell(row=row_num, column=8, value=driven_by)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_current_state_sheet(wb, items: List[Dict], styles: Dict):
    """Create the Current State / Inventory worksheet."""
    ws = wb.create_sheet("Current State")

    headers = ["#", "Domain", "Category", "Item Name", "Description",
               "Status", "Maturity", "Standalone Viability", "Key Characteristics"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=item.get('domain', ''))
        ws.cell(row=row_num, column=3, value=item.get('category', ''))
        ws.cell(row=row_num, column=4, value=item.get('item_name', ''))
        ws.cell(row=row_num, column=5, value=item.get('description', ''))
        ws.cell(row=row_num, column=6, value=item.get('status', ''))
        ws.cell(row=row_num, column=7, value=item.get('maturity', ''))
        ws.cell(row=row_num, column=8, value=item.get('standalone_viability', ''))

        chars = item.get('key_characteristics', [])
        if isinstance(chars, str):
            try:
                import json
                chars = json.loads(chars)
            except (json.JSONDecodeError, TypeError):
                pass
        if isinstance(chars, list):
            chars = ', '.join(chars)
        ws.cell(row=row_num, column=9, value=chars)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_assumptions_sheet(wb, assumptions: List[Dict], styles: Dict):
    """Create the Assumptions worksheet."""
    ws = wb.create_sheet("Assumptions")

    headers = ["#", "Domain", "Assumption", "Basis", "Confidence",
               "Impact", "Validation Needed", "Validation Status"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, assumption in enumerate(assumptions, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=assumption.get('domain', ''))
        ws.cell(row=row_num, column=3, value=assumption.get('assumption_text', ''))
        ws.cell(row=row_num, column=4, value=assumption.get('basis', ''))
        ws.cell(row=row_num, column=5, value=assumption.get('confidence', ''))
        ws.cell(row=row_num, column=6, value=assumption.get('impact', ''))
        ws.cell(row=row_num, column=7, value=assumption.get('validation_needed', ''))
        ws.cell(row=row_num, column=8, value=assumption.get('validation_status', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


# =============================================================================
# APPLICATION-SPECIFIC EXPORT FUNCTIONS (Phase 4-6 Enhancement)
# =============================================================================
# These functions export directly from AnalysisStore for team review.

def export_applications_to_excel(
    analysis_store: Any,
    output_path: str,
    include_buyer_comparison: bool = True
) -> Dict[str, Any]:
    """
    Export application inventory and related analysis to Excel.

    Phase 4-6 Enhancement: Dedicated export for application analysis.

    Creates worksheets for:
    - Applications: Full application inventory
    - Capability Coverage: Business capability analysis
    - Buyer Overlaps: Target vs Buyer comparison (if include_buyer_comparison=True)
    - Completeness: Data completeness metrics

    Args:
        analysis_store: AnalysisStore instance with populated data
        output_path: Path to save Excel file
        include_buyer_comparison: Whether to include buyer comparison tabs

    Returns:
        Dict with output_path, sheet counts, and completeness metrics
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ImportError("openpyxl required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    styles = _get_styles()

    # Create Applications sheet
    apps_count = _create_applications_sheet(wb, analysis_store, styles)

    # Create Capability Coverage sheet
    cap_count = _create_capability_coverage_sheet(wb, analysis_store, styles)

    # Create Buyer Overlaps sheet (if requested and data exists)
    overlap_count = 0
    buyer_count = 0
    if include_buyer_comparison:
        buyer_count = _create_buyer_applications_sheet(wb, analysis_store, styles)
        overlap_count = _create_overlaps_sheet(wb, analysis_store, styles)

    # Create Completeness Metrics sheet
    _create_completeness_sheet(wb, analysis_store, styles)

    # Remove default empty sheet if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    wb.save(output_path)
    logger.info(f"Exported applications to {output_path}")

    # Get completeness metrics
    metrics = analysis_store.get_application_completeness_metrics()

    return {
        'output_path': output_path,
        'counts': {
            'Applications': apps_count,
            'Capability Areas': cap_count,
            'Buyer Applications': buyer_count,
            'Overlaps': overlap_count
        },
        'completeness_score': metrics.get('completeness_score', 0),
        'total_applications': apps_count
    }


def _create_applications_sheet(wb, analysis_store: Any, styles: Dict) -> int:
    """Create the Applications inventory worksheet."""
    ws = wb.active
    ws.title = "Applications"

    apps = analysis_store.get_application_inventory()

    if not apps:
        ws['A1'] = "No applications recorded"
        return 0

    # Headers
    headers = [
        "ID", "Application Name", "Vendor", "Category", "Hosting Model",
        "Criticality", "Version", "Support Status", "User Count",
        "License Type", "Customization Level", "Integration Count",
        "Capability Areas", "Discovery Source", "Evidence Quote",
        "Fields Not Documented"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    # Data rows
    for row_num, app in enumerate(apps, 2):
        ws.cell(row=row_num, column=1, value=app.get('id', ''))
        ws.cell(row=row_num, column=2, value=app.get('application_name', ''))
        ws.cell(row=row_num, column=3, value=app.get('vendor', ''))
        ws.cell(row=row_num, column=4, value=app.get('application_category', ''))
        ws.cell(row=row_num, column=5, value=app.get('hosting_model', ''))

        # Criticality with formatting
        crit_cell = ws.cell(row=row_num, column=6, value=app.get('business_criticality', ''))
        _apply_severity_formatting(crit_cell, app.get('business_criticality'), styles)

        ws.cell(row=row_num, column=7, value=app.get('version', ''))
        ws.cell(row=row_num, column=8, value=app.get('support_status', ''))
        ws.cell(row=row_num, column=9, value=app.get('user_count', ''))
        ws.cell(row=row_num, column=10, value=app.get('license_type', ''))
        ws.cell(row=row_num, column=11, value=app.get('customization_level', ''))
        ws.cell(row=row_num, column=12, value=app.get('integration_count', ''))

        # Capability areas as comma-separated
        cap_areas = app.get('capability_areas_covered', [])
        ws.cell(row=row_num, column=13, value=', '.join(cap_areas) if cap_areas else '')

        ws.cell(row=row_num, column=14, value=app.get('discovery_source', ''))

        # Evidence quote
        evidence = app.get('source_evidence', {})
        ws.cell(row=row_num, column=15, value=evidence.get('exact_quote', ''))

        # Fields not documented
        missing = app.get('fields_not_documented', [])
        ws.cell(row=row_num, column=16, value=', '.join(missing) if missing else '')

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)

    return len(apps)


def _create_capability_coverage_sheet(wb, analysis_store: Any, styles: Dict) -> int:
    """Create the Capability Coverage worksheet."""
    ws = wb.create_sheet("Capability Coverage")

    coverage = analysis_store.capability_coverage

    if not coverage:
        ws['A1'] = "No capability coverage recorded"
        return 0

    # Headers
    headers = [
        "ID", "Capability Area", "Coverage Status", "Business Relevance",
        "Relevance Reasoning", "Apps Found", "Expected But Missing",
        "Gap Severity", "Maturity", "Follow-up Questions Count",
        "Confidence"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    # Data rows
    for row_num, cap in enumerate(coverage, 2):
        ws.cell(row=row_num, column=1, value=cap.get('id', ''))
        ws.cell(row=row_num, column=2, value=cap.get('capability_area', ''))

        # Coverage status
        status_cell = ws.cell(row=row_num, column=3, value=cap.get('coverage_status', ''))
        if cap.get('coverage_status') == 'Not_Found':
            status_cell.fill = styles['high_fill']
        elif cap.get('coverage_status') == 'Partially_Documented':
            status_cell.fill = styles['medium_fill']

        ws.cell(row=row_num, column=4, value=cap.get('business_relevance', ''))
        ws.cell(row=row_num, column=5, value=cap.get('relevance_reasoning', ''))

        # Apps found as comma-separated names
        apps_found = cap.get('applications_found', [])
        app_names = [a.get('app_name', '') for a in apps_found if isinstance(a, dict)]
        ws.cell(row=row_num, column=6, value=', '.join(app_names) if app_names else '')

        # Expected but missing
        missing = cap.get('expected_but_missing', [])
        ws.cell(row=row_num, column=7, value=', '.join(missing) if missing else '')

        # Gap severity with formatting
        gap_cell = ws.cell(row=row_num, column=8, value=cap.get('gap_severity', ''))
        _apply_severity_formatting(gap_cell, cap.get('gap_severity'), styles)

        ws.cell(row=row_num, column=9, value=cap.get('capability_maturity', ''))

        # Follow-up questions count
        questions = cap.get('follow_up_questions', [])
        ws.cell(row=row_num, column=10, value=len(questions))

        ws.cell(row=row_num, column=11, value=cap.get('confidence_level', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)

    return len(coverage)


def _create_buyer_applications_sheet(wb, analysis_store: Any, styles: Dict) -> int:
    """Create the Buyer Applications worksheet."""
    ws = wb.create_sheet("Buyer Applications")

    buyer_apps = analysis_store.get_buyer_applications()

    if not buyer_apps:
        ws['A1'] = "No buyer applications recorded"
        return 0

    # Headers
    headers = [
        "ID", "Application Name", "Vendor", "Category",
        "Capability Areas", "Information Source", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    # Data rows
    for row_num, app in enumerate(buyer_apps, 2):
        ws.cell(row=row_num, column=1, value=app.get('id', ''))
        ws.cell(row=row_num, column=2, value=app.get('application_name', ''))
        ws.cell(row=row_num, column=3, value=app.get('vendor', ''))
        ws.cell(row=row_num, column=4, value=app.get('application_category', ''))

        # Capability areas as comma-separated
        cap_areas = app.get('capability_areas_covered', [])
        ws.cell(row=row_num, column=5, value=', '.join(cap_areas) if cap_areas else '')

        ws.cell(row=row_num, column=6, value=app.get('information_source', ''))
        ws.cell(row=row_num, column=7, value=app.get('notes', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)

    return len(buyer_apps)


def _create_overlaps_sheet(wb, analysis_store: Any, styles: Dict) -> int:
    """Create the Application Overlaps worksheet."""
    ws = wb.create_sheet("Overlaps")

    overlaps = analysis_store.get_application_overlaps()

    if not overlaps:
        ws['A1'] = "No overlaps analyzed"
        return 0

    # Headers
    headers = [
        "ID", "Target App", "Target Category", "Buyer App",
        "Overlap Type", "Considerations", "Follow-up Questions",
        "Notes"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    # Data rows
    for row_num, ovl in enumerate(overlaps, 2):
        ws.cell(row=row_num, column=1, value=ovl.get('id', ''))
        ws.cell(row=row_num, column=2, value=ovl.get('target_app_name', ''))
        ws.cell(row=row_num, column=3, value=ovl.get('target_app_category', ''))
        ws.cell(row=row_num, column=4, value=ovl.get('buyer_app_name', '') or 'N/A')

        # Overlap type with color coding
        ovl_type = ovl.get('overlap_type', '')
        type_cell = ws.cell(row=row_num, column=5, value=ovl_type)
        if ovl_type == 'Same_Product':
            type_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif ovl_type == 'Same_Category_Different_Vendor':
            type_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        elif ovl_type == 'Target_Only':
            type_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        ws.cell(row=row_num, column=6, value=ovl.get('considerations', ''))

        # Follow-up questions as bullet list
        questions = ovl.get('follow_up_questions', [])
        q_text = '\n'.join([f"• {q.get('question', '')}" for q in questions]) if questions else ''
        ws.cell(row=row_num, column=7, value=q_text)

        ws.cell(row=row_num, column=8, value=ovl.get('notes', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)

    return len(overlaps)


def _create_completeness_sheet(wb, analysis_store: Any, styles: Dict):
    """Create the Completeness Metrics worksheet."""
    ws = wb.create_sheet("Completeness")

    # Get metrics
    metrics = analysis_store.get_application_completeness_metrics()
    cap_summary = analysis_store.get_capability_summary()

    # Title
    ws['A1'] = "Application Analysis Completeness Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Overall metrics
    ws['A4'] = "Overall Metrics"
    ws['A4'].font = Font(bold=True, size=12)

    ws['A5'] = "Total Applications:"
    ws['B5'] = metrics.get('total_applications', 0)

    ws['A6'] = "Completeness Score:"
    score = metrics.get('completeness_score', 0)
    ws['B6'] = f"{score}%"
    if score >= 70:
        ws['B6'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    elif score >= 40:
        ws['B6'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    else:
        ws['B6'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    ws['A7'] = "Apps with Missing Fields:"
    ws['B7'] = metrics.get('apps_with_missing_fields', 0)

    # Capability coverage
    ws['A9'] = "Capability Coverage"
    ws['A9'].font = Font(bold=True, size=12)

    ws['A10'] = "Total Capability Areas:"
    ws['B10'] = cap_summary.get('total_capability_areas', 0)

    ws['A11'] = "Areas Assessed:"
    ws['B11'] = cap_summary.get('areas_assessed', 0)

    ws['A12'] = "Coverage Percentage:"
    ws['B12'] = f"{cap_summary.get('completeness_percentage', 0)}%"

    ws['A13'] = "Critical Gaps:"
    critical_gaps = cap_summary.get('critical_gaps', [])
    ws['B13'] = ', '.join(critical_gaps) if critical_gaps else 'None'

    # Unknown value counts
    ws['A15'] = "Unknown Value Counts"
    ws['A15'].font = Font(bold=True, size=12)

    row = 16
    unknown_counts = metrics.get('unknown_value_counts', {})
    for field, count in unknown_counts.items():
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=count)
        if count > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        row += 1

    # Not assessed areas
    row += 1
    ws.cell(row=row, column=1, value="Areas Not Assessed")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    not_assessed = cap_summary.get('areas_not_assessed', [])
    if not_assessed:
        for area in not_assessed:
            ws.cell(row=row, column=1, value=area)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            row += 1
    else:
        ws.cell(row=row, column=1, value="All areas assessed")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    _auto_size_columns(ws)


# =============================================================================
# CLI TESTING
# =============================================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))

    from storage import Database, Repository

    logging.basicConfig(level=logging.INFO)

    db = Database()
    db.initialize_schema()
    repository = Repository(db)

    # Get latest run
    latest = repository.get_latest_run()
    if not latest:
        print("No runs found.")
        sys.exit(1)

    run_id = latest.run_id
    print(f"Exporting findings for run: {run_id}")

    output_path = f"findings_{run_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    result = export_findings_to_excel(run_id, repository, output_path)
    print(f"Exported to: {result}")
