"""
Web interface for IT Due Diligence Agent.

Provides a browser-based UI for reviewing and adjusting analysis outputs.
Run with: python -m web.app

Phase 1 Updates:
- Replaced global session with thread-safe SessionStore
- Added AnalysisTaskManager for background processing
- Connected analysis pipeline to web interface
- Fixed /analysis/status to report real progress
"""

import sys
import os
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session as flask_session
from interactive.session import Session
from tools_v2.reasoning_tools import COST_RANGE_VALUES

# Organization module imports
from models.organization_models import CompanyInfo, EmploymentType, RoleCategory
from models.organization_stores import OrganizationDataStore, OrganizationAnalysisResult
from services.organization_pipeline import OrganizationAnalysisPipeline

# Phase 1: New imports for task management and session handling
from web.task_manager import task_manager, AnalysisPhase, TaskStatus
from stores.session_store import session_store, get_or_create_session_id
from web.analysis_runner import run_analysis, run_analysis_simple

# Phase 2: Authentication imports
from flask_login import LoginManager, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from flask_talisman import Talisman

# Phase 3: Database imports
from web.database import db, migrate, init_db, create_all_tables

# Phase 3.1: Repository imports for A2 fix (documents to DB)
from web.repositories.document_repository import DocumentRepository

# Phase 4: Session and Task imports
from flask_session import Session as FlaskSession

# Configure Flask with static folder
app = Flask(__name__,
            static_folder='static',
            static_url_path='/static')

# Use environment variable for secret key, with fallback for MVP demo
SECRET_KEY = os.environ.get('FLASK_SECRET_KEY') or os.environ.get('SECRET_KEY') or 'itdd-mvp-demo-secret-2026-production'
app.secret_key = SECRET_KEY
app.config['SECRET_KEY'] = SECRET_KEY
app.config['WTF_CSRF_SECRET_KEY'] = SECRET_KEY

# =============================================================================
# Phase 3: Database Setup
# =============================================================================

# Check if database is enabled (for gradual migration)
USE_DATABASE = os.environ.get('USE_DATABASE', 'false').lower() == 'true'

# Check if db auth backend is used (requires database)
AUTH_BACKEND = os.environ.get('AUTH_BACKEND', 'db').lower()

# Initialize database if explicitly enabled OR if using db auth backend
NEED_DATABASE = USE_DATABASE or AUTH_BACKEND == 'db'

if NEED_DATABASE:
    init_db(app)
    # Create tables and run migrations (adds task_id column for UI resilience)
    create_all_tables(app)
    logger.info(f"Database initialized (USE_DATABASE={USE_DATABASE}, AUTH_BACKEND={AUTH_BACKEND})")

# =============================================================================
# Phase 4: Session & Task Configuration
# =============================================================================

# Session configuration - prioritize database sessions for Railway reliability
# Redis is optional; SQLAlchemy sessions work reliably with PostgreSQL
REDIS_URL = os.environ.get('REDIS_URL', 'redis://localhost:6379/0')
USE_REDIS_SESSIONS = os.environ.get('USE_REDIS_SESSIONS', 'false').lower() == 'true'


# =============================================================================
# Session Backend Health Checks (Spec 04)
# =============================================================================

def check_redis_health(redis_url: str, timeout: int = 2) -> bool:
    """
    Check if Redis is healthy before using it for sessions.

    Args:
        redis_url: Redis connection URL
        timeout: Connection timeout in seconds

    Returns:
        True if Redis is reachable and responsive, False otherwise
    """
    try:
        import redis
        client = redis.from_url(
            redis_url,
            socket_connect_timeout=timeout,
            socket_timeout=timeout
        )
        client.ping()
        client.close()
        return True
    except Exception as e:
        logger.warning(f"Redis health check failed: {e}")
        return False


# Helper to configure SQLAlchemy sessions
def _configure_db_sessions():
    """Configure SQLAlchemy-backed sessions (reliable on Railway)."""
    app.config['SESSION_TYPE'] = 'sqlalchemy'
    app.config['SESSION_SQLALCHEMY'] = db
    app.config['SESSION_SQLALCHEMY_TABLE'] = 'flask_sessions'
    app.config['SESSION_PERMANENT'] = True
    app.config['SESSION_USE_SIGNER'] = True
    app.config['SESSION_KEY_PREFIX'] = 'itdd:session:'
    app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours
    FlaskSession(app)
    logger.info("✅ Session backend: SQLAlchemy (PostgreSQL)")

# Helper to configure filesystem sessions
def _configure_fs_sessions():
    """Configure filesystem sessions (local development only).

    Note: We use Flask's built-in cookie sessions instead of Flask-Session
    to avoid SQLAlchemy metadata conflicts when the background analysis
    thread creates app contexts. Flask-Session's SQLAlchemy models can
    conflict with our own models even when using filesystem storage.
    """
    # Just use Flask's built-in session (cookie-based)
    # The secret key is already set, so sessions will work
    app.config['SESSION_PERMANENT'] = True
    app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours
    # Don't call FlaskSession(app) - use Flask's built-in sessions
    logger.info("Using Flask built-in cookie sessions (local dev)")

# Session priority: Redis (if enabled & working) > SQLAlchemy (if DB available) > Filesystem
session_configured = False

if USE_REDIS_SESSIONS and check_redis_health(REDIS_URL):
    try:
        import redis
        redis_client = redis.from_url(
            REDIS_URL,
            socket_connect_timeout=2,
            socket_timeout=2,
            retry_on_timeout=True,
            health_check_interval=30  # Check every 30 seconds
        )

        app.config['SESSION_TYPE'] = 'redis'
        app.config['SESSION_REDIS'] = redis_client
        app.config['SESSION_PERMANENT'] = True
        app.config['SESSION_USE_SIGNER'] = True
        app.config['SESSION_KEY_PREFIX'] = 'itdd:session:'
        app.config['PERMANENT_SESSION_LIFETIME'] = 86400
        FlaskSession(app)
        logger.info("✅ Session backend: Redis (healthy)")
        session_configured = True
    except Exception as e:
        logger.warning(f"Redis session configuration failed: {e}")
        # Fall through to try SQLAlchemy
elif USE_REDIS_SESSIONS:
    logger.warning("⚠️  Redis sessions requested but health check failed - falling back to database")

if not session_configured and USE_DATABASE:
    # Only use SQLAlchemy sessions with PostgreSQL (not SQLite - causes table conflicts)
    database_url = os.environ.get('DATABASE_URL', '')
    if 'postgresql' in database_url.lower():
        _configure_db_sessions()
        session_configured = True
        logger.info("Using PostgreSQL-backed sessions")
    else:
        logger.info("SQLite detected - skipping SQLAlchemy sessions (use filesystem instead)")

if not session_configured:
    # Last resort: cookie sessions (won't persist on Railway restarts, but no DB conflicts)
    _configure_fs_sessions()
    logger.info("Using cookie sessions for local development")

# Check if Celery is available for background tasks
USE_CELERY = os.environ.get('USE_CELERY', 'false').lower() == 'true'

if USE_CELERY:
    try:
        from web.celery_app import celery, is_celery_available
        if is_celery_available():
            logger.info("Celery tasks enabled")
        else:
            logger.warning("Celery broker not available, tasks will run synchronously")
            USE_CELERY = False
    except Exception as e:
        logger.warning(f"Celery not available: {e}")
        USE_CELERY = False

# =============================================================================
# Phase 5: Cloud Storage Setup
# =============================================================================

STORAGE_TYPE = os.environ.get('STORAGE_TYPE', 'local')

try:
    from web.storage import get_storage
    storage = get_storage()
    storage.initialize(STORAGE_TYPE)
    logger.info(f"Storage initialized: {storage.storage_type}")
except Exception as e:
    logger.warning(f"Storage initialization failed: {e}")
    storage = None

# =============================================================================
# Phase 2: Authentication Setup
# =============================================================================

# Initialize CSRF protection - DISABLED FOR MVP DEMO
# csrf = CSRFProtect(app)
app.config['WTF_CSRF_ENABLED'] = False

# Dummy csrf object so @csrf.exempt decorators don't break
class DummyCSRF:
    def exempt(self, f):
        return f
csrf = DummyCSRF()

# Configure CSRF exemptions for API routes
app.config['WTF_CSRF_CHECK_DEFAULT'] = False  # We'll check manually

# CSRF protection disabled for MVP demo
# @app.before_request
# def csrf_protect():
#     """Apply CSRF protection selectively - exempt API routes and file uploads."""
#     pass

# Initialize Flask-Talisman for security headers
# Only enforce HTTPS in production
FORCE_HTTPS = os.environ.get('FORCE_HTTPS', 'false').lower() == 'true'

# Content Security Policy
csp = {
    'default-src': "'self'",
    'script-src': ["'self'", "'unsafe-inline'"],  # Allow inline scripts for now (can tighten later)
    'style-src': ["'self'", "'unsafe-inline'"],   # Allow inline styles
    'img-src': ["'self'", "data:"],               # Allow data URIs for images
    'font-src': "'self'",
}

talisman = Talisman(
    app,
    force_https=FORCE_HTTPS,
    strict_transport_security=FORCE_HTTPS,
    session_cookie_secure=FORCE_HTTPS,
    content_security_policy=csp,
    # Disabled nonce requirement to allow inline scripts (unsafe-inline is in CSP)
    # content_security_policy_nonce_in=['script-src'],
    feature_policy={
        'geolocation': "'none'",
        'camera': "'none'",
        'microphone': "'none'",
    }
)

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'auth.login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# Phase 3: Feature flag for auth backend
AUTH_BACKEND = os.environ.get('AUTH_BACKEND', 'db').lower()

@login_manager.user_loader
def load_user(user_id):
    """Load user by ID for Flask-Login."""
    if AUTH_BACKEND == 'db':
        from web.services.auth_service import get_auth_service
        return get_auth_service().get_by_id(user_id)
    else:
        from web.models.user import get_user_store
        return get_user_store().get_by_id(user_id)

# Register authentication blueprint
from web.auth import auth_bp
app.register_blueprint(auth_bp)

# Register deals blueprint
from web.routes.deals import deals_bp
app.register_blueprint(deals_bp)

# Register inventory blueprint
from web.blueprints.inventory import inventory_bp
app.register_blueprint(inventory_bp)

# Register costs blueprint
from web.blueprints.costs import costs_bp
app.register_blueprint(costs_bp)

# Register PE reports blueprint
from web.blueprints.pe_reports import pe_reports_bp
app.register_blueprint(pe_reports_bp)

# Phase 3: Register CLI commands for user management
from web.cli import register_cli
register_cli(app)

# Check if authentication is required (can be disabled for development)
AUTH_REQUIRED = os.environ.get('AUTH_REQUIRED', 'true').lower() != 'false'

def auth_optional(f):
    """Decorator that makes auth optional based on AUTH_REQUIRED setting."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if AUTH_REQUIRED and not current_user.is_authenticated:
            return login_manager.unauthorized()
        return f(*args, **kwargs)
    return decorated_function

# Initialize user store and ensure admin exists
def init_auth():
    """Initialize authentication system."""
    if AUTH_BACKEND == 'db':
        # Database backend - NO auto-bootstrap (use CLI: flask create-admin)
        from web.services.auth_service import get_auth_service
        auth = get_auth_service()
        if not auth.admin_exists():
            logger.warning(
                "No admin user exists. Run 'flask create-admin --email <email>' to create one."
            )
    else:
        # Legacy file-based auth
        from web.models.user import get_user_store
        user_store = get_user_store()

        # Create default admin if no users exist
        if user_store.user_count() == 0:
            default_admin_email = os.environ.get('DEFAULT_ADMIN_EMAIL', 'admin@example.com')
            default_admin_password = os.environ.get('DEFAULT_ADMIN_PASSWORD', 'changeme123')
            user_store.ensure_admin_exists(default_admin_email, default_admin_password)
            logger.info(f"Created default admin user: {default_admin_email}")

# Initialize auth on first request
@app.before_request
def before_first_request():
    """Run initialization on first request."""
    if not hasattr(app, '_auth_initialized'):
        init_auth()
        app._auth_initialized = True

# Require authentication for all routes (except auth routes and static files)
@app.before_request
def require_authentication():
    """Require authentication for protected routes."""
    if not AUTH_REQUIRED:
        return None

    # Public routes that don't require authentication
    public_prefixes = ('/auth/', '/static/', '/api/health')
    if any(request.path.startswith(prefix) for prefix in public_prefixes):
        return None

    # Check if user is authenticated
    if not current_user.is_authenticated:
        return login_manager.unauthorized()

    return None


# =============================================================================
# Session Persistence - Auto-Restore Deal Context (Spec 03)
# =============================================================================

@app.before_request
def auto_restore_deal_context():
    """
    Automatically restore deal context from database when session is empty.

    This hook runs on every request for authenticated users. If the session
    doesn't have a current_deal_id but the user has a last_deal_id in the
    database, we restore it to the session.

    This ensures users don't see "Please select a deal" errors after:
    - Session expiry
    - Browser restart
    - Server restart
    - Multi-server handoff

    Note:
        Only restores if user has access to the deal (permissions checked).
        Does not restore if deal is deleted or access was revoked.
    """
    # Only run for authenticated users
    if not current_user.is_authenticated:
        return

    # Only restore if session is empty
    if flask_session.get('current_deal_id'):
        return

    # Attempt to restore from database
    try:
        last_deal = current_user.get_last_deal()

        if not last_deal:
            # User has no last deal recorded
            return

        # Verify user still has access (permissions may have changed)
        from web.permissions import user_can_access_deal
        if not user_can_access_deal(current_user, last_deal):
            # User lost access to this deal - clear it from database
            current_user.clear_last_deal()
            db.session.commit()
            logger.warning(f"User {current_user.id} lost access to deal {last_deal.id}, cleared last_deal_id")
            return

        # Restore to session
        flask_session['current_deal_id'] = last_deal.id
        flask_session.modified = True

        logger.info(f"Auto-restored deal {last_deal.id} for user {current_user.id} from database")

        # Optionally: Also load deal context into flask.g
        # (This matches the existing pattern in web/context.py)
        try:
            from web.context import load_deal_context
            load_deal_context()
        except Exception as context_error:
            # Non-fatal: session is restored even if g.deal_data fails
            logger.warning(f"Failed to load deal context into flask.g: {context_error}")

    except Exception as e:
        # Non-fatal: Don't break the request if restoration fails
        logger.error(f"Auto-restore deal context failed for user {current_user.id}: {e}")


# =============================================================================
# End Phase 2 Authentication Setup
# =============================================================================

# =============================================================================
# Phase 6: Multi-Tenancy Setup
# =============================================================================

USE_MULTI_TENANCY = os.environ.get('USE_MULTI_TENANCY', 'false').lower() == 'true'

if USE_MULTI_TENANCY:
    from web.middleware import set_tenant, create_default_tenant

    # Register tenant resolution middleware
    app.before_request(set_tenant)

    # Create default tenant if database is enabled
    if USE_DATABASE:
        with app.app_context():
            create_default_tenant(app)

    logger.info("Multi-tenancy enabled")

# =============================================================================
# End Phase 6 Multi-Tenancy Setup
# =============================================================================

# =============================================================================
# Phase 7: Enterprise Features
# =============================================================================

# 7.1 Structured Logging
USE_STRUCTURED_LOGGING = os.environ.get('USE_STRUCTURED_LOGGING', 'false').lower() == 'true'

if USE_STRUCTURED_LOGGING:
    try:
        from web.logging_config import init_logging, setup_request_logging
        init_logging()
        setup_request_logging(app)
        logger.info("Structured logging enabled")
    except Exception as e:
        logger.warning(f"Structured logging not available: {e}")

# 7.2 Error Tracking (Sentry)
USE_ERROR_TRACKING = os.environ.get('USE_ERROR_TRACKING', 'false').lower() == 'true'

if USE_ERROR_TRACKING:
    try:
        from web.error_tracking import init_sentry, setup_flask_error_handlers
        if init_sentry(app):
            setup_flask_error_handlers(app)
            logger.info("Error tracking (Sentry) enabled")
    except Exception as e:
        logger.warning(f"Error tracking not available: {e}")

# 7.3 Rate Limiting
USE_RATE_LIMITING = os.environ.get('USE_RATE_LIMITING', 'false').lower() == 'true'

if USE_RATE_LIMITING:
    try:
        from web.rate_limiting import init_rate_limiter, RateLimits
        rate_limiter = init_rate_limiter(app)
        if rate_limiter:
            logger.info("Rate limiting enabled")
    except Exception as e:
        logger.warning(f"Rate limiting not available: {e}")

# 7.4 Audit Logging
USE_AUDIT_LOGGING = os.environ.get('USE_AUDIT_LOGGING', 'true').lower() == 'true'

if USE_AUDIT_LOGGING:
    try:
        from web.audit_service import setup_audit_logging, audit_service
        setup_audit_logging(app)
        logger.info("Audit logging enabled")
    except Exception as e:
        logger.warning(f"Audit logging not available: {e}")

# =============================================================================
# End Phase 7 Enterprise Features
# =============================================================================

# Configure task manager and session store
from config_v2 import OUTPUT_DIR, DATA_DIR
task_manager.configure(
    results_dir=OUTPUT_DIR / "tasks",
    max_concurrent=2,
    timeout=1800  # 30 minutes
)
session_store.configure(
    storage_dir=DATA_DIR / "sessions",
    timeout_hours=24
)


# Custom Jinja filter to clean up Gap object strings in question text
import re

@app.template_filter('clean_gap_text')
def clean_gap_text(text):
    """
    Clean up Gap object strings that were accidentally serialized into question text.

    Converts: "Can you provide documentation for: Gap(gap_id='G-CYBER-001', ..., description='No explicit...', ...)"
    To: "Can you provide documentation for: No explicit..."
    """
    if not text or 'Gap(' not in text:
        return text

    # Pattern to extract description from Gap object string
    # Handles both single quotes and HTML-escaped quotes (&#39;)
    pattern = r"Gap\([^)]*description=['\"]([^'\"]+)['\"][^)]*\)"
    pattern_escaped = r"Gap\([^)]*description=(?:&#39;|')([^'&]+)(?:&#39;|')[^)]*\)"

    # Try to extract and replace with just the description
    match = re.search(pattern_escaped, text) or re.search(pattern, text)
    if match:
        description = match.group(1)
        # Replace the entire Gap(...) with just the description
        text = re.sub(pattern_escaped, description, text)
        text = re.sub(pattern, description, text)

    return text


def log_activity(action: str, details: dict):
    """Log an activity to the activity log file."""
    from config_v2 import OUTPUT_DIR
    from datetime import datetime
    import json as json_module

    activity_file = OUTPUT_DIR / "activity_log.json"

    # Load existing activities
    activities = []
    if activity_file.exists():
        try:
            with open(activity_file, 'r') as f:
                activities = json_module.load(f)
        except Exception:
            activities = []

    # Add new activity
    activities.append({
        "timestamp": datetime.now().isoformat(),
        "action": action,
        "details": details
    })

    # Keep only last 500 activities
    activities = activities[-500:]

    # Save
    try:
        with open(activity_file, 'w') as f:
            json_module.dump(activities, f, indent=2, default=str)
    except Exception:
        pass


def get_session():
    """Get the analysis session for the current user and current deal.

    IMPORTANT: This function is now DEAL-AWARE. It only loads data for the
    currently active deal, preventing data leakage between deals.
    """
    import logging
    logger = logging.getLogger(__name__)

    session_id = get_or_create_session_id(flask_session)
    user_session = session_store.get_session(session_id)

    # Get the current deal_id - this is critical for data isolation
    current_deal_id = flask_session.get('current_deal_id')

    # Check if cached session matches current deal
    if user_session and user_session.analysis_session:
        cached = user_session.analysis_session
        cached_deal_id = getattr(user_session, '_cached_deal_id', None)

        # If cached session is for a DIFFERENT deal, invalidate it
        if cached_deal_id and cached_deal_id != current_deal_id:
            logger.info(f"Deal mismatch: cached={cached_deal_id}, current={current_deal_id} - clearing cache")
            user_session.analysis_session = None
            user_session._cached_deal_id = None
            clear_organization_cache()
        elif cached.fact_store and len(cached.fact_store.facts) > 0:
            logger.debug(f"Using cached session for deal {current_deal_id} with {len(cached.fact_store.facts)} facts")
            return cached

    # If no deal selected, return empty session
    if not current_deal_id:
        logger.debug("No deal selected - returning empty session")
        analysis_session = session_store.get_or_create_analysis_session(session_id)
        return analysis_session if analysis_session else Session()

    # Try to load from database for the current deal
    try:
        from web.database import db, Fact, Finding, AnalysisRun

        # Get facts for this specific deal from database
        facts_query = Fact.query.filter_by(deal_id=current_deal_id, deleted_at=None).all()
        findings_query = Finding.query.filter_by(deal_id=current_deal_id, deleted_at=None).all()

        if facts_query or findings_query:
            logger.info(f"Loading {len(facts_query)} facts and {len(findings_query)} findings from DB for deal {current_deal_id}")

            # Create session with database data AND deal_id for isolation
            analysis_session = Session(deal_id=current_deal_id)

            # Convert DB facts to session format
            for db_fact in facts_query:
                from stores.fact_store import Fact as FactModel
                fact = FactModel(
                    fact_id=db_fact.id,
                    domain=db_fact.domain,
                    category=db_fact.category or '',
                    item=db_fact.item or '',
                    details=db_fact.details or {},
                    status=db_fact.status or 'documented',
                    evidence=db_fact.evidence or {},
                    entity=db_fact.entity or 'target',
                    analysis_phase=db_fact.analysis_phase or 'target_extraction',
                    is_integration_insight=db_fact.is_integration_insight or False,
                    source_document=db_fact.source_document or '',
                    confidence_score=db_fact.confidence_score or 0.5,
                    verified=db_fact.verified or False,
                    verification_status=db_fact.verification_status or 'pending',
                    needs_review=db_fact.needs_review or False,
                    needs_review_reason=db_fact.needs_review_reason or '',
                    related_domains=db_fact.related_domains or [],
                )
                analysis_session.fact_store.facts.append(fact)

            # Convert DB findings to session format (risks and work items)
            for db_finding in findings_query:
                if db_finding.finding_type == 'risk':
                    from tools_v2.reasoning_tools import Risk
                    risk = Risk(
                        finding_id=db_finding.id,
                        domain=db_finding.domain or 'general',
                        title=db_finding.title or '',
                        description=db_finding.description or '',
                        category=db_finding.category or '',
                        severity=db_finding.severity or 'medium',
                        integration_dependent=db_finding.integration_dependent or False,
                        mitigation=db_finding.mitigation or '',
                        based_on_facts=db_finding.based_on_facts or [],
                        confidence=db_finding.confidence or 'medium',
                        reasoning=db_finding.reasoning or '',
                        mna_lens=db_finding.mna_lens or '',
                        mna_implication=db_finding.mna_implication or '',
                        timeline=db_finding.timeline,
                    )
                    analysis_session.reasoning_store.risks.append(risk)
                elif db_finding.finding_type == 'work_item':
                    from tools_v2.reasoning_tools import WorkItem
                    work_item = WorkItem(
                        finding_id=db_finding.id,
                        domain=db_finding.domain or 'general',
                        title=db_finding.title or '',
                        description=db_finding.description or '',
                        phase=db_finding.phase or 'Day_100',
                        priority=db_finding.priority or 'medium',
                        owner_type=db_finding.owner_type or 'shared',
                        triggered_by=db_finding.based_on_facts or [],  # Use based_on_facts as triggered_by
                        based_on_facts=db_finding.based_on_facts or [],
                        confidence=db_finding.confidence or 'medium',
                        reasoning=db_finding.reasoning or '',
                        cost_estimate=db_finding.cost_estimate or 'unknown',
                        triggered_by_risks=db_finding.triggered_by_risks or [],
                        mna_lens=db_finding.mna_lens or '',
                        mna_implication=db_finding.mna_implication or '',
                        dependencies=db_finding.dependencies or [],
                    )
                    analysis_session.reasoning_store.work_items.append(work_item)

            # Load gaps from database into session
            from web.database import Gap as DbGap
            from stores.fact_store import Gap as GapModel
            gaps_query = DbGap.query.filter_by(deal_id=current_deal_id, deleted_at=None).all()
            logger.debug(f"Loading {len(gaps_query)} gaps from database for deal {current_deal_id}")
            for db_gap in gaps_query:
                gap = GapModel(
                    gap_id=db_gap.id,
                    domain=db_gap.domain or 'general',
                    category=db_gap.category or '',
                    description=db_gap.description or '',
                    importance=db_gap.importance or 'medium',
                    entity=db_gap.entity or 'target',
                )
                analysis_session.fact_store.gaps.append(gap)
            logger.info(f"Loaded {len(gaps_query)} gaps from database")

            # Cache with deal_id tracking
            if user_session:
                user_session.analysis_session = analysis_session
                user_session._cached_deal_id = current_deal_id

            return analysis_session

    except Exception as e:
        logger.warning(f"Error loading from database for deal {current_deal_id}: {e}")
        # Fall through to file-based loading

    # Fallback: Try to load from task results if available
    task_id = flask_session.get('current_task_id')
    if task_id:
        task = task_manager.get_task(task_id)
        if task and task.status == TaskStatus.COMPLETED and task.facts_file:
            logger.info(f"Loading session from task {task_id}: {task.facts_file}")
            analysis_session = session_store.load_session_from_results(
                session_id,
                task.facts_file,
                task.findings_file
            )
            if analysis_session and analysis_session.fact_store and len(analysis_session.fact_store.facts) > 0:
                logger.info(f"Loaded {len(analysis_session.fact_store.facts)} facts from task")
                if user_session:
                    user_session.analysis_session = analysis_session
                    user_session._cached_deal_id = current_deal_id
                return analysis_session
        elif not task:
            flask_session.pop('current_task_id', None)

    logger.debug(f"No data found for deal {current_deal_id} - returning empty session")
    analysis_session = session_store.get_or_create_analysis_session(session_id)
    if analysis_session:
        analysis_session.deal_id = current_deal_id  # Set deal_id on session
    if user_session:
        user_session._cached_deal_id = current_deal_id
    return analysis_session if analysis_session else Session(deal_id=current_deal_id)


@app.route('/')
def welcome():
    """Welcome/landing page."""
    return render_template('welcome.html')


@app.route('/health')
def health_check():
    """Health check endpoint for container orchestration and load balancers.

    Returns service health status and basic diagnostics.
    Used by Railway, Docker, and other deployment platforms.
    """
    from config_v2 import OUTPUT_DIR, DATA_DIR

    health_status = {
        "status": "healthy",
        "version": "1.0.0",
        "service": "it-diligence-agent",
        "checks": {
            "api_key_configured": bool(os.environ.get('ANTHROPIC_API_KEY')),
            "output_dir_exists": OUTPUT_DIR.exists(),
            "data_dir_exists": DATA_DIR.exists(),
            "output_dir_writable": os.access(OUTPUT_DIR, os.W_OK) if OUTPUT_DIR.exists() else False,
        },
        "services": {}
    }

    # Check database (Phase 3)
    if USE_DATABASE:
        try:
            from web.database import db
            db.session.execute(db.text('SELECT 1'))
            health_status["services"]["database"] = "connected"
        except Exception as e:
            health_status["services"]["database"] = f"error: {str(e)}"

    # Check Redis (Phase 4)
    if USE_REDIS_SESSIONS:
        try:
            import redis
            r = redis.from_url(REDIS_URL)
            r.ping()
            health_status["services"]["redis"] = "connected"
        except Exception as e:
            health_status["services"]["redis"] = f"error: {str(e)}"

    # Check Celery (Phase 4)
    if USE_CELERY:
        try:
            from web.celery_app import is_celery_available
            if is_celery_available():
                health_status["services"]["celery"] = "available"
            else:
                health_status["services"]["celery"] = "unavailable"
        except Exception as e:
            health_status["services"]["celery"] = f"error: {str(e)}"

    # Check Multi-tenancy (Phase 6)
    if USE_MULTI_TENANCY:
        try:
            from web.database import Tenant
            tenant_count = Tenant.query.count()
            health_status["services"]["multi_tenancy"] = f"enabled ({tenant_count} tenants)"
        except Exception as e:
            health_status["services"]["multi_tenancy"] = f"error: {str(e)}"

    # Check Enterprise Features (Phase 7)
    if USE_STRUCTURED_LOGGING:
        health_status["services"]["structured_logging"] = "enabled"

    if USE_ERROR_TRACKING:
        try:
            from web.error_tracking import SENTRY_AVAILABLE
            health_status["services"]["error_tracking"] = "enabled" if SENTRY_AVAILABLE else "sdk not installed"
        except Exception:
            health_status["services"]["error_tracking"] = "not configured"

    if USE_RATE_LIMITING:
        try:
            from web.rate_limiting import LIMITER_AVAILABLE
            health_status["services"]["rate_limiting"] = "enabled" if LIMITER_AVAILABLE else "limiter not installed"
        except Exception:
            health_status["services"]["rate_limiting"] = "not configured"

    if USE_AUDIT_LOGGING:
        health_status["services"]["audit_logging"] = "enabled"

    # Determine overall health
    # Note: Always return 200 for Railway health checks - app can function without API key for UI
    critical_checks = [
        health_status["checks"]["output_dir_exists"],
    ]

    if all(critical_checks):
        health_status["status"] = "healthy"
    else:
        health_status["status"] = "degraded"

    # Always return 200 so container stays up
    return jsonify(health_status), 200


@app.route('/upload')
@auth_optional
def upload_documents():
    """Document upload page."""
    # Check if there's an active deal - if so, pass it to the template
    active_deal = None
    try:
        from web.services.deal_service import get_deal_service
        deal_service = get_deal_service()
        active_deal = deal_service.get_active_deal_for_session()
    except Exception as e:
        logger.debug(f"No active deal: {e}")

    return render_template('upload.html', active_deal=active_deal)


@app.route('/upload/process', methods=['POST'])
@csrf.exempt  # Exempt for AJAX file uploads
@auth_optional
def process_upload():
    """Process uploaded documents and run analysis.

    Handles entity-separated uploads (target vs buyer documents).
    Integrates with DocumentStore for hash-based tracking.
    """
    from pathlib import Path
    from config_v2 import DATA_DIR, ensure_directories

    # Ensure directories exist
    ensure_directories()

    # Get current deal_id from session for data isolation
    current_deal_id = flask_session.get('current_deal_id')

    # Initialize DocumentStore with deal_id for proper isolation
    try:
        from stores.document_store import DocumentStore
        doc_store = DocumentStore.get_instance(deal_id=current_deal_id)
    except Exception as e:
        logger.error(f"Failed to initialize DocumentStore: {e}")
        doc_store = None

    # Create legacy upload directory (for compatibility)
    upload_dir = DATA_DIR / 'uploads'
    upload_dir.mkdir(exist_ok=True)

    # Get authority levels from form
    target_authority = int(request.form.get('target_authority', 1))
    buyer_authority = int(request.form.get('buyer_authority', 1))

    # Process TARGET documents
    target_files = request.files.getlist('target_documents')
    target_saved = []
    target_doc_ids = []

    for file in target_files:
        if file.filename:
            safe_filename = file.filename.replace('..', '').replace('/', '_')

            if doc_store:
                # Use DocumentStore for proper entity separation
                try:
                    file_bytes = file.read()
                    doc = doc_store.add_document_from_bytes(
                        file_bytes=file_bytes,
                        filename=safe_filename,
                        entity="target",
                        deal_id=current_deal_id,
                        authority_level=target_authority,
                        uploaded_by="web_upload"
                    )
                    target_doc_ids.append(doc.doc_id)
                    target_saved.append(doc.raw_file_path)
                    logger.info(f"Added target document: {safe_filename} (doc_id: {doc.doc_id})")

                    # A2 FIX: Also insert into database for UI persistence
                    if USE_DATABASE and current_deal_id:
                        try:
                            doc_repo = DocumentRepository()
                            doc_repo.create_document(
                                deal_id=current_deal_id,
                                filename=safe_filename,
                                file_hash=doc.hash_sha256,
                                storage_path=doc.raw_file_path,
                                entity="target",
                                file_size=doc.file_size_bytes,
                                mime_type=doc.mime_type,
                                authority_level=target_authority,
                                uploaded_by="web_upload"
                            )
                            logger.info(f"A2: Saved target document to DB: {safe_filename}")
                        except Exception as db_err:
                            logger.warning(f"A2: Failed to save document to DB: {db_err}")
                except Exception as e:
                    logger.error(f"DocumentStore error for {safe_filename}: {e}")
                    # Fallback to legacy path
                    file.seek(0)
                    filepath = upload_dir / f"target_{safe_filename}"
                    file.save(str(filepath))
                    target_saved.append(str(filepath))
            else:
                # Legacy behavior if DocumentStore not available
                filepath = upload_dir / f"target_{safe_filename}"
                file.save(str(filepath))
                target_saved.append(str(filepath))

    # Process BUYER documents
    buyer_files = request.files.getlist('buyer_documents')
    buyer_saved = []
    buyer_doc_ids = []

    for file in buyer_files:
        if file.filename:
            safe_filename = file.filename.replace('..', '').replace('/', '_')

            if doc_store:
                try:
                    file_bytes = file.read()
                    doc = doc_store.add_document_from_bytes(
                        file_bytes=file_bytes,
                        filename=safe_filename,
                        entity="buyer",
                        deal_id=current_deal_id,
                        authority_level=buyer_authority,
                        uploaded_by="web_upload"
                    )
                    buyer_doc_ids.append(doc.doc_id)
                    buyer_saved.append(doc.raw_file_path)
                    logger.info(f"Added buyer document: {safe_filename} (doc_id: {doc.doc_id})")

                    # A2 FIX: Also insert into database for UI persistence
                    if USE_DATABASE and current_deal_id:
                        try:
                            doc_repo = DocumentRepository()
                            doc_repo.create_document(
                                deal_id=current_deal_id,
                                filename=safe_filename,
                                file_hash=doc.hash_sha256,
                                storage_path=doc.raw_file_path,
                                entity="buyer",
                                file_size=doc.file_size_bytes,
                                mime_type=doc.mime_type,
                                authority_level=buyer_authority,
                                uploaded_by="web_upload"
                            )
                            logger.info(f"A2: Saved buyer document to DB: {safe_filename}")
                        except Exception as db_err:
                            logger.warning(f"A2: Failed to save document to DB: {db_err}")
                except Exception as e:
                    logger.error(f"DocumentStore error for {safe_filename}: {e}")
                    file.seek(0)
                    filepath = upload_dir / f"buyer_{safe_filename}"
                    file.save(str(filepath))
                    buyer_saved.append(str(filepath))
            else:
                filepath = upload_dir / f"buyer_{safe_filename}"
                file.save(str(filepath))
                buyer_saved.append(str(filepath))

    # Validate: must have target documents
    if not target_saved:
        flash('No TARGET documents were uploaded. At least one target document is required.', 'error')
        return redirect(url_for('upload_documents'))

    # Combine for analysis (target first, then buyer)
    all_saved_files = target_saved + buyer_saved

    # Get deal context
    deal_type = request.form.get('deal_type', 'bolt_on')
    target_name = request.form.get('target_name', 'Target Company')
    buyer_name = request.form.get('buyer_name', '')
    industry = request.form.get('industry', '')
    employee_count = request.form.get('employee_count', '')

    # Get selected domains (default to all if none selected)
    domains = request.form.getlist('domains')
    if not domains:
        from config_v2 import DOMAINS
        domains = DOMAINS

    # ==========================================================================
    # DEAL INTEGRATION: Get or create Deal in database
    # ==========================================================================
    deal_id = None
    try:
        from web.services.deal_service import get_deal_service
        deal_service = get_deal_service()

        # Check if there's an active deal in session
        active_deal = deal_service.get_active_deal_for_session()

        if active_deal:
            # Use existing deal
            deal_id = str(active_deal.id)
            logger.info(f"Using existing deal: {deal_id} ({active_deal.name})")
        else:
            # Create new deal from upload form data
            user = get_current_user()
            deal_name = f"{target_name} Analysis" if target_name else "New Analysis"

            new_deal, error = deal_service.create_deal(
                user_id=user.id if user else None,
                name=deal_name,
                target_name=target_name or "Unknown Target",
                buyer_name=buyer_name or None,
                deal_type=deal_type,
                industry=industry or None,
            )

            if new_deal:
                deal_id = str(new_deal.id)
                deal_service.set_active_deal_for_session(deal_id)
                logger.info(f"Created new deal: {deal_id} ({new_deal.name})")
            else:
                logger.warning(f"Failed to create deal: {error}")
    except Exception as e:
        logger.warning(f"Deal integration skipped: {e}")

    # Create deal context dict with entity info
    deal_context = {
        'deal_id': deal_id,  # Database deal ID for persistence
        'deal_type': deal_type,
        'target_name': target_name,
        'buyer_name': buyer_name,
        'industry': industry,
        'employee_count': employee_count,
        # Entity tracking for analysis
        'target_doc_ids': target_doc_ids,
        'buyer_doc_ids': buyer_doc_ids,
        'target_doc_count': len(target_saved),
        'buyer_doc_count': len(buyer_saved),
    }

    # Create analysis task
    task = task_manager.create_task(
        file_paths=all_saved_files,
        domains=domains,
        deal_context=deal_context
    )

    # Store task ID in Flask session
    flask_session['current_task_id'] = task.task_id
    flask_session['analysis_file_count'] = len(all_saved_files)
    flask_session['target_doc_count'] = len(target_saved)
    flask_session['buyer_doc_count'] = len(buyer_saved)

    # Associate task with user session
    session_id = get_or_create_session_id(flask_session)
    session_store.set_analysis_task(session_id, task.task_id)

    # Start background analysis
    # Pass deal_id to enable Celery task execution (when available)
    try:
        task_manager.start_task(task.task_id, run_analysis, deal_id=deal_id)
    except ImportError:
        task_manager.start_task(task.task_id, run_analysis_simple, deal_id=deal_id)

    logger.info(
        f"Started analysis task {task.task_id}: "
        f"{len(target_saved)} target docs, {len(buyer_saved)} buyer docs"
    )

    # Audit log the analysis start
    if USE_AUDIT_LOGGING:
        try:
            from web.audit_service import audit_log, AuditAction
            audit_log(
                action=AuditAction.ANALYSIS_START,
                resource_type='analysis',
                resource_id=task.task_id,
                details={
                    'target_documents': len(target_saved),
                    'buyer_documents': len(buyer_saved),
                    'domains': domains,
                    'target_name': target_name,
                }
            )
        except Exception as e:
            logger.debug(f"Audit logging failed: {e}")

    # Force session save before redirect (fixes race condition with Redis)
    flask_session.modified = True

    # Redirect to processing page with task_id as backup (in case session not ready)
    return redirect(url_for('processing', task_id=task.task_id))


@app.route('/processing')
@auth_optional
def processing():
    """Processing/loading page shown while analysis runs."""
    file_count = flask_session.get('analysis_file_count', 0)

    # Get task_id from session, with query param as fallback (handles Redis race condition)
    task_id = flask_session.get('current_task_id') or request.args.get('task_id')

    # If task_id came from query param, store it in session for subsequent status polls
    if task_id and not flask_session.get('current_task_id'):
        flask_session['current_task_id'] = task_id
        flask_session.modified = True
        logger.info(f"Restored task_id from query param: {task_id}")

    # Get initial task status if available
    initial_status = None
    if task_id:
        initial_status = task_manager.get_task_status(task_id)

    return render_template('processing.html',
                          file_count=file_count,
                          task_id=task_id,
                          initial_status=initial_status)


@app.route('/analysis/status')
@auth_optional
def analysis_status():
    """Check analysis progress. Returns JSON for polling."""
    # Get task_id from session, with query param as fallback (handles Redis race condition)
    task_id = flask_session.get('current_task_id') or request.args.get('task_id')

    # If task_id came from query param, store it in session for subsequent polls
    if task_id and not flask_session.get('current_task_id'):
        flask_session['current_task_id'] = task_id
        flask_session.modified = True

    if not task_id:
        # No task_id - this is an error state, don't pretend old results are valid
        # The task_id should come from the URL query param (set when analysis starts)
        logger.warning("analysis_status called without task_id - possible session/Redis issue")
        return jsonify({
            'complete': False,
            'success': False,
            'status': 'no_task',
            'message': 'No analysis task found. Please start a new analysis.',
            'error': 'missing_task_id'
        })

    # Get task status from task manager
    status = task_manager.get_task_status(task_id)

    if not status:
        # Task not found in task_manager - this means:
        # 1. Task completed and was cleaned up, OR
        # 2. Server restarted and lost in-memory task state
        # Check if there's data in the database for this deal
        logger.warning(f"Task {task_id} not found in task_manager - checking database for results")
        flask_session.pop('current_task_id', None)

        # Check if we have data in the database for the current deal
        current_deal_id = flask_session.get('current_deal_id')
        if current_deal_id:
            try:
                from web.database import Fact, Finding
                facts_count = Fact.query.filter_by(deal_id=current_deal_id, deleted_at=None).count()
                findings_count = Finding.query.filter_by(deal_id=current_deal_id, deleted_at=None).count()

                if facts_count > 0 or findings_count > 0:
                    # Data exists in database - analysis likely completed before restart
                    logger.info(f"Found {facts_count} facts and {findings_count} findings in DB for deal {current_deal_id}")
                    return jsonify({
                        'complete': True,
                        'success': True,
                        'status': 'recovered',
                        'message': f'Analysis data recovered: {facts_count} facts, {findings_count} findings',
                        'redirect': '/dashboard'
                    })
            except Exception as e:
                logger.warning(f"Error checking database for deal {current_deal_id}: {e}")

        # No data found - analysis was interrupted
        return jsonify({
            'complete': False,
            'success': False,
            'status': 'task_not_found',
            'message': 'Analysis was interrupted (server restart). Please start a new analysis.',
            'error': 'task_not_found'
        })

    # If completed successfully, load results into session and invalidate all caches
    if status['complete'] and status['success']:
        session_id = get_or_create_session_id(flask_session)
        if status.get('facts_file'):
            # First invalidate all caches to ensure fresh data
            invalidate_all_caches()

            # Then load the new results
            session_store.load_session_from_results(
                session_id,
                status['facts_file'],
                status.get('findings_file')
            )

    return jsonify(status)


@app.route('/analysis/cancel', methods=['POST'])
@auth_optional
def cancel_analysis():
    """Cancel a running analysis."""
    task_id = flask_session.get('current_task_id')

    if not task_id:
        return jsonify({'success': False, 'message': 'No task to cancel'})

    success = task_manager.cancel_task(task_id)

    return jsonify({
        'success': success,
        'message': 'Task cancelled' if success else 'Could not cancel task'
    })


@app.route('/api/analysis/diff')
@auth_optional
def get_analysis_diff():
    """
    Get the diff for a specific run or the latest diff for current deal.

    Phase 1a: Analysis Reasoning Layer - What Changed.

    Query params:
        run_id: Specific run ID to get diff for
        deal_id: Override deal (defaults to session deal)

    Returns:
        JSON with diff metrics and summary
    """
    from services.run_diff_service import run_diff_service

    run_id = request.args.get('run_id')
    deal_id = request.args.get('deal_id') or flask_session.get('current_deal_id')

    if not deal_id:
        return jsonify({'error': 'No deal selected'}), 400

    try:
        if run_id:
            diff = run_diff_service.get_diff_for_run(run_id)
        else:
            diff = run_diff_service.get_latest_diff(deal_id)

        if not diff:
            return jsonify({
                'found': False,
                'message': 'No diff available (first run or diff not generated)'
            })

        return jsonify({
            'found': True,
            'diff': diff.to_dict(),
            'summary': diff.summary_text()
        })

    except Exception as e:
        logger.error(f"Failed to get analysis diff: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/analysis/diffs')
@auth_optional
def get_analysis_diffs():
    """
    Get all diffs for the current deal.

    Phase 1a: Analysis Reasoning Layer.

    Query params:
        deal_id: Override deal (defaults to session deal)
        limit: Max number of diffs to return (default 10)

    Returns:
        JSON array of diffs
    """
    from services.run_diff_service import run_diff_service

    deal_id = request.args.get('deal_id') or flask_session.get('current_deal_id')
    limit = request.args.get('limit', 10, type=int)

    if not deal_id:
        return jsonify({'error': 'No deal selected'}), 400

    try:
        diffs = run_diff_service.get_diffs_for_deal(deal_id, limit=limit)
        return jsonify({
            'count': len(diffs),
            'diffs': [d.to_dict() for d in diffs]
        })

    except Exception as e:
        logger.error(f"Failed to get analysis diffs: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/facts/<fact_id>/reasoning')
@auth_optional
def get_fact_reasoning(fact_id):
    """
    Get the reasoning for a specific fact.

    Phase 1b: Analysis Reasoning Layer - Why This Matters.

    Returns:
        JSON with reasoning_text, source_refs, signal_type
    """
    from services.fact_reasoning_service import fact_reasoning_service

    try:
        reasoning = fact_reasoning_service.get_reasoning_for_fact(fact_id)
        if not reasoning:
            return jsonify({
                'found': False,
                'message': 'No reasoning available for this fact'
            })

        return jsonify({
            'found': True,
            'reasoning': reasoning
        })

    except Exception as e:
        logger.error(f"Failed to get fact reasoning: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/facts/reasoning')
@auth_optional
def get_all_fact_reasoning():
    """
    Get all fact reasoning for the current deal.

    Phase 1b: Analysis Reasoning Layer.

    Query params:
        deal_id: Override deal (defaults to session deal)
        signal_type: Filter by signal type

    Returns:
        JSON array of reasoning records
    """
    from services.fact_reasoning_service import fact_reasoning_service

    deal_id = request.args.get('deal_id') or flask_session.get('current_deal_id')
    signal_type = request.args.get('signal_type')

    if not deal_id:
        return jsonify({'error': 'No deal selected'}), 400

    try:
        reasonings = fact_reasoning_service.get_reasoning_for_deal(deal_id)

        # Filter by signal type if specified
        if signal_type:
            reasonings = [r for r in reasonings if r.get('signal_type') == signal_type]

        return jsonify({
            'count': len(reasonings),
            'reasonings': reasonings
        })

    except Exception as e:
        logger.error(f"Failed to get fact reasoning: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/dashboard')
@auth_optional
def dashboard():
    """Main dashboard view.

    Phase 2+: Database-first implementation (no fallback).
    """
    from config_v2 import OUTPUT_DIR
    import json as json_module

    # Phase 2+: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view the dashboard.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        db_data = DealData()

        # Build summary from database
        raw_summary = db_data.get_dashboard_summary()
        cost_summary = db_data.get_cost_summary_by_phase()
        summary = {
            'facts': sum(raw_summary.get('fact_counts', {}).values()),
            'gaps': sum(raw_summary.get('gap_counts', {}).values()),
            'risks': raw_summary.get('risk_summary', {}).get('total', 0),
            'work_items': raw_summary.get('work_item_summary', {}).get('total', 0),
            # Include full risk_summary for template access
            'risk_summary': raw_summary.get('risk_summary', {
                'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'total': 0
            }),
            # Include cost summary by phase for work items table
            'cost_summary': cost_summary,
        }
        top_risks = db_data.get_top_risks(5)
        day1_items = [w for w in db_data.get_work_items() if w.phase == 'Day_1'][:5]

        # Get analysis metadata from database
        analysis_metadata = None
        run = db_data.get_analysis_run()
        if run:
            analysis_metadata = {
                'run_id': run.id,
                'run_number': run.run_number,
                'timestamp': run.completed_at.strftime('%Y-%m-%d') if run.completed_at else None,
                'domains': run.domains or [],
                'source': 'database',
                'fact_count': run.facts_created or 0,
                'finding_count': run.findings_created or 0,
            }

        # Get pending changes summary for incremental updates
        pending_summary = {"tier1": 0, "tier2": 0, "tier3": 0, "total": 0}
        pending_file = OUTPUT_DIR / "pending_changes.json"
        if pending_file.exists():
            try:
                with open(pending_file, 'r') as f:
                    pending = json_module.load(f)
                pending_summary = {
                    "tier1": len(pending.get("tier1", [])),
                    "tier2": len(pending.get("tier2", [])),
                    "tier3": len(pending.get("tier3", [])),
                    "total": len(pending.get("tier1", [])) + len(pending.get("tier2", [])) + len(pending.get("tier3", []))
                }
            except Exception:
                pass

        # Get entity summary (target vs buyer breakdown) from database
        def domain_breakdown(facts):
            breakdown = {}
            for f in facts:
                domain = getattr(f, 'domain', 'unknown')
                breakdown[domain] = breakdown.get(domain, 0) + 1
            return breakdown

        entity_summary = None
        all_facts = db_data.get_all_facts()
        if all_facts:
            target_facts = [f for f in all_facts if getattr(f, 'entity', 'target') == 'target']
            buyer_facts = [f for f in all_facts if getattr(f, 'entity', '') == 'buyer']

            # Get document counts from DocumentStore
            target_doc_count = 0
            buyer_doc_count = 0
            try:
                from stores.document_store import DocumentStore
                doc_store = DocumentStore.get_instance(deal_id=current_deal_id)
                stats = doc_store.get_statistics()
                target_doc_count = stats["by_entity"]["target"]
                buyer_doc_count = stats["by_entity"]["buyer"]
            except Exception:
                pass

            entity_summary = {
                "target": {
                    "fact_count": len(target_facts),
                    "document_count": target_doc_count,
                    "by_domain": domain_breakdown(target_facts)
                },
                "buyer": {
                    "fact_count": len(buyer_facts),
                    "document_count": buyer_doc_count,
                    "by_domain": domain_breakdown(buyer_facts)
                }
            }

        # Get inventory summary
        inventory_summary = None
        try:
            from web.blueprints.inventory import get_inventory_store
            inv_store = get_inventory_store()
            if len(inv_store) > 0:
                apps = inv_store.get_items(inventory_type="application", entity="target", status="active")
                inventory_summary = {
                    "total": len(inv_store),
                    "applications": len(apps),
                    "total_cost": sum(app.cost or 0 for app in apps),
                    "critical": len([a for a in apps if a.criticality and 'critical' in str(a.criticality).lower()]),
                }
        except Exception:
            pass

        logger.debug(f"Dashboard: Loaded data from database for deal {current_deal_id}")

        # Get latest run diff (Phase 1a: What Changed)
        run_diff = None
        try:
            from services.run_diff_service import run_diff_service
            diff = run_diff_service.get_latest_diff(current_deal_id)
            if diff:
                run_diff = {
                    'run_number': diff.current_run_number,
                    'previous_run_number': diff.previous_run_number,
                    'facts_created': diff.facts_created,
                    'facts_updated': diff.facts_updated,
                    'facts_deleted': diff.facts_deleted,
                    'facts_total_before': diff.facts_total_before,
                    'facts_total_after': diff.facts_total_after,
                    'risks_created': diff.risks_created,
                    'risks_updated': diff.risks_updated,
                    'severity_changes': diff.severity_changes or [],
                    'docs_added': diff.docs_added or [],
                    'entity_scope_after': diff.entity_scope_after,
                    'has_changes': (diff.facts_created > 0 or diff.facts_updated > 0 or
                                   diff.risks_created > 0 or diff.risks_updated > 0)
                }
        except Exception as e:
            logger.warning(f"Failed to load run diff: {e}")

    except Exception as e:
        logger.error(f"Dashboard: Database path failed: {e}")
        flash('Error loading dashboard. Please try again.', 'error')
        summary = {
            'facts': 0,
            'gaps': 0,
            'risks': 0,
            'work_items': 0,
            'risk_summary': {
                'critical': 0,
                'high': 0,
                'medium': 0,
                'low': 0,
                'total': 0
            },
        }
        top_risks = []
        day1_items = []
        analysis_metadata = None
        pending_summary = {"tier1": 0, "tier2": 0, "tier3": 0, "total": 0}
        entity_summary = None
        inventory_summary = None
        run_diff = None

    return render_template('dashboard.html',
                         summary=summary,
                         top_risks=top_risks,
                         day1_items=day1_items,
                         analysis_metadata=analysis_metadata,
                         pending_summary=pending_summary,
                         entity_summary=entity_summary,
                         inventory_summary=inventory_summary,
                         run_diff=run_diff,
                         session=None)


@app.route('/risks')
@auth_optional
def risks():
    """List all risks with pagination.

    Phase 2+: Database-first implementation using DealData.
    """
    # Get query parameters
    severity_filter = request.args.get('severity', '')
    domain_filter = request.args.get('domain', '')
    search_query = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view risks.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        # Get paginated risks with filters (SQL-level filtering)
        paginated_risks, total = data.get_findings_paginated(
            finding_type='risk',
            domain=domain_filter or None,
            severity=severity_filter or None,
            search=search_query or None,
            page=page,
            per_page=per_page,
            order_by_severity=True
        )

        # Calculate pagination
        total_pages = (total + per_page - 1) // per_page

        # Get all risks for domain dropdown (unfiltered count)
        all_risks = data.get_risks()
        domains = sorted(list(set(r.domain for r in all_risks if r.domain)))

        logger.debug(f"Risks route: Loaded {len(paginated_risks)} risks from database (total: {total})")

    except Exception as e:
        logger.error(f"Risks route: Database path failed: {e}")
        flash('Error loading risks. Please try again.', 'error')
        paginated_risks = []
        all_risks = []
        domains = []
        total = 0
        total_pages = 0

    return render_template('risks.html',
                         risks=paginated_risks,
                         all_risks=all_risks,
                         domains=domains,
                         severity_filter=severity_filter,
                         domain_filter=domain_filter,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         per_page=per_page)


@app.route('/risk/<risk_id>')
@auth_optional
def risk_detail(risk_id):
    """Risk detail view with connected items.

    Phase 2+: Database-first implementation.
    """
    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view risk details.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.repositories import FindingRepository, FactRepository
        from web.context import load_deal_context

        load_deal_context(current_deal_id)

        finding_repo = FindingRepository()
        fact_repo = FactRepository()

        risk = finding_repo.get_by_id(risk_id)
        if not risk or risk.finding_type != 'risk':
            flash(f'Risk {risk_id} not found', 'error')
            return redirect(url_for('risks'))

        # Get supporting facts
        facts = []
        for fact_id in (risk.based_on_facts or []):
            fact = fact_repo.get_by_id(fact_id)
            if fact:
                facts.append(fact)

        # Get related work items (those triggered by this risk)
        related_wi = []
        all_work_items = finding_repo.get_work_items(current_deal_id)
        for wi in all_work_items:
            triggered_risks = getattr(wi, 'triggered_by_risks', None) or []
            if risk_id in triggered_risks:
                related_wi.append(wi)

    except Exception as e:
        logger.error(f"Risk detail: Database path failed: {e}")
        flash('Error loading risk details. Please try again.', 'error')
        return redirect(url_for('risks'))

    # Find affected inventory items based on fact item names
    affected_items = []
    affected_cost_total = 0
    affected_users_total = 0
    try:
        from web.blueprints.inventory import get_inventory_store
        inv_store = get_inventory_store()
        if len(inv_store) > 0:
            all_items = inv_store.get_items(entity="target", status="active")
            # Match fact items to inventory items by name
            fact_item_names = [f.item.lower() for f in facts if f.item]
            for item in all_items:
                item_name_lower = item.name.lower()
                for fact_name in fact_item_names:
                    # Fuzzy match: fact item contains inventory name or vice versa
                    if fact_name in item_name_lower or item_name_lower in fact_name:
                        if item not in affected_items:
                            affected_items.append(item)
                            # Accumulate cost
                            if item.cost:
                                affected_cost_total += item.cost
                            # Accumulate users (parse from string like "1,407")
                            users_str = item.data.get('users', '0')
                            if users_str:
                                try:
                                    users_val = int(str(users_str).replace(',', ''))
                                    affected_users_total += users_val
                                except (ValueError, TypeError):
                                    pass
                        break
    except Exception as e:
        logger.warning(f"Could not load affected inventory items: {e}")

    # Extract M&A-specific context from risk if available
    mna_context = {
        'lens': getattr(risk, 'mna_lens', None),
        'implication': getattr(risk, 'mna_implication', None),
        'timeline': getattr(risk, 'timeline', None),
        'confidence': getattr(risk, 'confidence', None),
    }

    # MVP Enhancement: Extract reasoning explicitly for template access
    risk_reasoning = getattr(risk, 'reasoning', None) or ''
    risk_mitigation = getattr(risk, 'mitigation', None) or ''
    has_cost_data = getattr(risk, 'cost_buildup_json', None) is not None

    return render_template('risk_detail.html',
                         risk=risk,
                         facts=facts,
                         related_work_items=related_wi,
                         affected_items=affected_items,
                         affected_cost_total=affected_cost_total,
                         affected_users_total=affected_users_total,
                         mna_context=mna_context,
                         risk_reasoning=risk_reasoning,
                         risk_mitigation=risk_mitigation,
                         has_cost_data=has_cost_data)


@app.route('/risk/<risk_id>/adjust', methods=['POST'])
@auth_optional
def adjust_risk(risk_id):
    """Adjust a risk's severity.

    Phase 2+: Database-first implementation.
    """
    new_severity = request.form.get('severity')
    if not new_severity:
        flash('No severity provided', 'error')
        return redirect(url_for('risk_detail', risk_id=risk_id))

    try:
        from web.repositories import FindingRepository
        from web.database import db

        repo = FindingRepository()
        risk = repo.get_by_id(risk_id)

        if risk and risk.finding_type == 'risk':
            risk.severity = new_severity
            db.session.commit()
            flash(f'Updated {risk_id} severity to {new_severity}', 'success')
        else:
            flash(f'Risk {risk_id} not found', 'error')

    except Exception as e:
        logger.error(f"Adjust risk failed: {e}")
        flash(f'Failed to update {risk_id}', 'error')

    return redirect(url_for('risk_detail', risk_id=risk_id))


@app.route('/risk/<risk_id>/note', methods=['POST'])
@auth_optional
def add_risk_note(risk_id):
    """Add a note to a risk.

    Phase 2+: Database-first implementation.
    """
    note_text = request.form.get('note')
    if not note_text:
        return redirect(url_for('risk_detail', risk_id=risk_id))

    try:
        from web.repositories import FindingRepository
        from web.database import db
        from datetime import datetime

        repo = FindingRepository()
        risk = repo.get_by_id(risk_id)

        if risk and risk.finding_type == 'risk':
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            old_reasoning = risk.reasoning or ''
            risk.reasoning = f"{old_reasoning}\n\n[Note {timestamp}]: {note_text}"
            db.session.commit()
            flash(f'Note added to {risk_id}', 'success')
        else:
            flash(f'Risk {risk_id} not found', 'error')

    except Exception as e:
        logger.error(f"Add risk note failed: {e}")
        flash(f'Failed to add note to {risk_id}', 'error')

    return redirect(url_for('risk_detail', risk_id=risk_id))


@app.route('/work-items')
@auth_optional
def work_items():
    """List all work items with pagination.

    Phase 2+: Database-first implementation using DealData.
    """
    phase_filter = request.args.get('phase', '')
    domain_filter = request.args.get('domain', '')
    search_query = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view work items.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        # Get all work items for totals calculation (unfiltered by page)
        all_work_items = data.get_work_items()

        # Apply domain and search filters manually for totals (phase-independent)
        filtered_items = all_work_items
        if domain_filter:
            filtered_items = [w for w in filtered_items if w.domain == domain_filter]
        if search_query:
            sq = search_query.lower()
            filtered_items = [w for w in filtered_items
                            if sq in (w.title or '').lower()
                            or sq in (w.description or '').lower()]

        # Group by phase for totals
        by_phase = {'Day_1': [], 'Day_100': [], 'Post_100': []}
        for wi in filtered_items:
            if wi.phase in by_phase:
                by_phase[wi.phase].append(wi)

        # Calculate totals
        totals = {}
        for phase_name, items_list in by_phase.items():
            low = sum(COST_RANGE_VALUES.get(w.cost_estimate, {}).get('low', 0) for w in items_list)
            high = sum(COST_RANGE_VALUES.get(w.cost_estimate, {}).get('high', 0) for w in items_list)
            totals[phase_name] = {'low': low, 'high': high, 'count': len(items_list)}

        # Get paginated items with all filters
        paginated_items, total = data.get_findings_paginated(
            finding_type='work_item',
            domain=domain_filter or None,
            phase=phase_filter or None,
            search=search_query or None,
            page=page,
            per_page=per_page
        )

        total_pages = (total + per_page - 1) // per_page

        # Re-group paginated items for display
        paginated_by_phase = {'Day_1': [], 'Day_100': [], 'Post_100': []}
        for wi in paginated_items:
            if wi.phase in paginated_by_phase:
                paginated_by_phase[wi.phase].append(wi)

        # Get domains for dropdown
        domains = sorted(list(set(w.domain for w in all_work_items if w.domain)))

        logger.debug(f"Work items route: Loaded {len(paginated_items)} items from database (total: {total})")

    except Exception as e:
        logger.error(f"Work items route: Database path failed: {e}")
        flash('Error loading work items. Please try again.', 'error')
        paginated_by_phase = {'Day_1': [], 'Day_100': [], 'Post_100': []}
        by_phase = {'Day_1': [], 'Day_100': [], 'Post_100': []}
        totals = {p: {'low': 0, 'high': 0, 'count': 0} for p in ['Day_1', 'Day_100', 'Post_100']}
        domains = []
        total = 0
        total_pages = 0

    return render_template('work_items.html',
                         by_phase=paginated_by_phase,
                         all_by_phase=by_phase,
                         totals=totals,
                         domains=domains,
                         phase_filter=phase_filter,
                         domain_filter=domain_filter,
                         cost_ranges=COST_RANGE_VALUES,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         per_page=per_page)


@app.route('/work-item/<wi_id>')
@auth_optional
def work_item_detail(wi_id):
    """Work item detail view.

    Phase 2+: Database-first implementation.
    """
    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view work item details.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.repositories import FindingRepository, FactRepository
        from web.context import load_deal_context

        load_deal_context(current_deal_id)

        finding_repo = FindingRepository()
        fact_repo = FactRepository()

        wi = finding_repo.get_by_id(wi_id)
        if not wi or wi.finding_type != 'work_item':
            flash(f'Work item {wi_id} not found', 'error')
            return redirect(url_for('work_items'))

        # Get triggering facts
        facts = []
        triggered_by = getattr(wi, 'triggered_by', None) or getattr(wi, 'based_on_facts', None) or []
        for fact_id in triggered_by:
            fact = fact_repo.get_by_id(fact_id)
            if fact:
                facts.append(fact)

        # Get triggering risks
        risks = []
        triggered_risks = getattr(wi, 'triggered_by_risks', None) or []
        for risk_id in triggered_risks:
            risk = finding_repo.get_by_id(risk_id)
            if risk and risk.finding_type == 'risk':
                risks.append(risk)

        cost_range = COST_RANGE_VALUES.get(wi.cost_estimate, {'low': 0, 'high': 0})

    except Exception as e:
        logger.error(f"Work item detail: Database path failed: {e}")
        flash('Error loading work item details. Please try again.', 'error')
        return redirect(url_for('work_items'))

    return render_template('work_item_detail.html',
                         wi=wi,
                         facts=facts,
                         risks=risks,
                         cost_range=cost_range)


@app.route('/open-questions')
@auth_optional
def open_questions():
    """Display open questions for the deal team."""
    from pathlib import Path
    from config_v2 import OUTPUT_DIR
    import json

    # Get deal context for display
    s = get_session()
    deal_context = s.deal_context if hasattr(s, 'deal_context') else {}

    # A3 FIX: Get current deal_id and filter questions by deal
    current_deal_id = flask_session.get('current_deal_id')

    # Find open questions file for THIS deal only
    questions = []
    if current_deal_id:
        # Look for deal-specific questions first (new format with deal_id)
        deal_questions_files = sorted(
            OUTPUT_DIR.glob(f"open_questions_{current_deal_id}_*.json"),
            reverse=True
        )
        if deal_questions_files:
            try:
                with open(deal_questions_files[0]) as f:
                    questions = json.load(f)
                logger.info(f"A3: Loaded questions from {deal_questions_files[0].name}")
            except Exception as e:
                flash(f'Error loading open questions: {e}', 'error')
    else:
        # No deal selected - try legacy format (for backwards compatibility)
        questions_files = sorted(OUTPUT_DIR.glob("open_questions_*.json"), reverse=True)
        # Filter out deal-specific files (they have UUID pattern)
        legacy_files = [f for f in questions_files if len(f.stem.split('_')) == 3]  # open_questions_TIMESTAMP
        if legacy_files:
            try:
                with open(legacy_files[0]) as f:
                    questions = json.load(f)
            except Exception as e:
                flash(f'Error loading open questions: {e}', 'error')

    # If no file exists, generate questions based on current context
    if not questions and deal_context.get('industry'):
        try:
            from tools_v2.open_questions import generate_open_questions_for_deal
            questions = generate_open_questions_for_deal(
                industry=deal_context.get('industry'),
                sub_industry=deal_context.get('sub_industry'),
                deal_type=deal_context.get('deal_type', 'bolt_on'),
                gaps=list(s.fact_store.gaps) if hasattr(s, 'fact_store') else []
            )
        except Exception as e:
            flash(f'Error generating questions: {e}', 'error')

    # Group questions by priority
    by_priority = {'critical': [], 'important': [], 'nice_to_have': []}
    for q in questions:
        priority = q.get('priority', 'important')
        if priority in by_priority:
            by_priority[priority].append(q)

    # Group by category for alternative view
    by_category = {}
    for q in questions:
        cat = q.get('category', 'operational')
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(q)

    return render_template('open_questions.html',
                         questions=questions,
                         by_priority=by_priority,
                         by_category=by_category,
                         deal_context=deal_context,
                         total=len(questions))


@app.route('/work-item/<wi_id>/adjust', methods=['POST'])
@auth_optional
def adjust_work_item(wi_id):
    """Adjust a work item.

    Phase 2+: Database-first implementation.
    """
    field = request.form.get('field')
    value = request.form.get('value')

    if not field or not value:
        flash('Missing field or value', 'error')
        return redirect(url_for('work_item_detail', wi_id=wi_id))

    try:
        from web.repositories import FindingRepository
        from web.database import db

        repo = FindingRepository()
        wi = repo.get_by_id(wi_id)

        if wi and wi.finding_type == 'work_item':
            # Update the field dynamically
            if hasattr(wi, field):
                setattr(wi, field, value)
                db.session.commit()
                flash(f'Updated {wi_id} {field} to {value}', 'success')
            else:
                flash(f'Invalid field: {field}', 'error')
        else:
            flash(f'Work item {wi_id} not found', 'error')

    except Exception as e:
        logger.error(f"Adjust work item failed: {e}")
        flash(f'Failed to update {wi_id}', 'error')

    return redirect(url_for('work_item_detail', wi_id=wi_id))


@app.route('/facts')
@auth_optional
def facts():
    """List all facts with pagination.

    Phase 2+: Database-first implementation (no fallback).
    """
    domain_filter = request.args.get('domain', '')
    category_filter = request.args.get('category', '')
    # B1 FIX: Default to 'target' entity for proper buyer/target separation
    # BUT: When searching for specific fact ID, search across all entities
    search_query = request.args.get('q', '').lower()
    if search_query:
        # If searching (e.g., from fact link in risk drawer), ignore entity filter
        entity_filter = request.args.get('entity', '')  # Allow all entities when searching
    else:
        # Normal browsing defaults to target
        entity_filter = request.args.get('entity', 'target')

    page = request.args.get('page', 1, type=int)
    per_page = 50

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view facts.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        # Use repository pagination with entity filtering (filtering done in SQL)
        paginated_facts, total = data.get_facts_paginated(
            domain=domain_filter or None,
            entity=entity_filter or None,  # Filter by entity at SQL level (None = all entities)
            status=None,
            search=search_query or None,
            page=page,
            per_page=per_page
        )

        # Apply category filter (may need to add to repository later)
        if category_filter:
            paginated_facts = [f for f in paginated_facts if f.category == category_filter]

        total_pages = (total + per_page - 1) // per_page

        # Get unique domains/categories from all facts for filter dropdowns
        all_facts = data.get_all_facts()  # Keep getting all for filter dropdowns
        domains = list(set(f.domain for f in all_facts))
        categories = list(set(f.category for f in all_facts if f.category))

        return render_template('facts.html',
                             facts=paginated_facts,
                             all_facts=all_facts,
                             domains=domains,
                             categories=categories,
                             domain_filter=domain_filter,
                             category_filter=category_filter,
                             entity_filter=entity_filter,
                             search_query=search_query,
                             page=page,
                             per_page=per_page,
                             total_pages=total_pages,
                             total=total,
                             data_source='database')
    except Exception as e:
        logger.error(f"Facts page: Database path failed: {e}")
        flash('Error loading facts. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/fact/<fact_id>')
@auth_optional
def fact_detail(fact_id):
    """Fact detail view with items that cite it.

    Phase 2+: Database-first implementation.
    """
    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view fact details.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.repositories import FactRepository, FindingRepository
        from web.context import load_deal_context

        load_deal_context(current_deal_id)

        fact_repo = FactRepository()
        finding_repo = FindingRepository()

        fact = fact_repo.get_by_id(fact_id)
        if not fact:
            flash(f'Fact {fact_id} not found', 'error')
            return redirect(url_for('facts'))

        # Find risks citing this fact
        citing_risks = []
        all_risks = finding_repo.get_risks(current_deal_id)
        for risk in all_risks:
            based_on = getattr(risk, 'based_on_facts', None) or []
            if fact_id in based_on:
                citing_risks.append(risk)

        # Find work items citing this fact
        citing_wi = []
        all_work_items = finding_repo.get_work_items(current_deal_id)
        for wi in all_work_items:
            triggered_by = getattr(wi, 'triggered_by', None) or getattr(wi, 'based_on_facts', None) or []
            if fact_id in triggered_by:
                citing_wi.append(wi)

    except Exception as e:
        logger.error(f"Fact detail: Database path failed: {e}")
        flash('Error loading fact details. Please try again.', 'error')
        return redirect(url_for('facts'))

    return render_template('fact_detail.html',
                         fact=fact,
                         citing_risks=citing_risks,
                         citing_work_items=citing_wi)


@app.route('/gaps')
@auth_optional
def gaps():
    """List all gaps with pagination.

    Phase 2+: Database-first implementation (no fallback).
    """
    domain_filter = request.args.get('domain', '')
    importance_filter = request.args.get('importance', '')
    search_query = request.args.get('q', '').lower()
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view gaps.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.repositories import GapRepository
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        repo = GapRepository()

        gaps_list = repo.get_by_deal(
            deal_id=current_deal_id,
            domain=domain_filter if domain_filter else None,
            importance=importance_filter if importance_filter else None
        )
        domains = repo.get_domains(current_deal_id)

        # Apply search filter
        if search_query:
            gaps_list = [g for g in gaps_list if search_query in (g.description or '').lower()]

        # Sort by importance
        importance_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        gaps_list = sorted(gaps_list, key=lambda g: importance_order.get(g.importance, 4))

        # Pagination
        total = len(gaps_list)
        total_pages = (total + per_page - 1) // per_page
        start = (page - 1) * per_page
        end = start + per_page
        paginated_gaps = gaps_list[start:end]

        logger.debug(f"Gaps route: Loaded {len(paginated_gaps)} gaps from database (total: {total})")

    except Exception as e:
        logger.error(f"Gaps route: Database path failed: {e}")
        flash('Error loading gaps. Please try again.', 'error')
        paginated_gaps = []
        gaps_list = []
        domains = []
        total = 0
        total_pages = 0

    return render_template('gaps.html',
                         gaps=paginated_gaps,
                         all_gaps=gaps_list,
                         domains=domains,
                         domain_filter=domain_filter,
                         importance_filter=importance_filter,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         per_page=per_page)


@app.route('/export-vdr')
@auth_optional
def export_vdr():
    """Export VDR request list.

    Phase 2+: Database-first implementation.
    """
    from datetime import datetime
    from config_v2 import OUTPUT_DIR

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to export VDR requests.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()
        gaps = data.get_gaps()
    except Exception as e:
        logger.error(f"Export VDR: Database path failed: {e}")
        flash('Error loading gaps. Please try again.', 'error')
        return redirect(url_for('gaps'))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Generate VDR markdown
    vdr_content = "# VDR Request List\n\n"
    for gap in gaps:
        importance = getattr(gap, 'importance', 'medium') or 'medium'
        gap_id = getattr(gap, 'id', '') or getattr(gap, 'gap_id', '')
        vdr_content += f"## [{importance.upper()}] {gap_id}\n"
        vdr_content += f"**Domain:** {gap.domain}\n"
        vdr_content += f"**Category:** {gap.category}\n"
        vdr_content += f"**Description:** {gap.description}\n\n"

    vdr_file = OUTPUT_DIR / f"vdr_requests_{timestamp}.md"
    with open(vdr_file, 'w') as f:
        f.write(vdr_content)

    flash(f'VDR requests exported to {vdr_file.name}', 'success')
    return redirect(url_for('gaps'))


@app.route('/context', methods=['GET', 'POST'])
@auth_optional
def context():
    """Manage deal context.

    Phase 2+: Database-first implementation.
    Stores context in Deal.context JSON field.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to manage context.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.repositories import DealRepository
        from web.database import db

        repo = DealRepository()
        deal = repo.get_by_id(current_deal_id)

        if not deal:
            flash('Deal not found.', 'error')
            return redirect(url_for('deals.deals_list_page'))

        # Ensure context is a dict
        deal_context = deal.context or {}
        if not isinstance(deal_context, dict):
            deal_context = {}

        context_notes = deal_context.get('notes', [])
        if not isinstance(context_notes, list):
            context_notes = []

        if request.method == 'POST':
            action = request.form.get('action')
            if action == 'add':
                text = request.form.get('context_text')
                if text:
                    # Add new context note
                    context_notes.append({
                        'text': text,
                        'added_at': datetime.now().isoformat()
                    })
                    deal_context['notes'] = context_notes
                    deal.context = deal_context
                    db.session.commit()
                    flash('Deal context added', 'success')
            elif action == 'clear':
                deal_context['notes'] = []
                deal.context = deal_context
                db.session.commit()
                flash('Deal context cleared', 'success')
            return redirect(url_for('context'))

        # Format context for display (backward compatible with template)
        display_context = '\n\n'.join([
            note.get('text', '') for note in context_notes
        ]) if context_notes else ''

    except Exception as e:
        logger.error(f"Context page: Database path failed: {e}")
        flash('Error managing context. Please try again.', 'error')
        return redirect(url_for('dashboard'))

    return render_template('context.html', deal_context=display_context)


@app.route('/export', methods=['POST'])
@csrf.exempt  # Form submission
@auth_optional
def export():
    """Export current session data to files.

    Phase 2+: Database-first implementation.
    Exports data from database to JSON files.
    """
    import json
    from datetime import datetime
    from config_v2 import OUTPUT_DIR

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to export.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        gaps = data.get_gaps()
        risks = data.get_risks()
        work_items = data.get_work_items()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        saved_files = []

        # Export facts (include entity for filtering)
        facts_data = [{
            'id': getattr(f, 'id', '') or getattr(f, 'fact_id', ''),
            'entity': getattr(f, 'entity', 'target'),
            'domain': f.domain,
            'category': f.category,
            'item': f.item,
            'details': f.details,
            'status': f.status,
        } for f in facts]
        facts_file = OUTPUT_DIR / f"facts_{timestamp}.json"
        with open(facts_file, 'w') as f:
            json.dump(facts_data, f, indent=2, default=str)
        saved_files.append(facts_file)

        # Export gaps (include entity for filtering)
        gaps_data = [{
            'id': getattr(g, 'id', '') or getattr(g, 'gap_id', ''),
            'entity': getattr(g, 'entity', 'target'),
            'domain': g.domain,
            'category': g.category,
            'description': g.description,
            'importance': getattr(g, 'importance', 'medium'),
        } for g in gaps]
        gaps_file = OUTPUT_DIR / f"gaps_{timestamp}.json"
        with open(gaps_file, 'w') as f:
            json.dump(gaps_data, f, indent=2, default=str)
        saved_files.append(gaps_file)

        # Export risks
        risks_data = [{
            'id': getattr(r, 'id', '') or getattr(r, 'finding_id', ''),
            'domain': r.domain,
            'title': r.title,
            'description': r.description,
            'severity': r.severity,
        } for r in risks]
        risks_file = OUTPUT_DIR / f"risks_{timestamp}.json"
        with open(risks_file, 'w') as f:
            json.dump(risks_data, f, indent=2, default=str)
        saved_files.append(risks_file)

        # Export work items
        work_items_data = [{
            'id': getattr(w, 'id', '') or getattr(w, 'finding_id', ''),
            'domain': w.domain,
            'title': w.title,
            'description': w.description,
            'phase': w.phase,
        } for w in work_items]
        work_items_file = OUTPUT_DIR / f"work_items_{timestamp}.json"
        with open(work_items_file, 'w') as f:
            json.dump(work_items_data, f, indent=2, default=str)
        saved_files.append(work_items_file)

    except Exception as e:
        logger.error(f"Export: Database path failed: {e}")
        flash('Error exporting data. Please try again.', 'error')
        return redirect(url_for('dashboard'))

    # Audit log the export
    if USE_AUDIT_LOGGING:
        try:
            from web.audit_service import audit_log, AuditAction
            audit_log(
                action=AuditAction.EXPORT_CREATE,
                resource_type='export',
                resource_id=timestamp,
                details={
                    'files_exported': len(saved_files),
                    'filenames': [str(f) for f in saved_files[:5]],  # First 5 only
                    'deal_id': current_deal_id,
                }
            )
        except Exception as e:
            logger.debug(f"Audit logging failed: {e}")

    flash(f'Exported to {len(saved_files)} files', 'success')
    return redirect(url_for('dashboard'))


@app.route('/search')
@auth_optional
def search():
    """Search across all items.

    Phase 2+: Database-first implementation using repository search methods.
    """
    query = request.args.get('q', '')
    results = {'risks': [], 'work_items': [], 'facts': [], 'gaps': []}

    # Phase 2: Database-first - require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        if query:
            flash('Please select a deal to search.', 'info')
        return render_template('search.html', query=query, results=results, total=0)

    if query:
        try:
            from web.repositories import FactRepository, FindingRepository, GapRepository
            from web.context import load_deal_context

            load_deal_context(current_deal_id)

            fact_repo = FactRepository()
            finding_repo = FindingRepository()
            gap_repo = GapRepository()

            # Search risks (findings with type 'risk')
            risk_results = finding_repo.get_by_deal(
                deal_id=current_deal_id,
                finding_type='risk'
            )
            for risk in risk_results:
                if query.lower() in (risk.title or '').lower() or query.lower() in (risk.description or '').lower():
                    results['risks'].append(risk)

            # Search work items (findings with type 'work_item')
            wi_results = finding_repo.get_by_deal(
                deal_id=current_deal_id,
                finding_type='work_item'
            )
            for wi in wi_results:
                if query.lower() in (wi.title or '').lower() or query.lower() in (wi.description or '').lower():
                    results['work_items'].append(wi)

            # Search facts
            fact_results = fact_repo.get_by_deal(deal_id=current_deal_id)
            for fact in fact_results:
                item_text = getattr(fact, 'item', '') or ''
                details_text = str(getattr(fact, 'details', {}) or {})
                if query.lower() in item_text.lower() or query.lower() in details_text.lower():
                    results['facts'].append(fact)

            # Search gaps
            gap_results = gap_repo.get_by_deal(deal_id=current_deal_id)
            for gap in gap_results:
                description = getattr(gap, 'description', '') or ''
                if query.lower() in description.lower():
                    results['gaps'].append(gap)

            logger.debug(f"Search '{query}': {len(results['risks'])} risks, {len(results['work_items'])} work items, {len(results['facts'])} facts, {len(results['gaps'])} gaps")

        except Exception as e:
            logger.error(f"Search failed: {e}")
            flash('Error performing search. Please try again.', 'error')

    total = sum(len(v) for v in results.values())

    return render_template('search.html', query=query, results=results, total=total)


# =============================================================================
# Organization Module Routes
# =============================================================================

# DEAL-SCOPED organization analysis cache
# Key: deal_id -> cached data
_org_cache_by_deal: dict = {}


def get_organization_analysis():
    """Get or run the organization analysis.

    Spec 04: InventoryStore FIRST (deduplicated), then database via DealData, then session fallback.
    IMPORTANT: This cache is now DEAL-SCOPED. Each deal has its own cache entry
    to prevent data leakage between deals.
    """
    global _org_cache_by_deal

    import logging
    logger = logging.getLogger(__name__)

    # Get current deal_id - this is CRITICAL for isolation
    current_deal_id = flask_session.get('current_deal_id') or 'no_deal'

    # PRIMARY: InventoryStore (deduplicated, entity-scoped) — Spec 04
    try:
        from web.blueprints.inventory import get_inventory_store
        from services.organization_bridge import build_organization_result_from_inventory_store

        inv_store = get_inventory_store()
        org_items = inv_store.get_items(inventory_type="organization", entity="target")
        if org_items:
            # Get target name
            from flask import g
            target_name = getattr(g, 'deal', None)
            target_name = target_name.target_name if target_name and hasattr(target_name, 'target_name') else 'Target'

            result, status = build_organization_result_from_inventory_store(
                inv_store,
                target_name=target_name,
                deal_id=current_deal_id if current_deal_id != 'no_deal' else "",
            )

            if status == "success" and result.total_it_headcount > 0:
                logger.info(f"Organization (InventoryStore primary): {result.total_it_headcount} headcount from {len(org_items)} items")
                # Update DEAL-SCOPED cache with inventory source
                _org_cache_by_deal[current_deal_id] = {
                    'fingerprint': hash(tuple(sorted(item.item_id for item in org_items if hasattr(item, 'item_id')))),
                    'org_count': len(org_items),
                    'result': result,
                    'source': 'inventory'
                }
                return result
    except Exception as e:
        import traceback
        logger.warning(f"InventoryStore org path failed, falling back to DB: {e}")
        logger.debug(traceback.format_exc())

    # FALLBACK: Try database via DealData
    if current_deal_id and current_deal_id != 'no_deal':
        try:
            from web.deal_data import DealData, create_store_adapters_from_deal_data
            from web.context import load_deal_context

            load_deal_context(current_deal_id)
            data = DealData()

            # Get TARGET facts only for organization analysis (not buyer facts)
            all_facts = data.get_all_facts(entity='target')
            org_facts = [f for f in all_facts if f.domain == "organization"]

            if all_facts:
                total_facts = len(all_facts)
                current_org_count = len(org_facts)

                # Generate fingerprint based on fact IDs
                fact_ids = sorted([f.id for f in all_facts])
                session_fingerprint = hash(tuple(fact_ids))

                logger.info(f"Phase 2 DB: {total_facts} total facts, {current_org_count} org facts for deal {current_deal_id}")

                # Get cached data for THIS DEAL only
                deal_cache = _org_cache_by_deal.get(current_deal_id, {})
                cached_fingerprint = deal_cache.get('fingerprint')
                cached_org_count = deal_cache.get('org_count', 0)
                cached_result = deal_cache.get('result')

                # Check if we need to rebuild
                needs_rebuild = (
                    cached_result is None or
                    cached_fingerprint != session_fingerprint or
                    cached_org_count != current_org_count
                )

                if needs_rebuild and total_facts > 0:
                    logger.info(f"Phase 2: Rebuilding organization data from DB for deal {current_deal_id}")

                    from services.organization_bridge import build_organization_result

                    # Get target name from deal
                    from flask import g
                    target_name = getattr(g, 'deal', None)
                    target_name = target_name.target_name if target_name and hasattr(target_name, 'target_name') else 'Target'

                    # Create adapters for bridge function (CRITICAL: filter by target entity)
                    fact_adapter, reasoning_adapter = create_store_adapters_from_deal_data(data, entity='target')

                    logger.info(f"Building org result for target: {target_name}")
                    result, status = build_organization_result(fact_adapter, reasoning_adapter, target_name, deal_id=current_deal_id)

                    # Update DEAL-SCOPED cache
                    _org_cache_by_deal[current_deal_id] = {
                        'fingerprint': session_fingerprint,
                        'org_count': current_org_count,
                        'result': result,
                        'source': 'database' if status == "success" else 'database_no_org'
                    }

                    if status == "success" and result.total_it_headcount > 0:
                        logger.info(f"SUCCESS: Built org analysis from DB with {result.total_it_headcount} headcount")
                        return result
                    elif status == "no_org_facts":
                        logger.warning(f"Analysis in DB has {total_facts} facts but no organization facts")
                        return result
                    else:
                        return result

                elif cached_result is not None:
                    logger.debug(f"Using cached org result for deal {current_deal_id}")
                    return cached_result

        except Exception as e:
            import traceback
            logger.warning(f"Phase 2 DB path failed, falling back to session: {e}")
            logger.debug(traceback.format_exc())

    # Fallback: Use session-based path (legacy)
    try:
        s = get_session()
        logger.info(f"get_organization_analysis (fallback): deal={current_deal_id}, session has fact_store={s.fact_store is not None}")

        if s and s.fact_store:
            total_facts = len(s.fact_store.facts)
            org_facts = [f for f in s.fact_store.facts if f.domain == "organization"]
            current_org_count = len(org_facts)

            # Generate a session fingerprint based on fact content
            fact_ids = sorted([f.fact_id for f in s.fact_store.facts]) if s.fact_store.facts else []
            session_fingerprint = hash(tuple(fact_ids)) if fact_ids else 0

            logger.info(f"Session has {total_facts} total facts, {current_org_count} org facts")

            # Get cached data for THIS DEAL only
            deal_cache = _org_cache_by_deal.get(current_deal_id, {})
            cached_fingerprint = deal_cache.get('fingerprint')
            cached_org_count = deal_cache.get('org_count', 0)
            cached_result = deal_cache.get('result')

            logger.info(f"Deal {current_deal_id}: fingerprint={session_fingerprint}, cached={cached_fingerprint}")

            # Check if we need to rebuild for THIS DEAL
            needs_rebuild = (
                cached_result is None or
                cached_fingerprint != session_fingerprint or
                cached_org_count != current_org_count
            )

            if needs_rebuild and total_facts > 0:
                logger.info(f"Rebuilding organization data for deal {current_deal_id}")

                from services.organization_bridge import build_organization_result
                deal_context = s.deal_context or {}
                target_name = deal_context.get('target_name', 'Target') if isinstance(deal_context, dict) else 'Target'

                logger.info(f"Building org result for target: {target_name}")
                result, status = build_organization_result(s.fact_store, s.reasoning_store, target_name, deal_id=current_deal_id)

                # Update DEAL-SCOPED cache
                _org_cache_by_deal[current_deal_id] = {
                    'fingerprint': session_fingerprint,
                    'org_count': current_org_count,
                    'result': result,
                    'source': 'analysis' if status == "success" else 'analysis_no_org'
                }

                if status == "success" and result.total_it_headcount > 0:
                    logger.info(f"SUCCESS: Built org analysis with {result.total_it_headcount} headcount from {current_org_count} org facts")
                    return result
                elif status == "no_org_facts":
                    logger.warning(f"Analysis ran with {total_facts} facts but no organization facts found")
                    return result
                else:
                    logger.warning(f"build_organization_result returned status={status}, headcount={result.total_it_headcount}")
                    return result

            elif cached_result is not None:
                # Cache is valid for THIS DEAL, return cached result
                logger.debug(f"Using cached org result for deal {current_deal_id} ({cached_org_count} org facts)")
                return cached_result

    except Exception as e:
        import traceback
        logger.error(f"Could not build org from facts: {e}")
        logger.error(traceback.format_exc())

    # Fall back to empty result (no demo data) if we have nothing for this deal
    deal_cache = _org_cache_by_deal.get(current_deal_id, {})
    if deal_cache.get('result') is None:
        logger.info(f"No organization data available for deal {current_deal_id} - showing empty state")
        empty_result = _create_empty_organization_result()
        _org_cache_by_deal[current_deal_id] = {
            'fingerprint': 0,
            'org_count': 0,
            'result': empty_result,
            'source': 'no_data'
        }
        return empty_result

    return deal_cache.get('result')


def get_org_data_source():
    """Get whether org data is from analysis or demo for current deal."""
    global _org_cache_by_deal
    current_deal_id = flask_session.get('current_deal_id') or 'no_deal'
    deal_cache = _org_cache_by_deal.get(current_deal_id, {})
    return deal_cache.get('source', 'unknown')


def clear_organization_cache():
    """Clear the organization cache for current deal to force rebuild."""
    global _org_cache_by_deal
    current_deal_id = flask_session.get('current_deal_id') or 'no_deal'
    if current_deal_id in _org_cache_by_deal:
        del _org_cache_by_deal[current_deal_id]
        import logging
        logging.getLogger(__name__).info(f"Cleared org cache for deal {current_deal_id}")


def invalidate_all_caches():
    """Invalidate all caches for current deal - call this after new analysis completes."""
    import logging
    logger = logging.getLogger(__name__)

    current_deal_id = flask_session.get('current_deal_id') or 'no_deal'
    logger.info(f"Invalidating all caches for deal {current_deal_id}")

    # Clear organization cache for this deal
    clear_organization_cache()

    # Clear session store caches
    session_id = flask_session.get('session_id')
    if session_id:
        user_session = session_store.get_session(session_id)
        if user_session:
            user_session.analysis_session = None
            user_session._cache_mtime = 0
            user_session._cached_deal_id = None


def _create_demo_organization_data():
    """Create demo organization data for UI testing."""
    from models.organization_models import (
        StaffMember, RoleSummary, CategorySummary,
        MSPService, MSPRelationship, SharedServiceDependency,
        TSAItem, StaffingNeed, DependencyLevel
    )
    from models.organization_stores import (
        OrganizationDataStore, StaffingComparisonResult,
        CategoryComparison, MissingRole, RatioComparison,
        MSPSummary, SharedServicesSummary, OrganizationAnalysisResult
    )
    from datetime import datetime
    import uuid

    def gen_id(prefix):
        return f"{prefix}_{uuid.uuid4().hex[:8]}"

    # Create data store
    store = OrganizationDataStore()

    # Add demo staff members
    demo_staff = [
        StaffMember(
            id=gen_id("staff"), name="John Smith", role_title="IT Director",
            role_category=RoleCategory.LEADERSHIP, department="IT",
            employment_type=EmploymentType.FTE, base_compensation=180000,
            entity="target", hire_date="2018-03-15", location="HQ",
            is_key_person=True, key_person_reason="Single point of IT leadership"
        ),
        StaffMember(
            id=gen_id("staff"), name="Sarah Johnson", role_title="Sr. Systems Administrator",
            role_category=RoleCategory.INFRASTRUCTURE, department="IT Infrastructure",
            employment_type=EmploymentType.FTE, base_compensation=95000,
            entity="target", hire_date="2019-06-01", location="HQ",
            is_key_person=True, key_person_reason="Critical infrastructure knowledge"
        ),
        StaffMember(
            id=gen_id("staff"), name="Mike Chen", role_title="Network Engineer",
            role_category=RoleCategory.INFRASTRUCTURE, department="IT Infrastructure",
            employment_type=EmploymentType.FTE, base_compensation=85000,
            entity="target", hire_date="2021-01-10", location="HQ"
        ),
        StaffMember(
            id=gen_id("staff"), name="Emily Davis", role_title="Help Desk Analyst",
            role_category=RoleCategory.SERVICE_DESK, department="IT Support",
            employment_type=EmploymentType.FTE, base_compensation=55000,
            entity="target", hire_date="2022-04-01", location="HQ"
        ),
        StaffMember(
            id=gen_id("staff"), name="David Wilson", role_title="Help Desk Analyst",
            role_category=RoleCategory.SERVICE_DESK, department="IT Support",
            employment_type=EmploymentType.CONTRACTOR, base_compensation=60000,
            entity="target", hire_date="2023-01-15", location="Remote"
        ),
        StaffMember(
            id=gen_id("staff"), name="Lisa Brown", role_title="Application Developer",
            role_category=RoleCategory.APPLICATIONS, department="IT Development",
            employment_type=EmploymentType.FTE, base_compensation=110000,
            entity="target", hire_date="2020-08-01", location="HQ"
        ),
        StaffMember(
            id=gen_id("staff"), name="Tom Garcia", role_title="Security Analyst",
            role_category=RoleCategory.SECURITY, department="IT Security",
            employment_type=EmploymentType.CONTRACTOR, base_compensation=95000,
            entity="target", hire_date="2023-06-01", location="Remote"
        ),
    ]

    for staff in demo_staff:
        store.add_staff_member(staff)

    # Add demo MSP relationships
    msp1 = MSPRelationship(
        id=gen_id("msp"),
        vendor_name="SecureIT Partners",
        services=[
            MSPService(service_name="24/7 SOC Monitoring", fte_equivalent=2.0,
                      coverage="24x7", criticality="critical", internal_backup=False),
            MSPService(service_name="Vulnerability Scanning", fte_equivalent=0.5,
                      coverage="business_hours", criticality="high", internal_backup=False),
        ],
        contract_value_annual=180000,
        contract_expiry="2025-12-31",
        notice_period_days=90,
        dependency_level=DependencyLevel.FULL
    )

    msp2 = MSPRelationship(
        id=gen_id("msp"),
        vendor_name="CloudManage Inc",
        services=[
            MSPService(service_name="Cloud Infrastructure Management", fte_equivalent=1.5,
                      coverage="business_hours", criticality="high", internal_backup=True),
            MSPService(service_name="Backup Management", fte_equivalent=0.5,
                      coverage="business_hours", criticality="critical", internal_backup=False),
        ],
        contract_value_annual=120000,
        contract_expiry="2024-06-30",
        notice_period_days=60,
        dependency_level=DependencyLevel.PARTIAL
    )

    store.add_msp_relationship(msp1)
    store.add_msp_relationship(msp2)

    # Add demo shared services (using correct field names from model)
    ss1 = SharedServiceDependency(
        id=gen_id("ss"),
        service_name="Active Directory / Identity Management",
        provider="Parent Corp",
        fte_equivalent=1.0,
        criticality="critical",
        will_transfer=False,
        tsa_candidate=True
    )

    ss2 = SharedServiceDependency(
        id=gen_id("ss"),
        service_name="Email & Collaboration (M365)",
        provider="Parent Corp",
        fte_equivalent=0.5,
        criticality="high",
        will_transfer=False,
        tsa_candidate=True
    )

    ss3 = SharedServiceDependency(
        id=gen_id("ss"),
        service_name="ERP System Support",
        provider="Parent Corp",
        fte_equivalent=0.5,
        criticality="high",
        will_transfer=True,
        tsa_candidate=False
    )

    store.add_shared_service(ss1)
    store.add_shared_service(ss2)
    store.add_shared_service(ss3)

    # Create benchmark comparison result (using correct field names from model)
    comparison = StaffingComparisonResult(
        benchmark_profile_id="mid_market_manufacturing",
        benchmark_profile_name="Mid-Market Manufacturing",
        comparison_date=datetime.now().isoformat(),
        total_actual=7,
        total_expected_min=8,
        total_expected_typical=10,
        total_expected_max=12,
        overall_status="understaffed",
        category_comparisons=[
            CategoryComparison(
                category="leadership",
                category_display="IT Leadership",
                actual_count=1,
                benchmark_min=1,
                benchmark_typical=1,
                benchmark_max=2,
                variance=0,
                status="in_range"
            ),
            CategoryComparison(
                category="infrastructure",
                category_display="Infrastructure & Operations",
                actual_count=2,
                benchmark_min=2,
                benchmark_typical=3,
                benchmark_max=4,
                variance=0,
                status="in_range"
            ),
            CategoryComparison(
                category="security",
                category_display="Security",
                actual_count=1,
                benchmark_min=1,
                benchmark_typical=1,
                benchmark_max=2,
                variance=0,
                status="in_range"
            ),
            CategoryComparison(
                category="service_desk",
                category_display="Service Desk / End User Support",
                actual_count=2,
                benchmark_min=2,
                benchmark_typical=2,
                benchmark_max=3,
                variance=0,
                status="in_range"
            ),
            CategoryComparison(
                category="applications",
                category_display="Applications",
                actual_count=1,
                benchmark_min=2,
                benchmark_typical=2,
                benchmark_max=3,
                variance=-1,
                status="understaffed"
            ),
        ],
        missing_roles=[
            MissingRole(
                role_title="DBA / Data Engineer",
                category=RoleCategory.DATA,
                importance="high",
                impact="No dedicated database expertise; relying on developers for DB tasks",
                recommendation="Consider hiring or contracting a DBA, especially if data-intensive operations grow",
                current_coverage="Ad-hoc by application developers"
            ),
        ],
        ratio_comparisons=[
            RatioComparison(
                ratio_name="IT Staff to Employee Ratio",
                actual_value=0.035,
                benchmark_min=0.03,
                benchmark_typical=0.04,
                benchmark_max=0.05,
                status="in_range"
            ),
        ],
        key_findings=[
            "Overall staffing is at the lower end of the expected range",
            "Application Development team appears understaffed for company size",
            "Security coverage relies heavily on MSP - key dependency",
            "No dedicated DBA role identified",
        ],
        recommendations=[
            "Consider adding 1 application developer to support growth",
            "Evaluate MSP dependency for security operations",
            "Plan for database administration coverage",
        ]
    )

    # Create MSP summary (using correct field names from model)
    msp_summary = MSPSummary(
        total_msp_count=2,
        total_fte_equivalent=4.5,
        total_annual_cost=300000,
        total_replacement_cost=450000,
        high_risk_count=1,
        critical_services_count=2,
        services_without_backup=3,
        earliest_expiry="2024-06-30"
    )

    # Create shared services summary (using correct field names from model)
    ss_summary = SharedServicesSummary(
        total_dependencies=3,
        total_fte_equivalent=2.0,
        transferring_fte=0.5,
        non_transferring_fte=1.5,
        hidden_headcount_need=1.5,
        hidden_cost_annual=150000,
        tsa_candidate_count=2,
        critical_dependencies=1
    )

    # Create TSA recommendations (using correct field names from model)
    tsa_recs = [
        TSAItem(
            id=gen_id("tsa"),
            service="Active Directory / Identity Management",
            provider="Parent Corp",
            duration_months=12,
            estimated_monthly_cost=10000,
            exit_criteria="Standalone AD or Azure AD tenant operational",
            replacement_plan="Deploy new Azure AD tenant, migrate users and groups"
        ),
        TSAItem(
            id=gen_id("tsa"),
            service="Email & Collaboration (M365)",
            provider="Parent Corp",
            duration_months=6,
            estimated_monthly_cost=7000,
            exit_criteria="Separate M365 tenant with mail migration complete",
            replacement_plan="Provision new M365 tenant, migrate mailboxes and SharePoint"
        ),
    ]

    # Create staffing needs (using correct field names from model)
    staffing_needs = [
        StaffingNeed(
            id=gen_id("need"),
            role="Identity & Access Management Specialist",
            category=RoleCategory.SECURITY,
            fte_count=1.0,
            urgency="day_100",
            estimated_salary=95000,
            reason="Required to manage standalone identity infrastructure post-TSA",
            source="shared_services"
        ),
    ]

    # Create final result (using correct field names from model)
    result = OrganizationAnalysisResult(
        benchmark_comparison=comparison,
        msp_summary=msp_summary,
        shared_services_summary=ss_summary,
        tsa_recommendations=tsa_recs,
        hiring_recommendations=staffing_needs
    )

    # Store the data store reference as a custom attribute for template access
    result.data_store = store

    return result


def _create_empty_organization_result():
    """Create an empty organization result for when no data is available."""
    from models.organization_stores import (
        OrganizationDataStore, StaffingComparisonResult,
        MSPSummary, SharedServicesSummary, OrganizationAnalysisResult
    )
    from datetime import datetime

    # Create empty data store
    store = OrganizationDataStore()

    # Create empty comparison
    comparison = StaffingComparisonResult(
        benchmark_profile_id="none",
        benchmark_profile_name="No Data",
        comparison_date=datetime.now().isoformat(),
        total_actual=0,
        total_expected_min=0,
        total_expected_typical=0,
        total_expected_max=0,
        overall_status="no_data"
    )

    # Create empty summaries
    msp_summary = MSPSummary(
        total_msp_count=0,
        total_fte_equivalent=0,
        total_annual_cost=0,
        high_risk_count=0,
        critical_services_count=0,
        services_without_backup=0
    )

    ss_summary = SharedServicesSummary(
        total_dependencies=0,
        total_fte_equivalent=0,
        transferring_fte=0,
        non_transferring_fte=0,
        hidden_headcount_need=0,
        hidden_cost_annual=0,
        tsa_candidate_count=0,
        critical_dependencies=0
    )

    # Create empty result
    result = OrganizationAnalysisResult(
        benchmark_comparison=comparison,
        msp_summary=msp_summary,
        shared_services_summary=ss_summary,
        tsa_recommendations=[],
        hiring_recommendations=[]
    )

    result.data_store = store
    return result


@app.route('/organization/refresh')
@auth_optional
def organization_refresh():
    """Refresh organization data from current analysis.

    This forces a complete cache invalidation and reload from the latest files.
    """
    # Invalidate ALL caches to ensure we get fresh data
    invalidate_all_caches()

    # Now get fresh session data (will reload from newest files)
    s = get_session()
    if s and s.fact_store:
        total_facts = len(s.fact_store.facts)
        org_facts = len([f for f in s.fact_store.facts if f.domain == "organization"])
        flash(f'Organization data refreshed - found {org_facts} org facts (of {total_facts} total)', 'success')
    else:
        flash('Organization cache cleared - no analysis session found', 'warning')

    return redirect(url_for('organization_overview'))


@app.route('/organization')
@auth_optional
def organization_overview():
    """Organization module overview."""
    result = get_organization_analysis()
    data_source = get_org_data_source()

    return render_template('organization/overview.html',
                         result=result,
                         store=result.data_store,
                         comparison=result.benchmark_comparison,
                         msp_summary=result.msp_summary,
                         ss_summary=result.shared_services_summary,
                         data_source=data_source)


@app.route('/organization/staffing')
@auth_optional
def organization_staffing():
    """Staffing tree view."""
    result = get_organization_analysis()
    store = result.data_store

    # Debug info for troubleshooting
    data_source = get_org_data_source()
    s = get_session()
    org_facts_count = 0
    total_facts_count = 0
    org_categories = {}
    org_fact_samples = []
    if s and s.fact_store:
        total_facts_count = len(s.fact_store.facts)
        org_facts = [f for f in s.fact_store.facts if f.domain == "organization"]
        org_facts_count = len(org_facts)
        # Count facts by category to help diagnose why staff_members might be empty
        for f in org_facts:
            cat = f.category or 'unknown'
            org_categories[cat] = org_categories.get(cat, 0) + 1
            # Sample first 3 facts for debugging
            if len(org_fact_samples) < 3:
                org_fact_samples.append({'item': f.item, 'category': f.category, 'has_details': bool(f.details)})

    debug_info = {
        'data_source': data_source,
        'total_facts': total_facts_count,
        'org_facts': org_facts_count,
        'org_categories': org_categories,  # Shows breakdown by category
        'org_samples': org_fact_samples,   # Sample facts for debugging
        'staff_members': len(store.staff_members) if store else 0,
        'deal_id': flask_session.get('current_deal_id', 'none')
    }
    logger.info(f"Organization staffing debug: {debug_info}")

    # Build categories dict for template by grouping staff members
    categories = {}
    role_groups = {}  # role_title -> list of members

    # Group by role first
    for member in store.staff_members:
        if member.role_title not in role_groups:
            role_groups[member.role_title] = []
        role_groups[member.role_title].append(member)

    # Now group roles by category
    for role_title, members in role_groups.items():
        if not members:
            continue
        cat = members[0].role_category

        if cat not in categories:
            categories[cat] = {
                'display_name': cat.value.replace('_', ' ').title(),
                'total_headcount': 0,
                'total_compensation': 0,
                'roles': []
            }

        role_data = {
            'role_title': role_title,
            'headcount': len(members),
            'avg_compensation': sum(m.base_compensation for m in members) / len(members),
            'members': members
        }
        categories[cat]['roles'].append(role_data)
        categories[cat]['total_headcount'] += len(members)
        categories[cat]['total_compensation'] += sum(m.base_compensation for m in members)

    # Convert keys to strings for template
    categories_by_name = {cat.value: data for cat, data in categories.items()}

    # Get MSP relationships for the org chart
    msp_relationships = store.msp_relationships if hasattr(store, 'msp_relationships') else []

    return render_template('organization/staffing_tree.html',
                         categories=categories_by_name,
                         total_headcount=store.get_target_headcount(),
                         total_compensation=store.total_compensation,
                         fte_count=store.total_internal_fte,
                         contractor_count=store.total_contractor,
                         key_person_count=len(store.get_key_persons()),
                         msp_relationships=msp_relationships,
                         debug_info=debug_info)


@app.route('/organization/benchmark')
@auth_optional
def organization_benchmark():
    """Benchmark comparison view."""
    result = get_organization_analysis()

    return render_template('organization/benchmark.html',
                         comparison=result.benchmark_comparison)


@app.route('/organization/msp')
@auth_optional
def organization_msp():
    """MSP dependency analysis view."""
    result = get_organization_analysis()
    store = result.data_store

    return render_template('organization/msp_analysis.html',
                         msp_relationships=store.msp_relationships,
                         msp_summary=result.msp_summary)


@app.route('/organization/shared-services')
@auth_optional
def organization_shared_services():
    """Shared services analysis view."""
    result = get_organization_analysis()
    store = result.data_store

    return render_template('organization/shared_services.html',
                         dependencies=store.shared_service_dependencies,
                         summary=result.shared_services_summary,
                         tsa_recommendations=result.tsa_recommendations,
                         staffing_needs=result.hiring_recommendations)


# Organization manual analysis endpoint removed - organization data now comes directly from document analysis


# =============================================================================
# Applications Inventory Routes
# =============================================================================

@app.route('/applications')
@auth_optional
def applications_overview():
    """Applications inventory overview.

    Spec 04: InventoryStore FIRST (deduplicated), with fallback to database facts (legacy).
    Supports ?entity=buyer or ?entity=target (default) query parameter.
    """
    from services.applications_bridge import build_applications_inventory, build_applications_from_inventory_store, ApplicationsInventory
    from web.deal_data import DealData, wrap_db_facts
    from web.context import load_deal_context

    # Debug info to help diagnose missing applications
    debug_info = {
        'inventory_store_count': 0,
        'inventory_apps_target': 0,
        'inventory_apps_all': 0,
        'fact_store_app_facts': 0,
        'fact_store_app_by_entity': {},
        'data_source': 'none'
    }

    current_deal_id = flask_session.get('current_deal_id')

    # Buyer entity support: allow ?entity=buyer or ?entity=target (default)
    requested_entity = request.args.get('entity', 'target')
    if requested_entity not in ('target', 'buyer'):
        requested_entity = 'target'

    # PRIMARY: InventoryStore (deduplicated, entity-scoped) — Spec 04
    try:
        from web.blueprints.inventory import get_inventory_store
        inv_store = get_inventory_store()
        debug_info['inventory_store_count'] = len(inv_store)
        if len(inv_store) > 0:
            # Check all apps vs entity-scoped apps
            all_apps = inv_store.get_items(inventory_type="application", status="active")
            entity_apps = inv_store.get_items(inventory_type="application", entity=requested_entity, status="active")
            debug_info['inventory_apps_all'] = len(all_apps) if all_apps else 0
            debug_info['inventory_apps_target'] = len(entity_apps) if entity_apps else 0

            if entity_apps:
                inventory, status = build_applications_from_inventory_store(inv_store)
                data_source = "inventory"
                debug_info['data_source'] = data_source
                logger.info(f"Applications (InventoryStore primary): {len(entity_apps)} {requested_entity} apps, debug={debug_info}")
                return render_template('applications/overview.html',
                                      inventory=inventory,
                                      status=status,
                                      data_source=data_source,
                                      debug_info=debug_info)
    except Exception as e:
        logger.warning(f"Could not load from InventoryStore: {e}")

    # FALLBACK: Database facts (legacy runs without InventoryStore)
    if current_deal_id:
        try:
            load_deal_context(current_deal_id)
            data = DealData()

            # Get facts for the requested entity
            all_facts = data.get_all_facts(entity=requested_entity)
            app_facts = [f for f in all_facts if f.domain == "applications"]
            debug_info['fact_store_app_facts'] = len(app_facts)

            for f in app_facts:
                entity = f.entity or 'target'
                debug_info['fact_store_app_by_entity'][entity] = debug_info['fact_store_app_by_entity'].get(entity, 0) + 1

            if app_facts:
                # Wrap DB facts in adapter for bridge function
                fact_adapter = wrap_db_facts(all_facts)
                inventory, status = build_applications_inventory(fact_adapter)
                data_source = "database_legacy"
                debug_info['data_source'] = data_source
                logger.info(f"Applications (DB legacy fallback): {debug_info}")
                return render_template('applications/overview.html',
                                      inventory=inventory,
                                      status=status,
                                      data_source=data_source,
                                      debug_info=debug_info)
        except Exception as e:
            logger.warning(f"Could not load from database: {e}")

    # EMPTY: No data from either source
    inventory = ApplicationsInventory()
    status = "no_facts"
    data_source = "none"

    debug_info['data_source'] = data_source
    logger.info(f"Applications debug: {debug_info}")

    return render_template('applications/overview.html',
                          inventory=inventory,
                          status=status,
                          data_source=data_source,
                          debug_info=debug_info)


@app.route('/applications/<category>')
@auth_optional
def applications_category(category):
    """View applications in a specific category.

    Phase 2: Now reads from database via DealData.
    """
    from services.applications_bridge import build_applications_inventory, ApplicationsInventory
    from web.deal_data import DealData, wrap_db_facts
    from web.context import load_deal_context

    inventory = None
    status = "no_facts"

    # Phase 2: Read from database via DealData
    current_deal_id = flask_session.get('current_deal_id')

    if current_deal_id:
        try:
            load_deal_context(current_deal_id)
            data = DealData()
            # Get TARGET facts only for applications category (not buyer facts)
            all_facts = data.get_all_facts(entity='target')

            if all_facts:
                fact_adapter = wrap_db_facts(all_facts)
                inventory, status = build_applications_inventory(fact_adapter)
        except Exception as e:
            logger.warning(f"Could not load from database: {e}")

    if inventory is None:
        inventory = ApplicationsInventory()

    # Get category summary
    cat_summary = inventory.by_category.get(category)
    if not cat_summary:
        flash(f'Category "{category}" not found', 'warning')
        return redirect(url_for('applications_overview'))

    return render_template('applications/category.html',
                          category=cat_summary,
                          inventory=inventory)


# =============================================================================
# Infrastructure Inventory Routes
# =============================================================================

@app.route('/infrastructure')
@auth_optional
def infrastructure_overview():
    """Infrastructure inventory overview.

    Spec 04: InventoryStore FIRST (deduplicated), with fallback to database facts (legacy).
    """
    from services.infrastructure_bridge import build_infrastructure_inventory, build_infrastructure_from_inventory_store, InfrastructureInventory
    from web.deal_data import DealData, wrap_db_facts
    from web.context import load_deal_context

    current_deal_id = flask_session.get('current_deal_id')

    # Buyer entity support: allow ?entity=buyer or ?entity=target (default)
    requested_entity = request.args.get('entity', 'target')
    if requested_entity not in ('target', 'buyer'):
        requested_entity = 'target'

    # PRIMARY: InventoryStore (deduplicated, entity-scoped) — Spec 04
    try:
        from web.blueprints.inventory import get_inventory_store
        inv_store = get_inventory_store()
        if len(inv_store) > 0:
            infra_items = inv_store.get_items(inventory_type="infrastructure", entity=requested_entity, status="active")
            if infra_items:
                inventory, status = build_infrastructure_from_inventory_store(inv_store)
                data_source = "inventory"
                logger.info(f"Infrastructure (InventoryStore primary): {len(infra_items)} {requested_entity} items")
                return render_template('infrastructure/overview.html',
                                      inventory=inventory,
                                      status=status,
                                      data_source=data_source)
    except Exception as e:
        logger.warning(f"Could not load from InventoryStore: {e}")

    # FALLBACK: Database facts (legacy runs without InventoryStore)
    if current_deal_id:
        try:
            load_deal_context(current_deal_id)
            data = DealData()
            # Get facts for the requested entity
            all_facts = data.get_all_facts(entity=requested_entity)

            if all_facts:
                fact_adapter = wrap_db_facts(all_facts)
                inventory, status = build_infrastructure_inventory(fact_adapter)
                data_source = "database_legacy"
                logger.info(f"Infrastructure (DB legacy fallback): {len(all_facts)} facts")
                return render_template('infrastructure/overview.html',
                                      inventory=inventory,
                                      status=status,
                                      data_source=data_source)
        except Exception as e:
            logger.warning(f"Could not load from database: {e}")

    # EMPTY: No data from either source
    inventory = InfrastructureInventory()
    status = "no_facts"
    data_source = "none"

    return render_template('infrastructure/overview.html',
                          inventory=inventory,
                          status=status,
                          data_source=data_source)


@app.route('/infrastructure/<category>')
@auth_optional
def infrastructure_category(category):
    """View infrastructure in a specific category.

    Phase 2: Now reads from database via DealData.
    """
    from services.infrastructure_bridge import build_infrastructure_inventory, build_infrastructure_from_inventory_store, InfrastructureInventory
    from web.deal_data import DealData, wrap_db_facts
    from web.context import load_deal_context

    inventory = None
    status = "no_facts"

    # First try InventoryStore (structured data from imports)
    try:
        from web.blueprints.inventory import get_inventory_store
        inv_store = get_inventory_store()
        if len(inv_store) > 0:
            infra_items = inv_store.get_items(inventory_type="infrastructure", entity="target", status="active")
            if infra_items:
                inventory, status = build_infrastructure_from_inventory_store(inv_store)
    except Exception as e:
        logger.warning(f"Could not load from InventoryStore: {e}")

    # Phase 2: Read from database via DealData
    if inventory is None:
        current_deal_id = flask_session.get('current_deal_id')
        if current_deal_id:
            try:
                load_deal_context(current_deal_id)
                data = DealData()
                # Get TARGET facts only for infrastructure category (not buyer facts)
                all_facts = data.get_all_facts(entity='target')

                if all_facts:
                    fact_adapter = wrap_db_facts(all_facts)
                    inventory, status = build_infrastructure_inventory(fact_adapter)
            except Exception as e:
                logger.warning(f"Could not load from database: {e}")

    if inventory is None:
        inventory = InfrastructureInventory()

    # Get category summary
    cat_summary = inventory.by_category.get(category)
    if not cat_summary:
        flash(f'Category "{category}" not found', 'warning')
        return redirect(url_for('infrastructure_overview'))

    return render_template('infrastructure/category.html',
                          category=cat_summary,
                          inventory=inventory)


# =============================================================================
# Fact Review Queue (Phase 1 - Validation Workflow)
# =============================================================================

@app.route('/review')
@auth_optional
def review_queue_page():
    """
    Fact validation review queue page.

    Phase 2+: Database-first implementation.
    Displays facts needing human review, prioritized by importance.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to view the review queue.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.repositories import FactRepository
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        repo = FactRepository()

        # Get review queue (paginated)
        result = repo.get_review_queue(deal_id=current_deal_id, per_page=50)
        queue = result.get('items', [])
        total_needs_review = result.get('total', 0)

        # Get all facts for stats
        all_facts = repo.get_by_deal(deal_id=current_deal_id)

        # Build stats from database
        verified_count = len([f for f in all_facts if f.verified])
        stats = {
            'total_facts': len(all_facts),
            'needs_review': total_needs_review,
            'verified': verified_count,
            'unverified': len([f for f in all_facts if not f.verified]),
            'completion_rate': verified_count / len(all_facts) if len(all_facts) > 0 else 0,
            'by_domain': {}
        }

        # Group by domain
        for fact in all_facts:
            domain = fact.domain
            if domain not in stats['by_domain']:
                stats['by_domain'][domain] = {'total': 0, 'needs_review': 0, 'verified': 0}
            stats['by_domain'][domain]['total'] += 1
            if getattr(fact, 'needs_review', False):
                stats['by_domain'][domain]['needs_review'] += 1
            if fact.verified:
                stats['by_domain'][domain]['verified'] += 1

        domains = list(stats['by_domain'].keys())
        data_source = 'database' if len(all_facts) > 0 else 'none'

    except Exception as e:
        logger.error(f"Review queue page: Database path failed: {e}")
        flash('Error loading review queue. Please try again.', 'error')
        return redirect(url_for('dashboard'))

    return render_template(
        'review/queue.html',
        stats=stats,
        queue=queue,
        domains=domains,
        data_source=data_source
    )


# =============================================================================
# Error Handlers
# =============================================================================

@app.errorhandler(404)
def not_found_error(error):
    """Handle 404 errors."""
    return render_template('error.html',
                          error_code=404,
                          error_title="Page Not Found",
                          error_message="The page you're looking for doesn't exist."), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors."""
    return render_template('error.html',
                          error_code=500,
                          error_title="Internal Server Error",
                          error_message="Something went wrong on our end. Please try again."), 500


@app.errorhandler(Exception)
def handle_exception(e):
    """Handle uncaught exceptions."""
    # Log the error
    import traceback
    print(f"Unhandled exception: {e}")
    traceback.print_exc()

    # Return user-friendly error
    if request.accept_mimetypes.best == 'application/json':
        return jsonify({
            'error': True,
            'message': str(e),
            'type': type(e).__name__
        }), 500

    return render_template('error.html',
                          error_code=500,
                          error_title="Error",
                          error_message=str(e)), 500


# =============================================================================
# API Endpoints
# =============================================================================

@app.route('/api/health')
def api_health_detailed():
    """Detailed health check endpoint with configuration info."""
    from config_v2 import validate_environment, get_config_summary

    env_status = validate_environment()
    config = get_config_summary()

    return jsonify({
        'status': 'healthy' if env_status['valid'] else 'degraded',
        'service': 'it-dd-agent',
        'version': '2.0',
        'environment': {
            'valid': env_status['valid'],
            'api_key_configured': env_status['api_key_configured'],
            'warnings': env_status['warnings'],
            'errors': env_status['errors']
        },
        'config': {
            'domains': config['domains'],
            'parallel_enabled': config['parallel_enabled'],
        }
    })


@app.route('/api/health/session')
def session_health():
    """
    Health check endpoint for session backend (Spec 04).

    Returns session backend status and health information.
    Used for monitoring and load balancer health checks.

    Returns:
        JSON with backend type, health status, and metrics
    """
    backend = app.config.get('SESSION_TYPE', 'unknown')
    status = {
        'backend': backend,
        'healthy': False,
        'details': {}
    }

    if backend == 'redis':
        try:
            redis_client = app.config.get('SESSION_REDIS')
            if redis_client:
                # Ping Redis
                redis_client.ping()

                # Get basic stats
                info = redis_client.info('stats')
                status['healthy'] = True
                status['details'] = {
                    'total_connections': info.get('total_connections_received', 0),
                    'connected_clients': info.get('connected_clients', 0),
                    'used_memory_human': info.get('used_memory_human', 'unknown'),
                }
            else:
                status['details'] = {'error': 'Redis client not configured'}
        except Exception as e:
            status['healthy'] = False
            status['details'] = {'error': str(e)}

    elif backend == 'sqlalchemy':
        try:
            # Check if session table exists and query count
            from sqlalchemy import inspect
            inspector = inspect(db.engine)

            if 'flask_sessions' in inspector.get_table_names():
                result = db.session.execute(db.text(
                    "SELECT COUNT(*) FROM flask_sessions"
                )).scalar()
                status['healthy'] = True
                status['details'] = {
                    'active_sessions': result,
                    'table_exists': True
                }
            else:
                status['healthy'] = False
                status['details'] = {
                    'error': 'flask_sessions table not found',
                    'table_exists': False
                }
        except Exception as e:
            status['healthy'] = False
            status['details'] = {'error': str(e)}

    elif backend == 'filesystem' or backend is None:
        # Filesystem/cookie sessions - always "healthy" but note limitations
        status['healthy'] = True
        status['details'] = {
            'note': 'Cookie-based sessions (not recommended for production)',
            'persistent': False
        }

    else:
        status['details'] = {'error': f'Unknown backend type: {backend}'}

    return jsonify(status), 200 if status['healthy'] else 503


@app.route('/api/session/info')
def session_info():
    """Get current session/deal information.

    Phase 2+: Database-first implementation.
    """
    current_deal_id = flask_session.get('current_deal_id')

    # Get task status if running
    task_id = flask_session.get('current_task_id')
    task_status = None
    if task_id:
        task_status = task_manager.get_task_status(task_id)

    if not current_deal_id:
        return jsonify({
            'summary': {'facts': 0, 'gaps': 0, 'risks': 0, 'work_items': 0},
            'task_id': task_id,
            'task_status': task_status,
            'has_results': False,
            'deal_id': None
        })

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        # Get summary from database
        dashboard = data.get_dashboard_summary()
        summary = {
            'facts': dashboard.get('fact_counts', {}).get('total', 0) if isinstance(dashboard.get('fact_counts'), dict) else len(data.get_all_facts()),
            'gaps': len(data.get_gaps()),
            'risks': dashboard.get('risk_summary', {}).get('total', 0),
            'work_items': dashboard.get('work_item_summary', {}).get('total', 0),
            'risk_summary': dashboard.get('risk_summary', {}),
            'work_item_summary': dashboard.get('work_item_summary', {}),
        }

        return jsonify({
            'summary': summary,
            'task_id': task_id,
            'task_status': task_status,
            'has_results': summary.get('facts', 0) > 0 or summary.get('risks', 0) > 0,
            'deal_id': current_deal_id,
            'source': 'database'
        })
    except Exception as e:
        logger.error(f"Session info: Database path failed: {e}")
        return jsonify({
            'summary': {'facts': 0, 'gaps': 0, 'risks': 0, 'work_items': 0},
            'task_id': task_id,
            'task_status': task_status,
            'has_results': False,
            'deal_id': current_deal_id,
            'error': str(e)
        })


@app.route('/api/pipeline/status')
def pipeline_status():
    """Check analysis pipeline component availability."""
    from web.analysis_runner import check_pipeline_availability

    status = check_pipeline_availability()

    return jsonify({
        'ready': status['api_key'] and status['pdf_parser'],
        'components': {
            'api_key': status['api_key'],
            'pdf_parser': status['pdf_parser'],
            'fact_store': status['fact_store'],
            'discovery_agents': status['discovery_agents'],
            'reasoning_agents': status['reasoning_agents'],
        },
        'errors': status['errors'] if status['errors'] else None
    })


# =============================================================================
# Phase 2: Deal Status Endpoint (Database-First)
# =============================================================================

@app.route('/api/deal/<deal_id>/status')
@auth_optional
def get_deal_analysis_status(deal_id):
    """
    Get analysis status for a deal from the database.

    Phase 2 endpoint: reads from AnalysisRun table instead of in-memory task manager.
    This survives server restarts and provides consistent status.

    Returns:
        JSON with status, progress, counts, and timestamps
    """
    from web.repositories import AnalysisRunRepository

    run_repo = AnalysisRunRepository()

    # Get latest run (any status) to show in-progress or completed analysis
    run = run_repo.get_latest(deal_id)

    if run:
        return jsonify({
            'status': run.status,
            'progress': run.progress or 0,
            'current_step': run.current_step or '',
            'facts_created': run.facts_created or 0,
            'findings_created': run.findings_created or 0,
            'run_id': run.id,
            'run_number': run.run_number,
            'started_at': run.started_at.isoformat() if run.started_at else None,
            'completed_at': run.completed_at.isoformat() if run.completed_at else None,
            'duration_seconds': run.duration_seconds,
            'error_message': run.error_message if run.status == 'failed' else None,
        })

    return jsonify({
        'status': 'no_analysis',
        'message': 'No analysis has been run for this deal'
    })


@app.route('/api/deal/<deal_id>/runs')
@auth_optional
def get_deal_analysis_runs(deal_id):
    """
    Get analysis run history for a deal.

    Phase 2 endpoint: lists all runs for historical tracking.
    """
    from web.repositories import AnalysisRunRepository

    run_repo = AnalysisRunRepository()
    runs = run_repo.get_by_deal(deal_id, limit=10)

    return jsonify({
        'runs': [
            {
                'run_id': r.id,
                'run_number': r.run_number,
                'status': r.status,
                'run_type': r.run_type,
                'progress': r.progress or 0,
                'facts_created': r.facts_created or 0,
                'findings_created': r.findings_created or 0,
                'started_at': r.started_at.isoformat() if r.started_at else None,
                'completed_at': r.completed_at.isoformat() if r.completed_at else None,
                'duration_seconds': r.duration_seconds,
            }
            for r in runs
        ],
        'total': len(runs)
    })


# =============================================================================
# Deal Selection API (Spec 02 - Session Persistence Fix)
# =============================================================================

@app.route('/api/deals/list')
@login_required
def list_user_deals():
    """
    Get list of deals accessible by the current user.

    Returns deals owned by user, or in same tenant if multi-tenancy enabled.
    Used by deal selector UI component.

    Returns:
        JSON array of deals with id, name, target_name, status, etc.
    """
    from web.permissions import get_user_deals
    from web.database import Deal

    try:
        deals_query = get_user_deals(current_user, include_tenant_deals=True)
        deals = deals_query.order_by(Deal.last_accessed_at.desc()).limit(50).all()

        return jsonify({
            'success': True,
            'deals': [
                {
                    'id': deal.id,
                    'name': deal.name or deal.target_name,
                    'target_name': deal.target_name,
                    'buyer_name': deal.buyer_name,
                    'status': deal.status,
                    'created_at': deal.created_at.isoformat() if deal.created_at else None,
                    'last_accessed_at': deal.last_accessed_at.isoformat() if deal.last_accessed_at else None,
                    'is_owner': deal.owner_id == current_user.id,
                }
                for deal in deals
            ],
            'total': len(deals)
        })

    except Exception as e:
        logger.error(f"Failed to list deals for user {current_user.id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to load deals'
        }), 500


@app.route('/api/deals/<deal_id>/select', methods=['POST'])
@login_required
def select_deal(deal_id):
    """
    Select a deal as the current working deal.

    This endpoint updates BOTH:
    1. flask_session['current_deal_id'] (ephemeral session)
    2. User.last_deal_id (persistent database)

    This dual persistence ensures deal selection survives session loss.

    Args:
        deal_id: The deal ID to select

    Returns:
        JSON with success status and deal info

    Example:
        POST /api/deals/abc-123/select
        Response: {"success": true, "deal": {...}}
    """
    from web.permissions import user_can_access_deal
    from web.database import Deal
    from datetime import datetime

    try:
        # Get deal
        deal = Deal.query.get(deal_id)
        if not deal or deal.is_deleted:
            return jsonify({
                'success': False,
                'error': 'Deal not found'
            }), 404

        # Check permissions
        if not user_can_access_deal(current_user, deal):
            logger.warning(
                f"User {current_user.id} attempted to access deal {deal_id} without permission"
            )
            return jsonify({
                'success': False,
                'error': 'Access denied'
            }), 403

        # Update session (ephemeral)
        flask_session['current_deal_id'] = deal_id
        flask_session.modified = True

        # Update database (persistent - Spec 01)
        current_user.update_last_deal(deal_id)

        # Update deal access time
        deal.last_accessed_at = datetime.utcnow()

        db.session.commit()

        # Audit log (if enabled)
        if USE_AUDIT_LOGGING:
            try:
                from web.audit_service import audit_service
                audit_service.log(
                    action='select_deal',
                    resource_type='deal',
                    resource_id=deal_id,
                    user_id=current_user.id,
                    details={
                        'deal_name': deal.name or deal.target_name,
                        'method': 'api'
                    }
                )
            except Exception as audit_error:
                logger.warning(f"Audit logging failed: {audit_error}")

        logger.info(f"User {current_user.id} selected deal {deal_id} ({deal.target_name})")

        return jsonify({
            'success': True,
            'deal': {
                'id': deal.id,
                'name': deal.name or deal.target_name,
                'target_name': deal.target_name,
                'buyer_name': deal.buyer_name,
                'status': deal.status,
            },
            'message': f'Selected deal: {deal.name or deal.target_name}'
        })

    except Exception as e:
        logger.error(f"Failed to select deal {deal_id}: {e}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': 'Failed to select deal'
        }), 500


@app.route('/api/audit')
@auth_optional
def get_audit_logs():
    """Get audit log entries (Phase 7)."""
    if not USE_AUDIT_LOGGING:
        return jsonify({'error': 'Audit logging not enabled'}), 400

    try:
        from web.audit_service import audit_service

        # Parse query parameters
        action = request.args.get('action')
        resource_type = request.args.get('resource_type')
        limit = min(int(request.args.get('limit', 100)), 500)
        offset = int(request.args.get('offset', 0))

        events = audit_service.query(
            action=action,
            resource_type=resource_type,
            limit=limit,
            offset=offset
        )

        return jsonify({
            'count': len(events),
            'limit': limit,
            'offset': offset,
            'events': events
        })
    except Exception as e:
        logger.error(f"Error fetching audit logs: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/rate-limit/status')
def rate_limit_status():
    """Get current rate limit status (Phase 7)."""
    if not USE_RATE_LIMITING:
        return jsonify({'enabled': False})

    try:
        from web.rate_limiting import get_rate_limit_status
        return jsonify(get_rate_limit_status())
    except Exception as e:
        return jsonify({'enabled': False, 'error': str(e)})


@app.route('/api/cleanup', methods=['POST'])
def cleanup_tasks():
    """Cleanup old tasks and sessions."""
    tasks_cleaned = 0
    sessions_cleaned = 0

    # Cleanup old tasks
    task_manager.cleanup_old_tasks(max_age_hours=24)

    # Cleanup expired sessions
    sessions_cleaned = session_store.cleanup_expired()

    return jsonify({
        'success': True,
        'tasks_cleaned': tasks_cleaned,
        'sessions_cleaned': sessions_cleaned
    })


@app.route('/api/export/json')
@auth_optional
def export_json():
    """Export all analysis data as JSON.

    Phase 2+: Database-first implementation.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export data.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        gaps = data.get_gaps()
        risks = data.get_risks()
        work_items = data.get_work_items()

        # Build summary from database counts
        summary = data.get_dashboard_summary()

        # Compile all data
        export_data = {
            'summary': {
                'facts': len(facts),
                'gaps': len(gaps),
                'risks': len(risks),
                'work_items': len(work_items),
                'risk_summary': summary.get('risk_summary', {}),
                'work_item_summary': summary.get('work_item_summary', {}),
            },
            'facts': [
                {
                    'id': getattr(f, 'id', '') or getattr(f, 'fact_id', ''),
                    'domain': f.domain,
                    'category': f.category,
                    'item': f.item,
                    'details': f.details,
                    'status': f.status,
                    'entity': getattr(f, 'entity', 'target'),
                }
                for f in facts
            ],
            'gaps': [
                {
                    'id': getattr(g, 'id', '') or getattr(g, 'gap_id', ''),
                    'domain': g.domain,
                    'category': g.category,
                    'description': g.description,
                    'importance': getattr(g, 'importance', 'medium'),
                }
                for g in gaps
            ],
            'risks': [
                {
                    'id': getattr(r, 'id', '') or getattr(r, 'finding_id', ''),
                    'domain': r.domain,
                    'title': r.title,
                    'description': r.description,
                    'severity': r.severity,
                    'category': getattr(r, 'category', ''),
                    'mitigation': getattr(r, 'mitigation', ''),
                    'based_on_facts': getattr(r, 'based_on_facts', []),
                }
                for r in risks
            ],
            'work_items': [
                {
                    'id': getattr(w, 'id', '') or getattr(w, 'finding_id', ''),
                    'domain': w.domain,
                    'title': w.title,
                    'description': w.description,
                    'phase': w.phase,
                    'priority': getattr(w, 'priority', 0),
                    'owner_type': getattr(w, 'owner_type', ''),
                    'cost_estimate': getattr(w, 'cost_estimate', ''),
                    'based_on_facts': getattr(w, 'based_on_facts', []),
                }
                for w in work_items
            ],
        }

        return jsonify(export_data)

    except Exception as e:
        logger.error(f"Export JSON: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500


@app.route('/api/facts')
@auth_optional
def api_facts():
    """Get facts as JSON with optional filtering.

    Phase 2+: Database-first implementation (no fallback).
    """
    domain = request.args.get('domain')
    category = request.args.get('category')
    entity = request.args.get('entity')

    # Database-first: require deal selection
    deal_id = flask_session.get('current_deal_id')
    if not deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to view facts.', 'facts': []}), 400

    try:
        from web.repositories.fact_repository import FactRepository
        repo = FactRepository()
        db_facts = repo.get_by_deal(
            deal_id=deal_id,
            domain=domain,
            entity=entity,
            category=category
        )
        return jsonify({
            'count': len(db_facts),
            'source': 'database',
            'deal_id': deal_id,
            'facts': [f.to_dict() for f in db_facts]
        })
    except Exception as e:
        logger.error(f"API facts: Database query failed: {e}")
        return jsonify({'error': 'Database query failed', 'message': str(e), 'facts': []}), 500


@app.route('/api/facts/<fact_id>/verify', methods=['POST'])
@auth_optional
def verify_fact(fact_id):
    """Mark a fact as verified by a human reviewer.

    Phase 2+: Database-first implementation (no fallback).
    """
    data = request.get_json() or {}
    verified_by = data.get('verified_by', 'web_user')
    verification_note = data.get('note', '')

    try:
        from web.repositories import FactRepository

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if fact:
            repo.verify_fact(fact, verified_by, verification_note)
            return jsonify({
                'status': 'success',
                'source': 'database',
                'message': f'Fact {fact_id} verified by {verified_by}',
                'fact_id': fact_id,
                'verified': True,
                'verified_by': verified_by
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

    except Exception as e:
        logger.error(f"Verify fact failed: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/facts/<fact_id>/unverify', methods=['POST'])
@auth_optional
def unverify_fact(fact_id):
    """Remove verification status from a fact.

    Phase 2+: Database-first implementation (no fallback).
    """
    try:
        from web.repositories import FactRepository
        from web.database import db

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if fact:
            fact.verified = False
            fact.verified_by = None
            fact.verified_at = None
            fact.verification_note = None
            db.session.commit()

            return jsonify({
                'status': 'success',
                'source': 'database',
                'message': f'Verification removed from {fact_id}',
                'fact_id': fact_id,
                'verified': False
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

    except Exception as e:
        logger.error(f"Unverify fact failed: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/facts/verification-stats')
@auth_optional
def verification_stats():
    """Get verification statistics for all facts.

    Phase 2+: Database-first implementation.
    """
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'total': 0, 'verified': 0, 'unverified': 0})

    try:
        from web.repositories import FactRepository

        repo = FactRepository()
        all_facts = repo.get_by_deal(deal_id=current_deal_id)

        verified = sum(1 for f in all_facts if getattr(f, 'verified', False))
        total = len(all_facts)

        return jsonify({
            'total': total,
            'verified': verified,
            'unverified': total - verified,
            'verification_rate': round(verified / total * 100, 1) if total > 0 else 0,
            'source': 'database'
        })

    except Exception as e:
        logger.error(f"Verification stats failed: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/facts/confidence-summary')
@auth_optional
def confidence_summary():
    """Get confidence score summary for all facts.

    Phase 2+: Database-first implementation.
    """
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected'})

    try:
        from web.repositories import FactRepository

        repo = FactRepository()
        all_facts = repo.get_by_deal(deal_id=current_deal_id)

        # Group by confidence level
        by_confidence = {'high': 0, 'medium': 0, 'low': 0}
        for fact in all_facts:
            confidence = getattr(fact, 'confidence', 'medium') or 'medium'
            if confidence in by_confidence:
                by_confidence[confidence] += 1

        return jsonify({
            'by_confidence': by_confidence,
            'total': len(all_facts),
            'source': 'database'
        })

    except Exception as e:
        logger.error(f"Confidence summary failed: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/entity-summary')
@auth_optional
def entity_summary():
    """Get summary of facts by entity (target vs buyer).

    Phase 2+: Database-first implementation.
    Returns counts and breakdown by domain for each entity.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({
            "status": "error",
            "message": "No deal selected"
        }), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()
        all_facts = data.get_all_facts()

        # Count by entity
        target_facts = [f for f in all_facts if getattr(f, 'entity', 'target') == 'target']
        buyer_facts = [f for f in all_facts if getattr(f, 'entity', '') == 'buyer']

        # Count by domain for each entity
        def domain_breakdown(facts):
            breakdown = {}
            for f in facts:
                domain = f.domain
                breakdown[domain] = breakdown.get(domain, 0) + 1
            return breakdown

        # Get document counts from DocumentStore if available
        doc_stats = {"target": 0, "buyer": 0}
        try:
            from stores.document_store import DocumentStore
            doc_store = DocumentStore.get_instance(deal_id=current_deal_id)
            stats = doc_store.get_statistics()
            doc_stats["target"] = stats["by_entity"]["target"]
            doc_stats["buyer"] = stats["by_entity"]["buyer"]
        except Exception:
            pass

        return jsonify({
            "status": "success",
            "summary": {
                "target": {
                    "fact_count": len(target_facts),
                    "document_count": doc_stats["target"],
                    "by_domain": domain_breakdown(target_facts)
                },
                "buyer": {
                    "fact_count": len(buyer_facts),
                    "document_count": doc_stats["buyer"],
                    "by_domain": domain_breakdown(buyer_facts)
                },
                "total_facts": len(all_facts)
            }
        })
    except Exception as e:
        logger.error(f"Entity summary: Database path failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/documents/store')
@auth_optional
def api_document_store():
    """Get document store statistics and list."""
    current_deal_id = flask_session.get('current_deal_id')
    try:
        from stores.document_store import DocumentStore
        doc_store = DocumentStore.get_instance(deal_id=current_deal_id)

        entity = request.args.get('entity')
        authority = request.args.get('authority')

        if authority:
            authority = int(authority)

        docs = doc_store.list_documents(entity=entity, authority_level=authority)
        stats = doc_store.get_statistics()

        return jsonify({
            "status": "success",
            "statistics": stats,
            "documents": [d.to_dict() for d in docs[:100]]  # Limit to 100
        })
    except Exception as e:
        logger.error(f"Error getting document store: {e}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# =============================================================================
# REVIEW QUEUE API (Phase 1 - Validation Workflow)
# =============================================================================

@app.route('/api/review/queue')
@auth_optional
def review_queue():
    """
    Get facts needing review, sorted by priority.

    Phase 2+: Database-first implementation.

    Query params:
        domain: Filter by domain
        limit: Max facts to return (default 50)
        include_skipped: Include previously skipped facts (default false)
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'count': 0, 'facts': []}), 400

    try:
        from web.repositories import FactRepository

        repo = FactRepository()
        limit = int(request.args.get('limit', 50))

        # Get review queue from database
        result = repo.get_review_queue(deal_id=current_deal_id, per_page=limit)
        facts = result.get('items', [])

        # Filter by domain if specified
        domain = request.args.get('domain')
        if domain:
            facts = [f for f in facts if f.domain == domain]

        return jsonify({
            'count': len(facts),
            'facts': [
                {
                    'fact_id': getattr(f, 'id', '') or getattr(f, 'fact_id', ''),
                    'domain': f.domain,
                    'category': f.category,
                    'item': f.item,
                    'details': f.details,
                    'evidence': getattr(f, 'evidence', {}),
                    'source_document': getattr(f, 'source_document', ''),
                    'confidence_score': getattr(f, 'confidence_score', 0.5),
                    'review_priority': getattr(f, 'review_priority', 'medium'),
                    'verification_status': getattr(f, 'verification_status', 'pending'),
                    'verification_note': getattr(f, 'verification_note', ''),
                    'entity': getattr(f, 'entity', 'target')
                }
                for f in facts
            ]
        })
    except Exception as e:
        logger.error(f"API review queue: Database path failed: {e}")
        return jsonify({'error': 'Database query failed', 'count': 0, 'facts': []}), 500


@app.route('/api/review/stats')
@auth_optional
def review_stats():
    """Get comprehensive review queue statistics.

    Phase 2+: Database-first implementation.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected'}), 400

    try:
        from web.repositories import FactRepository

        repo = FactRepository()
        all_facts = repo.get_by_deal(deal_id=current_deal_id)

        # Build stats from database
        stats = {
            'total_facts': len(all_facts),
            'confirmed': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'verified']),
            'incorrect': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'rejected']),
            'needs_info': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'needs_info']),
            'skipped': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'skipped']),
            'by_domain': {}
        }

        # Calculate completion rate
        reviewed = stats['confirmed'] + stats['incorrect'] + stats['skipped']
        stats['completion_rate'] = reviewed / stats['total_facts'] if stats['total_facts'] > 0 else 0

        # Group by domain
        for fact in all_facts:
            domain = fact.domain
            if domain not in stats['by_domain']:
                stats['by_domain'][domain] = {'total': 0, 'confirmed': 0}
            stats['by_domain'][domain]['total'] += 1
            if getattr(fact, 'verification_status', '') == 'verified':
                stats['by_domain'][domain]['confirmed'] += 1

        # Confidence breakdown
        high = len([f for f in all_facts if getattr(f, 'confidence_score', 0) >= 0.8])
        medium = len([f for f in all_facts if 0.5 <= getattr(f, 'confidence_score', 0) < 0.8])
        low = len([f for f in all_facts if getattr(f, 'confidence_score', 0) < 0.5])
        stats['confidence_breakdown'] = {'high': high, 'medium': medium, 'low': low}

        # Average confidence
        if all_facts:
            stats['average_confidence'] = sum(getattr(f, 'confidence_score', 0.5) for f in all_facts) / len(all_facts)
        else:
            stats['average_confidence'] = 0

        stats['remaining_critical'] = len([f for f in all_facts if getattr(f, 'needs_review', False)])
        stats['verified_today'] = 0  # Would need timestamp filtering

        return jsonify(stats)
    except Exception as e:
        logger.error(f"API review stats: Database path failed: {e}")
        return jsonify({'error': 'Database query failed'}), 500


@app.route('/api/review/export-report')
@auth_optional
def export_verification_report():
    """Export verification report as markdown.

    Phase 2+: Database-first implementation.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected'}), 400

    try:
        from web.repositories import FactRepository

        repo = FactRepository()
        all_facts = repo.get_by_deal(deal_id=current_deal_id)

        # Build stats
        stats = {
            'total_facts': len(all_facts),
            'confirmed': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'verified']),
            'incorrect': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'rejected']),
            'needs_info': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'needs_info']),
            'skipped': len([f for f in all_facts if getattr(f, 'verification_status', '') == 'skipped']),
            'by_domain': {}
        }

        reviewed = stats['confirmed'] + stats['incorrect'] + stats['skipped']
        stats['completion_rate'] = reviewed / stats['total_facts'] if stats['total_facts'] > 0 else 0

        for fact in all_facts:
            domain = fact.domain
            if domain not in stats['by_domain']:
                stats['by_domain'][domain] = {'total': 0, 'confirmed': 0}
            stats['by_domain'][domain]['total'] += 1
            if getattr(fact, 'verification_status', '') == 'verified':
                stats['by_domain'][domain]['confirmed'] += 1

        high = len([f for f in all_facts if getattr(f, 'confidence_score', 0) >= 0.8])
        medium = len([f for f in all_facts if 0.5 <= getattr(f, 'confidence_score', 0) < 0.8])
        low = len([f for f in all_facts if getattr(f, 'confidence_score', 0) < 0.5])
        stats['confidence_breakdown'] = {'high': high, 'medium': medium, 'low': low}

        if all_facts:
            stats['average_confidence'] = sum(getattr(f, 'confidence_score', 0.5) for f in all_facts) / len(all_facts)
        else:
            stats['average_confidence'] = 0

        stats['remaining_critical'] = len([f for f in all_facts if getattr(f, 'needs_review', False)])
        stats['verified_today'] = 0

    except Exception as e:
        logger.error(f"Export verification report: Database path failed: {e}")
        return jsonify({'error': 'Database query failed'}), 500

    from datetime import datetime
    date = datetime.now().strftime("%Y-%m-%d")

    report = f"""# Fact Verification Report

**Generated:** {date}

## Summary

| Metric | Count |
|--------|-------|
| Total Facts | {stats['total_facts']} |
| Confirmed | {stats['confirmed']} |
| Flagged Incorrect | {stats['incorrect']} |
| Needs Information | {stats['needs_info']} |
| Skipped | {stats['skipped']} |
| **Completion Rate** | **{int(stats['completion_rate'] * 100)}%** |

## Progress by Domain

| Domain | Confirmed | Total | Rate |
|--------|-----------|-------|------|
"""
    for domain, dstats in stats.get('by_domain', {}).items():
        rate = int((dstats['confirmed'] / dstats['total']) * 100) if dstats['total'] > 0 else 0
        report += f"| {domain.replace('_', ' ')} | {dstats['confirmed']} | {dstats['total']} | {rate}% |\n"

    conf = stats.get('confidence_breakdown', {})
    report += f"""
## Confidence Distribution

- **High Confidence (80%+):** {conf.get('high', 0)} facts
- **Medium Confidence (50-80%):** {conf.get('medium', 0)} facts
- **Low Confidence (<50%):** {conf.get('low', 0)} facts

## Quality Notes

- Average confidence score: {int((stats.get('average_confidence', 0)) * 100)}%
- Critical items remaining: {stats.get('remaining_critical', 0)}
- Verified today: {stats.get('verified_today', 0)}

---
*Report generated by IT Due Diligence Agent*
"""

    from flask import Response
    return Response(
        report,
        mimetype='text/markdown',
        headers={'Content-Disposition': f'attachment; filename=verification_report_{date}.md'}
    )


@app.route('/api/facts/<fact_id>/status', methods=['POST'])
@auth_optional
def update_fact_status(fact_id):
    """
    Update verification status for a fact.

    Phase 2+: Database-first implementation (no fallback).

    JSON body:
        status: pending, confirmed, incorrect, needs_info, skipped
        reviewer_id: Who is making this update
        note: Optional verification note
    """
    data = request.get_json() or {}
    status = data.get('status')
    reviewer_id = data.get('reviewer_id', 'web_user')
    note = data.get('note', '')

    if not status:
        return jsonify({
            'status': 'error',
            'message': 'status field is required'
        }), 400

    valid_statuses = ['pending', 'confirmed', 'incorrect', 'needs_info', 'skipped']
    if status not in valid_statuses:
        return jsonify({
            'status': 'error',
            'message': f'Invalid status. Must be one of: {valid_statuses}'
        }), 400

    try:
        from web.repositories import FactRepository
        from web.database import db
        from datetime import datetime

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if not fact:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

        # Update verification status
        fact.verification_status = status
        fact.verified_by = reviewer_id
        fact.verified_at = datetime.utcnow()
        fact.verification_note = note
        fact.verified = (status == 'confirmed')

        db.session.commit()

        return jsonify({
            'status': 'success',
            'source': 'database',
            'message': f'Fact {fact_id} status updated to {status}',
            'fact_id': fact_id,
            'verification_status': status,
            'verified': fact.verified
        })

    except Exception as e:
        logger.error(f"Update fact status failed: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/facts/<fact_id>/skip', methods=['POST'])
@auth_optional
def skip_fact(fact_id):
    """Mark a fact as skipped (will revisit later).

    Phase 2+: Database-first implementation (no fallback).
    """
    data = request.get_json() or {}
    reviewer_id = data.get('reviewer_id', 'web_user')
    note = data.get('note', 'Skipped for later review')

    try:
        from web.repositories import FactRepository
        from web.database import db
        from datetime import datetime

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if not fact:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

        fact.verification_status = 'skipped'
        fact.verified_by = reviewer_id
        fact.verified_at = datetime.utcnow()
        fact.verification_note = note
        db.session.commit()

        return jsonify({
            'status': 'success',
            'source': 'database',
            'message': f'Fact {fact_id} skipped',
            'fact_id': fact_id
        })

    except Exception as e:
        logger.error(f"Skip fact failed: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/facts/<fact_id>/flag-incorrect', methods=['POST'])
@auth_optional
def flag_incorrect(fact_id):
    """Mark a fact as incorrect.

    Phase 2+: Database-first implementation (no fallback).
    """
    data = request.get_json() or {}
    reviewer_id = data.get('reviewer_id', 'web_user')
    reason = data.get('reason', '')

    if not reason:
        return jsonify({
            'status': 'error',
            'message': 'reason field is required when flagging as incorrect'
        }), 400

    try:
        from web.repositories import FactRepository
        from web.database import db
        from datetime import datetime

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if not fact:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

        fact.verification_status = 'incorrect'
        fact.verified_by = reviewer_id
        fact.verified_at = datetime.utcnow()
        fact.verification_note = f"INCORRECT: {reason}"
        fact.verified = False
        db.session.commit()

        return jsonify({
            'status': 'success',
            'source': 'database',
            'message': f'Fact {fact_id} flagged as incorrect',
            'fact_id': fact_id
        })

    except Exception as e:
        logger.error(f"Flag incorrect failed: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Fact {fact_id} not found'
        }), 404


@app.route('/api/facts/<fact_id>/needs-info', methods=['POST'])
@auth_optional
def needs_more_info(fact_id):
    """Mark a fact as needing more information and optionally create a question.

    Phase 2+: Database-first implementation (no fallback).
    """
    data = request.get_json() or {}
    reviewer_id = data.get('reviewer_id', 'web_user')
    question = data.get('question', '')
    recipient = data.get('recipient', '')

    # Build the note
    note_parts = ["NEEDS INFO"]
    if question:
        note_parts.append(f"Question: {question}")
    if recipient:
        note_parts.append(f"Ask: {recipient}")
    note = " | ".join(note_parts)

    try:
        from web.repositories import FactRepository
        from web.database import db
        from datetime import datetime

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if not fact:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

        fact.verification_status = 'needs_info'
        fact.verified_by = reviewer_id
        fact.verified_at = datetime.utcnow()
        fact.verification_note = note
        db.session.commit()

        # Create an OpenQuestion if we have a question
        question_id = None
        if question:
            try:
                from tools_v2.open_questions import OpenQuestionsStore
                questions_store = OpenQuestionsStore()
                question_id = questions_store.add_question(
                    domain=fact.domain,
                    question=question,
                    context=f"Generated from fact review. Fact: {fact.item}",
                    suggested_recipient=recipient,
                    source_fact_id=fact_id
                )
            except Exception as e:
                logger.warning(f"Could not create OpenQuestion: {e}")

        return jsonify({
            'status': 'success',
            'source': 'database',
            'message': f'Fact {fact_id} marked as needing more info',
            'fact_id': fact_id,
            'question_id': question_id
        })

    except Exception as e:
        logger.error(f"Needs more info failed: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/facts/<fact_id>/note', methods=['PATCH'])
@auth_optional
def update_fact_note(fact_id):
    """Update just the note for a fact without changing status.

    Phase 2+: Database-first implementation (no fallback).
    """
    data = request.get_json() or {}
    note = data.get('note', '')
    reviewer_id = data.get('reviewer_id', 'web_user')

    try:
        from web.repositories import FactRepository
        from web.database import db
        from datetime import datetime

        repo = FactRepository()
        fact = repo.get_by_id(fact_id)

        if not fact:
            return jsonify({
                'status': 'error',
                'message': f'Fact {fact_id} not found'
            }), 404

        fact.verification_note = note
        fact.verified_by = reviewer_id
        fact.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'status': 'success',
            'source': 'database',
            'message': f'Note updated for {fact_id}',
            'fact_id': fact_id
        })

    except Exception as e:
        logger.error(f"Update fact note failed: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/risks')
@auth_optional
def api_risks():
    """Get risks as JSON with optional filtering.

    Phase 2+: Database-first implementation (no fallback).
    """
    domain = request.args.get('domain')
    severity = request.args.get('severity')

    deal_id = flask_session.get('current_deal_id')
    if not deal_id:
        return jsonify({
            'error': 'No deal selected',
            'count': 0,
            'risks': []
        }), 400

    try:
        from web.repositories import FindingRepository
        from web.context import load_deal_context

        load_deal_context(deal_id)
        repo = FindingRepository()
        db_findings = repo.get_by_deal(
            deal_id=deal_id,
            finding_type='risk',
            domain=domain,
            severity=severity
        )
        return jsonify({
            'count': len(db_findings),
            'source': 'database',
            'deal_id': deal_id,
            'risks': [f.to_dict() for f in db_findings]
        })
    except Exception as e:
        logger.error(f"API risks failed: {e}")
        return jsonify({
            'error': str(e),
            'count': 0,
            'deal_id': deal_id,
            'risks': []
        }), 500


# =============================================================================
# NARRATIVE API ROUTES (for Report View toggle)
# =============================================================================

@app.route('/api/narrative/<domain>')
@auth_optional
def api_get_narrative(domain):
    """
    Get or generate narrative for a domain.

    Returns cached narrative if available, otherwise generates new one.
    """
    from web.narrative_service import (
        generate_domain_narrative,
        get_placeholder_narrative
    )
    from config_v2 import ANTHROPIC_API_KEY

    s = get_session()
    session_id = get_or_create_session_id(flask_session)

    # Check if we have data to generate narrative from
    if not s or not s.fact_store or not s.fact_store.facts:
        return jsonify(get_placeholder_narrative(domain))

    # Check if domain has facts
    domain_facts = [f for f in s.fact_store.facts if f.domain == domain]
    if not domain_facts:
        return jsonify(get_placeholder_narrative(domain))

    # Generate narrative
    result = generate_domain_narrative(
        domain=domain,
        fact_store=s.fact_store,
        reasoning_store=s.reasoning_store,
        api_key=ANTHROPIC_API_KEY,
        force_regenerate=False,
        session_id=session_id
    )

    return jsonify(result)


@app.route('/api/narrative/<domain>/regenerate', methods=['POST'])
@auth_optional
def api_regenerate_narrative(domain):
    """
    Force regenerate narrative for a domain.

    Bypasses cache and generates fresh narrative.
    """
    from web.narrative_service import (
        generate_domain_narrative,
        clear_narrative_cache,
        get_placeholder_narrative
    )
    from config_v2 import ANTHROPIC_API_KEY

    s = get_session()
    session_id = get_or_create_session_id(flask_session)

    # Check if we have data to generate narrative from
    if not s or not s.fact_store or not s.fact_store.facts:
        return jsonify(get_placeholder_narrative(domain))

    # Clear cache for this domain
    clear_narrative_cache(session_id, domain)

    # Generate fresh narrative
    result = generate_domain_narrative(
        domain=domain,
        fact_store=s.fact_store,
        reasoning_store=s.reasoning_store,
        api_key=ANTHROPIC_API_KEY,
        force_regenerate=True,
        session_id=session_id
    )

    return jsonify(result)


@app.route('/api/export/excel')
@auth_optional
def export_excel():
    """Export analysis data as Excel file.

    Phase 2+: Database-first implementation.
    """
    from io import BytesIO
    from flask import send_file

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export data.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        gaps = data.get_gaps()
        risks = data.get_risks()
        work_items = data.get_work_items()
        summary = data.get_dashboard_summary()
    except Exception as e:
        logger.error(f"Export Excel: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    # Create workbook
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    risk_summary = summary.get('risk_summary', {})
    summary_data = [
        ["IT Due Diligence Analysis Summary"],
        [""],
        ["Metric", "Value"],
        ["Facts Discovered", len(facts)],
        ["Information Gaps", len(gaps)],
        ["Risks Identified", len(risks)],
        ["Work Items", len(work_items)],
        [""],
        ["Risk Summary"],
        ["Critical", risk_summary.get('critical', 0)],
        ["High", risk_summary.get('high', 0)],
        ["Medium", risk_summary.get('medium', 0)],
        ["Low", risk_summary.get('low', 0)],
    ]
    for row in summary_data:
        ws_summary.append(row)

    # Risks sheet
    ws_risks = wb.create_sheet("Risks")
    risk_headers = ["ID", "Domain", "Severity", "Title", "Description", "Mitigation"]
    ws_risks.append(risk_headers)
    for col, header in enumerate(risk_headers, 1):
        cell = ws_risks.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for risk in risks:
        risk_id = getattr(risk, 'id', '') or getattr(risk, 'finding_id', '')
        ws_risks.append([
            risk_id,
            risk.domain,
            (risk.severity or '').upper(),
            risk.title,
            risk.description or "",
            getattr(risk, 'mitigation', '') or ""
        ])

    # Work Items sheet
    ws_items = wb.create_sheet("Work Items")
    item_headers = ["ID", "Domain", "Phase", "Priority", "Title", "Description", "Cost Estimate", "Owner"]
    ws_items.append(item_headers)
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for wi in work_items:
        wi_id = getattr(wi, 'id', '') or getattr(wi, 'finding_id', '')
        ws_items.append([
            wi_id,
            wi.domain,
            wi.phase,
            getattr(wi, 'priority', 0),
            wi.title,
            wi.description or "",
            getattr(wi, 'cost_estimate', ''),
            getattr(wi, 'owner_type', '')
        ])

    # Facts sheet
    ws_facts = wb.create_sheet("Facts")
    fact_headers = ["ID", "Domain", "Category", "Item", "Status"]
    ws_facts.append(fact_headers)
    for col, header in enumerate(fact_headers, 1):
        cell = ws_facts.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for fact in facts:
        fact_id = getattr(fact, 'id', '') or getattr(fact, 'fact_id', '')
        ws_facts.append([
            fact_id,
            fact.domain,
            fact.category,
            fact.item,
            fact.status
        ])

    # Gaps sheet
    ws_gaps = wb.create_sheet("Gaps")
    gap_headers = ["ID", "Domain", "Category", "Importance", "Description"]
    ws_gaps.append(gap_headers)
    for col, header in enumerate(gap_headers, 1):
        cell = ws_gaps.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for gap in gaps:
        gap_id = getattr(gap, 'id', '') or getattr(gap, 'gap_id', '')
        ws_gaps.append([
            gap_id,
            gap.domain,
            gap.category,
            getattr(gap, 'importance', 'medium'),
            gap.description
        ])

    # Auto-adjust column widths
    for ws in [ws_summary, ws_risks, ws_items, ws_facts, ws_gaps]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from datetime import datetime
    filename = f"it_due_diligence_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/export/dossiers/<domain>')
@app.route('/api/export/dossiers/<domain>/<entity>')
@auth_optional
def export_dossiers(domain, entity=None):
    """Export comprehensive dossiers for a specific domain.

    Phase 2+: Database-first implementation.
    Phase 8: Entity filtering and comparison view support.

    Args:
        domain: Domain to export (applications, infrastructure, etc.)
        entity: Optional entity filter ("target", "buyer", or "comparison")
    """
    from flask import send_file, request
    from io import BytesIO
    from datetime import datetime

    format_type = request.args.get('format', 'html')  # html, md, json
    # Allow entity in query param as well as path
    if not entity:
        entity = request.args.get('entity')  # None means show all

    try:
        from services.inventory_dossier import DossierBuilder, DossierMarkdownExporter, DossierJSONExporter, DossierHTMLExporter, ExportGate
    except ImportError as e:
        return jsonify({'error': f'Dossier service not available: {e}'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export dossiers.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        risks = data.get_risks()
        work_items = data.get_work_items()
    except Exception as e:
        logger.error(f"Export Dossiers: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    # Build facts as flat list (DossierBuilder expects List[Dict], not Dict[str, List])
    facts_list = [
        {
            'fact_id': getattr(fact, 'id', '') or getattr(fact, 'fact_id', ''),
            'domain': fact.domain,
            'category': fact.category,
            'item': fact.item,
            'details': fact.details,
            'status': fact.status,
            'evidence': getattr(fact, 'evidence', '') or '',
            'entity': getattr(fact, 'entity', 'target'),
        }
        for fact in facts
    ]

    findings_data = {
        'risks': [
            {
                'finding_id': getattr(r, 'id', '') or getattr(r, 'finding_id', ''),
                'domain': r.domain,
                'title': r.title,
                'description': r.description,
                'severity': r.severity,
                'category': getattr(r, 'category', ''),
                'mitigation': getattr(r, 'mitigation', ''),
                'based_on_facts': getattr(r, 'based_on_facts', []),
            }
            for r in risks
        ],
        'work_items': [
            {
                'finding_id': getattr(wi, 'id', '') or getattr(wi, 'finding_id', ''),
                'domain': wi.domain,
                'title': wi.title,
                'description': wi.description,
                'phase': wi.phase,
                'priority': getattr(wi, 'priority', 0),
                'cost_estimate': getattr(wi, 'cost_estimate', ''),
                'owner_type': getattr(wi, 'owner_type', ''),
                'based_on_facts': getattr(wi, 'based_on_facts', []),
            }
            for wi in work_items
        ]
    }

    # Build dossiers
    builder = DossierBuilder(facts_list, findings_data)

    # Map domain names
    domain_map = {
        'applications': 'applications',
        'infrastructure': 'infrastructure',
        'cybersecurity': 'cybersecurity',
        'network': 'network',
        'identity': 'identity_access',
        'identity_access': 'identity_access',
        'organization': 'organization',
        'all': 'all'
    }

    actual_domain = domain_map.get(domain.lower(), domain)

    # Phase 8: Handle comparison view
    if entity == 'comparison':
        from services.delta_comparator import DeltaComparator, render_delta_html

        if actual_domain == 'all':
            return jsonify({'error': 'Comparison view not supported for all domains. Select a specific domain.'}), 400

        # Build dossiers for both entities
        target_dossiers = builder.build_domain_dossiers(actual_domain, entity='target')
        buyer_dossiers = builder.build_domain_dossiers(actual_domain, entity='buyer')

        # Compute delta
        comparator = DeltaComparator()
        deltas = comparator.match_items(target_dossiers, buyer_dossiers)

        # Render to HTML
        content = render_delta_html(deltas, actual_domain)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        buffer = BytesIO(content.encode('utf-8'))
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='text/html',
            as_attachment=True,
            download_name=f"delta_{actual_domain}_{timestamp}.html"
        )

    # Normal export: filter by entity if specified
    if actual_domain == 'all':
        # Export all domains
        all_dossiers = {}
        for d in ['applications', 'infrastructure', 'cybersecurity', 'network', 'identity_access', 'organization']:
            all_dossiers[d] = builder.build_domain_dossiers(d, entity=entity)
    else:
        all_dossiers = {actual_domain: builder.build_domain_dossiers(actual_domain, entity=entity)}

    # Phase 9: Run export gate check
    export_gate = ExportGate()
    readiness = export_gate.check_readiness(all_dossiers)
    readiness_banner = export_gate.render_readiness_banner(readiness)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if format_type == 'md':
        content_parts = []
        # Add readiness header for markdown
        if not readiness.is_ready:
            content_parts.append("# ⛔ EXPORT BLOCKED - DATA QUALITY ISSUES\n\n" +
                               '\n'.join(f"- {issue}" for issue in readiness.blocking_issues) +
                               "\n\n---\n")
        elif readiness.warnings:
            content_parts.append("# ⚠️ DATA QUALITY WARNINGS\n\n" +
                               '\n'.join(f"- {w}" for w in readiness.warnings) +
                               "\n\n---\n")
        for d, dossiers in all_dossiers.items():
            content_parts.append(DossierMarkdownExporter.export_to_string(dossiers, d))
        content = '\n\n---\n\n'.join(content_parts)

        buffer = BytesIO(content.encode('utf-8'))
        filename = f"dossiers_{domain}_{timestamp}.md"
        mimetype = 'text/markdown'

    elif format_type == 'json':
        import json
        all_data = {
            'export_readiness': {
                'is_ready': readiness.is_ready,
                'blocking_issues': readiness.blocking_issues,
                'warnings': readiness.warnings,
                'unknown_entity_count': readiness.unknown_entity_count,
                'unresolved_conflicts': readiness.unresolved_conflicts,
                'average_completeness': readiness.average_completeness,
            },
            'domains': {}
        }
        for d, dossiers in all_dossiers.items():
            all_data['domains'][d] = json.loads(DossierJSONExporter.export_to_string(dossiers, d))
        content = json.dumps(all_data, indent=2)

        buffer = BytesIO(content.encode('utf-8'))
        filename = f"dossiers_{domain}_{timestamp}.json"
        mimetype = 'application/json'

    else:  # html
        content_parts = []
        for d, dossiers in all_dossiers.items():
            content_parts.append(DossierHTMLExporter.export_to_string(dossiers, d))
        # Combine HTML exports with readiness banner
        if len(content_parts) == 1:
            # Inject readiness banner into single domain HTML
            content = content_parts[0]
            if readiness_banner:
                # Insert banner after <body> tag
                content = content.replace('<body>', f'<body>\n{readiness_banner}\n', 1)
        else:
            # Create a combined HTML with navigation and readiness banner
            content = f"""<!DOCTYPE html>
<html><head><title>IT Due Diligence - All Dossiers</title>
<style>
body {{ font-family: -apple-system, sans-serif; margin: 20px; }}
.domain-nav {{ background: #f5f5f5; padding: 10px; margin-bottom: 20px; border-radius: 8px; }}
.domain-nav a {{ margin-right: 15px; color: #0066cc; text-decoration: none; }}
.domain-section {{ margin-bottom: 40px; padding-top: 20px; border-top: 2px solid #ddd; }}
</style></head><body>
{readiness_banner}
<h1>IT Due Diligence - Complete Dossiers</h1>
<div class="domain-nav">
{''.join(f'<a href="#domain-{d}">{d.replace("_", " ").title()}</a>' for d in all_dossiers.keys())}
</div>
{''.join(f'<div id="domain-{d}" class="domain-section">{html}</div>' for d, html in zip(all_dossiers.keys(), content_parts))}
</body></html>"""

        buffer = BytesIO(content.encode('utf-8'))
        filename = f"dossiers_{domain}_{timestamp}.html"
        mimetype = 'text/html'

    buffer.seek(0)
    return send_file(
        buffer,
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/export/inventory/<domain>')
@auth_optional
def export_inventory(domain):
    """Export inventory for a specific domain using the export service.

    Phase 2+: Database-first implementation.
    """
    from flask import send_file, request
    from io import BytesIO
    from datetime import datetime

    format_type = request.args.get('format', 'excel')  # excel, md, csv

    try:
        from services.export_service import ExportService
    except ImportError as e:
        return jsonify({'error': f'Export service not available: {e}'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export inventory.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        risks = data.get_risks()
        work_items = data.get_work_items()
    except Exception as e:
        logger.error(f"Export Inventory: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    # Build facts as flat list
    facts_list = [
        {
            'fact_id': getattr(fact, 'id', '') or getattr(fact, 'fact_id', ''),
            'domain': fact.domain,
            'category': fact.category,
            'item': fact.item,
            'details': fact.details,
            'status': fact.status,
            'evidence': getattr(fact, 'evidence', '') or '',
            'entity': getattr(fact, 'entity', 'target'),
        }
        for fact in facts
    ]

    # Build findings as flat list (ExportService expects List[Dict], not Dict)
    findings_list = []
    for r in risks:
        findings_list.append({
            'finding_id': getattr(r, 'id', '') or getattr(r, 'finding_id', ''),
            'domain': r.domain,
            'title': r.title,
            'description': r.description,
            'severity': r.severity,
            'mitigation': getattr(r, 'mitigation', ''),
            'based_on_facts': getattr(r, 'based_on_facts', []),
            'type': 'risk',
        })
    for wi in work_items:
        findings_list.append({
            'finding_id': getattr(wi, 'id', '') or getattr(wi, 'finding_id', ''),
            'domain': wi.domain,
            'title': wi.title,
            'description': wi.description,
            'phase': wi.phase,
            'based_on_facts': getattr(wi, 'based_on_facts', []),
            'type': 'work_item',
        })

    export_service = ExportService(facts_list, findings_list)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    domain_map = {
        'applications': 'applications',
        'infrastructure': 'infrastructure',
        'cybersecurity': 'cybersecurity',
        'network': 'network',
        'organization': 'organization',
    }

    actual_domain = domain_map.get(domain.lower())
    if not actual_domain:
        return jsonify({'error': f'Unknown domain: {domain}'}), 400

    if format_type == 'excel':
        buffer = export_service.export_domain_excel(actual_domain)
        filename = f"{actual_domain}_inventory_{timestamp}.xlsx"
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif format_type == 'md':
        content = export_service.export_domain_markdown(actual_domain)
        buffer = BytesIO(content.encode('utf-8'))
        filename = f"{actual_domain}_inventory_{timestamp}.md"
        mimetype = 'text/markdown'
    elif format_type == 'csv':
        content = export_service.export_domain_csv(actual_domain)
        buffer = BytesIO(content.encode('utf-8'))
        filename = f"{actual_domain}_inventory_{timestamp}.csv"
        mimetype = 'text/csv'
    else:
        return jsonify({'error': f'Unknown format: {format_type}'}), 400

    buffer.seek(0)
    return send_file(
        buffer,
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/export/work-items')
@auth_optional
def export_work_items():
    """Export work items as Markdown.

    Phase 2+: Database-first implementation.
    """
    from flask import send_file
    from io import BytesIO
    from datetime import datetime

    try:
        from services.export_service import ExportService
    except ImportError as e:
        return jsonify({'error': f'Export service not available: {e}'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export work items.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        work_items = data.get_work_items()
    except Exception as e:
        logger.error(f"Export Work Items: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    # Build data
    facts_list = [
        {
            'fact_id': getattr(fact, 'id', '') or getattr(fact, 'fact_id', ''),
            'domain': fact.domain,
            'category': fact.category,
            'item': fact.item,
            'details': fact.details,
        }
        for fact in facts
    ]

    findings_list = []
    for wi in work_items:
        findings_list.append({
            'finding_id': getattr(wi, 'id', '') or getattr(wi, 'finding_id', ''),
            'domain': wi.domain,
            'title': wi.title,
            'description': wi.description,
            'phase': wi.phase,
            'priority': getattr(wi, 'priority', 0),
            'cost_estimate': getattr(wi, 'cost_estimate', ''),
            'owner_type': getattr(wi, 'owner_type', ''),
            'based_on_facts': getattr(wi, 'based_on_facts', []),
            'type': 'work_item',
        })

    export_service = ExportService(facts_list, findings_list)
    content = export_service.export_work_items_markdown()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    buffer = BytesIO(content.encode('utf-8'))
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='text/markdown',
        as_attachment=True,
        download_name=f"work_items_{timestamp}.md"
    )


@app.route('/api/export/risks')
@auth_optional
def export_risks():
    """Export risks as Markdown.

    Phase 2+: Database-first implementation.
    """
    from flask import send_file
    from io import BytesIO
    from datetime import datetime

    try:
        from services.export_service import ExportService
    except ImportError as e:
        return jsonify({'error': f'Export service not available: {e}'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export risks.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        risks = data.get_risks()
    except Exception as e:
        logger.error(f"Export Risks: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    facts_list = [
        {
            'fact_id': getattr(fact, 'id', '') or getattr(fact, 'fact_id', ''),
            'domain': fact.domain,
            'category': fact.category,
            'item': fact.item,
            'details': fact.details,
        }
        for fact in facts
    ]

    findings_list = []
    for r in risks:
        findings_list.append({
            'finding_id': getattr(r, 'id', '') or getattr(r, 'finding_id', ''),
            'domain': r.domain,
            'title': r.title,
            'description': r.description,
            'severity': r.severity,
            'mitigation': getattr(r, 'mitigation', ''),
            'based_on_facts': getattr(r, 'based_on_facts', []),
            'type': 'risk',
        })

    export_service = ExportService(facts_list, findings_list)
    content = export_service.export_risks_markdown()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    buffer = BytesIO(content.encode('utf-8'))
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='text/markdown',
        as_attachment=True,
        download_name=f"risks_{timestamp}.md"
    )


@app.route('/api/export/vdr-requests')
@auth_optional
def export_vdr_requests():
    """Export VDR/information gap requests as Markdown.

    Phase 2+: Database-first implementation.
    """
    from flask import send_file
    from io import BytesIO
    from datetime import datetime

    try:
        from services.export_service import ExportService
    except ImportError as e:
        return jsonify({'error': f'Export service not available: {e}'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export VDR requests.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        gaps = data.get_gaps()
    except Exception as e:
        logger.error(f"Export VDR Requests: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    # Build gaps list
    gaps_list = [
        {
            'gap_id': getattr(gap, 'id', '') or getattr(gap, 'gap_id', ''),
            'domain': gap.domain,
            'category': gap.category,
            'description': gap.description,
            'importance': getattr(gap, 'importance', 'medium'),
        }
        for gap in gaps
    ]

    export_service = ExportService([], [])
    content = export_service.export_vdr_requests_markdown(gaps_list)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    buffer = BytesIO(content.encode('utf-8'))
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='text/markdown',
        as_attachment=True,
        download_name=f"vdr_requests_{timestamp}.md"
    )


@app.route('/api/export/executive-summary')
@auth_optional
def export_executive_summary():
    """Export executive summary as Markdown.

    Phase 2+: Database-first implementation.
    """
    from flask import send_file
    from io import BytesIO
    from datetime import datetime

    try:
        from services.export_service import ExportService
    except ImportError as e:
        return jsonify({'error': f'Export service not available: {e}'}), 500

    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({'error': 'No deal selected', 'message': 'Please select a deal to export executive summary.'}), 400

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context
        from web.repositories import DealRepository

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        risks = data.get_risks()
        work_items = data.get_work_items()
        gaps = data.get_gaps()

        # Get deal info for context
        deal_repo = DealRepository()
        deal = deal_repo.get_by_id(current_deal_id)
    except Exception as e:
        logger.error(f"Export Executive Summary: Database path failed: {e}")
        return jsonify({'error': 'Export failed', 'message': str(e)}), 500

    # Build facts list
    facts_list = [
        {
            'fact_id': getattr(fact, 'id', '') or getattr(fact, 'fact_id', ''),
            'domain': fact.domain,
            'category': fact.category,
            'item': fact.item,
            'details': fact.details,
        }
        for fact in facts
    ]

    # Build findings list
    findings_list = []
    for r in risks:
        findings_list.append({
            'finding_id': getattr(r, 'id', '') or getattr(r, 'finding_id', ''),
            'domain': r.domain,
            'title': r.title,
            'description': r.description,
            'severity': r.severity,
            'mitigation': getattr(r, 'mitigation', ''),
            'type': 'risk',
        })
    for wi in work_items:
        findings_list.append({
            'finding_id': getattr(wi, 'id', '') or getattr(wi, 'finding_id', ''),
            'domain': wi.domain,
            'title': wi.title,
            'description': wi.description,
            'phase': wi.phase,
            'priority': getattr(wi, 'priority', 0),
            'type': 'work_item',
        })

    # Build gaps list
    gaps_list = [
        {
            'gap_id': getattr(gap, 'id', '') or getattr(gap, 'gap_id', ''),
            'domain': gap.domain,
            'category': gap.category,
            'description': gap.description,
            'importance': getattr(gap, 'importance', 'medium'),
        }
        for gap in gaps
    ]

    # Get deal context from database
    deal_context = {}
    if deal:
        deal_context = {
            'target_name': deal.name or '',
            'buyer_name': '',  # Not currently stored in Deal model
        }

    export_service = ExportService(facts_list, findings_list)
    content = export_service.export_executive_summary(gaps_list, deal_context)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    buffer = BytesIO(content.encode('utf-8'))
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='text/markdown',
        as_attachment=True,
        download_name=f"executive_summary_{timestamp}.md"
    )


@app.route('/exports')
@auth_optional
def exports_page():
    """Export center page with all export options.

    Phase 2+: Database-first implementation.
    """
    # Database-first: require deal selection
    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        flash('Please select a deal to access exports.', 'info')
        return redirect(url_for('deals.deals_list_page'))

    try:
        from web.deal_data import DealData
        from web.context import load_deal_context

        load_deal_context(current_deal_id)
        data = DealData()

        facts = data.get_all_facts()
        risks = data.get_risks()
        work_items = data.get_work_items()
        gaps = data.get_gaps()

        # Build summary
        dashboard = data.get_dashboard_summary()
        summary = {
            'facts': len(facts),
            'gaps': len(gaps),
            'risks': len(risks),
            'work_items': len(work_items),
            'risk_summary': dashboard.get('risk_summary', {}),
            'work_item_summary': dashboard.get('work_item_summary', {}),
        }

        # Get domain stats
        domain_stats = {}
        for fact in facts:
            domain = fact.domain
            if domain not in domain_stats:
                domain_stats[domain] = {'facts': 0, 'risks': 0, 'work_items': 0}
            domain_stats[domain]['facts'] += 1

        for risk in risks:
            domain = risk.domain
            if domain not in domain_stats:
                domain_stats[domain] = {'facts': 0, 'risks': 0, 'work_items': 0}
            domain_stats[domain]['risks'] += 1

        for wi in work_items:
            domain = wi.domain
            if domain not in domain_stats:
                domain_stats[domain] = {'facts': 0, 'risks': 0, 'work_items': 0}
            domain_stats[domain]['work_items'] += 1

    except Exception as e:
        logger.error(f"Exports page: Database path failed: {e}")
        flash('Error loading export data. Please try again.', 'error')
        return redirect(url_for('dashboard'))

    return render_template('exports.html',
                         summary=summary,
                         domain_stats=domain_stats)


# =============================================================================
# Document Management & Incremental Updates
# =============================================================================

@app.route('/documents')
@auth_optional
def documents_page():
    """Document management and incremental update page."""
    s = get_session()

    # A2 FIX: Query documents from database (source of truth)
    docs = []
    stats = {'total_documents': 0, 'total_facts_linked': 0, 'analysis_runs': 0}
    runs = []

    current_deal_id = flask_session.get('current_deal_id')

    if USE_DATABASE and current_deal_id:
        try:
            from enum import Enum

            # Create a simple status enum wrapper for template compatibility
            class StatusWrapper:
                def __init__(self, status_str):
                    self.value = status_str

            doc_repo = DocumentRepository()
            db_docs = doc_repo.get_by_deal(current_deal_id, current_only=True)

            # Convert DB documents to object-like dicts for template
            total_facts_linked = 0
            for db_doc in db_docs:
                # Create an object-like wrapper the template can use with dot notation
                class DocWrapper:
                    pass
                doc_obj = DocWrapper()
                doc_obj.doc_id = db_doc.id
                doc_obj.filename = db_doc.filename
                doc_obj.entity = db_doc.entity
                doc_obj.status = StatusWrapper(db_doc.status or 'pending')
                doc_obj.document_type = db_doc.document_type or 'document'
                doc_obj.version = db_doc.version or 1
                # Count facts linked to this document from FactStore
                if s and s.fact_store:
                    doc_facts = s.fact_store.get_facts_by_source(db_doc.filename)
                    doc_obj.fact_count = len(doc_facts)
                    total_facts_linked += len(doc_facts)
                else:
                    doc_obj.fact_count = 0
                doc_obj.last_processed_at = db_doc.processed_at.isoformat() if db_doc.processed_at else None
                docs.append(doc_obj)

            # Calculate stats from DB docs
            stats['total_documents'] = len(db_docs)
            stats['total_facts_linked'] = total_facts_linked
            stats['analysis_runs'] = len(s.fact_store.get_source_documents()) if s and s.fact_store else 0

            logger.info(f"A2: Loaded {len(db_docs)} documents from DB for deal {current_deal_id}")
        except Exception as e:
            logger.warning(f"A2: Failed to load documents from DB: {e}")
            # Fall back to session-based registry
            if not hasattr(s, 'document_registry'):
                from tools_v2.document_registry import DocumentRegistry
                s.document_registry = DocumentRegistry()
            docs = s.document_registry.get_all_documents()
            stats = s.document_registry.get_stats()
    else:
        # Fallback: use session-based registry when DB not enabled
        if not hasattr(s, 'document_registry'):
            from tools_v2.document_registry import DocumentRegistry
            s.document_registry = DocumentRegistry()
        docs = s.document_registry.get_all_documents()
        stats = s.document_registry.get_stats()
        runs = s.document_registry.get_analysis_runs()

    # Get "What's New" data if we have a fact merger
    whats_new = {"new": [], "updated": [], "conflicts": []}
    pending_conflicts = []
    if hasattr(s, 'fact_merger'):
        whats_new = s.fact_merger.get_whats_new()
        pending_conflicts = s.fact_merger.get_pending_conflicts()

    # Get pending changes from document processor
    pending_changes = getattr(s, 'pending_changes', {"tier1": [], "tier2": [], "tier3": []})

    # Also check document processor for pending changes
    if hasattr(s, 'document_processor'):
        for tier in ["tier1", "tier2", "tier3"]:
            proc_changes = s.document_processor.pending_changes.get(tier, [])
            if proc_changes:
                if tier not in pending_changes:
                    pending_changes[tier] = []
                pending_changes[tier].extend(proc_changes)
        # Sync back to session
        s.pending_changes = pending_changes

    tier_stats = {
        "tier1": len(pending_changes.get("tier1", [])),
        "tier2": len(pending_changes.get("tier2", [])),
        "tier3": len(pending_changes.get("tier3", [])),
        "total": sum(len(pending_changes.get(t, [])) for t in ["tier1", "tier2", "tier3"])
    }

    return render_template('documents/manage.html',
                          documents=docs,
                          stats=stats,
                          analysis_runs=runs,
                          whats_new=whats_new,
                          pending_conflicts=pending_conflicts,
                          tier_stats=tier_stats,
                          data_source='analysis' if docs else 'none')


@app.route('/api/documents')
@auth_optional
def api_documents():
    """Get all registered documents."""
    from tools_v2.document_registry import DocumentRegistry
    from config_v2 import OUTPUT_DIR

    current_deal_id = flask_session.get('current_deal_id')

    # A2 FIX: Query from database when available
    if USE_DATABASE and current_deal_id:
        try:
            doc_repo = DocumentRepository()
            db_docs = doc_repo.get_by_deal(current_deal_id, current_only=True)

            docs = []
            for db_doc in db_docs:
                docs.append({
                    'id': db_doc.id,
                    'filename': db_doc.filename,
                    'entity': db_doc.entity,
                    'status': db_doc.status,
                    'uploaded_at': db_doc.uploaded_at.isoformat() if db_doc.uploaded_at else '',
                    'file_size': db_doc.file_size or 0,
                    'authority_level': db_doc.authority_level or 1,
                    'page_count': db_doc.page_count or 0
                })

            stats = {
                'total': len(db_docs),
                'target': len([d for d in db_docs if d.entity == 'target']),
                'buyer': len([d for d in db_docs if d.entity == 'buyer']),
                'pending': len([d for d in db_docs if d.status == 'pending']),
                'completed': len([d for d in db_docs if d.status == 'completed'])
            }

            return jsonify({"documents": docs, "stats": stats})
        except Exception as e:
            logger.warning(f"A2: Failed to load documents from DB for API: {e}")
            # Fall through to session-based

    # Fallback to session-based registry
    s = get_session()

    if not hasattr(s, 'document_registry'):
        registry_file = OUTPUT_DIR / "document_registry.json"
        if registry_file.exists():
            try:
                s.document_registry = DocumentRegistry()
                s.document_registry.load_from_file(str(registry_file))
            except Exception as e:
                logger.warning(f"Failed to load document registry: {e}")
                return jsonify({"documents": [], "stats": {}})
        else:
            return jsonify({"documents": [], "stats": {}})

    docs = [d.to_dict() for d in s.document_registry.get_all_documents()]
    stats = s.document_registry.get_stats()

    return jsonify({"documents": docs, "stats": stats})


@app.route('/api/documents/backfill', methods=['POST'])
@auth_optional
def api_backfill_documents():
    """A2 FIX: Backfill documents from manifest.json to database.

    This syncs any documents that were uploaded before the DB insert was added.
    Idempotent: skips documents that already exist in DB by hash.
    """
    if not USE_DATABASE:
        return jsonify({"error": "Database not enabled"}), 400

    current_deal_id = flask_session.get('current_deal_id')
    if not current_deal_id:
        return jsonify({"error": "No active deal selected"}), 400

    from stores.document_store import DocumentStore

    try:
        doc_store = DocumentStore.get_instance(deal_id=current_deal_id)
        manifest_docs = doc_store.get_all_documents()

        doc_repo = DocumentRepository()
        synced = 0
        skipped = 0

        for doc in manifest_docs:
            # Check if already in DB by hash
            existing = doc_repo.get_by_hash(current_deal_id, doc.hash_sha256)
            if existing:
                skipped += 1
                continue

            # Insert into DB
            doc_repo.create_document(
                deal_id=current_deal_id,
                filename=doc.filename,
                file_hash=doc.hash_sha256,
                storage_path=doc.raw_file_path,
                entity=doc.entity,
                file_size=doc.file_size_bytes,
                mime_type=doc.mime_type,
                authority_level=doc.authority_level,
                uploaded_by=doc.uploaded_by
            )
            synced += 1

        logger.info(f"A2 Backfill: synced={synced}, skipped={skipped} for deal {current_deal_id}")
        return jsonify({
            "success": True,
            "synced": synced,
            "skipped": skipped,
            "message": f"Backfilled {synced} documents ({skipped} already existed)"
        })

    except Exception as e:
        logger.error(f"A2 Backfill error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/documents/upload', methods=['POST'])
@auth_optional
def api_upload_documents():
    """Upload new documents for incremental analysis."""
    from tools_v2.document_registry import DocumentRegistry, ChangeType
    from tools_v2.fact_merger import FactMerger
    from tools_v2.document_processor import DocumentProcessor, ProcessingPriority
    from config_v2 import OUTPUT_DIR

    s = get_session()

    # Registry file path for persistence
    registry_file = OUTPUT_DIR / "document_registry.json"

    # Initialize document registry - load from file if exists
    if not hasattr(s, 'document_registry'):
        s.document_registry = DocumentRegistry()
        if registry_file.exists():
            try:
                s.document_registry.load_from_file(str(registry_file))
                logger.info(f"Loaded document registry with {len(s.document_registry.documents)} documents")
            except Exception as e:
                logger.warning(f"Failed to load document registry: {e}")

    if not hasattr(s, 'fact_merger'):
        s.fact_merger = FactMerger(s.fact_store)

    if not hasattr(s, 'document_processor'):
        s.document_processor = DocumentProcessor(
            document_registry=s.document_registry,
            fact_store=s.fact_store,
            fact_merger=s.fact_merger
        )
        # Load pending changes from file
        s.document_processor.load_pending_changes()
        # Start background worker
        s.document_processor.start_worker()

    if 'files' not in request.files:
        return jsonify({"status": "error", "message": "No files provided"}), 400

    files = request.files.getlist('files')
    results = []

    # Create uploads directory if needed
    uploads_dir = DATA_DIR / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)

    for file in files:
        if file.filename:
            # Save file to disk
            safe_filename = file.filename.replace('/', '_').replace('\\', '_')
            file_path = uploads_dir / safe_filename
            file.save(str(file_path))

            # Read content for hash
            with open(file_path, 'rb') as f:
                content = f.read()
            content_text = content.decode('utf-8', errors='ignore')

            # Register document
            record, change_type = s.document_registry.register_document(
                filename=file.filename,
                content=content_text,
                file_path=str(file_path),
                file_size=len(content)
            )

            # Queue for processing if new or updated
            if change_type in [ChangeType.NEW, ChangeType.UPDATED]:
                priority = ProcessingPriority.NORMAL
                s.document_processor.queue_document(
                    doc_id=record.doc_id,
                    filename=file.filename,
                    file_path=str(file_path),
                    priority=priority
                )

            results.append({
                "filename": file.filename,
                "doc_id": record.doc_id,
                "change_type": change_type.value,
                "version": record.version,
                "queued": change_type in [ChangeType.NEW, ChangeType.UPDATED]
            })

    # Save document registry to file for persistence
    try:
        s.document_registry.save_to_file(str(registry_file))
    except Exception as e:
        logger.error(f"Failed to save document registry: {e}")

    return jsonify({
        "status": "success",
        "documents": results,
        "message": f"Registered {len(results)} documents"
    })


@app.route('/api/documents/processing/status')
def processing_status():
    """Get processing status for all documents in queue."""
    s = get_session()

    if not hasattr(s, 'document_processor'):
        return jsonify({"queue": [], "stats": {}})

    all_status = s.document_processor.get_all_status()
    stats = s.document_processor.get_queue_stats()

    return jsonify({
        "queue": all_status,
        "stats": stats
    })


@app.route('/api/documents/<doc_id>/processing-status')
def document_processing_status(doc_id):
    """Get processing status for a specific document."""
    s = get_session()

    if not hasattr(s, 'document_processor'):
        return jsonify({"status": "error", "message": "No processor initialized"}), 400

    status = s.document_processor.get_status(doc_id)
    if not status:
        return jsonify({"status": "not_found", "message": "Document not in processing queue"})

    return jsonify({"status": "success", "processing": status})


@app.route('/api/documents/<doc_id>/process', methods=['POST'])
def process_document_now(doc_id):
    """Manually trigger processing for a document."""
    from tools_v2.document_processor import ProcessingPriority

    s = get_session()

    if not hasattr(s, 'document_registry'):
        return jsonify({"status": "error", "message": "No documents registered"}), 400

    doc = s.document_registry.get_document(doc_id)
    if not doc:
        return jsonify({"status": "error", "message": "Document not found"}), 404

    # Initialize processor if needed
    if not hasattr(s, 'document_processor'):
        from tools_v2.document_processor import DocumentProcessor
        from tools_v2.fact_merger import FactMerger
        if not hasattr(s, 'fact_merger'):
            s.fact_merger = FactMerger(s.fact_store)
        s.document_processor = DocumentProcessor(
            document_registry=s.document_registry,
            fact_store=s.fact_store,
            fact_merger=s.fact_merger
        )
        s.document_processor.start_worker()

    # Get priority from request
    data = request.get_json() or {}
    priority_str = data.get('priority', 'normal').upper()
    priority = getattr(ProcessingPriority, priority_str, ProcessingPriority.NORMAL)

    # Queue document
    s.document_processor.queue_document(
        doc_id=doc.doc_id,
        filename=doc.filename,
        file_path=doc.file_path,
        priority=priority
    )

    return jsonify({
        "status": "success",
        "message": f"Document queued for processing",
        "doc_id": doc_id
    })


@app.route('/api/documents/<doc_id>')
@auth_optional
def get_document(doc_id):
    """Get document details."""
    s = get_session()

    if not hasattr(s, 'document_registry'):
        return jsonify({"status": "error", "message": "No documents registered"}), 404

    doc = s.document_registry.get_document(doc_id)
    if not doc:
        return jsonify({"status": "error", "message": "Document not found"}), 404

    return jsonify({"status": "success", "document": doc.to_dict()})


@app.route('/api/documents/<doc_id>/facts')
@auth_optional
def get_document_facts(doc_id):
    """Get all facts linked to a document."""
    s = get_session()

    if not hasattr(s, 'document_registry'):
        return jsonify({"facts": []})

    fact_ids = s.document_registry.get_facts_for_document(doc_id)
    facts = [s.fact_store.get_fact(fid).to_dict() for fid in fact_ids if s.fact_store.get_fact(fid)]

    return jsonify({"facts": facts, "count": len(facts)})


@app.route('/api/documents/whats-new')
def whats_new():
    """Get facts that are new or changed since last review."""
    s = get_session()

    if not hasattr(s, 'fact_merger'):
        return jsonify({"new": [], "updated": [], "conflicts": []})

    data = s.fact_merger.get_whats_new()

    return jsonify({
        "new": [f.to_dict() for f in data["new"]],
        "updated": [f.to_dict() for f in data["updated"]],
        "conflicts": [f.to_dict() for f in data["conflicts"]],
        "new_count": len(data["new"]),
        "updated_count": len(data["updated"]),
        "conflict_count": len(data["conflicts"])
    })


@app.route('/api/documents/conflicts')
def get_conflicts():
    """Get pending merge conflicts."""
    s = get_session()

    if not hasattr(s, 'fact_merger'):
        return jsonify({"conflicts": []})

    conflicts = s.fact_merger.get_pending_conflicts()
    return jsonify({
        "conflicts": [c.to_dict() for c in conflicts],
        "count": len(conflicts)
    })


@app.route('/api/documents/conflicts/<conflict_id>/resolve', methods=['POST'])
def resolve_conflict(conflict_id):
    """Resolve a merge conflict."""
    s = get_session()

    if not hasattr(s, 'fact_merger'):
        return jsonify({"status": "error", "message": "No merger initialized"}), 400

    data = request.get_json() or {}
    resolution = data.get('resolution')  # keep_existing, use_new, merge
    resolved_by = data.get('resolved_by', 'web_user')
    notes = data.get('notes', '')

    if not resolution:
        return jsonify({"status": "error", "message": "resolution is required"}), 400

    success = s.fact_merger.resolve_conflict(conflict_id, resolution, resolved_by, notes)

    if success:
        return jsonify({"status": "success", "message": "Conflict resolved"})
    else:
        return jsonify({"status": "error", "message": "Failed to resolve conflict"}), 400


@app.route('/api/documents/clear-markers', methods=['POST'])
def clear_change_markers():
    """Clear [NEW] and [UPDATED] markers after user has reviewed."""
    s = get_session()

    if hasattr(s, 'fact_merger'):
        s.fact_merger.clear_change_markers()
        return jsonify({"status": "success", "message": "Change markers cleared"})

    return jsonify({"status": "error", "message": "No merger initialized"}), 400


@app.route('/api/documents/analysis-runs')
def get_analysis_runs():
    """Get history of analysis runs."""
    s = get_session()

    if not hasattr(s, 'document_registry'):
        return jsonify({"runs": []})

    runs = s.document_registry.get_analysis_runs()
    return jsonify({"runs": runs})


# === Tiered Review API Endpoints (Phase 3) ===

@app.route('/api/changes/pending')
def get_pending_changes():
    """Get all pending changes organized by tier."""
    from config_v2 import OUTPUT_DIR
    import json as json_module

    s = get_session()

    # Load pending changes from file (primary source of truth)
    pending_file = OUTPUT_DIR / "pending_changes.json"
    if pending_file.exists():
        try:
            with open(pending_file, 'r') as f:
                changes = json_module.load(f)
        except Exception as e:
            logger.error(f"Failed to load pending changes: {e}")
            changes = {"tier1": [], "tier2": [], "tier3": []}
    else:
        changes = {"tier1": [], "tier2": [], "tier3": []}

    stats = {
        "total": len(changes.get("tier1", [])) + len(changes.get("tier2", [])) + len(changes.get("tier3", [])),
        "tier1": len(changes.get("tier1", [])),
        "tier2": len(changes.get("tier2", [])),
        "tier3": len(changes.get("tier3", []))
    }

    return jsonify({
        "tier1": changes.get("tier1", []),
        "tier2": changes.get("tier2", []),
        "tier3": changes.get("tier3", []),
        "stats": stats
    })


@app.route('/api/changes/auto-apply', methods=['POST'])
def auto_apply_changes():
    """Auto-apply all Tier 1 eligible changes."""
    s = get_session()

    if not hasattr(s, 'pending_changes') or not hasattr(s, 'fact_merger'):
        return jsonify({"status": "error", "message": "No pending changes"}), 400

    tier1_changes = s.pending_changes.get("tier1", [])
    auto_eligible = [c for c in tier1_changes if c.get("classification", {}).get("auto_apply_eligible")]

    applied = []
    for change in auto_eligible:
        fact_data = change.get("fact", {})
        try:
            from tools_v2.fact_merger import MergeAction
            action, fact_id = s.fact_merger.merge_fact(fact_data, fact_data.get("source_document", ""))
            if action in [MergeAction.ADD, MergeAction.UPDATE]:
                applied.append({
                    "fact_id": fact_id,
                    "item": fact_data.get("item"),
                    "action": action.value
                })
        except Exception as e:
            logger.error(f"Auto-apply failed for fact: {e}")

    # Remove applied changes from pending
    remaining = [c for c in tier1_changes if c not in auto_eligible or c.get("fact", {}).get("temp_id") not in [a.get("temp_id") for a in applied]]
    s.pending_changes["tier1"] = remaining

    # Save fact store
    if hasattr(s, 'fact_store'):
        facts_files = sorted(OUTPUT_DIR.glob("facts_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
        if facts_files:
            s.fact_store.save(str(facts_files[0]))

    return jsonify({
        "status": "success",
        "applied_count": len(applied),
        "applied": applied
    })


@app.route('/api/changes/batch-accept', methods=['POST'])
def batch_accept_changes():
    """Accept multiple changes at once (Tier 2 batch review)."""
    from config_v2 import OUTPUT_DIR
    import json as json_module
    from tools_v2.fact_merger import FactMerger, MergeAction

    s = get_session()

    data = request.get_json() or {}
    change_ids = data.get("change_ids", [])
    tier = data.get("tier", "tier2")

    # Load pending changes from file
    pending_file = OUTPUT_DIR / "pending_changes.json"
    if not pending_file.exists():
        return jsonify({"status": "error", "message": "No pending changes file"}), 404

    try:
        with open(pending_file, 'r') as f:
            pending_changes = json_module.load(f)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to load pending changes: {e}"}), 500

    # Initialize fact_merger if needed
    if not hasattr(s, 'fact_merger') and hasattr(s, 'fact_store'):
        s.fact_merger = FactMerger(s.fact_store)

    if not hasattr(s, 'fact_merger'):
        return jsonify({"status": "error", "message": "Cannot initialize merger"}), 400

    tier_changes = pending_changes.get(tier, [])

    accepted = []
    rejected_ids = set()

    for change in tier_changes:
        fact_data = change.get("fact", {})
        temp_id = fact_data.get("temp_id")

        if not change_ids or temp_id in change_ids:
            try:
                action, fact_id = s.fact_merger.merge_fact(fact_data, fact_data.get("source_document", ""))
                if action in [MergeAction.ADD, MergeAction.UPDATE]:
                    accepted.append({
                        "fact_id": fact_id,
                        "temp_id": temp_id,
                        "item": fact_data.get("item"),
                        "action": action.value
                    })
            except Exception as e:
                logger.error(f"Batch accept failed: {e}")
                rejected_ids.add(temp_id)

    # Remove accepted changes from pending
    accepted_temp_ids = {a["temp_id"] for a in accepted}
    pending_changes[tier] = [
        c for c in tier_changes
        if c.get("fact", {}).get("temp_id") not in accepted_temp_ids
    ]

    # Save updated pending changes to file
    with open(pending_file, 'w') as f:
        json_module.dump(pending_changes, f, indent=2, default=str)

    # Save fact store
    if hasattr(s, 'fact_store'):
        facts_files = sorted(OUTPUT_DIR.glob("facts_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
        if facts_files:
            s.fact_store.save(str(facts_files[0]))

    return jsonify({
        "status": "success",
        "accepted_count": len(accepted),
        "accepted": accepted
    })


@app.route('/api/changes/<change_id>/accept', methods=['POST'])
def accept_single_change(change_id):
    """Accept a single change (Tier 3 individual review)."""
    from config_v2 import OUTPUT_DIR
    import json as json_module
    from tools_v2.fact_merger import FactMerger, MergeAction

    s = get_session()

    data = request.get_json() or {}
    resolution = data.get("resolution", "accept")  # accept, reject, modify
    modifications = data.get("modifications", {})
    notes = data.get("notes", "")

    # Load pending changes from file (primary source of truth)
    pending_file = OUTPUT_DIR / "pending_changes.json"
    if not pending_file.exists():
        return jsonify({"status": "error", "message": "No pending changes file"}), 404

    try:
        with open(pending_file, 'r') as f:
            pending_changes = json_module.load(f)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to load pending changes: {e}"}), 500

    # Find the change across all tiers
    change = None
    change_tier = None
    for tier in ["tier1", "tier2", "tier3"]:
        for c in pending_changes.get(tier, []):
            if c.get("fact", {}).get("temp_id") == change_id:
                change = c
                change_tier = tier
                break
        if change:
            break

    if not change:
        return jsonify({"status": "error", "message": "Change not found"}), 404

    fact_data = change.get("fact", {})

    # Remove from pending changes file
    pending_changes[change_tier] = [
        c for c in pending_changes[change_tier]
        if c.get("fact", {}).get("temp_id") != change_id
    ]

    if resolution == "reject":
        # Save updated pending changes
        with open(pending_file, 'w') as f:
            json_module.dump(pending_changes, f, indent=2, default=str)
        # Log activity
        log_activity("change_rejected", {
            "change_id": change_id,
            "tier": change_tier,
            "item": fact_data.get("item", ""),
            "domain": fact_data.get("domain", ""),
            "notes": notes
        })
        return jsonify({"status": "success", "action": "rejected"})

    # Apply modifications if any
    if modifications:
        fact_data.update(modifications)

    # Initialize fact_merger
    if not hasattr(s, 'fact_merger') and hasattr(s, 'fact_store'):
        s.fact_merger = FactMerger(s.fact_store)

    if not hasattr(s, 'fact_merger'):
        return jsonify({"status": "error", "message": "Cannot initialize merger"}), 400

    # Merge the fact
    try:
        action, fact_id = s.fact_merger.merge_fact(fact_data, fact_data.get("source_document", ""))

        # Add review note
        if notes and fact_id and hasattr(s, 'fact_store'):
            fact = s.fact_store.get_fact(fact_id)
            if fact:
                fact.verification_note = f"{fact.verification_note or ''} [Review: {notes}]"

        # Save updated pending changes to file
        with open(pending_file, 'w') as f:
            json_module.dump(pending_changes, f, indent=2, default=str)

        # Save fact store
        if hasattr(s, 'fact_store'):
            facts_files = sorted(OUTPUT_DIR.glob("facts_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
            if facts_files:
                s.fact_store.save(str(facts_files[0]))

        # Log activity
        log_activity("change_accepted", {
            "change_id": change_id,
            "tier": change_tier,
            "item": fact_data.get("item", ""),
            "domain": fact_data.get("domain", ""),
            "fact_id": fact_id,
            "action": action.value,
            "notes": notes
        })

        return jsonify({
            "status": "success",
            "action": action.value,
            "fact_id": fact_id
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/review/batch')
@auth_optional
def batch_review_page():
    """Batch review page for Tier 2 changes."""
    from config_v2 import OUTPUT_DIR
    import json as json_module

    s = get_session()

    # Load pending changes from file
    pending_file = OUTPUT_DIR / "pending_changes.json"
    if pending_file.exists():
        try:
            with open(pending_file, 'r') as f:
                pending = json_module.load(f)
        except Exception:
            pending = {"tier1": [], "tier2": [], "tier3": []}
    else:
        pending = {"tier1": [], "tier2": [], "tier3": []}

    tier2_changes = pending.get("tier2", [])

    # Group by source document
    by_document = {}
    for change in tier2_changes:
        doc = change.get("fact", {}).get("source_document", "Unknown")
        if doc not in by_document:
            by_document[doc] = []
        by_document[doc].append(change)

    return render_template('review/batch.html',
                          changes=tier2_changes,
                          by_document=by_document,
                          total_count=len(tier2_changes))


@app.route('/review/conflicts')
@auth_optional
def conflicts_review_page():
    """Individual review page for Tier 3 conflicts."""
    from config_v2 import OUTPUT_DIR
    import json as json_module

    s = get_session()

    # Load pending changes from file
    pending_file = OUTPUT_DIR / "pending_changes.json"
    if pending_file.exists():
        try:
            with open(pending_file, 'r') as f:
                pending = json_module.load(f)
        except Exception:
            pending = {"tier1": [], "tier2": [], "tier3": []}
    else:
        pending = {"tier1": [], "tier2": [], "tier3": []}

    tier3_changes = pending.get("tier3", [])

    # Get existing facts for comparison
    conflicts_with_existing = []
    for change in tier3_changes:
        existing_id = change.get("existing_fact_id")
        existing_fact = None
        if existing_id and hasattr(s, 'fact_store'):
            existing_fact = s.fact_store.get_fact(existing_id)

        conflicts_with_existing.append({
            "change": change,
            "existing": existing_fact.to_dict() if existing_fact else None
        })

    return render_template('review/conflicts.html',
                          conflicts=conflicts_with_existing,
                          total_count=len(tier3_changes))


# === Activity Log API ===

@app.route('/api/activity-log')
def get_activity_log():
    """Get recent activity log entries."""
    from config_v2 import OUTPUT_DIR
    import json as json_module

    activity_file = OUTPUT_DIR / "activity_log.json"
    limit = request.args.get('limit', 50, type=int)

    if not activity_file.exists():
        return jsonify({"activities": [], "total": 0})

    try:
        with open(activity_file, 'r') as f:
            activities = json_module.load(f)

        # Return most recent first
        activities = list(reversed(activities[-limit:]))

        return jsonify({
            "activities": activities,
            "total": len(activities)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# === Stale Items API (Phase 4) ===

@app.route('/api/stale-items')
def get_stale_items():
    """Get all stale items that need review."""
    from tools_v2.dependency_tracker import DependencyTracker
    from tools_v2.inventory_updater import InventoryUpdater

    s = get_session()

    if not hasattr(s, 'dependency_tracker'):
        s.dependency_tracker = DependencyTracker()

    updater = InventoryUpdater(
        fact_store=s.fact_store if hasattr(s, 'fact_store') else None,
        dependency_tracker=s.dependency_tracker,
        session=s
    )

    dashboard = updater.get_stale_dashboard()
    return jsonify(dashboard)


@app.route('/api/stale-items/<item_type>/<item_id>/review', methods=['POST'])
def review_stale_item(item_type, item_id):
    """Mark a stale item as reviewed."""
    from tools_v2.dependency_tracker import DependencyTracker, DependentItemType

    s = get_session()

    if not hasattr(s, 'dependency_tracker'):
        return jsonify({"status": "error", "message": "No tracker initialized"}), 400

    data = request.get_json() or {}
    reviewer = data.get("reviewer", "web_user")
    notes = data.get("notes", "")

    try:
        item_type_enum = DependentItemType(item_type)
    except ValueError:
        return jsonify({"status": "error", "message": f"Invalid item type: {item_type}"}), 400

    success = s.dependency_tracker.mark_reviewed(
        item_type=item_type_enum,
        item_id=item_id,
        reviewed_by=reviewer,
        notes=notes
    )

    if success:
        return jsonify({"status": "success", "message": "Item marked as reviewed"})
    else:
        return jsonify({"status": "error", "message": "Item not found"}), 404


@app.route('/api/stale-items/bulk-revalidate', methods=['POST'])
def bulk_revalidate_stale():
    """Bulk revalidate stale items."""
    from tools_v2.dependency_tracker import DependencyTracker, DependentItemType
    from tools_v2.inventory_updater import InventoryUpdater

    s = get_session()

    if not hasattr(s, 'dependency_tracker'):
        s.dependency_tracker = DependencyTracker()

    data = request.get_json() or {}
    item_type_str = data.get("item_type")  # Optional filter
    reviewer = data.get("reviewer", "web_user")
    notes = data.get("notes", "Bulk revalidation")

    updater = InventoryUpdater(
        fact_store=s.fact_store if hasattr(s, 'fact_store') else None,
        dependency_tracker=s.dependency_tracker,
        session=s
    )

    item_type = None
    if item_type_str:
        try:
            item_type = DependentItemType(item_type_str)
        except ValueError:
            return jsonify({"status": "error", "message": f"Invalid item type: {item_type_str}"}), 400

    count = updater.bulk_revalidate(item_type, reviewer, notes) if item_type else 0

    # If no type specified, do all types
    if not item_type:
        for t in DependentItemType:
            count += updater.bulk_revalidate(t, reviewer, notes)

    return jsonify({
        "status": "success",
        "revalidated_count": count
    })


@app.route('/api/propagate-changes', methods=['POST'])
def propagate_fact_changes():
    """Propagate fact changes to downstream items."""
    from tools_v2.dependency_tracker import DependencyTracker
    from tools_v2.inventory_updater import InventoryUpdater

    s = get_session()

    data = request.get_json() or {}
    fact_ids = data.get("fact_ids", [])
    change_type = data.get("change_type", "updated")

    if not fact_ids:
        return jsonify({"status": "error", "message": "No fact_ids provided"}), 400

    if not hasattr(s, 'dependency_tracker'):
        s.dependency_tracker = DependencyTracker()

    updater = InventoryUpdater(
        fact_store=s.fact_store if hasattr(s, 'fact_store') else None,
        dependency_tracker=s.dependency_tracker,
        session=s
    )

    result = updater.propagate_changes(fact_ids, change_type)

    return jsonify({
        "status": "success",
        "result": result.to_dict()
    })


# === Settings API (Phase 5) ===

@app.route('/api/settings/incremental')
def get_incremental_settings():
    """Get incremental update settings."""
    s = get_session()

    # Get or initialize settings
    if not hasattr(s, 'incremental_settings'):
        s.incremental_settings = {
            "auto_apply_enabled": True,
            "auto_apply_min_confidence": 0.9,
            "critical_domains": ["security", "compliance", "data"],
            "batch_review_threshold": 0.6,
            "auto_propagate": False
        }

    return jsonify(s.incremental_settings)


@app.route('/api/settings/incremental', methods=['POST'])
def update_incremental_settings():
    """Update incremental update settings."""
    s = get_session()

    data = request.get_json() or {}

    if not hasattr(s, 'incremental_settings'):
        s.incremental_settings = {}

    # Update settings
    valid_keys = [
        "auto_apply_enabled",
        "auto_apply_min_confidence",
        "critical_domains",
        "batch_review_threshold",
        "auto_propagate"
    ]

    for key in valid_keys:
        if key in data:
            s.incremental_settings[key] = data[key]

    # Apply to tier classifier if exists
    if hasattr(s, 'document_processor'):
        from tools_v2.tier_classifier import TierClassifier
        # Classifier settings will be read on next classification

    return jsonify({
        "status": "success",
        "settings": s.incremental_settings
    })


# =============================================================================
# VALIDATION DASHBOARD ROUTES (Sprint 4 - Human Review Loop)
# =============================================================================

@app.route('/validation')
@auth_optional
def validation_dashboard():
    """
    Validation dashboard showing overall confidence and domain status.

    Part of the human review loop for fact validation.
    """
    s = get_session()

    # Get basic stats from fact store
    stats = s.fact_store.get_review_stats() if hasattr(s.fact_store, 'get_review_stats') else {}

    # Calculate overall confidence
    total_facts = stats.get('total_facts', 0)
    confirmed = stats.get('confirmed', 0)
    overall_confidence = int((confirmed / total_facts * 100)) if total_facts > 0 else 0

    # Determine confidence level for styling
    if overall_confidence >= 80:
        confidence_level = 'high'
    elif overall_confidence >= 60:
        confidence_level = 'medium'
    elif overall_confidence >= 40:
        confidence_level = 'low'
    else:
        confidence_level = 'critical'

    # Build domain status
    domains = {}
    for domain, dstats in stats.get('by_domain', {}).items():
        total = dstats.get('total', 0)
        validated = dstats.get('confirmed', 0)
        flagged = dstats.get('incorrect', 0) + dstats.get('needs_info', 0)

        domain_confidence = int((validated / total * 100)) if total > 0 else 0

        if domain_confidence >= 80:
            dlevel = 'high'
        elif domain_confidence >= 60:
            dlevel = 'medium'
        elif domain_confidence >= 40:
            dlevel = 'low'
        else:
            dlevel = 'critical'

        domains[domain] = {
            'total_facts': total,
            'validated': validated,
            'flagged': flagged,
            'confidence': domain_confidence,
            'confidence_level': dlevel,
            'validated_pct': int((validated / total * 100)) if total > 0 else 0,
            'critical_count': 0,  # Would come from validation store
            'high_count': 0,
            'warning_count': flagged
        }

    # Recent activity (placeholder - would come from audit store)
    recent_activity = []

    return render_template(
        'validation/dashboard.html',
        overall_confidence=overall_confidence,
        confidence_level=confidence_level,
        total_facts=total_facts,
        flagged_count=stats.get('incorrect', 0) + stats.get('needs_info', 0),
        validated_count=confirmed,
        validated_pct=int((confirmed / total_facts * 100)) if total_facts > 0 else 0,
        corrections_count=0,  # Would come from correction store
        queue_count=stats.get('needs_info', 0) + stats.get('incorrect', 0),
        domains=domains,
        recent_activity=recent_activity
    )


@app.route('/validation/queue')
@auth_optional
def validation_review_queue():
    """
    Validation-aware review queue with flag-based filtering.

    Uses the enhanced validation templates with correction support.
    """
    s = get_session()

    # Get filter parameters
    filter_domain = request.args.get('domain', '')
    filter_severity = request.args.get('severity', '')

    # Get review queue
    queue = s.fact_store.get_review_queue(
        domain=filter_domain if filter_domain else None,
        limit=100
    )

    # Transform facts to validation format
    items = []
    for fact in queue:
        items.append({
            'fact_id': fact.fact_id,
            'item': fact.item,
            'domain': fact.domain,
            'category': fact.category,
            'highest_severity': 'warning' if fact.confidence_score < 0.7 else 'info',
            'flags': [
                {
                    'severity': 'warning' if fact.confidence_score < 0.7 else 'info',
                    'message': fact.verification_note or 'Low confidence extraction'
                }
            ] if fact.confidence_score < 0.8 else [],
            'flags_count': 1 if fact.confidence_score < 0.8 else 0,
            'ai_confidence': fact.confidence_score or 0.5,
            'evidence_verified': fact.verification_status == 'confirmed',
            'evidence_matched_text': fact.evidence.get('exact_quote', '') if fact.evidence else '',
            'human_reviewed': fact.verification_status in ['confirmed', 'incorrect'],
            'details': fact.details
        })

    return render_template(
        'validation/review_queue.html',
        items=items,
        filter_domain=filter_domain,
        filter_severity=filter_severity,
        reviewer_name=''  # Would come from session/auth
    )


# =============================================================================
# Phase 5: Document Storage Endpoints
# =============================================================================

@app.route('/documents/file/<path:key>')
@auth_optional
def serve_document(key):
    """
    Serve a document file from storage.

    For local storage: streams the file directly.
    For S3 storage: redirects to a presigned URL.
    """
    if storage is None:
        return jsonify({"error": "Storage not initialized"}), 500

    try:
        if storage.storage_type == 's3':
            # Redirect to presigned URL for S3
            url = storage.get_document_url(key, expires_in=3600)
            return redirect(url)
        else:
            # Stream file for local storage
            file_obj, stored_file = storage.get_document(key)
            from flask import send_file
            return send_file(
                file_obj,
                mimetype=stored_file.content_type,
                as_attachment=True,
                download_name=stored_file.filename
            )
    except FileNotFoundError:
        return jsonify({"error": "File not found"}), 404
    except Exception as e:
        logger.error(f"Error serving document {key}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/storage/upload', methods=['POST'])
@auth_optional
@csrf.exempt  # File uploads use multipart form
def api_storage_upload():
    """
    Upload a document to storage.

    Expects multipart form with:
    - file: The file to upload
    - deal_id: The deal to associate with
    - entity: 'target' or 'buyer'
    """
    if storage is None:
        return jsonify({"error": "Storage not initialized"}), 500

    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    deal_id = request.form.get('deal_id')
    entity = request.form.get('entity', 'target')

    if not deal_id:
        return jsonify({"error": "deal_id is required"}), 400

    try:
        # Upload to storage
        stored_file = storage.upload_document(
            file=file.stream,
            deal_id=deal_id,
            entity=entity,
            filename=file.filename,
            content_type=file.content_type
        )

        # If database is enabled, also create Document record
        if USE_DATABASE:
            from web.database import db, Document
            doc = Document(
                deal_id=deal_id,
                filename=stored_file.filename,
                file_hash=stored_file.hash,
                file_size=stored_file.size,
                mime_type=stored_file.content_type,
                entity=entity,
                storage_path=stored_file.key,
                storage_type=storage.storage_type,
                status='pending'
            )
            db.session.add(doc)
            db.session.commit()

            return jsonify({
                "success": True,
                "document": {
                    "id": doc.id,
                    "key": stored_file.key,
                    "filename": stored_file.filename,
                    "size": stored_file.size,
                    "hash": stored_file.hash,
                }
            })

        return jsonify({
            "success": True,
            "document": stored_file.to_dict()
        })

    except Exception as e:
        logger.error(f"Upload failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/storage/documents/<deal_id>')
@auth_optional
def api_storage_list_documents(deal_id):
    """List all documents for a deal from storage."""
    if storage is None:
        return jsonify({"error": "Storage not initialized"}), 500

    entity = request.args.get('entity')

    try:
        files = storage.list_deal_documents(deal_id, entity)
        return jsonify({
            "deal_id": deal_id,
            "entity": entity,
            "documents": [f.to_dict() for f in files]
        })
    except Exception as e:
        logger.error(f"Error listing documents: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/storage/url/<path:key>')
@auth_optional
def api_storage_get_url(key):
    """Get a URL for accessing a document."""
    if storage is None:
        return jsonify({"error": "Storage not initialized"}), 500

    expires_in = request.args.get('expires', 3600, type=int)

    try:
        url = storage.get_document_url(key, expires_in=expires_in)
        return jsonify({
            "key": key,
            "url": url,
            "expires_in": expires_in
        })
    except Exception as e:
        logger.error(f"Error getting URL for {key}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/storage/status')
def api_storage_status():
    """Get storage service status."""
    if storage is None:
        return jsonify({
            "status": "not_initialized",
            "type": None
        })

    return jsonify({
        "status": "active",
        "type": storage.storage_type,
        "config": {
            "storage_type": STORAGE_TYPE,
            "s3_configured": bool(os.environ.get('S3_ACCESS_KEY')),
        }
    })


# =============================================================================
# Run Management API (Phase B)
# =============================================================================

@app.route('/api/runs')
@auth_optional
def api_list_runs():
    """List all analysis runs."""
    try:
        from services.run_manager import get_run_manager
        manager = get_run_manager()

        include_archived = request.args.get('include_archived', 'false').lower() == 'true'
        runs = manager.list_runs(include_archived=include_archived)

        # Add stats for each run
        for run in runs:
            stats = manager.get_run_stats(run['run_id'])
            run['stats'] = stats

        return jsonify({
            'runs': runs,
            'total': len(runs),
            'latest': manager.get_latest_run()
        })
    except Exception as e:
        logger.error(f"Error listing runs: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/runs/<run_id>')
@auth_optional
def api_get_run(run_id):
    """Get details for a specific run."""
    try:
        from services.run_manager import get_run_manager
        manager = get_run_manager()

        metadata = manager.get_run_metadata(run_id)
        if metadata is None:
            return jsonify({'error': 'Run not found'}), 404

        stats = manager.get_run_stats(run_id)

        return jsonify({
            'metadata': metadata.to_dict(),
            'stats': stats,
            'is_latest': run_id == manager.get_latest_run()
        })
    except Exception as e:
        logger.error(f"Error getting run {run_id}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/runs/<run_id>/load', methods=['POST'])
@auth_optional
def api_load_run(run_id):
    """Load a run into the current session."""
    try:
        from interactive.session import Session

        # Load session from run
        analysis_session = Session.load_from_run(run_id)

        # Store in session store
        session_id = get_or_create_session_id(flask_session)
        user_session = session_store.get_session(session_id)
        if user_session:
            user_session.analysis_session = analysis_session
            user_session.touch()

        return jsonify({
            'success': True,
            'run_id': run_id,
            'facts': len(analysis_session.fact_store.facts),
            'risks': len(analysis_session.reasoning_store.risks),
            'work_items': len(analysis_session.reasoning_store.work_items)
        })
    except Exception as e:
        logger.error(f"Error loading run {run_id}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/runs/current')
@auth_optional
def api_current_run():
    """Get current run info for the session."""
    try:
        s = get_session()
        run_id = s.get_current_run_id() if hasattr(s, 'get_current_run_id') else None

        if run_id is None:
            from services.run_manager import get_run_manager
            run_id = get_run_manager().get_latest_run()

        return jsonify({
            'run_id': run_id,
            'has_data': len(s.fact_store.facts) > 0
        })
    except Exception as e:
        logger.error(f"Error getting current run: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/runs/save', methods=['POST'])
@auth_optional
def api_save_run():
    """Save current session to a run."""
    try:
        s = get_session()
        saved_files = s.save_to_run()

        return jsonify({
            'success': True,
            'files': {k: str(v) for k, v in saved_files.items()},
            'run_id': s.get_current_run_id()
        })
    except Exception as e:
        logger.error(f"Error saving run: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/runs/<run_id>/archive', methods=['POST'])
@auth_optional
def api_archive_run(run_id):
    """Archive a run."""
    try:
        from services.run_manager import get_run_manager
        manager = get_run_manager()

        success = manager.archive_run(run_id)

        return jsonify({
            'success': success,
            'run_id': run_id
        })
    except Exception as e:
        logger.error(f"Error archiving run {run_id}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/runs')
@auth_optional
def runs_page():
    """Run history page."""
    try:
        from services.run_manager import get_run_manager
        manager = get_run_manager()

        runs = manager.list_runs()
        latest_run = manager.get_latest_run()

        # Get current session stats
        s = get_session()
        current_stats = {
            'facts': len(s.fact_store.facts),
            'risks': len(s.reasoning_store.risks),
            'work_items': len(s.reasoning_store.work_items),
            'gaps': len(s.fact_store.gaps)
        }

        return render_template('runs.html',
                             runs=runs,
                             latest_run=latest_run,
                             current_stats=current_stats)
    except Exception as e:
        logger.error(f"Error rendering runs page: {e}")
        return render_template('error.html', error=str(e))


@app.route('/admin/cleanup')
@auth_optional
def admin_cleanup():
    """Admin page to view and delete old deals."""
    from web.database import Deal, Fact, Finding, AnalysisRun, Document

    deals = Deal.query.order_by(Deal.created_at.desc()).all()

    deals_data = []
    for deal in deals:
        facts_count = Fact.query.filter_by(deal_id=deal.id, deleted_at=None).count()
        findings_count = Finding.query.filter_by(deal_id=deal.id, deleted_at=None).count()
        runs_count = AnalysisRun.query.filter_by(deal_id=deal.id).count()
        docs_count = Document.query.filter_by(deal_id=deal.id).count()

        deals_data.append({
            'id': deal.id,
            'name': deal.name,
            'created_at': deal.created_at,
            'facts': facts_count,
            'findings': findings_count,
            'runs': runs_count,
            'docs': docs_count
        })

    return render_template('admin/cleanup.html', deals=deals_data)


@app.route('/admin/cleanup/delete/<deal_id>', methods=['POST'])
@auth_optional
def admin_delete_deal(deal_id):
    """Delete a specific deal."""
    from web.database import Deal, Fact, Finding, Gap, Document, AnalysisRun

    deal = Deal.query.get(deal_id)
    if not deal:
        flash(f'Deal {deal_id} not found', 'error')
        return redirect(url_for('admin_cleanup'))

    try:
        deal_name = deal.name

        # Explicitly delete related records to avoid FK constraint issues
        # SQLAlchemy relationships don't always handle cascade correctly
        Gap.query.filter_by(deal_id=deal_id).delete()
        Fact.query.filter_by(deal_id=deal_id).delete()
        Finding.query.filter_by(deal_id=deal_id).delete()
        Document.query.filter_by(deal_id=deal_id).delete()
        AnalysisRun.query.filter_by(deal_id=deal_id).delete()

        # Now delete the deal itself
        db.session.delete(deal)
        db.session.commit()

        flash(f'✅ Deleted deal "{deal_name}" and all associated data', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting deal: {e}', 'error')

    return redirect(url_for('admin_cleanup'))


@app.route('/admin/cleanup/delete-all', methods=['POST'])
@auth_optional
def admin_delete_all_deals():
    """Delete ALL deals."""
    from web.database import Deal

    deals = Deal.query.all()
    count = len(deals)

    try:
        for deal in deals:
            db.session.delete(deal)
        db.session.commit()
        flash(f'✅ Deleted {count} deals and all associated data', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error during cleanup: {e}', 'error')

    return redirect(url_for('admin_cleanup'))


# Register CLI commands for database management
from web.cleanup_old_deals import register_commands
register_commands(app)


if __name__ == '__main__':
    print("\n" + "="*60)
    print("  IT Due Diligence Agent - Web Interface")
    print("="*60)
    print("\n  Starting server at: http://127.0.0.1:5001")
    print("  Press Ctrl+C to stop\n")
    app.run(debug=True, port=5001)
