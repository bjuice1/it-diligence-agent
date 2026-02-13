"""
Celery Tasks for Background Processing

Phase 4: Distributed task processing.
Tasks run in Celery workers, not the web process.
"""

import os
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

from celery import shared_task, current_task

logger = logging.getLogger(__name__)


@shared_task(bind=True, name='web.tasks.run_analysis')
def run_analysis_task(
    self,
    deal_id: str,
    domains: List[str],
    entity: str = 'target',
    user_id: Optional[str] = None
) -> Dict[str, Any]:
    """
    Run due diligence analysis as a background task.

    Args:
        deal_id: The deal to analyze
        domains: List of domains to analyze
        entity: 'target' or 'buyer'
        user_id: User who started the analysis

    Returns:
        Dict with analysis results summary
    """
    from web.database import db, Deal, Fact, Finding

    logger.info(f"Starting analysis task for deal {deal_id}, domains: {domains}")

    # Update progress
    self.update_state(
        state='PROGRESS',
        meta={
            'progress': 0,
            'phase': 'initializing',
            'message': 'Starting analysis...'
        }
    )

    try:
        # Get deal
        deal = Deal.query.get(deal_id)
        if not deal:
            raise ValueError(f"Deal {deal_id} not found")

        total_domains = len(domains)
        facts_created = 0
        findings_created = 0

        for i, domain in enumerate(domains):
            # Calculate progress
            progress = int((i / total_domains) * 100)

            self.update_state(
                state='PROGRESS',
                meta={
                    'progress': progress,
                    'phase': f'analyzing_{domain}',
                    'message': f'Analyzing {domain}...',
                    'current_domain': domain,
                    'domains_completed': i,
                    'total_domains': total_domains,
                }
            )

            # Run the actual analysis for this domain
            # This would call into the existing analysis pipeline
            domain_facts, domain_findings = _analyze_domain(
                deal_id=deal_id,
                domain=domain,
                entity=entity,
                progress_callback=lambda p, m: self.update_state(
                    state='PROGRESS',
                    meta={
                        'progress': progress + int(p / total_domains),
                        'phase': f'analyzing_{domain}',
                        'message': m,
                    }
                )
            )

            facts_created += domain_facts
            findings_created += domain_findings

        # Final update
        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 100,
                'phase': 'complete',
                'message': 'Analysis complete'
            }
        )

        result = {
            'deal_id': deal_id,
            'domains_analyzed': domains,
            'entity': entity,
            'facts_created': facts_created,
            'findings_created': findings_created,
            'completed_at': datetime.utcnow().isoformat(),
        }

        logger.info(f"Analysis complete for deal {deal_id}: {facts_created} facts, {findings_created} findings")
        return result

    except Exception as e:
        logger.error(f"Analysis failed for deal {deal_id}: {e}")
        raise


def _analyze_domain(
    deal_id: str,
    domain: str,
    entity: str,
    progress_callback=None
) -> tuple:
    """
    Analyze a single domain.

    Returns:
        Tuple of (facts_created, findings_created)
    """
    # This is a placeholder that integrates with the existing analysis pipeline
    # In production, this would call into the actual analysis code

    try:
        # Import the existing analysis runner
        from web.analysis_runner import run_analysis_simple
        from stores.session_store import session_store

        # For now, return placeholder counts
        # The actual integration would:
        # 1. Load documents for the deal
        # 2. Run discovery agents for the domain
        # 3. Run reasoning agents
        # 4. Save results to database

        if progress_callback:
            progress_callback(50, f'Processing {domain}...')

        # Placeholder - actual implementation would create real facts/findings
        return (0, 0)

    except Exception as e:
        logger.error(f"Domain analysis failed for {domain}: {e}")
        return (0, 0)


@shared_task(bind=True, name='web.tasks.process_document')
def process_document_task(
    self,
    document_id: str,
    deal_id: str,
    extract_text: bool = True
) -> Dict[str, Any]:
    """
    Process an uploaded document (extract text, analyze content).

    Args:
        document_id: The document to process
        deal_id: The deal this document belongs to
        extract_text: Whether to extract text from the document

    Returns:
        Dict with processing results
    """
    from web.database import db, Document

    logger.info(f"Processing document {document_id}")

    self.update_state(
        state='PROGRESS',
        meta={
            'progress': 0,
            'phase': 'loading',
            'message': 'Loading document...'
        }
    )

    try:
        document = Document.query.get(document_id)
        if not document:
            raise ValueError(f"Document {document_id} not found")

        # Update status
        document.status = 'processing'
        db.session.commit()

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 20,
                'phase': 'extracting',
                'message': 'Extracting text...'
            }
        )

        # Extract text if requested
        extracted_text = ''
        page_count = 0

        if extract_text:
            extracted_text, page_count = _extract_document_text(document.storage_path)

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 80,
                'phase': 'saving',
                'message': 'Saving results...'
            }
        )

        # Update document record
        document.extracted_text = extracted_text
        document.page_count = page_count
        document.status = 'completed'
        document.processed_at = datetime.utcnow()
        db.session.commit()

        result = {
            'document_id': document_id,
            'filename': document.filename,
            'page_count': page_count,
            'text_length': len(extracted_text),
            'status': 'completed',
        }

        logger.info(f"Document {document_id} processed: {page_count} pages")
        return result

    except Exception as e:
        logger.error(f"Document processing failed for {document_id}: {e}")

        # Update status to failed
        try:
            document = Document.query.get(document_id)
            if document:
                document.status = 'failed'
                document.processing_error = str(e)
                db.session.commit()
        except Exception:
            pass

        raise


def _extract_document_text(storage_path: str) -> tuple:
    """
    Extract text from a document.

    Returns:
        Tuple of (extracted_text, page_count)
    """
    import fitz  # PyMuPDF

    try:
        doc = fitz.open(storage_path)
        text_parts = []

        for page in doc:
            text_parts.append(page.get_text())

        page_count = len(doc)
        doc.close()

        return ('\n'.join(text_parts), page_count)

    except Exception as e:
        logger.error(f"Text extraction failed: {e}")
        return ('', 0)


@shared_task(name='web.tasks.cleanup_old_tasks')
def cleanup_old_tasks() -> Dict[str, int]:
    """
    Periodic task to clean up old task results and temporary data.

    Runs every hour via Celery Beat.
    """
    logger.info("Running task cleanup...")

    cleaned = {
        'task_results': 0,
        'temp_files': 0,
    }

    # Clean up old temporary files
    try:
        from config_v2 import OUTPUT_DIR
        import os
        from pathlib import Path

        temp_dir = OUTPUT_DIR / 'temp'
        if temp_dir.exists():
            cutoff = datetime.utcnow() - timedelta(hours=24)

            for file_path in temp_dir.iterdir():
                if file_path.is_file():
                    mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if mtime < cutoff:
                        file_path.unlink()
                        cleaned['temp_files'] += 1

    except Exception as e:
        logger.error(f"Temp file cleanup failed: {e}")

    logger.info(f"Cleanup complete: {cleaned}")
    return cleaned


@shared_task(bind=True, name='web.tasks.export_report')
def export_report_task(
    self,
    deal_id: str,
    format: str = 'excel',
    include_sections: List[str] = None
) -> Dict[str, Any]:
    """
    Generate and export a due diligence report.

    Args:
        deal_id: The deal to export
        format: Export format ('excel', 'pdf', 'json')
        include_sections: Sections to include (None = all)

    Returns:
        Dict with export file path and metadata
    """
    from web.database import Deal

    logger.info(f"Exporting report for deal {deal_id}, format: {format}")

    self.update_state(
        state='PROGRESS',
        meta={
            'progress': 0,
            'phase': 'gathering',
            'message': 'Gathering data...'
        }
    )

    try:
        deal = Deal.query.get(deal_id)
        if not deal:
            raise ValueError(f"Deal {deal_id} not found")

        # Default sections
        if include_sections is None:
            include_sections = ['summary', 'facts', 'risks', 'work_items', 'recommendations']

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 30,
                'phase': 'generating',
                'message': f'Generating {format} report...'
            }
        )

        # Generate report based on format
        if format == 'excel':
            file_path = _generate_excel_report(deal, include_sections)
        elif format == 'json':
            file_path = _generate_json_report(deal, include_sections)
        else:
            raise ValueError(f"Unsupported format: {format}")

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 100,
                'phase': 'complete',
                'message': 'Report ready'
            }
        )

        result = {
            'deal_id': deal_id,
            'format': format,
            'file_path': str(file_path),
            'generated_at': datetime.utcnow().isoformat(),
        }

        logger.info(f"Report exported for deal {deal_id}: {file_path}")
        return result

    except Exception as e:
        logger.error(f"Report export failed for deal {deal_id}: {e}")
        raise


def _generate_excel_report(deal, sections: List[str]) -> str:
    """Generate Excel report."""
    from config_v2 import OUTPUT_DIR
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Add basic content
    ws['A1'] = "IT Due Diligence Report"
    ws['A2'] = f"Target: {deal.target_name}"
    ws['A3'] = f"Generated: {datetime.utcnow().isoformat()}"

    # Save
    filename = f"report_{deal.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = OUTPUT_DIR / filename
    wb.save(file_path)

    return str(file_path)


def _generate_json_report(deal, sections: List[str]) -> str:
    """Generate JSON report."""
    import json
    from config_v2 import OUTPUT_DIR
    from web.database import Fact, Finding

    report = {
        'deal': deal.to_dict(),
        'generated_at': datetime.utcnow().isoformat(),
    }

    if 'facts' in sections:
        report['facts'] = [f.to_dict() for f in Fact.query.filter_by(deal_id=deal.id).all()]

    if 'risks' in sections:
        report['risks'] = [f.to_dict() for f in Finding.query.filter_by(
            deal_id=deal.id, finding_type='risk'
        ).all()]

    if 'work_items' in sections:
        report['work_items'] = [f.to_dict() for f in Finding.query.filter_by(
            deal_id=deal.id, finding_type='work_item'
        ).all()]

    filename = f"report_{deal.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    file_path = OUTPUT_DIR / filename

    with open(file_path, 'w') as f:
        json.dump(report, f, indent=2)

    return str(file_path)


@shared_task(bind=True, name='web.tasks.process_document')
def process_document_task(
    self,
    doc_id: str,
    file_path: str,
    filename: str,
    deal_id: Optional[str] = None
) -> Dict[str, Any]:
    """
    Process a single document as a background Celery task.

    This replaces the threading-based DocumentProcessor for production.
    Extracts facts, classifies into tiers, and persists to database.

    Args:
        doc_id: Document ID from registry
        file_path: Path to the uploaded file
        filename: Original filename
        deal_id: Deal ID for database persistence

    Returns:
        Dict with processing results
    """
    from web.database import db, Fact, Document
    from tools_v2.document_processor import ContentExtractor
    from tools_v2.incremental_extractor import IncrementalExtractor
    from tools_v2.tier_classifier import TierClassifier, Tier
    from stores.fact_store import FactStore
    from tools_v2.fact_merger import FactMerger
    import time

    logger.info(f"Processing document: {filename} (doc_id={doc_id})")

    # Update progress
    self.update_state(
        state='PROGRESS',
        meta={
            'progress': 0,
            'phase': 'extracting',
            'message': f'Extracting content from {filename}...'
        }
    )

    start_time = time.time()
    result = {
        'doc_id': doc_id,
        'filename': filename,
        'facts_extracted': 0,
        'facts_added': 0,
        'facts_updated': 0,
        'pending_review': 0,
        'errors': []
    }

    try:
        # Stage 1: Extract content
        extractor = ContentExtractor()
        extraction = extractor.extract(file_path)

        if not extraction.success:
            raise Exception(f"Content extraction failed: {extraction.error}")

        logger.info(f"Extracted {extraction.word_count} words from {extraction.page_count} pages")

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 20,
                'phase': 'analyzing',
                'message': f'Analyzing content ({extraction.word_count} words)...'
            }
        )

        # Stage 2: Extract facts using IncrementalExtractor
        # NOTE: We need a FactStore instance. Since we're in Celery worker,
        # we'll create a temporary one and persist results to DB at the end.
        fact_store = FactStore(deal_id=deal_id)

        fact_extractor = IncrementalExtractor(fact_store=fact_store)
        context = fact_extractor.build_context()

        extracted_facts = fact_extractor.extract_facts_from_content(
            content=extraction.content,
            source_document=filename,
            context=context
        )

        result['facts_extracted'] = len(extracted_facts)
        logger.info(f"Extracted {len(extracted_facts)} facts")

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 50,
                'phase': 'classifying',
                'message': f'Classifying {len(extracted_facts)} facts...'
            }
        )

        # Stage 3: Classify facts into review tiers
        fact_merger = FactMerger(fact_store)
        classifier = TierClassifier(fact_store=fact_store, fact_merger=fact_merger)

        facts_as_dicts = [f.to_dict() for f in extracted_facts]
        classified, stats = classifier.classify_batch(facts_as_dicts)

        tier1_count = len(classified.get(Tier.AUTO_APPLY, []))
        tier2_count = len(classified.get(Tier.BATCH_REVIEW, []))
        tier3_count = len(classified.get(Tier.INDIVIDUAL_REVIEW, []))

        logger.info(f"Classified: {tier1_count} auto-apply, {tier2_count} batch, {tier3_count} individual")

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 70,
                'phase': 'persisting',
                'message': 'Saving to database...'
            }
        )

        # Stage 4: Persist auto-eligible facts to database
        tier1_items = classified.get(Tier.AUTO_APPLY, [])
        auto_eligible_facts = [
            item["fact"] for item in tier1_items
            if item.get("classification", {}).get("auto_apply_eligible", False)
        ]

        facts_added = 0
        if auto_eligible_facts and deal_id:
            # Persist to database
            for fact_dict in auto_eligible_facts:
                try:
                    db_fact = Fact(
                        id=fact_dict.get('fact_id'),
                        deal_id=deal_id,
                        domain=fact_dict.get('domain', 'general'),
                        category=fact_dict.get('category', ''),
                        entity=fact_dict.get('entity', 'target'),
                        item=fact_dict.get('item', ''),
                        status=fact_dict.get('status', 'documented'),
                        details=fact_dict.get('details', {}),
                        evidence=fact_dict.get('evidence', {}),
                        source_document=filename,
                        source_quote=fact_dict.get('evidence', {}).get('exact_quote', ''),
                        confidence_score=fact_dict.get('confidence_score', 0.5),
                    )
                    db.session.add(db_fact)
                    facts_added += 1
                except Exception as e:
                    logger.warning(f"Failed to persist fact: {e}")
                    result['errors'].append(f"Fact persistence error: {str(e)}")

            db.session.commit()
            logger.info(f"Persisted {facts_added} auto-eligible facts to database")

        result['facts_added'] = facts_added
        result['pending_review'] = tier2_count + tier3_count + (tier1_count - facts_added)

        # Stage 5: Update document record in database
        if deal_id:
            doc_record = Document.query.filter_by(deal_id=deal_id, filename=filename).first()
            if doc_record:
                doc_record.status = 'processed'
                doc_record.facts_count = facts_added
                doc_record.processed_at = datetime.utcnow()
                db.session.commit()

        # Complete
        elapsed_ms = int((time.time() - start_time) * 1000)
        logger.info(f"Document processing complete: {filename} ({elapsed_ms}ms)")

        self.update_state(
            state='PROGRESS',
            meta={
                'progress': 100,
                'phase': 'complete',
                'message': f'Processing complete: {facts_added} facts added, {result["pending_review"]} pending review'
            }
        )

        return result

    except Exception as e:
        logger.error(f"Document processing failed for {filename}: {e}")
        import traceback
        logger.error(traceback.format_exc())

        # Update document status to failed
        if deal_id:
            try:
                doc_record = Document.query.filter_by(deal_id=deal_id, filename=filename).first()
                if doc_record:
                    doc_record.status = 'failed'
                    doc_record.error_message = str(e)
                    db.session.commit()
            except Exception as db_err:
                logger.error(f"Failed to update document status: {db_err}")

        raise  # Re-raise so Celery marks task as FAILURE
