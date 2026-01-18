"""
Excel Export Module

Exports IT Due Diligence findings to client-ready Excel workbooks.

Features:
- Multiple worksheets (Summary, Risks, Gaps, Work Items, Recommendations, etc.)
- Professional formatting with conditional formatting for severity/priority
- Evidence citations and source references
- Sortable/filterable columns
"""

import os
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Import openpyxl components at module level
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = ColorScaleRule = FormulaRule = None


def export_findings_to_excel(
    run_id: str,
    repository: Any,
    output_path: str,
    include_evidence: bool = True
) -> Dict[str, Any]:
    """
    Export all findings for a run to a formatted Excel workbook.

    Args:
        run_id: Analysis run ID
        repository: Repository instance for data access
        output_path: Path to save Excel file
        include_evidence: Whether to include source evidence columns

    Returns:
        Dict with output_path, counts per sheet, and total
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ImportError("openpyxl required for Excel export. Run: pip install openpyxl")

    # Fetch all data
    data = _fetch_all_findings(run_id, repository)

    wb = openpyxl.Workbook()

    # Define styles
    styles = _get_styles()

    # Create worksheets
    _create_summary_sheet(wb, data, styles)
    _create_risks_sheet(wb, data['risks'], styles, include_evidence)
    _create_gaps_sheet(wb, data['gaps'], styles, include_evidence)
    _create_work_items_sheet(wb, data['work_items'], styles)
    _create_recommendations_sheet(wb, data['recommendations'], styles)
    _create_current_state_sheet(wb, data['current_state'], styles)
    _create_assumptions_sheet(wb, data['assumptions'], styles)

    # Remove default empty sheet if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    wb.save(output_path)
    logger.info(f"Exported findings to {output_path}")

    # Return summary info
    counts = {
        'Risks': len(data['risks']),
        'Gaps': len(data['gaps']),
        'Work Items': len(data['work_items']),
        'Recommendations': len(data['recommendations']),
        'Current State': len(data['current_state']),
        'Assumptions': len(data['assumptions'])
    }
    total = sum(counts.values())

    return {
        'output_path': output_path,
        'counts': counts,
        'total': total,
        'run_id': run_id
    }


def _fetch_all_findings(run_id: str, repository: Any) -> Dict[str, List]:
    """Fetch all findings for a run from the repository."""
    data = {
        'run_id': run_id,
        'risks': [],
        'gaps': [],
        'work_items': [],
        'recommendations': [],
        'current_state': [],
        'assumptions': [],
        'questions': [],
        'strategic_considerations': []
    }

    try:
        # Get run info
        run = repository.get_run(run_id)
        if run:
            data['run_info'] = run.to_dict() if hasattr(run, 'to_dict') else run

        # Get all findings
        conn = repository.db.connect()

        # Risks
        rows = conn.execute(
            "SELECT * FROM risks WHERE run_id = ? ORDER BY severity DESC, likelihood DESC",
            (run_id,)
        ).fetchall()
        data['risks'] = [dict(row) for row in rows]

        # Gaps
        rows = conn.execute(
            "SELECT * FROM gaps WHERE run_id = ? ORDER BY priority DESC",
            (run_id,)
        ).fetchall()
        data['gaps'] = [dict(row) for row in rows]

        # Work Items
        rows = conn.execute(
            "SELECT * FROM work_items WHERE run_id = ? ORDER BY phase, priority_score DESC",
            (run_id,)
        ).fetchall()
        data['work_items'] = [dict(row) for row in rows]

        # Recommendations
        rows = conn.execute(
            "SELECT * FROM recommendations WHERE run_id = ? ORDER BY priority DESC",
            (run_id,)
        ).fetchall()
        data['recommendations'] = [dict(row) for row in rows]

        # Current State / Inventory
        rows = conn.execute(
            "SELECT * FROM inventory_items WHERE run_id = ? ORDER BY domain, category",
            (run_id,)
        ).fetchall()
        data['current_state'] = [dict(row) for row in rows]

        # Assumptions
        rows = conn.execute(
            "SELECT * FROM assumptions WHERE run_id = ? ORDER BY confidence DESC",
            (run_id,)
        ).fetchall()
        data['assumptions'] = [dict(row) for row in rows]

        # Questions
        rows = conn.execute(
            "SELECT * FROM questions WHERE run_id = ? ORDER BY priority DESC, status",
            (run_id,)
        ).fetchall()
        data['questions'] = [dict(row) for row in rows]

        # Strategic Considerations
        rows = conn.execute(
            "SELECT * FROM strategic_considerations WHERE run_id = ? ORDER BY domain",
            (run_id,)
        ).fetchall()
        data['strategic_considerations'] = [dict(row) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching findings: {e}")
        raise

    return data


def _get_styles():
    """Define Excel styles for consistent formatting."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'header_fill': PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),

        'subheader_font': Font(bold=True, size=10),
        'subheader_fill': PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid"),

        'data_alignment': Alignment(vertical="top", wrap_text=True),
        'center_alignment': Alignment(horizontal="center", vertical="top"),

        'border': thin_border,

        # Severity/Priority colors
        'critical_fill': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        'critical_font': Font(color="FFFFFF", bold=True),
        'high_fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'medium_fill': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'low_fill': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),

        # Phase colors
        'day1_fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'day100_fill': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'post100_fill': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
        'optional_fill': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid"),
    }


def _apply_header_style(ws, row, styles, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']


def _apply_severity_formatting(cell, value, styles):
    """Apply conditional formatting based on severity/priority value."""
    if not value:
        return

    value_lower = str(value).lower()
    if value_lower == 'critical':
        cell.fill = styles['critical_fill']
        cell.font = styles['critical_font']
    elif value_lower == 'high':
        cell.fill = styles['high_fill']
    elif value_lower == 'medium':
        cell.fill = styles['medium_fill']
    elif value_lower == 'low':
        cell.fill = styles['low_fill']


def _apply_phase_formatting(cell, value, styles):
    """Apply conditional formatting based on phase value."""
    if not value:
        return

    value_lower = str(value).lower().replace('_', '')
    if 'day1' in value_lower or value_lower == 'day_1':
        cell.fill = styles['day1_fill']
    elif 'day100' in value_lower or value_lower == 'day_100':
        cell.fill = styles['day100_fill']
    elif 'post' in value_lower:
        cell.fill = styles['post100_fill']
    elif 'optional' in value_lower:
        cell.fill = styles['optional_fill']


def _auto_size_columns(ws, min_width=10, max_width=60):
    """Auto-size columns based on content."""
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_summary_sheet(wb, data: Dict, styles: Dict):
    """Create the Summary worksheet."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws['A1'] = "IT Due Diligence Analysis Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    # Run info
    ws['A3'] = "Run ID:"
    ws['B3'] = data.get('run_id', 'N/A')
    ws['A4'] = "Generated:"
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    run_info = data.get('run_info', {})
    ws['A5'] = "Run Name:"
    ws['B5'] = run_info.get('run_name', 'N/A')
    ws['A6'] = "Status:"
    ws['B6'] = run_info.get('status', 'N/A')

    # Counts summary
    ws['A8'] = "Findings Summary"
    ws['A8'].font = Font(bold=True, size=12)

    counts = [
        ("Risks", len(data['risks'])),
        ("Gaps", len(data['gaps'])),
        ("Work Items", len(data['work_items'])),
        ("Recommendations", len(data['recommendations'])),
        ("Current State Items", len(data['current_state'])),
        ("Assumptions", len(data['assumptions'])),
        ("Questions", len(data['questions'])),
    ]

    row = 9
    for label, count in counts:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Risk severity breakdown
    row += 1
    ws.cell(row=row, column=1, value="Risk Severity Breakdown")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    severity_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    for risk in data['risks']:
        sev = str(risk.get('severity', 'medium')).lower()
        if sev in severity_counts:
            severity_counts[sev] += 1

    for sev, count in severity_counts.items():
        ws.cell(row=row, column=1, value=sev.title())
        ws.cell(row=row, column=2, value=count)
        _apply_severity_formatting(ws.cell(row=row, column=1), sev, styles)
        row += 1

    # Work item phase breakdown
    row += 1
    ws.cell(row=row, column=1, value="Work Items by Phase")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    phase_counts = {'Day_1': 0, 'Day_100': 0, 'Post_100': 0, 'Optional': 0}
    for wi in data['work_items']:
        phase = wi.get('phase', 'Day_100')
        if phase in phase_counts:
            phase_counts[phase] += 1

    for phase, count in phase_counts.items():
        ws.cell(row=row, column=1, value=phase.replace('_', ' '))
        ws.cell(row=row, column=2, value=count)
        _apply_phase_formatting(ws.cell(row=row, column=1), phase, styles)
        row += 1

    _auto_size_columns(ws)


def _create_risks_sheet(wb, risks: List[Dict], styles: Dict, include_evidence: bool):
    """Create the Risks worksheet."""
    ws = wb.create_sheet("Risks")

    # Headers
    headers = ["#", "Domain", "Risk Description", "Severity", "Likelihood",
               "Risk Type", "Integration Dependent", "Deal Impact",
               "Mitigation", "Cost Impact"]

    if include_evidence:
        headers.extend(["Evidence Quote", "Evidence Type", "Confidence"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    # Data rows
    for row_num, risk in enumerate(risks, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=risk.get('domain', ''))
        ws.cell(row=row_num, column=3, value=risk.get('risk_description', ''))

        # Severity with formatting
        sev_cell = ws.cell(row=row_num, column=4, value=risk.get('severity', ''))
        _apply_severity_formatting(sev_cell, risk.get('severity'), styles)

        ws.cell(row=row_num, column=5, value=risk.get('likelihood', ''))
        ws.cell(row=row_num, column=6, value=risk.get('risk_type', ''))
        ws.cell(row=row_num, column=7, value='Yes' if risk.get('integration_dependent') else 'No')

        # Deal impact (may be JSON)
        deal_impact = risk.get('deal_impact', [])
        if isinstance(deal_impact, str):
            try:
                import json
                deal_impact = json.loads(deal_impact)
            except:
                pass
        if isinstance(deal_impact, list):
            deal_impact = ', '.join(deal_impact)
        ws.cell(row=row_num, column=8, value=deal_impact)

        ws.cell(row=row_num, column=9, value=risk.get('mitigation', ''))
        ws.cell(row=row_num, column=10, value=risk.get('cost_impact_estimate', ''))

        if include_evidence:
            # Parse source_evidence if it's JSON
            evidence = risk.get('source_evidence', {})
            if isinstance(evidence, str):
                try:
                    import json
                    evidence = json.loads(evidence)
                except:
                    evidence = {}

            ws.cell(row=row_num, column=11, value=evidence.get('exact_quote', '') if evidence else '')
            ws.cell(row=row_num, column=12, value=evidence.get('evidence_type', '') if evidence else '')
            ws.cell(row=row_num, column=13, value=evidence.get('confidence_level', '') if evidence else '')

        # Apply alignment
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    _auto_size_columns(ws)


def _create_gaps_sheet(wb, gaps: List[Dict], styles: Dict, include_evidence: bool):
    """Create the Gaps worksheet."""
    ws = wb.create_sheet("Gaps")

    headers = ["#", "Domain", "Gap Description", "Why Needed", "Priority",
               "Gap Type", "Suggested Source", "Cost Impact", "Question Status"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, gap in enumerate(gaps, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=gap.get('domain', ''))
        ws.cell(row=row_num, column=3, value=gap.get('gap_description', ''))
        ws.cell(row=row_num, column=4, value=gap.get('why_needed', ''))

        # Priority with formatting
        pri_cell = ws.cell(row=row_num, column=5, value=gap.get('priority', ''))
        _apply_severity_formatting(pri_cell, gap.get('priority'), styles)

        ws.cell(row=row_num, column=6, value=gap.get('gap_type', ''))
        ws.cell(row=row_num, column=7, value=gap.get('suggested_source', ''))
        ws.cell(row=row_num, column=8, value=gap.get('cost_impact', ''))
        ws.cell(row=row_num, column=9, value=gap.get('question_status', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_work_items_sheet(wb, work_items: List[Dict], styles: Dict):
    """Create the Work Items worksheet."""
    ws = wb.create_sheet("Work Items")

    headers = ["#", "Domain", "Title", "Description", "Category", "Phase",
               "Phase Rationale", "Effort Estimate", "Cost Estimate",
               "Skills Required", "Dependencies", "Triggered By"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, wi in enumerate(work_items, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=wi.get('domain', ''))
        ws.cell(row=row_num, column=3, value=wi.get('title', ''))
        ws.cell(row=row_num, column=4, value=wi.get('description', ''))
        ws.cell(row=row_num, column=5, value=wi.get('category', ''))

        # Phase with formatting
        phase_cell = ws.cell(row=row_num, column=6, value=wi.get('phase', ''))
        _apply_phase_formatting(phase_cell, wi.get('phase'), styles)

        ws.cell(row=row_num, column=7, value=wi.get('phase_rationale', ''))
        ws.cell(row=row_num, column=8, value=wi.get('effort_estimate', ''))
        ws.cell(row=row_num, column=9, value=wi.get('cost_estimate_range', ''))

        # Parse JSON arrays
        def parse_list(val):
            if isinstance(val, str):
                try:
                    import json
                    val = json.loads(val)
                except:
                    pass
            if isinstance(val, list):
                return ', '.join(str(v) for v in val)
            return val or ''

        ws.cell(row=row_num, column=10, value=parse_list(wi.get('skills_required')))
        ws.cell(row=row_num, column=11, value=parse_list(wi.get('depends_on')))
        ws.cell(row=row_num, column=12, value=parse_list(wi.get('triggered_by')))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_recommendations_sheet(wb, recommendations: List[Dict], styles: Dict):
    """Create the Recommendations worksheet."""
    ws = wb.create_sheet("Recommendations")

    headers = ["#", "Domain", "Recommendation", "Rationale", "Priority",
               "Timing", "Investment Required", "Driven By"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, rec in enumerate(recommendations, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=rec.get('domain', ''))
        ws.cell(row=row_num, column=3, value=rec.get('recommendation_text', ''))
        ws.cell(row=row_num, column=4, value=rec.get('rationale', ''))

        pri_cell = ws.cell(row=row_num, column=5, value=rec.get('priority', ''))
        _apply_severity_formatting(pri_cell, rec.get('priority'), styles)

        ws.cell(row=row_num, column=6, value=rec.get('timing', ''))
        ws.cell(row=row_num, column=7, value=rec.get('investment_required', ''))

        driven_by = rec.get('driven_by', [])
        if isinstance(driven_by, str):
            try:
                import json
                driven_by = json.loads(driven_by)
            except:
                pass
        if isinstance(driven_by, list):
            driven_by = ', '.join(driven_by)
        ws.cell(row=row_num, column=8, value=driven_by)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_current_state_sheet(wb, items: List[Dict], styles: Dict):
    """Create the Current State / Inventory worksheet."""
    ws = wb.create_sheet("Current State")

    headers = ["#", "Domain", "Category", "Item Name", "Description",
               "Status", "Maturity", "Standalone Viability", "Key Characteristics"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=item.get('domain', ''))
        ws.cell(row=row_num, column=3, value=item.get('category', ''))
        ws.cell(row=row_num, column=4, value=item.get('item_name', ''))
        ws.cell(row=row_num, column=5, value=item.get('description', ''))
        ws.cell(row=row_num, column=6, value=item.get('status', ''))
        ws.cell(row=row_num, column=7, value=item.get('maturity', ''))
        ws.cell(row=row_num, column=8, value=item.get('standalone_viability', ''))

        chars = item.get('key_characteristics', [])
        if isinstance(chars, str):
            try:
                import json
                chars = json.loads(chars)
            except:
                pass
        if isinstance(chars, list):
            chars = ', '.join(chars)
        ws.cell(row=row_num, column=9, value=chars)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


def _create_assumptions_sheet(wb, assumptions: List[Dict], styles: Dict):
    """Create the Assumptions worksheet."""
    ws = wb.create_sheet("Assumptions")

    headers = ["#", "Domain", "Assumption", "Basis", "Confidence",
               "Impact", "Validation Needed", "Validation Status"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, styles, len(headers))

    for row_num, assumption in enumerate(assumptions, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=assumption.get('domain', ''))
        ws.cell(row=row_num, column=3, value=assumption.get('assumption_text', ''))
        ws.cell(row=row_num, column=4, value=assumption.get('basis', ''))
        ws.cell(row=row_num, column=5, value=assumption.get('confidence', ''))
        ws.cell(row=row_num, column=6, value=assumption.get('impact', ''))
        ws.cell(row=row_num, column=7, value=assumption.get('validation_needed', ''))
        ws.cell(row=row_num, column=8, value=assumption.get('validation_status', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = styles['data_alignment']
            ws.cell(row=row_num, column=col).border = styles['border']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_size_columns(ws)


# =============================================================================
# CLI TESTING
# =============================================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))

    from storage import Database, Repository

    logging.basicConfig(level=logging.INFO)

    db = Database()
    db.initialize_schema()
    repository = Repository(db)

    # Get latest run
    latest = repository.get_latest_run()
    if not latest:
        print("No runs found.")
        sys.exit(1)

    run_id = latest.run_id
    print(f"Exporting findings for run: {run_id}")

    output_path = f"findings_{run_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    result = export_findings_to_excel(run_id, repository, output_path)
    print(f"Exported to: {result}")
