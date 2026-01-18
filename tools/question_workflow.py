"""
Question Workflow Module

Handles the complete question tracking lifecycle:
1. Parse questions from input documents (answered + unanswered)
2. Support domain agents identifying new questions during analysis
3. Export questions to Excel for client follow-up
4. Import answers from Excel
5. Update gaps/assumptions when questions are answered
"""

import os
import re
import json
import logging
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from dataclasses import dataclass, asdict

logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT QUESTION FORMAT SPECIFICATION
# =============================================================================
"""
Expected question format in input documents:

## Questions - Answered
| # | Question | Answer | Source | Date |
|---|----------|--------|--------|------|
| 1 | What is the DR RTO? | 4 hours for critical systems | CTO Interview | 2026-01-05 |
| 2 | How many SAP users? | 450 named users | IT Director Email | 2026-01-03 |

## Questions - Unanswered
| # | Question | Context | Priority | Asked Date |
|---|----------|---------|----------|------------|
| 1 | What is the cyber insurance coverage limit? | Needed for risk assessment | High | 2026-01-02 |
| 2 | Who owns the data center lease? | Determine transfer requirements | Medium | 2026-01-04 |

Alternative formats also supported:
- Markdown lists with "Answered:" and "Unanswered:" headers
- JSON sections with "answered_questions" and "unanswered_questions" arrays
"""


# =============================================================================
# QUESTION PARSING
# =============================================================================

@dataclass
class ParsedQuestion:
    """Question extracted from document"""
    question_text: str
    is_answered: bool = False
    answer_text: Optional[str] = None
    answer_source: Optional[str] = None
    answer_date: Optional[str] = None
    context: Optional[str] = None
    priority: str = "medium"
    asked_date: Optional[str] = None
    domain: Optional[str] = None

    def to_dict(self) -> Dict:
        return asdict(self)


class DocumentQuestionParser:
    """Parse questions from various document formats"""

    # Patterns for detecting question sections
    ANSWERED_SECTION_PATTERNS = [
        r'#+\s*Questions?\s*[-–]\s*Answered',
        r'#+\s*Answered\s*Questions?',
        r'ANSWERED\s*QUESTIONS?:',
        r'\*\*Answered\s*Questions?\*\*',
    ]

    UNANSWERED_SECTION_PATTERNS = [
        r'#+\s*Questions?\s*[-–]\s*Unanswered',
        r'#+\s*Unanswered\s*Questions?',
        r'#+\s*Open\s*Questions?',
        r'#+\s*Pending\s*Questions?',
        r'UNANSWERED\s*QUESTIONS?:',
        r'OPEN\s*QUESTIONS?:',
        r'\*\*Unanswered\s*Questions?\*\*',
        r'\*\*Open\s*Questions?\*\*',
    ]

    def parse_document(self, content: str, domain: str = None) -> List[ParsedQuestion]:
        """
        Parse all questions from document content.

        Args:
            content: Document text content
            domain: Optional domain to assign to all questions

        Returns:
            List of ParsedQuestion objects
        """
        questions = []

        # Try JSON parsing first
        json_questions = self._parse_json_questions(content, domain)
        if json_questions:
            questions.extend(json_questions)

        # Parse markdown table format
        table_questions = self._parse_table_questions(content, domain)
        questions.extend(table_questions)

        # Parse list format
        list_questions = self._parse_list_questions(content, domain)
        questions.extend(list_questions)

        # Deduplicate based on question text
        seen = set()
        unique_questions = []
        for q in questions:
            key = q.question_text.strip().lower()
            if key not in seen:
                seen.add(key)
                unique_questions.append(q)

        logger.info(f"Parsed {len(unique_questions)} unique questions from document")
        return unique_questions

    def _parse_json_questions(self, content: str, domain: str = None) -> List[ParsedQuestion]:
        """Parse questions from JSON blocks in document"""
        questions = []

        # Look for JSON code blocks
        json_pattern = r'```json\s*(.*?)```'
        matches = re.findall(json_pattern, content, re.DOTALL)

        for match in matches:
            try:
                data = json.loads(match)

                # Check for answered_questions array
                if 'answered_questions' in data:
                    for q in data['answered_questions']:
                        questions.append(ParsedQuestion(
                            question_text=q.get('question', q.get('text', '')),
                            is_answered=True,
                            answer_text=q.get('answer', ''),
                            answer_source=q.get('source', q.get('answer_source', '')),
                            answer_date=q.get('date', q.get('answer_date', '')),
                            context=q.get('context', ''),
                            priority=q.get('priority', 'medium'),
                            domain=domain
                        ))

                # Check for unanswered_questions array
                if 'unanswered_questions' in data or 'open_questions' in data:
                    unanswered = data.get('unanswered_questions', data.get('open_questions', []))
                    for q in unanswered:
                        questions.append(ParsedQuestion(
                            question_text=q.get('question', q.get('text', '')),
                            is_answered=False,
                            context=q.get('context', ''),
                            priority=q.get('priority', 'medium'),
                            asked_date=q.get('asked_date', q.get('date', '')),
                            domain=domain
                        ))

            except json.JSONDecodeError:
                continue

        return questions

    def _parse_table_questions(self, content: str, domain: str = None) -> List[ParsedQuestion]:
        """Parse questions from markdown tables"""
        questions = []
        lines = content.split('\n')

        current_section = None  # 'answered' or 'unanswered'
        in_table = False
        header_indices = {}

        for i, line in enumerate(lines):
            # Check for section headers
            for pattern in self.ANSWERED_SECTION_PATTERNS:
                if re.search(pattern, line, re.IGNORECASE):
                    current_section = 'answered'
                    in_table = False
                    header_indices = {}
                    break

            for pattern in self.UNANSWERED_SECTION_PATTERNS:
                if re.search(pattern, line, re.IGNORECASE):
                    current_section = 'unanswered'
                    in_table = False
                    header_indices = {}
                    break

            if current_section is None:
                continue

            # Detect table header row
            if '|' in line and not in_table:
                cells = [c.strip().lower() for c in line.split('|')]
                if any(x in cells for x in ['question', 'questions', '#']):
                    in_table = True
                    # Map column indices
                    for idx, cell in enumerate(cells):
                        if cell in ['question', 'questions']:
                            header_indices['question'] = idx
                        elif cell in ['answer', 'response']:
                            header_indices['answer'] = idx
                        elif cell in ['source', 'answer source']:
                            header_indices['source'] = idx
                        elif cell in ['date', 'answer date', 'asked date']:
                            header_indices['date'] = idx
                        elif cell in ['context', 'notes']:
                            header_indices['context'] = idx
                        elif cell in ['priority']:
                            header_indices['priority'] = idx
                    continue

            # Skip separator row
            if in_table and re.match(r'\|[\s\-:]+\|', line):
                continue

            # Parse data row
            if in_table and '|' in line:
                cells = [c.strip() for c in line.split('|')]

                # Check if this looks like a data row (not header/separator)
                if len(cells) < 3 or all(not c for c in cells):
                    continue

                question_text = cells[header_indices.get('question', 1)] if header_indices.get('question') else ''

                if not question_text or question_text.startswith('-'):
                    continue

                if current_section == 'answered':
                    questions.append(ParsedQuestion(
                        question_text=question_text,
                        is_answered=True,
                        answer_text=cells[header_indices['answer']] if header_indices.get('answer') else '',
                        answer_source=cells[header_indices['source']] if header_indices.get('source') else '',
                        answer_date=cells[header_indices['date']] if header_indices.get('date') else '',
                        context=cells[header_indices['context']] if header_indices.get('context') else '',
                        priority=cells[header_indices['priority']] if header_indices.get('priority') else 'medium',
                        domain=domain
                    ))
                else:
                    questions.append(ParsedQuestion(
                        question_text=question_text,
                        is_answered=False,
                        context=cells[header_indices['context']] if header_indices.get('context') else '',
                        priority=cells[header_indices['priority']] if header_indices.get('priority') else 'medium',
                        asked_date=cells[header_indices['date']] if header_indices.get('date') else '',
                        domain=domain
                    ))

            # End of table detection (new section or blank line after table data)
            if in_table and (line.strip() == '' or line.startswith('#')):
                if not '|' in line:
                    in_table = False

        return questions

    def _parse_list_questions(self, content: str, domain: str = None) -> List[ParsedQuestion]:
        """Parse questions from bullet/numbered lists"""
        questions = []
        lines = content.split('\n')

        current_section = None

        for i, line in enumerate(lines):
            # Check for section headers
            for pattern in self.ANSWERED_SECTION_PATTERNS:
                if re.search(pattern, line, re.IGNORECASE):
                    current_section = 'answered'
                    break

            for pattern in self.UNANSWERED_SECTION_PATTERNS:
                if re.search(pattern, line, re.IGNORECASE):
                    current_section = 'unanswered'
                    break

            if current_section is None:
                continue

            # Parse list items
            list_match = re.match(r'^[\s]*[-*•]\s*(.+)$', line)
            numbered_match = re.match(r'^[\s]*\d+[\.\)]\s*(.+)$', line)

            item_text = None
            if list_match:
                item_text = list_match.group(1)
            elif numbered_match:
                item_text = numbered_match.group(1)

            if item_text and '?' in item_text:
                # This looks like a question
                # Try to extract question and answer if format is "Q: ... A: ..."
                qa_match = re.match(r'(?:Q:|Question:)?\s*(.+\?)\s*(?:A:|Answer:)?\s*(.*)$', item_text, re.IGNORECASE)

                if qa_match and current_section == 'answered':
                    questions.append(ParsedQuestion(
                        question_text=qa_match.group(1).strip(),
                        is_answered=True,
                        answer_text=qa_match.group(2).strip() if qa_match.group(2) else '',
                        domain=domain
                    ))
                elif current_section == 'unanswered':
                    questions.append(ParsedQuestion(
                        question_text=item_text.strip(),
                        is_answered=False,
                        domain=domain
                    ))

        return questions


# =============================================================================
# QUESTION TOOL FOR DOMAIN AGENTS
# =============================================================================

QUESTION_TOOL_DEFINITION = {
    "name": "flag_question",
    "description": "Flag a question that needs to be asked to management/seller. Use when you identify information gaps that require follow-up.",
    "input_schema": {
        "type": "object",
        "properties": {
            "question": {
                "type": "string",
                "description": "The specific question to ask"
            },
            "context": {
                "type": "string",
                "description": "Why this question is important / what decision it impacts"
            },
            "priority": {
                "type": "string",
                "enum": ["critical", "high", "medium", "low"],
                "description": "Priority of getting this answered"
            },
            "related_to": {
                "type": "string",
                "description": "Related gap ID, assumption ID, or topic area"
            },
            "suggested_source": {
                "type": "string",
                "description": "Suggested person/role to ask (e.g., 'CTO', 'IT Director', 'CFO')"
            }
        },
        "required": ["question", "context", "priority"]
    }
}


def execute_question_tool(tool_input: Dict) -> Dict:
    """Execute the flag_question tool"""
    if not isinstance(tool_input, dict):
        logger.error(f"Invalid tool input: expected dict, got {type(tool_input)}")
        return {"status": "error", "message": "Invalid input type"}

    return {
        "status": "recorded",
        "question": tool_input.get("question"),
        "context": tool_input.get("context"),
        "priority": tool_input.get("priority"),
        "related_to": tool_input.get("related_to"),
        "suggested_source": tool_input.get("suggested_source")
    }


# =============================================================================
# EXCEL EXPORT/IMPORT
# =============================================================================

def export_questions_to_excel(
    questions: List[Dict],
    output_path: str,
    include_answered: bool = True
) -> str:
    """
    Export questions to Excel for client follow-up.

    Args:
        questions: List of question dicts from database
        output_path: Path to save Excel file
        include_answered: Whether to include already-answered questions

    Returns:
        Path to created Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        logger.warning("openpyxl not installed, falling back to CSV export")
        return _export_questions_to_csv(questions, output_path, include_answered)

    wb = openpyxl.Workbook()

    # Create Open Questions sheet
    ws_open = wb.active
    ws_open.title = "Open Questions"

    # Headers
    headers = ["#", "Domain", "Question", "Context", "Priority", "Related To", "Suggested Source", "Status", "Asked Date"]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws_open.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows - Open questions
    row_num = 2
    open_count = 0
    for q in questions:
        status = q.get('status', 'draft')
        if status not in ['answered', 'closed'] or include_answered:
            if status in ['answered', 'closed']:
                continue  # Skip answered for open sheet

            open_count += 1
            ws_open.cell(row=row_num, column=1, value=open_count)
            ws_open.cell(row=row_num, column=2, value=q.get('domain', ''))
            ws_open.cell(row=row_num, column=3, value=q.get('question_text', ''))
            ws_open.cell(row=row_num, column=4, value=q.get('context', ''))
            ws_open.cell(row=row_num, column=5, value=q.get('priority', 'medium'))
            ws_open.cell(row=row_num, column=6, value=q.get('linked_gap_id', '') or q.get('linked_assumption_id', ''))
            ws_open.cell(row=row_num, column=7, value=q.get('suggested_source', ''))
            ws_open.cell(row=row_num, column=8, value=status)
            ws_open.cell(row=row_num, column=9, value=q.get('sent_date', '') or q.get('created_at', ''))
            row_num += 1

    # Add Answer column for input
    ws_open.cell(row=1, column=10, value="Answer (Fill In)")
    ws_open.cell(row=1, column=10).font = header_font
    ws_open.cell(row=1, column=10).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    ws_open.cell(row=1, column=11, value="Answer Source")
    ws_open.cell(row=1, column=11).font = header_font
    ws_open.cell(row=1, column=11).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # Create Answered Questions sheet
    if include_answered:
        ws_answered = wb.create_sheet("Answered Questions")

        answered_headers = ["#", "Domain", "Question", "Answer", "Answer Source", "Answer Date", "Priority"]
        for col, header in enumerate(answered_headers, 1):
            cell = ws_answered.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        row_num = 2
        answered_count = 0
        for q in questions:
            if q.get('status') in ['answered', 'closed']:
                answered_count += 1
                ws_answered.cell(row=row_num, column=1, value=answered_count)
                ws_answered.cell(row=row_num, column=2, value=q.get('domain', ''))
                ws_answered.cell(row=row_num, column=3, value=q.get('question_text', ''))
                ws_answered.cell(row=row_num, column=4, value=q.get('answer_text', ''))
                ws_answered.cell(row=row_num, column=5, value=q.get('answer_source', ''))
                ws_answered.cell(row=row_num, column=6, value=q.get('answer_date', ''))
                ws_answered.cell(row=row_num, column=7, value=q.get('priority', ''))
                row_num += 1

    # Auto-size columns
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    # Save
    wb.save(output_path)
    logger.info(f"Exported {open_count} open questions to {output_path}")

    return output_path


def _export_questions_to_csv(questions: List[Dict], output_path: str, include_answered: bool) -> str:
    """Fallback CSV export when openpyxl not available"""
    import csv

    csv_path = output_path.replace('.xlsx', '.csv')

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["#", "Domain", "Question", "Context", "Priority", "Status", "Answer", "Answer Source"])

        count = 0
        for q in questions:
            status = q.get('status', 'draft')
            if status in ['answered', 'closed'] and not include_answered:
                continue
            count += 1
            writer.writerow([
                count,
                q.get('domain', ''),
                q.get('question_text', ''),
                q.get('context', ''),
                q.get('priority', 'medium'),
                status,
                q.get('answer_text', ''),
                q.get('answer_source', '')
            ])

    logger.info(f"Exported {count} questions to {csv_path}")
    return csv_path


def import_answers_from_excel(
    file_path: str,
    repository: Any = None
) -> Dict[str, Any]:
    """
    Import answers from Excel file.

    Expected format: Excel with "Open Questions" sheet containing:
    - Column 3: Question text (to match)
    - Column 10: Answer (Fill In)
    - Column 11: Answer Source

    Args:
        file_path: Path to Excel file with answers
        repository: Repository instance for database updates

    Returns:
        Dict with import results
    """
    results = {
        'file': file_path,
        'questions_found': 0,
        'answers_imported': 0,
        'errors': [],
        'imported_answers': []
    }

    try:
        import openpyxl
    except ImportError:
        results['errors'].append("openpyxl not installed - cannot read Excel files")
        return results

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        results['errors'].append(f"Failed to open file: {e}")
        return results

    # Look for Open Questions sheet
    if "Open Questions" not in wb.sheetnames:
        results['errors'].append("Sheet 'Open Questions' not found in workbook")
        return results

    ws = wb["Open Questions"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:  # Skip empty rows
            continue

        results['questions_found'] += 1

        question_text = row[2]  # Column 3 (0-indexed: 2)
        answer_text = row[9] if len(row) > 9 else None  # Column 10
        answer_source = row[10] if len(row) > 10 else None  # Column 11

        if answer_text and str(answer_text).strip():
            results['imported_answers'].append({
                'question_text': question_text,
                'answer_text': str(answer_text).strip(),
                'answer_source': str(answer_source).strip() if answer_source else 'Excel Import'
            })
            results['answers_imported'] += 1

    wb.close()

    # If repository provided, update database
    if repository and results['imported_answers']:
        for answer in results['imported_answers']:
            try:
                # Find matching question in database
                # This would need the repository to have a search-by-text method
                logger.info(f"Would update question: {answer['question_text'][:50]}...")
            except Exception as e:
                results['errors'].append(f"DB update failed for question: {e}")

    logger.info(f"Imported {results['answers_imported']} answers from {file_path}")
    return results


# =============================================================================
# QUESTION WORKFLOW INTEGRATION
# =============================================================================

class QuestionWorkflow:
    """
    Manages the full question tracking workflow.

    Coordinates:
    - Parsing questions from input documents
    - Storing questions in database
    - Tracking question status
    - Export/import for client follow-up
    """

    def __init__(self, repository: Any = None):
        """
        Initialize workflow.

        Args:
            repository: Repository instance for database operations
        """
        self.repository = repository
        self.parser = DocumentQuestionParser()

    def process_document_questions(
        self,
        document_content: str,
        run_id: str,
        document_id: str = None
    ) -> Dict[str, Any]:
        """
        Parse questions from document and store in database.

        Args:
            document_content: Full document text
            run_id: Current analysis run ID
            document_id: Optional source document ID

        Returns:
            Dict with processing results
        """
        results = {
            'parsed_count': 0,
            'answered_count': 0,
            'unanswered_count': 0,
            'stored_count': 0,
            'questions': []
        }

        # Parse questions from document
        parsed = self.parser.parse_document(document_content)
        results['parsed_count'] = len(parsed)

        for pq in parsed:
            if pq.is_answered:
                results['answered_count'] += 1
            else:
                results['unanswered_count'] += 1

            results['questions'].append(pq.to_dict())

            # Store in database if repository available
            if self.repository:
                try:
                    from storage.models import Question

                    q = Question(
                        question_id=f"Q-{datetime.now().strftime('%Y%m%d%H%M%S')}-{results['stored_count']:04d}",
                        run_id=run_id,
                        question_text=pq.question_text,
                        context=pq.context,
                        priority=pq.priority,
                        status='answered' if pq.is_answered else 'draft',
                        answer_text=pq.answer_text,
                        answer_date=pq.answer_date,
                        answer_source=pq.answer_source
                    )

                    self.repository.create_question(q)
                    results['stored_count'] += 1

                except Exception as e:
                    logger.warning(f"Failed to store question: {e}")

        logger.info(
            f"Processed {results['parsed_count']} questions from document: "
            f"{results['answered_count']} answered, {results['unanswered_count']} unanswered"
        )

        return results

    def add_agent_question(
        self,
        question_data: Dict,
        run_id: str,
        domain: str
    ) -> Optional[str]:
        """
        Add a question identified by domain agent during analysis.

        Args:
            question_data: Output from flag_question tool
            run_id: Current analysis run ID
            domain: Domain that identified the question

        Returns:
            Question ID if stored, None otherwise
        """
        if not question_data or not isinstance(question_data, dict):
            return None

        if not self.repository:
            logger.warning("No repository available to store question")
            return None

        try:
            from storage.models import Question, generate_id

            q = Question(
                question_id=generate_id("Q"),
                run_id=run_id,
                question_text=question_data.get('question', ''),
                context=question_data.get('context', ''),
                priority=question_data.get('priority', 'medium'),
                status='draft'
            )

            # Link to gap if related_to looks like a gap ID
            related = question_data.get('related_to', '')
            if related and related.startswith('G-'):
                q.linked_gap_id = related
            elif related and related.startswith('A-'):
                q.linked_assumption_id = related

            self.repository.create_question(q)
            logger.info(f"Stored agent-identified question: {q.question_id}")
            return q.question_id

        except Exception as e:
            logger.error(f"Failed to store agent question: {e}")
            return None

    def export_for_followup(
        self,
        run_id: str,
        output_path: str,
        status_filter: List[str] = None
    ) -> str:
        """
        Export questions to Excel for client follow-up.

        Args:
            run_id: Analysis run ID
            output_path: Path to save Excel file
            status_filter: Only include questions with these statuses

        Returns:
            Path to created file
        """
        if not self.repository:
            raise ValueError("Repository required for export")

        # Get all questions for run
        # This would need repository method to get all questions
        try:
            conn = self.repository.db.connect()
            rows = conn.execute(
                "SELECT * FROM questions WHERE run_id = ? ORDER BY priority, created_at",
                (run_id,)
            ).fetchall()

            questions = [dict(row) for row in rows]

            if status_filter:
                questions = [q for q in questions if q.get('status') in status_filter]

            return export_questions_to_excel(questions, output_path)

        except Exception as e:
            logger.error(f"Failed to export questions: {e}")
            raise

    def import_answers(self, file_path: str, run_id: str) -> Dict[str, Any]:
        """
        Import answers from Excel and update database.

        Args:
            file_path: Path to Excel file with answers
            run_id: Analysis run ID

        Returns:
            Import results
        """
        results = import_answers_from_excel(file_path, self.repository)

        if self.repository and results['imported_answers']:
            # Update questions in database
            for answer in results['imported_answers']:
                try:
                    # Find question by text match
                    conn = self.repository.db.connect()
                    row = conn.execute(
                        "SELECT question_id FROM questions WHERE run_id = ? AND question_text = ?",
                        (run_id, answer['question_text'])
                    ).fetchone()

                    if row:
                        self.repository.answer_question(
                            row['question_id'],
                            answer['answer_text'],
                            answer['answer_source']
                        )
                        logger.info(f"Updated answer for question {row['question_id']}")

                except Exception as e:
                    results['errors'].append(f"Failed to update: {e}")

        return results

    def get_summary(self, run_id: str) -> Dict[str, Any]:
        """
        Get question tracking summary for a run.

        Args:
            run_id: Analysis run ID

        Returns:
            Summary dict with counts and lists
        """
        if not self.repository:
            return {'error': 'No repository available'}

        return self.repository.get_question_status_summary(run_id)


# =============================================================================
# CLI TESTING
# =============================================================================

if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)

    # Test document parsing
    test_content = """
# IT Due Diligence Document

## Questions - Answered
| # | Question | Answer | Source | Date |
|---|----------|--------|--------|------|
| 1 | What is the DR RTO? | 4 hours for critical systems | CTO Interview | 2026-01-05 |
| 2 | How many SAP users? | 450 named users | IT Director Email | 2026-01-03 |
| 3 | What is the backup frequency? | Daily incremental, weekly full | IT Documentation | 2026-01-02 |

## Questions - Unanswered
| # | Question | Context | Priority |
|---|----------|---------|----------|
| 1 | What is the cyber insurance coverage limit? | Needed for risk assessment | High |
| 2 | Who owns the data center lease? | Determine transfer requirements | Medium |
| 3 | What's the SAP licensing model? | Impact on deal structure | Critical |

## Other Content
This is other document content that should not be parsed as questions.
"""

    parser = DocumentQuestionParser()
    questions = parser.parse_document(test_content)

    print(f"\nParsed {len(questions)} questions:\n")
    for q in questions:
        status = "ANSWERED" if q.is_answered else "OPEN"
        print(f"[{status}] {q.question_text}")
        if q.is_answered:
            print(f"  Answer: {q.answer_text}")
        print(f"  Priority: {q.priority}")
        print()
